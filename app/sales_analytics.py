from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import Iterable

import pandas as pd


COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "date": (
        "дата",
        "дата продажи",
        "дата отчета",
        "дата отчёта",
        "период",
        "документ дата",
        "дата документа",
        "sale date",
        "date",
    ),
    "product": (
        "номенклатура",
        "номенклатура товара",
        "товар",
        "наименование",
        "наименование товара",
        "название товара",
        "продукт",
        "product",
        "item",
        "sku",
    ),
    "category": (
        "категория",
        "группа",
        "товарная группа",
        "category",
        "group",
    ),
    "manager": (
        "менеджер",
        "ответственный",
        "продавец",
        "manager",
        "salesperson",
    ),
    "revenue": (
        "выручка",
        "доход",
        "сумма продажи",
        "сумма продаж",
        "продажи",
        "сумма",
        "revenue",
        "sales amount",
        "amount",
    ),
    "cost": (
        "себестоимость",
        "сумма закупки",
        "закупочная стоимость",
        "cost",
        "cogs",
    ),
    "margin": (
        "маржа",
        "прибыль",
        "валовая прибыль",
        "margin",
        "gross profit",
        "profit",
    ),
    "quantity": (
        "количество",
        "кол во",
        "кол-во",
        "qty",
        "quantity",
        "шт",
    ),
    "unit_price": (
        "цена",
        "цена продажи",
        "price",
        "unit price",
    ),
    "unit_cost": (
        "себестоимость за единицу",
        "закупочная цена",
        "unit cost",
        "cost price",
    ),
}

DISPLAY_NAMES: dict[str, str] = {
    "date": "Дата",
    "product": "Товар",
    "category": "Категория",
    "manager": "Менеджер",
    "revenue": "Выручка",
    "cost": "Себестоимость",
    "margin": "Маржа",
    "quantity": "Количество",
    "unit_price": "Цена за единицу",
    "unit_cost": "Себестоимость за единицу",
}

RUSSIAN_MONTHS = {
    "январь": 1,
    "января": 1,
    "февраль": 2,
    "февраля": 2,
    "март": 3,
    "марта": 3,
    "апрель": 4,
    "апреля": 4,
    "май": 5,
    "мая": 5,
    "июнь": 6,
    "июня": 6,
    "июль": 7,
    "июля": 7,
    "август": 8,
    "августа": 8,
    "сентябрь": 9,
    "сентября": 9,
    "октябрь": 10,
    "октября": 10,
    "ноябрь": 11,
    "ноября": 11,
    "декабрь": 12,
    "декабря": 12,
}


@dataclass
class PreparedSalesData:
    data: pd.DataFrame
    warnings: list[str]


def normalize_column_name(name: str) -> str:
    normalized = re.sub(r"[_\-/]+", " ", str(name).strip().casefold())
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def guess_column_mapping(columns: Iterable[str]) -> dict[str, str | None]:
    normalized_columns = {column: normalize_column_name(column) for column in columns}
    guesses: dict[str, str | None] = {}

    for field, aliases in COLUMN_ALIASES.items():
        best_column: str | None = None
        best_score = -1
        normalized_aliases = [normalize_column_name(alias) for alias in aliases]

        for column, normalized_column in normalized_columns.items():
            score = 0
            for alias in normalized_aliases:
                alias_tokens = alias.split()

                if normalized_column == alias:
                    score = max(score, 100)
                elif alias in normalized_column:
                    score = max(score, 80)
                elif all(token in normalized_column for token in alias_tokens):
                    score = max(score, 60)

            if score > best_score:
                best_score = score
                best_column = column

        guesses[field] = best_column if best_score >= 60 else None

    return guesses


def list_excel_sheets(file_bytes: bytes) -> list[str]:
    workbook = pd.ExcelFile(BytesIO(file_bytes))
    return workbook.sheet_names


def _coerce_single_numeric(value: object) -> float | None:
    normalized = _normalize_number_string(value)
    if normalized is None:
        return None

    numeric = pd.to_numeric([normalized], errors="coerce")[0]
    if pd.isna(numeric):
        return None

    return float(numeric)


def _find_header_index(header_values: Iterable[object], aliases: Iterable[str]) -> int | None:
    normalized_aliases = [normalize_column_name(alias) for alias in aliases]
    best_index: int | None = None
    best_score = -1

    for index, value in enumerate(header_values):
        normalized_value = normalize_column_name(value) if value is not None else ""
        if not normalized_value:
            continue

        score = 0
        for alias in normalized_aliases:
            alias_tokens = alias.split()

            if normalized_value == alias:
                score = max(score, 100)
            elif alias in normalized_value:
                score = max(score, 80)
            elif alias_tokens and all(token in normalized_value for token in alias_tokens):
                score = max(score, 60)

        if score > best_score:
            best_score = score
            best_index = index

    return best_index if best_score >= 60 else None


def _extract_report_date_from_rows(rows: list[tuple[object, ...]]) -> pd.Timestamp | None:
    preview_chunks: list[str] = []

    for row in rows[:12]:
        for value in row[:4]:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                preview_chunks.append(text)

    preview_text = " ".join(preview_chunks)
    for match in re.findall(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b", preview_text):
        parsed = pd.to_datetime(match, errors="coerce", dayfirst=True)
        if pd.notna(parsed):
            return parsed.normalize()

    textual_match = re.search(
        r"\b(\d{1,2})\s+([а-яё]+)\s+(\d{4})(?:\s*г\.?)?\b",
        preview_text.casefold(),
    )
    if textual_match:
        day_text, month_text, year_text = textual_match.groups()
        month_number = RUSSIAN_MONTHS.get(month_text)
        if month_number is not None:
            parsed = pd.to_datetime(
                f"{int(year_text):04d}-{month_number:02d}-{int(day_text):02d}",
                errors="coerce",
            )
            if pd.notna(parsed):
                return parsed.normalize()

    return None


def _resolve_sheet_name(sheet_names: list[str], sheet_name: str | int | None) -> str:
    if sheet_name is None:
        return sheet_names[0]

    if isinstance(sheet_name, int):
        return sheet_names[sheet_name]

    return sheet_name


def _parse_1c_grouped_sales_report(file_bytes: bytes, sheet_name: str | int | None = 0) -> pd.DataFrame | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return None

    if not workbook.sheetnames:
        return None

    try:
        resolved_sheet_name = _resolve_sheet_name(workbook.sheetnames, sheet_name)
    except (IndexError, KeyError, TypeError):
        return None

    if resolved_sheet_name not in workbook.sheetnames:
        return None

    worksheet = workbook[resolved_sheet_name]
    rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    if not rows:
        return None

    header_row_number: int | None = None
    header_values: tuple[object, ...] | None = None

    for row_number, row_values in enumerate(rows, start=1):
        normalized_row = [normalize_column_name(value) if value is not None else "" for value in row_values]
        if not any(normalized_row):
            continue

        has_product = any(value == "номенклатура" for value in normalized_row)
        has_quantity = any("колич" in value for value in normalized_row if value)
        has_revenue = any(("доход" in value) or ("выруч" in value) for value in normalized_row if value)
        has_cost = any("себестоим" in value for value in normalized_row if value)

        if has_product and has_quantity and has_revenue and has_cost:
            header_row_number = row_number
            header_values = row_values
            break

    if header_row_number is None or header_values is None:
        return None

    product_index = _find_header_index(header_values, ("номенклатура", "товар", "наименование"))
    quantity_index = _find_header_index(header_values, ("количество", "кол-во", "qty"))
    revenue_index = _find_header_index(header_values, ("доход", "выручка"))
    cost_index = _find_header_index(header_values, ("себестоимость",))

    if product_index is None or quantity_index is None or revenue_index is None or cost_index is None:
        return None

    margin_index = _find_header_index(header_values, ("прибыль", "маржа"))
    discount_index = _find_header_index(header_values, ("скидка",))
    vat_index = _find_header_index(header_values, ("сумма ндс", "ндс"))
    sales_tax_index = _find_header_index(header_values, ("сумма нсп", "нсп"))
    total_index = _find_header_index(header_values, ("всего",))

    report_date = _extract_report_date_from_rows(rows[:header_row_number])
    if report_date is None:
        return None

    records: list[dict[str, object]] = []
    group_stack: dict[int, str] = {}
    saw_hierarchy = False
    skip_group_names = {"товары", "итого", "всего"}

    for row_number in range(header_row_number + 1, worksheet.max_row + 1):
        name_cell = worksheet.cell(row=row_number, column=product_index + 1)
        raw_name = name_cell.value
        name = str(raw_name).strip() if raw_name is not None else ""

        if not name:
            continue

        quantity = _coerce_single_numeric(worksheet.cell(row=row_number, column=quantity_index + 1).value)
        indent = name_cell.alignment.indent or 0
        outline_level = worksheet.row_dimensions[row_number].outlineLevel or 0
        level = max(int(indent // 2), int(outline_level), 0)
        normalized_name = normalize_column_name(name)

        if quantity is None:
            if level > 0:
                saw_hierarchy = True

            if normalized_name in skip_group_names:
                if level <= 0:
                    group_stack.clear()
                continue

            group_stack = {depth: label for depth, label in group_stack.items() if depth < level}
            group_stack[level] = name
            continue

        category_path_parts = [label for depth, label in sorted(group_stack.items()) if depth < max(level, 1)]
        category = category_path_parts[0] if category_path_parts else "Без категории"
        group_name = category_path_parts[-1] if category_path_parts else "Без категории"
        category_path = " > ".join(category_path_parts) if category_path_parts else "Без категории"

        records.append(
            {
                "Дата": report_date,
                "Номенклатура": name,
                "Категория": category,
                "Группа": group_name,
                "Путь категории": category_path,
                "Количество": quantity,
                "Доход": _coerce_single_numeric(worksheet.cell(row=row_number, column=revenue_index + 1).value),
                "Себестоимость": _coerce_single_numeric(worksheet.cell(row=row_number, column=cost_index + 1).value),
                "Прибыль": _coerce_single_numeric(worksheet.cell(row=row_number, column=margin_index + 1).value)
                if margin_index is not None
                else None,
                "Скидка": _coerce_single_numeric(worksheet.cell(row=row_number, column=discount_index + 1).value)
                if discount_index is not None
                else None,
                "Сумма НДС": _coerce_single_numeric(worksheet.cell(row=row_number, column=vat_index + 1).value)
                if vat_index is not None
                else None,
                "Сумма НСП": _coerce_single_numeric(worksheet.cell(row=row_number, column=sales_tax_index + 1).value)
                if sales_tax_index is not None
                else None,
                "Всего": _coerce_single_numeric(worksheet.cell(row=row_number, column=total_index + 1).value)
                if total_index is not None
                else None,
            }
        )

    if not records or not saw_hierarchy:
        return None

    return pd.DataFrame.from_records(records)


def load_input_file(
    file_bytes: bytes,
    filename: str,
    *,
    csv_separator: str = ";",
    csv_encoding: str = "utf-8",
    sheet_name: str | int | None = 0,
) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        return pd.read_csv(BytesIO(file_bytes), sep=csv_separator, encoding=csv_encoding)

    grouped_report = _parse_1c_grouped_sales_report(file_bytes, sheet_name=sheet_name)
    if grouped_report is not None:
        return grouped_report

    return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)


def _normalize_number_string(value: object) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, (int, float)):
        return str(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("\xa0", "").replace(" ", "")
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None

    comma_count = text.count(",")
    dot_count = text.count(".")

    if comma_count and dot_count:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif comma_count == 1 and dot_count == 0:
        text = text.replace(",", ".")
    elif comma_count > 1 and dot_count == 0:
        text = text.replace(",", "")
    elif dot_count > 1 and comma_count == 0:
        text = text.replace(".", "")

    return text


def coerce_numeric(series: pd.Series | None, fill_value: float | None = None) -> pd.Series | None:
    if series is None:
        return None

    normalized = series.map(_normalize_number_string)
    numeric = pd.to_numeric(normalized, errors="coerce")

    if fill_value is not None:
        numeric = numeric.fillna(fill_value)

    return numeric


def _series_from_mapping(frame: pd.DataFrame, mapping: dict[str, str | None], key: str) -> pd.Series | None:
    column_name = mapping.get(key)
    if not column_name:
        return None
    return frame[column_name]


def parse_dates(series: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(series):
        return pd.to_datetime(series, errors="coerce")

    as_text = series.astype(str).str.strip()
    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    iso_mask = as_text.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", na=False)

    if iso_mask.any():
        parsed.loc[iso_mask] = pd.to_datetime(as_text.loc[iso_mask], errors="coerce")

    if (~iso_mask).any():
        parsed.loc[~iso_mask] = pd.to_datetime(as_text.loc[~iso_mask], errors="coerce", dayfirst=True)

    return parsed


def prepare_sales_data(frame: pd.DataFrame, mapping: dict[str, str | None]) -> PreparedSalesData:
    resolved_mapping = dict(mapping)
    guessed_mapping = guess_column_mapping(frame.columns.astype(str).tolist())

    for field, guessed_column in guessed_mapping.items():
        if not resolved_mapping.get(field) and guessed_column in frame.columns:
            resolved_mapping[field] = guessed_column

    if not resolved_mapping.get("date") or not resolved_mapping.get("product"):
        raise ValueError("Нужно указать хотя бы колонки с датой и названием товара.")

    prepared = frame.copy()
    warnings: list[str] = []

    raw_date = _series_from_mapping(prepared, resolved_mapping, "date")
    raw_product = _series_from_mapping(prepared, resolved_mapping, "product")

    prepared["date"] = parse_dates(raw_date)
    prepared["product"] = raw_product.astype(str).str.strip()

    if category := resolved_mapping.get("category"):
        prepared["category"] = prepared[category].astype(str).str.strip()
    else:
        prepared["category"] = "Без категории"

    if manager := resolved_mapping.get("manager"):
        prepared["manager"] = prepared[manager].astype(str).str.strip()
    else:
        prepared["manager"] = "Не указан"

    quantity = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "quantity"), fill_value=0)
    if quantity is None:
        quantity = pd.Series(0, index=prepared.index, dtype="float64")

    unit_price = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "unit_price"))
    unit_cost = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "unit_cost"))
    revenue = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "revenue"))
    cost = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "cost"))
    margin = coerce_numeric(_series_from_mapping(prepared, resolved_mapping, "margin"))

    if revenue is None and unit_price is not None:
        revenue = unit_price * quantity
        warnings.append("Выручка была рассчитана как цена за единицу × количество.")

    if cost is None and unit_cost is not None:
        cost = unit_cost * quantity
        warnings.append("Себестоимость была рассчитана как себестоимость за единицу × количество.")

    if revenue is None and cost is not None and margin is not None:
        revenue = cost + margin
        warnings.append("Выручка была рассчитана как себестоимость + маржа.")

    if cost is None and revenue is not None and margin is not None:
        cost = revenue - margin
        warnings.append("Себестоимость была рассчитана как выручка - маржа.")

    if margin is None and revenue is not None and cost is not None:
        margin = revenue - cost

    if revenue is None:
        raise ValueError(
            "Не удалось определить выручку. Укажите колонку `Выручка` или связку `Цена за единицу + Количество`."
        )

    prepared["quantity"] = quantity.fillna(0)
    prepared["revenue"] = revenue

    if cost is not None:
        prepared["cost"] = cost
    else:
        prepared["cost"] = pd.Series(pd.NA, index=prepared.index, dtype="Float64")

    if margin is not None:
        prepared["margin"] = margin
    else:
        prepared["margin"] = pd.Series(pd.NA, index=prepared.index, dtype="Float64")

    prepared["margin_pct"] = (prepared["margin"] / prepared["revenue"]).where(prepared["revenue"] != 0) * 100
    prepared["month"] = prepared["date"].dt.to_period("M").dt.to_timestamp()
    prepared["month_label"] = prepared["month"].dt.strftime("%Y-%m")

    initial_rows = len(prepared)
    prepared = prepared.dropna(subset=["date", "product", "revenue"])
    prepared = prepared[prepared["product"] != ""]

    dropped_rows = initial_rows - len(prepared)
    if dropped_rows:
        warnings.append(f"Из анализа исключено строк: {dropped_rows}. Причина: пустая дата, товар или выручка.")

    if prepared.empty:
        raise ValueError("После очистки не осталось строк для анализа. Проверьте формат выгрузки.")

    return PreparedSalesData(data=prepared, warnings=warnings)


def build_overview_metrics(frame: pd.DataFrame) -> dict[str, float]:
    total_revenue = frame["revenue"].sum()
    total_cost = frame["cost"].sum(min_count=1)
    total_margin = frame["margin"].sum(min_count=1)

    return {
        "total_revenue": float(total_revenue),
        "total_cost": float(total_cost) if pd.notna(total_cost) else float("nan"),
        "total_margin": float(total_margin) if pd.notna(total_margin) else float("nan"),
        "margin_pct": float((total_margin / total_revenue) * 100) if total_revenue and pd.notna(total_margin) else float("nan"),
        "total_quantity": float(frame["quantity"].sum()),
        "line_count": float(len(frame)),
        "product_count": float(frame["product"].nunique()),
    }


def build_product_summary(frame: pd.DataFrame, group_column: str = "product") -> pd.DataFrame:
    summary = (
        frame.groupby(group_column, dropna=False)
        .agg(
            revenue=("revenue", "sum"),
            cost=("cost", "sum"),
            margin=("margin", "sum"),
            quantity=("quantity", "sum"),
            sales_lines=("product", "size"),
        )
        .reset_index()
        .rename(columns={group_column: "group_name"})
    )

    summary["margin_pct"] = (summary["margin"] / summary["revenue"]).where(summary["revenue"] != 0) * 100
    summary = summary.sort_values("revenue", ascending=False, ignore_index=True)
    return summary


def _classify_abc(cumulative_pct: float) -> str:
    if cumulative_pct <= 80:
        return "A"
    if cumulative_pct <= 95:
        return "B"
    return "C"


def build_abc_analysis(product_summary: pd.DataFrame, metric: str = "revenue") -> pd.DataFrame:
    abc = product_summary.copy()
    abc["abc_basis"] = abc[metric].clip(lower=0).fillna(0)
    abc = abc.sort_values("abc_basis", ascending=False, ignore_index=True)

    total = abc["abc_basis"].sum()
    if total > 0:
        abc["share_pct"] = abc["abc_basis"] / total * 100
        abc["cum_share_pct"] = abc["share_pct"].cumsum()
    else:
        abc["share_pct"] = 0.0
        abc["cum_share_pct"] = 0.0

    abc["abc_class"] = abc["cum_share_pct"].map(_classify_abc)
    return abc


def build_monthly_summary(frame: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        frame.groupby(["month", "month_label"], as_index=False)
        .agg(
            revenue=("revenue", "sum"),
            cost=("cost", "sum"),
            margin=("margin", "sum"),
            quantity=("quantity", "sum"),
            product_count=("product", "nunique"),
        )
        .sort_values("month")
        .reset_index(drop=True)
    )

    monthly["revenue_change_pct"] = monthly["revenue"].pct_change() * 100
    monthly["margin_change_pct"] = monthly["margin"].pct_change() * 100
    return monthly


def extract_return_rows(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame.iloc[0:0].copy()

    return_mask = (frame["revenue"].fillna(0) < 0) | (frame["quantity"].fillna(0) < 0)
    return frame.loc[return_mask].copy()


def build_returns_overview(frame: pd.DataFrame) -> dict[str, float]:
    returns = extract_return_rows(frame)
    positive_revenue = frame.loc[frame["revenue"].fillna(0) > 0, "revenue"].sum()
    return_revenue = returns["revenue"].abs().sum()
    return_margin = returns["margin"].abs().sum(min_count=1)
    return_quantity = returns["quantity"].abs().sum()

    return {
        "return_lines": float(len(returns)),
        "return_revenue": float(return_revenue),
        "return_margin": float(return_margin) if pd.notna(return_margin) else float("nan"),
        "return_quantity": float(return_quantity),
        "return_product_count": float(returns["product"].nunique()) if not returns.empty else 0.0,
        "return_month_count": float(returns["month_label"].nunique()) if not returns.empty else 0.0,
        "return_share_pct": float(return_revenue / positive_revenue * 100) if positive_revenue else float("nan"),
    }


def build_return_groups(frame: pd.DataFrame, group_column: str = "product") -> pd.DataFrame:
    returns = extract_return_rows(frame)
    if returns.empty or group_column not in returns.columns:
        return pd.DataFrame(
            columns=[
                "group_name",
                "return_revenue",
                "return_margin",
                "return_quantity",
                "return_lines",
                "last_return_date",
                "return_share_pct",
            ]
        )

    grouped = returns.copy()
    grouped["return_revenue"] = grouped["revenue"].abs()
    grouped["return_margin"] = grouped["margin"].abs()
    grouped["return_quantity"] = grouped["quantity"].abs()

    summary = (
        grouped.groupby(group_column, dropna=False)
        .agg(
            return_revenue=("return_revenue", "sum"),
            return_margin=("return_margin", "sum"),
            return_quantity=("return_quantity", "sum"),
            return_lines=("product", "size"),
            last_return_date=("date", "max"),
        )
        .reset_index()
        .rename(columns={group_column: "group_name"})
    )

    total_return_revenue = summary["return_revenue"].sum()
    summary["return_share_pct"] = (
        summary["return_revenue"] / total_return_revenue * 100 if total_return_revenue else 0.0
    )
    return summary.sort_values(["return_revenue", "return_lines"], ascending=[False, False], ignore_index=True)


def build_return_monthly_summary(frame: pd.DataFrame) -> pd.DataFrame:
    returns = extract_return_rows(frame)
    if returns.empty:
        return pd.DataFrame(
            columns=[
                "month",
                "month_label",
                "return_revenue",
                "return_margin",
                "return_quantity",
                "return_lines",
                "return_share_pct",
            ]
        )

    monthly = returns.copy()
    monthly["return_revenue"] = monthly["revenue"].abs()
    monthly["return_margin"] = monthly["margin"].abs()
    monthly["return_quantity"] = monthly["quantity"].abs()

    summary = (
        monthly.groupby(["month", "month_label"], as_index=False)
        .agg(
            return_revenue=("return_revenue", "sum"),
            return_margin=("return_margin", "sum"),
            return_quantity=("return_quantity", "sum"),
            return_lines=("product", "size"),
        )
        .sort_values("month")
        .reset_index(drop=True)
    )

    positive_monthly_revenue = (
        frame.loc[frame["revenue"].fillna(0) > 0]
        .groupby(["month", "month_label"], as_index=False)
        .agg(gross_revenue=("revenue", "sum"))
    )
    summary = summary.merge(positive_monthly_revenue, on=["month", "month_label"], how="left")
    summary["return_share_pct"] = summary["return_revenue"] / summary["gross_revenue"].replace(0, pd.NA) * 100
    summary = summary.drop(columns=["gross_revenue"])
    return summary


def build_plan_fact_summary(monthly_summary: pd.DataFrame, plan_summary: pd.DataFrame) -> pd.DataFrame:
    if monthly_summary.empty and plan_summary.empty:
        return pd.DataFrame()

    fact = monthly_summary.copy()
    plan = plan_summary.copy()

    if not fact.empty:
        fact["month"] = pd.to_datetime(fact["month"], errors="coerce").dt.normalize()
    if not plan.empty:
        plan["month"] = pd.to_datetime(plan["month"], errors="coerce").dt.normalize()

    merged = fact.merge(
        plan,
        on=["month", "month_label"],
        how="outer",
        suffixes=("_fact", "_plan_source"),
    ).sort_values("month").reset_index(drop=True)

    for metric in ("revenue", "margin", "quantity"):
        fact_column = metric
        plan_column = f"{metric}_plan"
        if fact_column not in merged:
            merged[fact_column] = 0.0
        if plan_column not in merged:
            merged[plan_column] = pd.NA

        merged[f"{metric}_gap"] = merged[fact_column] - merged[plan_column]
        merged[f"{metric}_execution_pct"] = (
            merged[fact_column] / merged[plan_column].replace(0, pd.NA) * 100
        )

    merged["has_plan"] = merged[["revenue_plan", "margin_plan", "quantity_plan"]].notna().any(axis=1)
    return merged


def build_plan_fact_by_salon(frame: pd.DataFrame, plan_frame: pd.DataFrame, month_label: str) -> pd.DataFrame:
    if frame.empty or "salon" not in frame.columns:
        return pd.DataFrame()

    month_data = frame[frame["month_label"] == month_label].copy()
    if month_data.empty:
        return pd.DataFrame()

    fact = (
        month_data.groupby("salon", as_index=False)
        .agg(
            revenue=("revenue", "sum"),
            margin=("margin", "sum"),
            quantity=("quantity", "sum"),
        )
        .rename(columns={"salon": "scope_name"})
    )

    plan = plan_frame.copy()
    if plan.empty:
        plan = pd.DataFrame(columns=["scope_name", "revenue_plan", "margin_plan", "quantity_plan"])
    else:
        plan = plan.rename(columns={"salon": "scope_name"})

    merged = fact.merge(plan, on="scope_name", how="left")

    for metric in ("revenue", "margin", "quantity"):
        plan_column = f"{metric}_plan"
        merged[f"{metric}_gap"] = merged[metric] - merged[plan_column]
        merged[f"{metric}_execution_pct"] = (
            merged[metric] / merged[plan_column].replace(0, pd.NA) * 100
        )

    return merged.sort_values(["revenue_execution_pct", "revenue"], ascending=[False, False], na_position="last").reset_index(drop=True)


def build_month_comparison(
    frame: pd.DataFrame,
    left_month: str,
    right_month: str,
    *,
    group_column: str = "product",
) -> pd.DataFrame:
    comparison = (
        frame[frame["month_label"].isin([left_month, right_month])]
        .groupby(["month_label", group_column], as_index=False)
        .agg(
            revenue=("revenue", "sum"),
            margin=("margin", "sum"),
            quantity=("quantity", "sum"),
        )
    )

    if comparison.empty:
        return pd.DataFrame()

    pivot = comparison.pivot(index=group_column, columns="month_label", values=["revenue", "margin", "quantity"]).fillna(0)
    pivot.columns = [f"{metric}_{month}" for metric, month in pivot.columns]
    result = pivot.reset_index().rename(columns={group_column: "group_name"})

    for metric in ("revenue", "margin", "quantity"):
        left_column = f"{metric}_{left_month}"
        right_column = f"{metric}_{right_month}"

        if left_column not in result:
            result[left_column] = 0.0
        if right_column not in result:
            result[right_column] = 0.0

        result[f"{metric}_delta"] = result[right_column] - result[left_column]
        result[f"{metric}_delta_pct"] = (
            result[f"{metric}_delta"] / result[left_column].replace(0, pd.NA) * 100
        )

    result = result.sort_values("revenue_delta", ascending=False, ignore_index=True)
    return result


def to_csv_bytes(frame: pd.DataFrame) -> bytes:
    return frame.to_csv(index=False).encode("utf-8-sig")


def build_yoy_comparison(frame: pd.DataFrame) -> pd.DataFrame:
    """Monthly summary with year annotation for YoY charts."""
    if frame.empty:
        return pd.DataFrame()
    monthly = (
        frame.groupby(["month", "month_label"], as_index=False)
        .agg(revenue=("revenue", "sum"), margin=("margin", "sum"), quantity=("quantity", "sum"))
    )
    monthly["year"] = pd.to_datetime(monthly["month"]).dt.year.astype(str)
    monthly["month_num"] = pd.to_datetime(monthly["month"]).dt.month
    monthly["month_abbr"] = pd.to_datetime(monthly["month"]).dt.strftime("%b")
    return monthly.sort_values(["year", "month_num"]).reset_index(drop=True)


def build_rfm_summary(frame: pd.DataFrame, group_column: str = "manager") -> pd.DataFrame:
    """RFM scoring grouped by specified column."""
    if frame.empty or group_column not in frame.columns:
        return pd.DataFrame()

    reference_date = frame["date"].max()

    rfm = (
        frame.groupby(group_column)
        .agg(
            recency=("date", lambda x: int((reference_date - x.max()).days)),
            frequency=("month_label", "nunique"),
            monetary=("revenue", "sum"),
        )
        .reset_index()
        .rename(columns={group_column: "group_name"})
    )

    def _safe_score(series: pd.Series, ascending: bool) -> pd.Series:
        n = series.nunique()
        if n < 2:
            return pd.Series(3, index=series.index, dtype=int)
        labels = list(range(5, 0, -1)) if ascending else list(range(1, 6))
        n_bins = min(5, n)
        try:
            return pd.qcut(series, q=n_bins, labels=labels[:n_bins], duplicates="drop").astype(int)
        except Exception:
            return pd.Series(3, index=series.index, dtype=int)

    rfm["r_score"] = _safe_score(rfm["recency"], ascending=True)
    rfm["f_score"] = _safe_score(rfm["frequency"], ascending=False)
    rfm["m_score"] = _safe_score(rfm["monetary"], ascending=False)
    rfm["rfm_score"] = rfm["r_score"] + rfm["f_score"] + rfm["m_score"]
    rfm["segment"] = rfm["rfm_score"].map(
        lambda s: "Лидер" if s >= 12 else ("Активный" if s >= 9 else ("Стабильный" if s >= 6 else "Слабый"))
    )
    return rfm.sort_values("rfm_score", ascending=False, ignore_index=True)


def detect_anomalies(
    monthly_summary: pd.DataFrame,
    column: str = "revenue",
    threshold: float = 2.0,
) -> list[tuple[str, str]]:
    """Return (month_label, description) for months deviating > threshold std devs."""
    if len(monthly_summary) < 4:
        return []
    values = monthly_summary[column].dropna()
    if len(values) < 4:
        return []
    mean = float(values.mean())
    std = float(values.std())
    if std == 0:
        return []
    result: list[tuple[str, str]] = []
    for _, row in monthly_summary.iterrows():
        val = row[column]
        if pd.isna(val):
            continue
        z = (float(val) - mean) / std
        if abs(z) > threshold:
            direction = "рост" if z > 0 else "падение"
            result.append((
                str(row["month_label"]),
                f"Аномальный {direction} — отклонение {abs(z):.1f}σ от среднего.",
            ))
    return result


def build_forecast(monthly_summary: pd.DataFrame, periods: int = 3) -> pd.DataFrame:
    """Linear-trend forecast for next N months."""
    if len(monthly_summary) < 3:
        return pd.DataFrame()
    import numpy as np

    x = np.arange(len(monthly_summary), dtype=float)
    rev = monthly_summary["revenue"].fillna(0).values.astype(float)
    s_r, i_r = np.polyfit(x, rev, 1)

    mar = monthly_summary["margin"].values.astype(float)
    valid = ~np.isnan(mar)
    if valid.sum() >= 3:
        s_m, i_m = np.polyfit(x[valid], mar[valid], 1)
    else:
        s_m, i_m = 0.0, 0.0

    last_month = pd.to_datetime(monthly_summary["month"].iloc[-1])
    future_dates = pd.date_range(last_month, periods=periods + 1, freq="MS")[1:]
    fx = np.arange(len(monthly_summary), len(monthly_summary) + periods, dtype=float)

    return pd.DataFrame({
        "month": future_dates,
        "month_label": [d.strftime("%Y-%m") + " ▸" for d in future_dates],
        "revenue": np.maximum(s_r * fx + i_r, 0.0),
        "margin": s_m * fx + i_m,
        "is_forecast": True,
    })


def build_heatmap_data(frame: pd.DataFrame) -> pd.DataFrame:
    """Aggregate revenue by ISO week × day-of-week."""
    if frame.empty:
        return pd.DataFrame()
    h = frame.copy()
    h["dow"] = h["date"].dt.dayofweek
    h["iso_week"] = h["date"].dt.isocalendar().week.astype(int)
    h["year"] = h["date"].dt.year
    h["week_label"] = h["year"].astype(str) + "-W" + h["iso_week"].astype(str).str.zfill(2)
    return (
        h.groupby(["week_label", "dow"], as_index=False)
        .agg(revenue=("revenue", "sum"), quantity=("quantity", "sum"))
        .sort_values(["week_label", "dow"])
        .reset_index(drop=True)
    )
