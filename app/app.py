from __future__ import annotations

import os
from datetime import date
from html import escape
from io import BytesIO
from pathlib import Path
import secrets
from textwrap import dedent

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components

from auth_store import (
    authenticate_user,
    authenticate_session,
    bootstrap_first_admin,
    create_auth_session,
    create_user,
    delete_user,
    delete_users_by_salon,
    find_user,
    has_admin_users,
    has_users,
    list_users,
    promote_first_manager_to_admin,
    reassign_salon_user,
    revoke_auth_session,
    set_user_password,
    update_user_role,
)
from data_quality import analyze_sales_quality, build_catalog_health
from plan_store import delete_monthly_plan, load_monthly_plans, normalize_plan_month, upsert_monthly_plan
from db import log_audit_event
from inventory_analytics import (
    guess_inventory_column_mapping,
    load_inventory_input_file,
    prepare_inventory_data,
)
from procurement_analytics import (
    build_procurement_forecast,
    build_procurement_overview,
    build_procurement_stock_risk_frames,
    build_procurement_supplier_summary,
)
from procurement_order_store import (
    PROCUREMENT_ORDER_STATUSES,
    PROCUREMENT_ORDER_STATUS_LABELS,
    ACTIVE_INBOUND_ORDER_STATUSES,
    apply_procurement_order_receipt,
    build_open_order_summary,
    build_procurement_order_overview,
    create_procurement_order,
    load_procurement_order_items,
    load_procurement_orders,
    update_procurement_order,
)
from procurement_store import (
    load_latest_inventory_snapshot,
    load_procurement_items,
    merge_procurement_upload,
    save_inventory_snapshot,
    upsert_procurement_items,
)
from sales_analytics import (
    DISPLAY_NAMES,
    SUPPLIER_KEYWORD_PATTERNS,
    build_abc_analysis,
    build_forecast,
    build_plan_fact_by_salon,
    build_plan_fact_summary,
    build_heatmap_data,
    build_month_comparison,
    build_monthly_summary,
    build_overview_metrics,
    build_product_summary,
    build_rfm_summary,
    build_return_groups,
    build_return_monthly_summary,
    build_returns_overview,
    build_yoy_comparison,
    detect_anomalies,
    extract_return_rows,
    guess_column_mapping,
    infer_supplier_from_text,
    list_excel_sheets,
    load_input_file,
    prepare_sales_data,
)
from salon_data_store import (
    count_uploads_for_salon,
    delete_salon,
    load_archive_data,
    load_manifest,
    load_salons,
    register_upload,
    save_salon,
)
from supplier_rules_store import (
    KEYWORD_RULE_COLUMNS,
    PRODUCT_ASSIGNMENT_COLUMNS,
    delete_supplier_product_assignments,
    load_supplier_keyword_rules,
    load_supplier_product_assignments,
    replace_supplier_keyword_rules,
    upsert_supplier_product_assignments,
)
from telegram_reports import build_supplier_order_file, send_telegram_report_pack, telegram_is_configured

# Design System Tokens
PRIMARY_COLOR = "#003461"
SECONDARY_COLOR = "#006c49"
BG_COLOR = "#F4F7F9"
BORDER_COLOR = "#E2E8F0"
SURFACE_COLOR = "#FFFFFF"
TEXT_PRIMARY = "#0f172a"
TEXT_SECONDARY = "#475569"
TEXT_MUTED = "#64748b"
CHART_GRID_COLOR = "#EAF0F4"
CHART_ZERO_COLOR = "#CBD5E1"
CHART_ACCENT_COLOR = "#D89A2B"
CHART_DANGER_COLOR = "#D94D3D"
CHART_INFO_COLOR = "#2673B8"
CHART_SOFT_BLUE = "#DDEBFA"
CHART_SOFT_TEAL = "#DDEFE8"
CHART_SOFT_AMBER = "#F6E5C4"
CHART_COLORS = [
    PRIMARY_COLOR,
    CHART_ACCENT_COLOR,
    SECONDARY_COLOR,
    CHART_INFO_COLOR,
    "#0E9AA7",
    CHART_DANGER_COLOR,
    "#7C5CC4",
    "#58667A",
]
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
APP_DIR = Path(__file__).resolve().parent
APP_FAVICON_PATH = APP_DIR / "assets" / "favicon.svg"

st.set_page_config(
    page_title="ArtDB — аналитика продаж",
    page_icon=APP_FAVICON_PATH if APP_FAVICON_PATH.exists() else "📊",
    layout="wide",
    initial_sidebar_state="auto",
)


DASHBOARD_CSS = f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700&display=swap');

    .stApp {{
        font-family: 'Inter', -apple-system, sans-serif;
        background-color: {BG_COLOR};
    }}

    h1, h2, h3, .dashboard-title, .panel-title {{
        font-family: 'Manrope', sans-serif !important;
        font-weight: 800 !important;
    }}

    .block-container {{
        padding-top: 1rem;
        padding-bottom: 1.6rem;
        max-width: 1600px;
    }}

    .sticky-rail-marker {{
        display: none !important;
        width: 0 !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }}

    .sticky-header-marker {{
        display: none !important;
        width: 0 !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }}

    @media (min-width: 1100px) {{
        [data-testid="stVerticalBlockBorderWrapper"].control-header-shell {{
            position: sticky;
            top: 0.65rem;
            z-index: 12;
        }}

        [data-testid="stVerticalBlockBorderWrapper"].control-header-shell > div {{
            background: rgba(255, 255, 255, 0.94);
            backdrop-filter: blur(12px);
            border-radius: 18px;
            box-shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
        }}

        [data-testid="column"].control-rail-column {{
            position: sticky;
            top: 0.9rem;
            align-self: start;
            z-index: 4;
        }}

        [data-testid="column"].control-rail-column > .control-rail-body {{
            max-height: calc(100vh - 1.2rem);
            overflow-y: auto;
            overflow-x: visible;
            padding-right: 0.25rem;
            scrollbar-width: thin;
            overscroll-behavior: contain;
        }}

        [data-testid="column"].control-rail-column > .control-rail-body::-webkit-scrollbar {{
            width: 8px;
        }}

        [data-testid="column"].control-rail-column > .control-rail-body::-webkit-scrollbar-thumb {{
            background: rgba(100, 116, 139, 0.35);
            border-radius: 999px;
        }}
    }}

    /* Bento Grid / Card Style */
    [data-testid="stVerticalBlockBorderWrapper"] {{
        background-color: {SURFACE_COLOR} !important;
        border: 1px solid {BORDER_COLOR} !important;
        border-radius: 14px !important;
        padding: 1.05rem !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04) !important;
    }}

    /* Hero Section */
    .dashboard-hero {{
        padding: 1.05rem 1.15rem;
        border-radius: 14px;
        border: 1px solid rgba(0, 108, 73, 0.16);
        background:
            linear-gradient(135deg, rgba(0, 108, 73, 0.08), rgba(255, 255, 255, 0) 34%),
            {SURFACE_COLOR};
        margin-bottom: 0.85rem;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }}

    .dashboard-hero-meta {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.7rem;
        flex-wrap: wrap;
    }}

    .dashboard-eyebrow {{
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: {SECONDARY_COLOR};
        margin-bottom: 0.5rem;
        font-weight: 700;
    }}

    .dashboard-title {{
        font-size: 1.62rem;
        color: {PRIMARY_COLOR};
        margin: 0;
        line-height: 1.12;
    }}

    .dashboard-subtitle {{
        margin: 0.55rem 0 0;
        color: {TEXT_SECONDARY};
        font-size: 0.94rem;
        line-height: 1.45;
        max-width: 68rem;
    }}

    .scope-strip {{
        display: flex;
        gap: 0.4rem;
        flex-wrap: wrap;
        align-items: flex-start;
        margin-top: 0.7rem;
    }}

    .scope-chip,
    .app-build-badge {{
        display: inline-flex;
        align-items: center;
        border-radius: 999px;
        padding: 0.28rem 0.62rem;
        border: 1px solid rgba(0, 52, 97, 0.10);
        background: rgba(244, 247, 249, 0.84);
        color: {TEXT_SECONDARY};
        font-size: 0.78rem;
        font-weight: 650;
        line-height: 1.25;
        max-width: 100%;
        overflow-wrap: anywhere;
    }}

    .app-build-badge {{
        color: {PRIMARY_COLOR};
        background: rgba(0, 52, 97, 0.06);
        white-space: nowrap;
    }}

    /* Metric Cards Grid */
    .metric-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(175px, 1fr));
        gap: 0.7rem;
        margin: 0.8rem 0 1rem;
    }}

    .metric-card {{
        padding: 0.9rem 0.95rem;
        border-radius: 12px;
        background: {SURFACE_COLOR};
        border: 1px solid {BORDER_COLOR};
        border-left: 4px solid rgba(0, 108, 73, 0.72);
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
        transition: transform 0.18s ease, border-color 0.18s ease, box-shadow 0.18s ease;
    }}

    .metric-card:hover {{
        transform: translateY(-1px);
        border-color: rgba(0, 52, 97, 0.22);
        box-shadow: 0 8px 20px rgba(15, 23, 42, 0.07);
    }}

    .metric-card.info {{
        border-left-color: {PRIMARY_COLOR};
    }}

    .metric-card.success {{
        border-left-color: {SECONDARY_COLOR};
    }}

    .metric-card.warning {{
        border-left-color: {CHART_ACCENT_COLOR};
        background: linear-gradient(180deg, rgba(255,251,235,0.72), rgba(255,255,255,0.98));
    }}

    .metric-card.danger {{
        border-left-color: #dc2626;
        background: linear-gradient(180deg, rgba(254,242,242,0.72), rgba(255,255,255,0.98));
    }}

    .metric-label {{
        color: {TEXT_MUTED};
        font-size: 0.72rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        margin-bottom: 0.42rem;
    }}

    .metric-value {{
        color: {PRIMARY_COLOR};
        font-size: 1.45rem;
        font-weight: 800;
        font-family: 'Manrope', sans-serif;
        font-variant-numeric: tabular-nums;
        line-height: 1.08;
    }}

    .metric-delta {{
        margin-top: 0.38rem;
        font-size: 0.78rem;
        font-weight: 700;
    }}

    .metric-delta.positive {{ color: {SECONDARY_COLOR}; }}
    .metric-delta.negative {{ color: #dc2626; }}

    /* Button Styling */
    .stButton > button,
    .stDownloadButton > button,
    div[data-testid="stFormSubmitButton"] > button {{
        border-radius: 8px;
        font-weight: 700;
        padding: 0.5rem 1rem;
        background-color: {PRIMARY_COLOR} !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 7px 18px rgba(0, 52, 97, 0.12);
    }}

    .stButton > button:hover,
    .stDownloadButton > button:hover,
    div[data-testid="stFormSubmitButton"] > button:hover {{
        opacity: 0.9;
        transform: translateY(-1px);
    }}

    @keyframes chartReveal {{
        0% {{
            opacity: 0;
            transform: translateY(18px) scale(0.985);
            filter: blur(6px);
        }}
        100% {{
            opacity: 1;
            transform: translateY(0) scale(1);
            filter: blur(0);
        }}
    }}

    @keyframes chartRevealStrong {{
        0% {{
            opacity: 0;
            transform: translateY(28px) scale(0.97);
            filter: blur(10px);
        }}
        60% {{
            opacity: 1;
            transform: translateY(-3px) scale(1.01);
            filter: blur(0);
        }}
        100% {{
            opacity: 1;
            transform: translateY(0) scale(1);
            filter: blur(0);
        }}
    }}

    [data-testid="stPlotlyChart"] {{
        animation: chartReveal 680ms cubic-bezier(0.22, 1, 0.36, 1);
        transform-origin: center top;
        will-change: opacity, transform, filter;
    }}

    [data-testid="stPlotlyChart"] > div {{
        border-radius: 12px;
        overflow: hidden;
    }}

    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"] {{
        border: 1px solid {BORDER_COLOR};
        border-radius: 12px;
        overflow: hidden;
        background: {SURFACE_COLOR};
    }}

    [data-testid="stDataFrame"] [role="gridcell"],
    [data-testid="stDataFrame"] [role="columnheader"],
    [data-testid="stDataEditor"] [role="gridcell"],
    [data-testid="stDataEditor"] [role="columnheader"] {{
        white-space: normal !important;
        word-break: break-word !important;
        line-height: 1.3 !important;
        font-size: 0.82rem !important;
    }}

    [data-testid="stDataFrame"] [role="columnheader"],
    [data-testid="stDataEditor"] [role="columnheader"] {{
        color: {TEXT_PRIMARY} !important;
        font-weight: 800 !important;
        background: #f8fafc !important;
    }}

    @media (prefers-reduced-motion: reduce) {{
        [data-testid="stPlotlyChart"] {{
            animation: none !important;
        }}
    }}

    /* Section Headers */
    .panel-title {{
        color: {PRIMARY_COLOR};
        font-size: 1.08rem;
        line-height: 1.2;
        margin-bottom: 0.28rem;
    }}

    .panel-caption {{
        color: {TEXT_SECONDARY};
        font-size: 0.84rem;
        margin-bottom: 0.78rem;
        line-height: 1.4;
    }}

    .section-intro,
    .section-marker,
    .nav-shell,
    .login-shell {{
        background: {SURFACE_COLOR};
        border: 1px solid {BORDER_COLOR};
        border-radius: 14px;
        padding: 0.85rem 0.95rem;
        margin: 0 0 0.78rem 0;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }}

    .section-intro-title,
    .section-marker-title,
    .nav-title,
    .insight-title,
    .auth-title {{
        font-family: 'Manrope', sans-serif;
        font-weight: 800;
        color: {PRIMARY_COLOR};
        line-height: 1.2;
    }}

    .section-intro-title,
    .section-marker-title {{
        font-size: 1.12rem;
        margin-bottom: 0.24rem;
    }}

    .section-intro-body,
    .section-marker-body,
    .panel-copy,
    .insight-body,
    .auth-copy {{
        color: {TEXT_SECONDARY};
        font-size: 0.86rem;
        line-height: 1.42;
    }}

    .section-marker-kicker,
    .nav-title,
    .auth-kicker {{
        color: {TEXT_MUTED};
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 700;
        margin-bottom: 0.45rem;
    }}

    .workspace-band,
    .journey-grid,
    .snapshot-strip,
    .admin-stat-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        gap: 0.65rem;
        margin: 0.7rem 0;
    }}

    .journey-grid {{
        grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    }}

    .workspace-band-item,
    .journey-card,
    .snapshot-card,
    .admin-stat,
    .insight-item {{
        background: linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(248,250,252,0.98) 100%);
        border: 1px solid {BORDER_COLOR};
        border-radius: 12px;
        padding: 0.8rem 0.9rem;
        min-height: 100%;
    }}

    .workspace-band-label,
    .snapshot-label,
    .admin-stat-label,
    .journey-step {{
        color: {TEXT_MUTED};
        font-size: 0.7rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 700;
    }}

    .workspace-band-value,
    .snapshot-value,
    .admin-stat-value {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.22rem;
        font-weight: 800;
        margin-top: 0.24rem;
        font-variant-numeric: tabular-nums;
        line-height: 1.08;
    }}

    .workspace-band-meta,
    .snapshot-delta,
    .journey-body,
    .insight-body {{
        color: {TEXT_SECONDARY};
        font-size: 0.8rem;
        line-height: 1.35;
        margin-top: 0.32rem;
    }}

    .snapshot-card {{
        border-left: 4px solid rgba(100, 116, 139, 0.32);
    }}

    .snapshot-card.success {{
        border-left-color: {SECONDARY_COLOR};
        background: linear-gradient(180deg, rgba(240,253,244,0.78), rgba(255,255,255,0.98));
    }}

    .snapshot-card.warning {{
        border-left-color: {CHART_ACCENT_COLOR};
        background: linear-gradient(180deg, rgba(255,251,235,0.84), rgba(255,255,255,0.98));
    }}

    .snapshot-card.danger {{
        border-left-color: #dc2626;
        background: linear-gradient(180deg, rgba(254,242,242,0.84), rgba(255,255,255,0.98));
    }}

    .snapshot-card.info {{
        border-left-color: {PRIMARY_COLOR};
    }}

    .journey-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1rem;
        font-weight: 800;
        margin-top: 0.5rem;
    }}

    .journey-hint,
    .snapshot-delta.neutral {{
        color: {TEXT_MUTED};
    }}

    .snapshot-delta.negative {{
        color: #dc2626;
    }}

    .insight-item + .insight-item {{
        margin-top: 0.85rem;
    }}

    .insight-title {{
        font-size: 0.98rem;
        margin-bottom: 0.3rem;
    }}

    .insight-compact-list {{
        display: grid;
        grid-template-columns: repeat(3, minmax(0, 1fr));
        gap: 0.55rem;
        margin-top: 0.35rem;
        align-items: stretch;
        grid-auto-rows: 1fr;
    }}

    .insight-compact-item {{
        background: linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(248,250,252,0.98) 100%);
        border: 1px solid {BORDER_COLOR};
        border-left: 4px solid rgba(100, 116, 139, 0.32);
        border-radius: 12px;
        padding: 0.62rem 0.72rem;
        min-height: 100%;
        height: 100%;
        display: flex;
        flex-direction: column;
        justify-content: flex-start;
    }}

    .insight-compact-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 0.84rem;
        font-weight: 800;
        line-height: 1.25;
    }}

    .insight-compact-body {{
        color: {TEXT_SECONDARY};
        font-size: 0.76rem;
        line-height: 1.34;
        margin-top: 0.22rem;
    }}

    .insight-compact-item.success {{
        border-left-color: {SECONDARY_COLOR};
    }}

    .insight-compact-item.warning {{
        border-left-color: {CHART_ACCENT_COLOR};
        background: linear-gradient(180deg, rgba(255,251,235,0.78), rgba(255,255,255,0.98));
    }}

    .insight-compact-item.danger {{
        border-left-color: #dc2626;
        background: linear-gradient(180deg, rgba(254,242,242,0.78), rgba(255,255,255,0.98));
    }}

    .insight-compact-item.info {{
        border-left-color: {PRIMARY_COLOR};
    }}

    .overview-command-panel {{
        border: 1px solid rgba(0, 52, 97, 0.12);
        border-radius: 16px;
        padding: 0.9rem;
        margin: 0.75rem 0 0.85rem;
        background:
            radial-gradient(circle at top left, rgba(0, 108, 73, 0.10), transparent 34%),
            linear-gradient(180deg, rgba(255,255,255,0.98), rgba(248,250,252,0.98));
        box-shadow: 0 12px 30px rgba(15, 23, 42, 0.05);
    }}

    .overview-command-header {{
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 0.75rem;
        margin-bottom: 0.7rem;
    }}

    .overview-command-kicker {{
        color: {TEXT_MUTED};
        font-size: 0.68rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 800;
        margin-bottom: 0.18rem;
    }}

    .overview-command-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.18rem;
        font-weight: 850;
        line-height: 1.12;
    }}

    .overview-command-meta {{
        display: inline-flex;
        align-items: center;
        white-space: nowrap;
        border-radius: 999px;
        padding: 0.28rem 0.56rem;
        background: rgba(0, 108, 73, 0.08);
        color: {SECONDARY_COLOR};
        font-size: 0.72rem;
        font-weight: 800;
    }}

    .overview-focus-stack {{
        display: grid;
        gap: 0.55rem;
        margin-top: 0.1rem;
    }}

    .overview-focus-card {{
        border: 1px solid {BORDER_COLOR};
        border-left: 4px solid rgba(100, 116, 139, 0.32);
        border-radius: 12px;
        padding: 0.72rem 0.78rem;
        background: linear-gradient(180deg, rgba(255,255,255,0.98), rgba(248,250,252,0.98));
    }}

    .overview-focus-card.success {{
        border-left-color: {SECONDARY_COLOR};
        background: linear-gradient(180deg, rgba(240,253,244,0.78), rgba(255,255,255,0.98));
    }}

    .overview-focus-card.warning {{
        border-left-color: {CHART_ACCENT_COLOR};
        background: linear-gradient(180deg, rgba(255,251,235,0.82), rgba(255,255,255,0.98));
    }}

    .overview-focus-card.danger {{
        border-left-color: #dc2626;
        background: linear-gradient(180deg, rgba(254,242,242,0.82), rgba(255,255,255,0.98));
    }}

    .overview-focus-card.info {{
        border-left-color: {PRIMARY_COLOR};
    }}

    .overview-focus-label {{
        color: {TEXT_MUTED};
        font-size: 0.66rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 800;
    }}

    .overview-focus-value {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.08rem;
        font-weight: 850;
        line-height: 1.12;
        margin-top: 0.2rem;
        overflow-wrap: anywhere;
    }}

    .overview-focus-body {{
        color: {TEXT_SECONDARY};
        font-size: 0.76rem;
        line-height: 1.34;
        margin-top: 0.25rem;
    }}

    .spotlight-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
        gap: 0.65rem;
        margin: 0.65rem 0 0.9rem;
    }}

    .spotlight-card {{
        padding: 0.85rem 0.95rem;
        border-radius: 12px;
        border: 1px solid {BORDER_COLOR};
        background: {SURFACE_COLOR};
        min-height: 120px;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }}

    .spotlight-card.accent {{
        border-left: 4px solid {SECONDARY_COLOR};
    }}

    .spotlight-card.warm {{
        border-left: 4px solid {CHART_ACCENT_COLOR};
    }}

    .spotlight-label {{
        color: {TEXT_MUTED};
        font-size: 0.7rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 700;
        line-height: 1.25;
    }}

    .spotlight-value {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.22rem;
        font-weight: 800;
        margin-top: 0.28rem;
        line-height: 1.12;
        overflow-wrap: anywhere;
    }}

    .spotlight-body {{
        color: {TEXT_SECONDARY};
        font-size: 0.8rem;
        line-height: 1.36;
        margin-top: 0.32rem;
    }}

    @media (max-width: 1180px) {{
        .insight-compact-list {{
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }}
    }}

    @media (max-width: 760px) {{
        .insight-compact-list {{
            grid-template-columns: 1fr;
        }}
    }}

    .feature-list {{
        display: flex;
        flex-wrap: wrap;
        gap: 0.55rem;
        margin-top: 1rem;
    }}

    .feature-chip {{
        padding: 0.45rem 0.7rem;
        border-radius: 999px;
        background: #f8fafc;
        border: 1px solid {BORDER_COLOR};
        color: {TEXT_SECONDARY};
        font-size: 0.82rem;
        font-weight: 600;
    }}

    .auth-title {{
        font-size: 1.8rem;
        margin: 0;
    }}

    .auth-copy {{
        margin-top: 0.75rem;
        margin-bottom: 0;
    }}

    /* Navigation */
    .nav-item {{
        padding: 0.75rem 1rem;
        border-radius: 10px;
        margin-bottom: 0.4rem;
        color: {TEXT_SECONDARY};
        font-weight: 600;
        transition: all 0.2s;
        border: 1px solid transparent;
    }}
    .nav-item.active {{
        background-color: #f1f5f9;
        color: {PRIMARY_COLOR};
        border-color: {BORDER_COLOR};
    }}

    .stTabs [data-baseweb="tab-list"] {{
        gap: 0.35rem;
        background: #f8fafc;
        border: 1px solid {BORDER_COLOR};
        border-radius: 12px;
        padding: 0.28rem;
    }}

    .stTabs [data-baseweb="tab"] {{
        border-radius: 9px;
        color: {TEXT_SECONDARY};
        font-weight: 700;
        padding: 0.42rem 0.72rem;
    }}

    .stTabs [aria-selected="true"] {{
        background: {SURFACE_COLOR};
        color: {PRIMARY_COLOR} !important;
        box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08);
    }}

    section[data-testid="stSidebar"] [data-testid="stExpander"] {{
        border: 1px solid {BORDER_COLOR};
        border-radius: 12px;
        background: #f8fafc;
    }}

    section[data-testid="stSidebar"] [role="radiogroup"] label {{
        border-radius: 10px;
        padding: 0.32rem 0.45rem;
    }}
</style>
"""

REFERENCE_THEME_CSS = f"""
<style>
    [data-testid="stHeader"] {{
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(12px);
        border-bottom: 1px solid {BORDER_COLOR};
    }}

    /* Sidebar adjustments */
    section[data-testid="stSidebar"] {{
        background-color: {SURFACE_COLOR} !important;
        border-right: 1px solid {BORDER_COLOR};
    }}

    /* Premium chart surface */
    [data-testid="stPlotlyChart"] {{
        position: relative;
        background:
            radial-gradient(circle at 10% 0%, rgba(0, 52, 97, 0.12), transparent 32%),
            radial-gradient(circle at 92% 8%, rgba(216, 154, 43, 0.12), transparent 30%),
            linear-gradient(180deg, rgba(255, 255, 255, 0.98), rgba(247, 250, 252, 0.94));
        border: 1px solid rgba(184, 199, 211, 0.78);
        border-radius: 20px;
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.85) inset,
            0 18px 42px rgba(15, 23, 42, 0.07),
            0 4px 12px rgba(15, 23, 42, 0.035);
        padding: 0.48rem 0.38rem 0.16rem;
        overflow: hidden;
        transition:
            transform 180ms ease,
            box-shadow 180ms ease,
            border-color 180ms ease;
    }}

    [data-testid="stPlotlyChart"]::before {{
        content: "";
        position: absolute;
        inset: 0;
        pointer-events: none;
        border-radius: inherit;
        background:
            linear-gradient(120deg, rgba(255,255,255,0.42), transparent 28%, rgba(255,255,255,0.12) 64%, transparent);
        opacity: 0.75;
    }}

    [data-testid="stPlotlyChart"]:hover {{
        transform: translateY(-2px);
        border-color: rgba(0, 52, 97, 0.22);
        box-shadow:
            0 1px 0 rgba(255, 255, 255, 0.9) inset,
            0 22px 50px rgba(15, 23, 42, 0.09),
            0 8px 18px rgba(0, 52, 97, 0.06);
    }}

    [data-testid="stPlotlyChart"] .js-plotly-plot,
    [data-testid="stPlotlyChart"] .plot-container,
    [data-testid="stPlotlyChart"] .svg-container {{
        border-radius: 16px;
    }}

    [data-testid="stPlotlyChart"] .modebar {{
        background: rgba(255, 255, 255, 0.82) !important;
        border: 1px solid rgba(203, 213, 225, 0.72);
        border-radius: 999px;
        box-shadow: 0 8px 18px rgba(15, 23, 42, 0.08);
        padding: 0.12rem 0.35rem;
        right: 0.65rem !important;
        top: 0.45rem !important;
        opacity: 0;
        transform: translateY(-4px);
        transition: opacity 180ms ease, transform 180ms ease;
    }}

    [data-testid="stPlotlyChart"]:hover .modebar,
    [data-testid="stPlotlyChart"]:focus-within .modebar {{
        opacity: 1;
        transform: translateY(0);
    }}

    [data-testid="stPlotlyChart"] .modebar-btn path {{
        fill: {TEXT_MUTED} !important;
    }}

    /* Input focus */
    div[data-baseweb="input"] > div:focus-within {{
        border-color: {PRIMARY_COLOR} !important;
    }}
</style>
"""


WORKSPACE_UI_CSS = f"""
<style>
    :root {{
        --artdb-navy: {PRIMARY_COLOR};
        --artdb-green: {SECONDARY_COLOR};
        --artdb-bg: {BG_COLOR};
        --artdb-surface: {SURFACE_COLOR};
        --artdb-border: {BORDER_COLOR};
        --artdb-muted: {TEXT_MUTED};
    }}

    .block-container {{
        max-width: 1680px;
        padding-top: 0.65rem;
        padding-bottom: 2rem;
    }}

    [data-testid="stVerticalBlockBorderWrapper"] {{
        border-radius: 8px !important;
        padding: 0.88rem !important;
        box-shadow: 0 1px 3px rgba(15, 23, 42, 0.04) !important;
    }}

    .app-shell-header {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
        padding: 0.72rem 0 0.82rem;
        border-bottom: 1px solid {BORDER_COLOR};
        margin-bottom: 0.72rem;
    }}

    .app-shell-brand {{
        display: flex;
        align-items: baseline;
        gap: 0.7rem;
        min-width: 0;
    }}

    .app-shell-name {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.34rem;
        font-weight: 800;
        line-height: 1;
    }}

    .app-shell-copy {{
        color: {TEXT_MUTED};
        font-size: 0.82rem;
        line-height: 1.3;
    }}

    .app-shell-user {{
        display: flex;
        align-items: center;
        justify-content: flex-end;
        gap: 0.65rem;
        text-align: right;
    }}

    .app-shell-user-name {{
        color: {TEXT_PRIMARY};
        font-size: 0.8rem;
        font-weight: 800;
        line-height: 1.2;
    }}

    .app-shell-user-meta {{
        color: {TEXT_MUTED};
        font-size: 0.7rem;
        line-height: 1.25;
        margin-top: 0.12rem;
    }}

    .app-shell-role {{
        display: inline-flex;
        align-items: center;
        min-height: 28px;
        padding: 0.24rem 0.52rem;
        border-radius: 999px;
        background: #EEF6F2;
        color: {SECONDARY_COLOR};
        font-size: 0.7rem;
        font-weight: 800;
        white-space: nowrap;
    }}

    .workspace-header {{
        display: flex;
        align-items: flex-start;
        justify-content: space-between;
        gap: 1rem;
        margin: 0.15rem 0 0.68rem;
        padding: 0.35rem 0 0.55rem;
    }}

    .workspace-header-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1.55rem;
        font-weight: 800;
        line-height: 1.15;
    }}

    .workspace-header-meta {{
        color: {TEXT_SECONDARY};
        font-size: 0.82rem;
        line-height: 1.4;
        margin-top: 0.28rem;
    }}

    .workspace-context {{
        display: flex;
        align-items: center;
        justify-content: flex-end;
        gap: 0.38rem;
        flex-wrap: wrap;
    }}

    .workspace-chip {{
        display: inline-flex;
        align-items: center;
        min-height: 28px;
        padding: 0.25rem 0.55rem;
        border: 1px solid {BORDER_COLOR};
        border-radius: 999px;
        background: {SURFACE_COLOR};
        color: {TEXT_SECONDARY};
        font-size: 0.74rem;
        font-weight: 650;
        line-height: 1.2;
    }}

    .filter-toolbar-heading {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.75rem;
        padding-bottom: 0.45rem;
    }}

    .filter-toolbar-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 0.95rem;
        font-weight: 800;
    }}

    .filter-toolbar-meta {{
        color: {TEXT_MUTED};
        font-size: 0.74rem;
    }}

    .filter-summary {{
        display: flex;
        flex-wrap: wrap;
        gap: 0.35rem;
        margin-top: 0.55rem;
    }}

    .filter-summary span {{
        padding: 0.22rem 0.48rem;
        border-radius: 999px;
        background: #EEF6F2;
        color: {SECONDARY_COLOR};
        font-size: 0.72rem;
        font-weight: 700;
    }}

    .metric-grid {{
        grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
        gap: 0.55rem;
        margin: 0.55rem 0 0.75rem;
    }}

    .metric-card,
    .snapshot-card,
    .insight-compact-item,
    .overview-focus-card,
    .spotlight-card {{
        border-radius: 8px;
        background: {SURFACE_COLOR};
        box-shadow: none;
    }}

    .metric-card {{
        padding: 0.74rem 0.8rem;
        min-height: 92px;
    }}

    .metric-card.warning,
    .metric-card.danger,
    .snapshot-card.warning,
    .snapshot-card.danger,
    .insight-compact-item.warning,
    .insight-compact-item.danger,
    .overview-focus-card.warning,
    .overview-focus-card.danger {{
        background: {SURFACE_COLOR};
    }}

    .metric-label {{
        letter-spacing: 0;
        text-transform: none;
        font-size: 0.73rem;
    }}

    .metric-value {{
        font-size: 1.3rem;
    }}

    .overview-action-center {{
        height: 100%;
        border: 1px solid {BORDER_COLOR};
        border-radius: 8px;
        background: {SURFACE_COLOR};
        overflow: hidden;
    }}

    .overview-action-header {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.7rem;
        padding: 0.78rem 0.85rem;
        border-bottom: 1px solid {BORDER_COLOR};
        background: #F8FAFC;
    }}

    .overview-action-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 1rem;
        font-weight: 800;
    }}

    .overview-action-count {{
        color: {TEXT_MUTED};
        font-size: 0.72rem;
        font-weight: 700;
    }}

    .overview-action-list {{
        display: grid;
    }}

    .overview-action-item {{
        display: grid;
        grid-template-columns: 8px minmax(0, 1fr);
        gap: 0.65rem;
        padding: 0.68rem 0.82rem;
        border-bottom: 1px solid #EEF2F6;
    }}

    .overview-action-item:last-child {{
        border-bottom: 0;
    }}

    .overview-action-indicator {{
        width: 8px;
        height: 8px;
        border-radius: 50%;
        margin-top: 0.25rem;
        background: {PRIMARY_COLOR};
    }}

    .overview-action-item.success .overview-action-indicator {{ background: {SECONDARY_COLOR}; }}
    .overview-action-item.warning .overview-action-indicator {{ background: {CHART_ACCENT_COLOR}; }}
    .overview-action-item.danger .overview-action-indicator {{ background: {CHART_DANGER_COLOR}; }}

    .overview-action-item-title {{
        color: {TEXT_PRIMARY};
        font-size: 0.8rem;
        font-weight: 800;
        line-height: 1.28;
    }}

    .overview-action-item-body {{
        color: {TEXT_SECONDARY};
        font-size: 0.73rem;
        line-height: 1.35;
        margin-top: 0.18rem;
    }}

    [data-testid="stPlotlyChart"] {{
        background: {SURFACE_COLOR};
        border-radius: 8px;
        box-shadow: none;
        padding: 0.2rem;
    }}

    [data-testid="stPlotlyChart"]::before {{
        display: none;
    }}

    [data-testid="stPlotlyChart"]:hover {{
        transform: none;
        box-shadow: none;
        border-color: rgba(0, 52, 97, 0.24);
    }}

    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"] {{
        border-radius: 8px;
    }}

    section[data-testid="stSidebar"] [role="radiogroup"] {{
        gap: 0.18rem;
    }}

    .sidebar-section-label {{
        color: {TEXT_MUTED};
        font-size: 0.68rem;
        font-weight: 800;
        letter-spacing: 0.06em;
        margin: 0.75rem 0 0.32rem;
        text-transform: uppercase;
    }}

    section[data-testid="stSidebar"] [role="radiogroup"] label {{
        min-height: 38px;
        padding: 0.38rem 0.48rem;
        border: 1px solid transparent;
        border-radius: 6px;
        transition: background-color 120ms ease, border-color 120ms ease;
    }}

    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {{
        background: #EEF6F2;
        border-color: rgba(0, 108, 73, 0.18);
    }}

    section[data-testid="stSidebar"] [data-testid="stExpander"] {{
        border-radius: 8px;
        background: {SURFACE_COLOR};
    }}

    .st-key-mobile_screen_navigation {{
        display: none;
    }}

    .st-key-mobile_bottom_navigation,
    .mobile-table-card-grid,
    .artdb-pwa-install {{
        display: none;
    }}

    .mobile-nav-heading {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.6rem;
        margin-bottom: 0.35rem;
    }}

    .mobile-nav-title {{
        color: {PRIMARY_COLOR};
        font-family: 'Manrope', sans-serif;
        font-size: 0.78rem;
        font-weight: 800;
    }}

    .mobile-nav-hint {{
        color: {TEXT_MUTED};
        font-size: 0.67rem;
        font-weight: 650;
        white-space: nowrap;
    }}

    @media (max-width: 900px) {{
        .app-shell-header,
        .workspace-header {{
            align-items: flex-start;
            flex-direction: column;
        }}

        .app-shell-user {{
            justify-content: flex-start;
            text-align: left;
        }}

        .workspace-context {{
            justify-content: flex-start;
        }}
    }}

    @media (max-width: 767px) {{
        .block-container {{
            padding: 0.45rem 0.65rem calc(6rem + env(safe-area-inset-bottom));
        }}

        header[data-testid="stHeader"] {{
            height: 3rem;
            background: rgba(244, 247, 249, 0.94);
            backdrop-filter: blur(10px);
        }}

        [data-testid="stSidebarCollapsedControl"] {{
            position: fixed;
            top: max(0.32rem, env(safe-area-inset-top));
            left: 0.45rem;
            z-index: 1002;
        }}

        [data-testid="stSidebarCollapsedControl"] button {{
            width: 44px;
            height: 44px;
            border: 1px solid {BORDER_COLOR};
            border-radius: 8px;
            background: {SURFACE_COLOR};
            box-shadow: 0 3px 12px rgba(15, 23, 42, 0.1);
        }}

        section[data-testid="stSidebar"] {{
            width: min(88vw, 360px) !important;
            min-width: min(88vw, 360px) !important;
            box-shadow: 12px 0 30px rgba(15, 23, 42, 0.16);
        }}

        section[data-testid="stSidebar"][aria-expanded="false"] {{
            width: 0 !important;
            min-width: 0 !important;
            overflow: hidden;
            transform: none !important;
            box-shadow: none;
        }}

        section[data-testid="stSidebar"] > div {{
            padding-bottom: calc(1.2rem + env(safe-area-inset-bottom));
        }}

        section[data-testid="stSidebar"] [role="radiogroup"] label {{
            min-height: 46px;
            padding: 0.52rem 0.62rem;
        }}

        section[data-testid="stSidebar"] [data-testid="stExpander"] summary {{
            min-height: 46px;
        }}

        .st-key-mobile_screen_navigation {{
            display: block;
            position: sticky;
            top: 0.25rem;
            z-index: 25;
            margin: 0 0 0.55rem;
            padding: 0.55rem 0.62rem 0.62rem;
            border: 1px solid {BORDER_COLOR};
            border-radius: 8px;
            background: rgba(255, 255, 255, 0.96);
            box-shadow: 0 4px 14px rgba(15, 23, 42, 0.08);
            backdrop-filter: blur(12px);
        }}

        .st-key-mobile_screen_navigation > [data-testid="stElementContainer"]:has(.mobile-nav-heading) {{
            min-height: 25px;
        }}

        .st-key-mobile_screen_navigation .mobile-nav-heading {{
            min-height: 20px;
            line-height: 1.2;
        }}

        .st-key-mobile_screen_navigation [data-baseweb="select"] > div {{
            min-height: 44px;
            border-radius: 6px;
        }}

        .st-key-mobile_bottom_navigation {{
            display: block;
            position: fixed;
            right: 0;
            bottom: 0;
            left: 0;
            z-index: 1000;
            padding: 0.42rem max(0.55rem, env(safe-area-inset-right)) calc(0.42rem + env(safe-area-inset-bottom)) max(0.55rem, env(safe-area-inset-left));
            border-top: 1px solid rgba(148, 163, 184, 0.32);
            background: rgba(255, 255, 255, 0.97);
            box-shadow: 0 -8px 24px rgba(15, 23, 42, 0.1);
            backdrop-filter: blur(16px);
        }}

        .st-key-mobile_bottom_navigation [data-testid="stHorizontalBlock"] {{
            display: grid !important;
            grid-template-columns: repeat(4, minmax(0, 1fr)) !important;
            gap: 0.3rem !important;
            align-items: stretch !important;
        }}

        .st-key-mobile_bottom_navigation [data-testid="stColumn"] {{
            width: 100% !important;
            min-width: 0 !important;
            max-width: none !important;
            flex: none !important;
        }}

        .st-key-mobile_bottom_navigation .stButton {{
            width: 100% !important;
        }}

        .st-key-mobile_bottom_navigation .stButton > button {{
            width: 100% !important;
            min-height: 52px;
            gap: 0.1rem;
            flex-direction: column;
            justify-content: center;
            padding: 0.25rem 0.12rem;
            border-color: transparent;
            border-radius: 8px;
            box-shadow: none;
        }}

        .st-key-mobile_bottom_navigation .stButton > button p {{
            font-size: 0.66rem;
            font-weight: 750;
            line-height: 1.05;
        }}

        .st-key-mobile_bottom_navigation .stButton > button span[data-testid="stIconMaterial"] {{
            font-size: 1.18rem;
        }}

        .artdb-pwa-install {{
            position: fixed;
            right: 0.7rem;
            bottom: calc(5rem + env(safe-area-inset-bottom));
            z-index: 1001;
            align-items: center;
            gap: 0.38rem;
            min-height: 42px;
            padding: 0.58rem 0.78rem;
            border: 0;
            border-radius: 999px;
            color: #FFFFFF;
            background: {PRIMARY_COLOR};
            box-shadow: 0 10px 24px rgba(0, 52, 97, 0.22);
            font-family: 'Manrope', sans-serif;
            font-size: 0.76rem;
            font-weight: 800;
        }}

        .app-shell-header {{
            align-items: center;
            flex-direction: row;
            min-height: 44px;
            padding: 0.42rem 0 0.55rem 3.25rem;
            margin-bottom: 0.45rem;
        }}

        .app-shell-brand {{
            align-items: flex-start;
            flex-direction: column;
            gap: 0.12rem;
        }}

        .app-shell-name {{
            font-size: 1.15rem;
        }}

        .app-shell-copy,
        .app-shell-user-meta,
        .app-shell-role {{
            display: none;
        }}

        .app-shell-user-name {{
            max-width: 38vw;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}

        .workspace-header {{
            gap: 0.45rem;
            margin: 0.05rem 0 0.5rem;
            padding: 0.2rem 0 0.35rem;
        }}

        .workspace-header-title {{
            font-size: 1.22rem;
        }}

        .workspace-header-meta {{
            font-size: 0.75rem;
            margin-top: 0.18rem;
        }}

        .workspace-context {{
            width: 100%;
            flex-wrap: nowrap;
            justify-content: flex-start;
            overflow-x: auto;
            padding: 0 0 0.25rem;
            scrollbar-width: none;
            scroll-snap-type: x proximity;
        }}

        .workspace-context::-webkit-scrollbar {{
            display: none;
        }}

        .workspace-chip {{
            flex: 0 0 auto;
            min-height: 30px;
            scroll-snap-align: start;
        }}

        [data-testid="stVerticalBlockBorderWrapper"] {{
            padding: 0.68rem !important;
        }}

        .metric-grid {{
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.42rem;
            margin: 0.45rem 0 0.65rem;
        }}

        .metric-card {{
            min-height: 86px;
            padding: 0.62rem;
        }}

        .metric-label {{
            font-size: 0.65rem;
            line-height: 1.2;
        }}

        .metric-value {{
            font-size: 1.08rem;
            overflow-wrap: anywhere;
        }}

        .metric-delta {{
            font-size: 0.68rem;
            line-height: 1.25;
        }}

        .workspace-band,
        .journey-grid,
        .snapshot-strip,
        .admin-stat-grid,
        .spotlight-grid {{
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.45rem;
        }}

        .insight-compact-list {{
            grid-template-columns: 1fr;
            grid-auto-rows: auto;
        }}

        .insight-compact-item,
        .spotlight-card {{
            min-height: auto;
            padding: 0.65rem;
        }}

        .overview-action-header {{
            padding: 0.65rem;
        }}

        .overview-action-item {{
            gap: 0.52rem;
            padding: 0.62rem 0.65rem;
        }}

        .section-intro {{
            min-height: auto;
            padding: 0.7rem;
        }}

        .section-intro h2,
        h1 {{
            font-size: 1.35rem !important;
        }}

        h2 {{
            font-size: 1.16rem !important;
        }}

        h3 {{
            font-size: 1rem !important;
        }}

        .panel-title {{
            font-size: 0.98rem;
        }}

        .panel-caption,
        [data-testid="stCaptionContainer"] {{
            font-size: 0.76rem;
        }}

        .stButton > button,
        .stDownloadButton > button,
        div[data-testid="stFormSubmitButton"] > button {{
            width: 100%;
            min-height: 44px;
            padding: 0.55rem 0.75rem;
        }}

        [data-baseweb="input"] > div,
        [data-baseweb="select"] > div,
        [data-testid="stDateInput"] input,
        [data-testid="stNumberInput"] input {{
            min-height: 44px;
        }}

        [data-testid="stFileUploaderDropzone"] {{
            min-height: 112px;
            padding: 0.72rem;
            border-radius: 8px;
        }}

        [data-testid="stFileUploaderDropzone"] button {{
            min-height: 42px;
        }}

        .stTabs [data-baseweb="tab-list"] {{
            flex-wrap: nowrap;
            overflow-x: auto;
            overscroll-behavior-x: contain;
            scrollbar-width: none;
        }}

        .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar {{
            display: none;
        }}

        .stTabs [data-baseweb="tab"] {{
            flex: 0 0 auto;
            min-height: 44px;
            white-space: nowrap;
        }}

        [data-testid="stDataFrame"],
        [data-testid="stDataEditor"] {{
            max-width: calc(100vw - 1.3rem);
            overflow-x: auto;
            -webkit-overflow-scrolling: touch;
        }}

        [data-testid="stDataFrame"] [role="gridcell"],
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataEditor"] [role="gridcell"],
        [data-testid="stDataEditor"] [role="columnheader"] {{
            font-size: 0.74rem !important;
        }}

        .mobile-table-card-grid {{
            display: grid;
            grid-template-columns: 1fr;
            gap: 0.48rem;
            margin: 0.35rem 0 0.6rem;
        }}

        .mobile-data-card {{
            padding: 0.68rem;
            border: 1px solid {BORDER_COLOR};
            border-radius: 8px;
            background: linear-gradient(145deg, #FFFFFF 0%, #F8FBFA 100%);
        }}

        .mobile-data-card-title {{
            color: {PRIMARY_COLOR};
            font-family: 'Manrope', sans-serif;
            font-size: 0.82rem;
            font-weight: 800;
            line-height: 1.3;
            overflow-wrap: anywhere;
        }}

        .mobile-data-card-fields {{
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 0.42rem 0.65rem;
            margin-top: 0.55rem;
        }}

        .mobile-data-card-field {{
            min-width: 0;
        }}

        .mobile-data-card-label {{
            color: {TEXT_MUTED};
            font-size: 0.62rem;
            font-weight: 750;
            letter-spacing: 0.035em;
            text-transform: uppercase;
        }}

        .mobile-data-card-value {{
            margin-top: 0.08rem;
            color: {TEXT_PRIMARY};
            font-size: 0.78rem;
            font-weight: 700;
            line-height: 1.25;
            overflow-wrap: anywhere;
        }}

        .mobile-table-card-note {{
            color: {TEXT_MUTED};
            font-size: 0.68rem;
            line-height: 1.3;
        }}

        [data-testid="stPlotlyChart"] {{
            min-height: 290px;
            padding: 0;
            touch-action: pan-y;
            animation: none !important;
        }}

        [data-testid="stPlotlyChart"] .modebar {{
            display: none !important;
        }}

        [data-testid="stExpander"] summary {{
            min-height: 46px;
        }}

        [data-testid="stAlert"] {{
            padding: 0.7rem;
        }}

        [data-testid="stHorizontalBlock"] {{
            gap: 0.55rem;
        }}
    }}
</style>
"""


st.markdown(DASHBOARD_CSS + REFERENCE_THEME_CSS + WORKSPACE_UI_CSS, unsafe_allow_html=True)


def inject_pwa_shell() -> None:
    components.html(
        """
        <script>
        (() => {
          const parentWindow = window.parent;
          const parentDoc = parentWindow && parentWindow.document;
          if (!parentDoc) return;

          const ensureLink = (rel, href, attributes = {}) => {
            let node = parentDoc.head.querySelector(`link[rel="${rel}"]`);
            if (!node) {
              node = parentDoc.createElement('link');
              node.rel = rel;
              parentDoc.head.appendChild(node);
            }
            node.href = href;
            Object.entries(attributes).forEach(([key, value]) => node.setAttribute(key, value));
          };

          const ensureMeta = (name, content) => {
            let node = parentDoc.head.querySelector(`meta[name="${name}"]`);
            if (!node) {
              node = parentDoc.createElement('meta');
              node.name = name;
              parentDoc.head.appendChild(node);
            }
            node.content = content;
          };

          const staticRoot = `${parentWindow.location.origin}/app/static`;
          ensureLink('manifest', `${staticRoot}/manifest.webmanifest`);
          ensureLink('icon', `${staticRoot}/artdb-icon.svg`, { type: 'image/svg+xml' });
          ensureLink('apple-touch-icon', `${staticRoot}/artdb-icon.svg`);
          ensureMeta('theme-color', '#003461');
          ensureMeta('mobile-web-app-capable', 'yes');
          ensureMeta('apple-mobile-web-app-capable', 'yes');
          ensureMeta('apple-mobile-web-app-status-bar-style', 'default');
          ensureMeta('apple-mobile-web-app-title', 'ArtDB');

          const viewport = parentDoc.head.querySelector('meta[name="viewport"]');
          if (viewport && !viewport.content.includes('viewport-fit=cover')) {
            viewport.content = `${viewport.content}, viewport-fit=cover`;
          }

          if ('serviceWorker' in parentWindow.navigator && parentWindow.isSecureContext) {
            parentWindow.navigator.serviceWorker
              .register(`${staticRoot}/service-worker.js`, { scope: '/app/static/' })
              .catch(() => undefined);
          }

          if (parentWindow.__artdbInstallPromptBound) return;
          parentWindow.__artdbInstallPromptBound = true;
          let deferredPrompt = null;

          const removeInstallButton = () => {
            parentDoc.getElementById('artdb-pwa-install')?.remove();
          };

          parentWindow.addEventListener('beforeinstallprompt', (event) => {
            event.preventDefault();
            deferredPrompt = event;
            removeInstallButton();

            const button = parentDoc.createElement('button');
            button.id = 'artdb-pwa-install';
            button.className = 'artdb-pwa-install';
            button.type = 'button';
            button.innerHTML = '<span aria-hidden="true">＋</span><span>Установить ArtDB</span>';
            button.addEventListener('click', async () => {
              if (!deferredPrompt) return;
              deferredPrompt.prompt();
              await deferredPrompt.userChoice;
              deferredPrompt = null;
              removeInstallButton();
            });
            parentDoc.body.appendChild(button);
          });

          parentWindow.addEventListener('appinstalled', removeInstallButton);
        })();
        </script>
        """,
        height=0,
        width=0,
    )


inject_pwa_shell()


def inject_chart_motion() -> None:
    components.html(
        """
        <script>
        (() => {
          const ua = navigator.userAgent || '';
          const isSafari = /^((?!chrome|android|crios|fxios|edgios).)*safari/i.test(ua);
          if (isSafari) {
            return;
          }

          if (window.matchMedia && window.matchMedia('(prefers-reduced-motion: reduce)').matches) {
            return;
          }

          const parentDoc = window.parent && window.parent.document;
          if (!parentDoc) return;

          let attempts = 0;
          const maxAttempts = 20;
          const intervalMs = 220;

          const animateBars = (chart, baseDelay) => {
            const bars = chart.querySelectorAll('.barlayer .trace .points path');
            bars.forEach((bar, index) => {
              bar.style.transformBox = 'fill-box';
              bar.style.transformOrigin = 'center bottom';
              bar.animate(
                [
                  { transform: 'scaleY(0.02)', opacity: 0.18 },
                  { transform: 'scaleY(1.04)', opacity: 1, offset: 0.72 },
                  { transform: 'scaleY(1)', opacity: 1 }
                ],
                {
                  duration: 820,
                  delay: baseDelay + index * 34,
                  easing: 'cubic-bezier(0.22, 1, 0.36, 1)',
                  fill: 'both'
                }
              );
            });
            return bars.length > 0;
          };

          const animateLines = (chart, baseDelay) => {
            const lines = chart.querySelectorAll('.scatterlayer .trace path.js-line');
            lines.forEach((line, index) => {
              if (typeof line.getTotalLength !== 'function') return;
              const length = line.getTotalLength();
              if (!length || !Number.isFinite(length)) return;
              line.style.strokeDasharray = `${length}`;
              line.style.strokeDashoffset = `${length}`;
              line.animate(
                [
                  { strokeDashoffset: length, opacity: 0.18 },
                  { strokeDashoffset: 0, opacity: 1 }
                ],
                {
                  duration: 980,
                  delay: baseDelay + index * 70,
                  easing: 'cubic-bezier(0.22, 1, 0.36, 1)',
                  fill: 'forwards'
                }
              );
            });
            return lines.length > 0;
          };

          const animateMarkers = (chart, baseDelay) => {
            const markers = chart.querySelectorAll('.scatterlayer .trace .points path, .scatterlayer .trace .points circle');
            markers.forEach((marker, index) => {
              marker.style.transformBox = 'fill-box';
              marker.style.transformOrigin = 'center center';
              marker.animate(
                [
                  { transform: 'scale(0.12)', opacity: 0 },
                  { transform: 'scale(1.08)', opacity: 1, offset: 0.68 },
                  { transform: 'scale(1)', opacity: 1 }
                ],
                {
                  duration: 520,
                  delay: baseDelay + 260 + index * 22,
                  easing: 'cubic-bezier(0.22, 1, 0.36, 1)',
                  fill: 'both'
                }
              );
            });
            return markers.length > 0;
          };

          const animateSlices = (chart, baseDelay) => {
            const slices = chart.querySelectorAll('.pielayer path.surface, .treemaplayer path, .iciclelayer path, .sunburstlayer path');
            slices.forEach((slice, index) => {
              slice.style.transformBox = 'fill-box';
              slice.style.transformOrigin = 'center center';
              slice.animate(
                [
                  { transform: 'scale(0.88)', opacity: 0.08 },
                  { transform: 'scale(1.03)', opacity: 1, offset: 0.72 },
                  { transform: 'scale(1)', opacity: 1 }
                ],
                {
                  duration: 760,
                  delay: baseDelay + index * 48,
                  easing: 'cubic-bezier(0.22, 1, 0.36, 1)',
                  fill: 'both'
                }
              );
            });
            return slices.length > 0;
          };

          const animateCharts = () => {
            const charts = parentDoc.querySelectorAll('[data-testid="stPlotlyChart"]');
            charts.forEach((chart, index) => {
              const marker = chart.dataset.codexChartMotionVersion;
              const version = 'v3';
              if (marker === version) return;
              const svg = chart.querySelector('svg');
              if (!svg) return;

              chart.dataset.codexChartMotionVersion = version;
              const baseDelay = index * 110;
              chart.style.willChange = 'opacity, transform, filter';
              chart.style.animation = 'none';
              chart.offsetHeight;
              chart.style.animation = `chartRevealStrong 860ms cubic-bezier(0.22, 1, 0.36, 1) ${baseDelay}ms both`;

              const hasBars = animateBars(chart, baseDelay);
              const hasLines = animateLines(chart, baseDelay);
              const hasMarkers = animateMarkers(chart, baseDelay);
              const hasSlices = animateSlices(chart, baseDelay);

              if (!hasBars && !hasLines && !hasMarkers && !hasSlices) {
                const plotRoot = chart.querySelector('.js-plotly-plot, .plotly, svg');
                if (plotRoot && typeof plotRoot.animate === 'function') {
                  plotRoot.animate(
                    [
                      { opacity: 0, transform: 'translateY(20px) scale(0.98)' },
                      { opacity: 1, transform: 'translateY(0) scale(1)' }
                    ],
                    {
                      duration: 720,
                      delay: baseDelay,
                      easing: 'cubic-bezier(0.22, 1, 0.36, 1)',
                      fill: 'both'
                    }
                  );
                }
              }
            });
            attempts += 1;
            if (attempts >= maxAttempts) clearInterval(timer);
          };

          const timer = setInterval(animateCharts, intervalMs);
          animateCharts();
        })();
        </script>
        """,
        height=0,
        width=0,
    )


inject_chart_motion()


def is_missing(value: object) -> bool:
    return value is None or bool(pd.isna(value))


def render_html_block(html: str) -> None:
    st.markdown(dedent(html).strip(), unsafe_allow_html=True)


def render_sticky_rail_marker() -> None:
    components.html(
        """
        <script>
        (() => {
          const ua = navigator.userAgent || '';
          const isSafari = /^((?!chrome|android|crios|fxios|edgios).)*safari/i.test(ua);
          if (isSafari) {
            return;
          }

          const host = window.frameElement;
          const parentWindow = window.parent;
          if (!host || !parentWindow || parentWindow.innerWidth < 1100) return;

          const column = host.closest('[data-testid="column"]');
          if (!column) return;

          column.classList.add('control-rail-column');
          const body = column.firstElementChild;
          if (body) body.classList.add('control-rail-body');

          host.style.display = 'none';
          host.style.width = '0';
          host.style.height = '0';
          host.style.border = '0';
        })();
        </script>
        """,
        height=0,
        width=0,
    )


def render_sticky_header_marker() -> None:
    components.html(
        """
        <script>
        (() => {
          const ua = navigator.userAgent || '';
          const isSafari = /^((?!chrome|android|crios|fxios|edgios).)*safari/i.test(ua);
          if (isSafari) {
            return;
          }

          const host = window.frameElement;
          if (!host) return;

          const wrapper =
            host.closest('[data-testid="stVerticalBlockBorderWrapper"]') ||
            host.closest('[data-testid="stVerticalBlock"]');
          if (!wrapper) return;

          wrapper.classList.add('control-header-shell');
          host.style.display = 'none';
          host.style.width = '0';
          host.style.height = '0';
          host.style.border = '0';
        })();
        </script>
        """,
        height=0,
        width=0,
    )


def schedule_widget_reset(resets: dict[str, object | None]) -> None:
    pending_resets = dict(st.session_state.get("_pending_widget_resets", {}))
    pending_resets.update(resets)
    st.session_state["_pending_widget_resets"] = pending_resets


def apply_pending_widget_resets() -> None:
    pending_resets = st.session_state.pop("_pending_widget_resets", None)
    if not pending_resets:
        return
    for key, value in pending_resets.items():
        if value is None:
            st.session_state.pop(key, None)
        else:
            st.session_state[key] = value


def format_money(value: float) -> str:
    if is_missing(value):
        return "н/д"
    return f"{float(value):,.0f} сом".replace(",", " ")


def format_number(value: float) -> str:
    if is_missing(value):
        return "н/д"
    return f"{float(value):,.0f}".replace(",", " ")


def format_percent(value: float) -> str:
    if is_missing(value):
        return "н/д"
    return f"{float(value):,.1f}%".replace(",", " ")


def percent_or_none(value: float) -> str | None:
    return None if is_missing(value) else format_percent(value)


def format_change_percent(value: float) -> str:
    if is_missing(value):
        return ""
    return f"{float(value):+,.1f}%".replace(",", " ")


def calculate_change_pct(current: float, previous: float) -> float | None:
    if is_missing(current) or is_missing(previous):
        return None
    previous_value = float(previous)
    if abs(previous_value) < 1e-9:
        return None
    return ((float(current) - previous_value) / previous_value) * 100.0


def parse_plan_input(value: str) -> float | None:
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    return float(text)


ROLE_LABELS = {
    "admin": "Администратор",
    "manager": "Руководитель",
    "salon": "Салон",
}


def role_label(role: str) -> str:
    return ROLE_LABELS.get(role, role)


def get_build_badge_label() -> str:
    commit = (os.getenv("RENDER_GIT_COMMIT") or os.getenv("GIT_COMMIT") or "").strip()
    if commit:
        return f"Сборка {commit[:7]}"
    return "Обновление 07.04"


def is_network_role(role: str) -> bool:
    return role in {"admin", "manager"}


def can_manage_access(current_user: dict[str, str]) -> bool:
    return current_user["role"] == "admin"


def can_manage_plans(current_user: dict[str, str]) -> bool:
    return current_user["role"] == "admin"


def can_manage_procurement(current_user: dict[str, str]) -> bool:
    return current_user["role"] == "admin"


MARGIN_HIDDEN_COLUMNS = {
    "cost",
    "margin",
    "margin_pct",
    "margin_change_pct",
    "return_margin",
    "margin_plan",
    "margin_gap",
    "margin_execution_pct",
    "unit_cost",
    "avg_margin_per_unit",
    "moving_avg_margin_3m",
}
MARGIN_HIDDEN_MAPPING_FIELDS = {"cost", "margin", "unit_cost"}


def can_view_margin(current_user: dict[str, str]) -> bool:
    return current_user["role"] in {"admin", "manager"}


def margin_safe_columns(columns: list[str], current_user: dict[str, str]) -> list[str]:
    if can_view_margin(current_user):
        return list(columns)
    return [column for column in columns if column not in MARGIN_HIDDEN_COLUMNS]


def margin_safe_rename_map(
    rename_map: dict[str, str] | None,
    current_user: dict[str, str],
) -> dict[str, str] | None:
    if rename_map is None or can_view_margin(current_user):
        return rename_map
    return {key: value for key, value in rename_map.items() if key not in MARGIN_HIDDEN_COLUMNS}


def margin_safe_frame(frame: pd.DataFrame, current_user: dict[str, str]) -> pd.DataFrame:
    if can_view_margin(current_user):
        return frame.copy()
    visible_columns = [column for column in frame.columns if column not in MARGIN_HIDDEN_COLUMNS]
    return frame[visible_columns].copy()


def margin_safe_mapping(selected_mapping: dict[str, str | None], current_user: dict[str, str]) -> dict[str, str | None]:
    if can_view_margin(current_user):
        return dict(selected_mapping)
    return {
        key: value
        for key, value in selected_mapping.items()
        if key not in MARGIN_HIDDEN_MAPPING_FIELDS
    }


def _hex_to_rgb(color: str) -> tuple[int, int, int]:
    cleaned = color.strip().lstrip("#")
    if len(cleaned) != 6:
        return (0, 0, 0)
    return tuple(int(cleaned[index : index + 2], 16) for index in (0, 2, 4))


def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
    return "#{:02X}{:02X}{:02X}".format(*[max(0, min(255, int(value))) for value in rgb])


def blend_hex_color(start: str, end: str, ratio: float) -> str:
    start_rgb = _hex_to_rgb(start)
    end_rgb = _hex_to_rgb(end)
    safe_ratio = max(0.0, min(1.0, float(ratio)))
    blended = tuple(
        start_rgb[channel] + (end_rgb[channel] - start_rgb[channel]) * safe_ratio
        for channel in range(3)
    )
    return _rgb_to_hex(blended)


def gradient_colors(count: int, start: str, end: str) -> list[str]:
    safe_count = max(int(count), 0)
    if safe_count <= 0:
        return []
    if safe_count == 1:
        return [end]
    return [blend_hex_color(start, end, index / (safe_count - 1)) for index in range(safe_count)]


def style_chart_traces(figure: go.Figure) -> None:
    figure.update_traces(
        selector=dict(type="bar"),
        opacity=0.96,
        marker_line_width=0.75,
        marker_line_color="rgba(255,255,255,0.7)",
        hoverlabel=dict(bordercolor="rgba(15,23,42,0.12)"),
    )
    figure.update_traces(
        selector=dict(type="scatter"),
        line=dict(width=3.25),
        marker=dict(line=dict(width=1.7, color=SURFACE_COLOR)),
        hoverlabel=dict(bordercolor="rgba(15,23,42,0.12)"),
    )
    figure.update_traces(
        selector=dict(type="waterfall"),
        opacity=0.94,
        connector=dict(line=dict(color=CHART_ZERO_COLOR, width=1)),
    )
    figure.update_traces(
        selector=dict(type="pie"),
        textfont=dict(size=12, color=TEXT_PRIMARY),
        marker=dict(line=dict(color=SURFACE_COLOR, width=2)),
        hole=0.38,
        pull=0.015,
    )
    figure.update_traces(
        selector=dict(type="treemap"),
        marker=dict(line=dict(color=SURFACE_COLOR, width=2)),
        textfont=dict(size=12),
        tiling=dict(pad=3),
    )
    figure.update_traces(
        selector=dict(type="heatmap"),
        colorscale=[
            [0.0, "#EEF6FA"],
            [0.45, "#8EC5D6"],
            [1.0, PRIMARY_COLOR],
        ],
    )


def polish_figure(figure: go.Figure, *, height: int | None = None) -> go.Figure:
    trace_types = {getattr(trace, "type", "") for trace in figure.data}
    use_unified_hover = bool(trace_types & {"bar", "scatter", "waterfall"}) and not bool(
        trace_types & {"pie", "treemap", "heatmap"}
    )
    title_text = str(getattr(figure.layout.title, "text", "") or "")
    style_chart_traces(figure)
    figure.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,0)",
        font=dict(family="Inter", color=TEXT_PRIMARY, size=12),
        title_text=title_text,
        title_font=dict(family="Manrope", size=17, color=TEXT_PRIMARY),
        title_x=0.02,
        title_xanchor="left",
        margin=dict(l=24, r=20, t=52, b=30),
        legend_title_text="",
        transition=dict(duration=520, easing="cubic-in-out"),
        colorway=CHART_COLORS,
        hovermode="x unified" if use_unified_hover else None,
        hoverlabel=dict(
            bgcolor="rgba(15, 23, 42, 0.96)",
            bordercolor="rgba(15, 23, 42, 0.96)",
            font_color="#FFFFFF",
            font_size=13,
            font_family="Inter",
            align="left",
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="left",
            x=0,
            bgcolor="rgba(255,255,255,0.62)",
            bordercolor="rgba(203,213,225,0.62)",
            borderwidth=1,
            font=dict(size=12, color=TEXT_SECONDARY),
            itemclick="toggleothers",
            itemdoubleclick="toggle",
        ),
        uniformtext=dict(minsize=10, mode="hide"),
        bargap=0.28,
        bargroupgap=0.08,
    )
    figure.update_xaxes(
        gridcolor=CHART_GRID_COLOR,
        zerolinecolor=CHART_ZERO_COLOR,
        linecolor="rgba(203,213,225,0.62)",
        showline=False,
        tickfont=dict(color=TEXT_MUTED),
        title_font=dict(color=TEXT_SECONDARY),
        ticks="",
        ticklabelposition="outside",
        automargin=True,
    )
    figure.update_yaxes(
        gridcolor=CHART_GRID_COLOR,
        zerolinecolor=CHART_ZERO_COLOR,
        linecolor="rgba(203,213,225,0.62)",
        showline=False,
        tickfont=dict(color=TEXT_MUTED),
        title_font=dict(color=TEXT_SECONDARY),
        ticks="",
        ticklabelposition="outside",
        automargin=True,
    )
    if height is not None:
        figure.update_layout(height=height)
    return figure


def compact_bar_chart_height(
    row_count: int,
    *,
    minimum: int = 240,
    maximum: int = 340,
    base: int = 110,
    row_step: int = 32,
) -> int:
    safe_rows = max(int(row_count), 1)
    return max(minimum, min(maximum, base + safe_rows * row_step))


def infer_dashboard_tone(*values: object) -> str:
    text = " ".join(str(value) for value in values if value is not None).lower()
    if not text:
        return "info"
    if any(isinstance(value, str) and value.strip().startswith("-") for value in values):
        return "danger"
    if any(isinstance(value, str) and value.strip().startswith("+") for value in values):
        return "success"

    danger_markers = (
        "дефицит",
        "без остатка",
        "риск",
        "ниже",
        "аномал",
        "ошиб",
        "критич",
        "неликвид",
    )
    warning_markers = (
        "возврат",
        "излиш",
        "спящий",
        "нет спроса",
        "план:",
        "предуп",
    )
    success_markers = (
        "лидер",
        "сильнейший",
        "главная",
        "маржа %",
        "покрыто",
        "рост",
        "+",
    )

    if any(marker in text for marker in danger_markers):
        return "danger"
    if any(marker in text for marker in warning_markers):
        return "warning"
    if any(marker in text for marker in success_markers):
        return "success"
    return "info"


def render_metric_cards(cards: list[dict[str, str]]) -> None:
    cards_html: list[str] = []
    for card in cards:
        delta = card.get("delta") or ""
        delta_class = "metric-delta"
        if delta.startswith("+"):
            delta_class += " positive"
        elif delta.startswith("-"):
            delta_class += " negative"

        # New: Progress bar if 'progress' key exists (0-100)
        progress_html = ""
        if "progress" in card:
            p_val = card["progress"]
            progress_html = f'''
            <div style="background: {BORDER_COLOR}; border-radius: 99px; height: 6px; margin-top: 0.75rem; overflow: hidden;">
                <div style="background: {SECONDARY_COLOR}; width: {p_val}%; height: 100%; border-radius: 99px;"></div>
            </div>
            '''

        delta_html = f'<div class="{delta_class}">{escape(delta)}</div>' if delta else ""
        inner = f'<div class="metric-label">{escape(card["label"])}</div><div class="metric-value">{escape(card["value"])}</div>{delta_html}{progress_html}'
        tone = card.get("tone") or infer_dashboard_tone(card.get("label", ""), delta)
        cards_html.append(f'<div class="metric-card {escape(str(tone))}">{inner}</div>')

    st.markdown(f'<div class="metric-grid">{"".join(cards_html)}</div>', unsafe_allow_html=True)


def render_section_intro(title: str, description: str) -> None:
    render_html_block(
        f"""
        <div class="section-intro">
            <div class="section-intro-title">{escape(title)}</div>
            <div class="section-intro-body">{escape(description)}</div>
        </div>
        """
    )


def render_panel_header(title: str, description: str) -> None:
    st.markdown(f'<div class="panel-title">{escape(title)}</div>', unsafe_allow_html=True)
    if str(description).strip():
        st.markdown(f'<div class="panel-caption">{escape(description)}</div>', unsafe_allow_html=True)


def render_section_marker(kicker: str, title: str, description: str) -> None:
    render_html_block(
        f"""
        <div class="section-marker">
            <div class="section-marker-kicker">{escape(kicker)}</div>
            <div class="section-marker-title">{escape(title)}</div>
            <div class="section-marker-body">{escape(description)}</div>
        </div>
        """
    )


def render_workspace_band(items: list[dict[str, str]]) -> None:
    cards_html: list[str] = []
    for item in items:
        meta = item.get("meta", "")
        meta_html = f'<div class="workspace-band-meta">{escape(meta)}</div>' if meta else ""
        cards_html.append(
            dedent(
                f"""
                <div class="workspace-band-item">
                    <div class="workspace-band-label">{escape(item["label"])}</div>
                    <div class="workspace-band-value">{escape(item["value"])}</div>
                    {meta_html}
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="workspace-band">{"".join(cards_html)}</div>')


def render_journey_cards(items: list[dict[str, str]]) -> None:
    cards_html: list[str] = []
    for index, item in enumerate(items, start=1):
        hint = item.get("hint", "")
        hint_html = f'<div class="journey-hint">{escape(hint)}</div>' if hint else ""
        cards_html.append(
            dedent(
                f"""
                <div class="journey-card">
                    <div class="journey-step">{index:02d}</div>
                    <div class="journey-title">{escape(item["title"])}</div>
                    <div class="journey-body">{escape(item["body"])}</div>
                    {hint_html}
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="journey-grid">{"".join(cards_html)}</div>')


def render_spotlight_cards(items: list[dict[str, str]]) -> None:
    cards_html: list[str] = []
    for item in items:
        tone = item.get("tone", "neutral")
        cards_html.append(
            dedent(
                f"""
                <div class="spotlight-card {escape(tone)}">
                    <div class="spotlight-label">{escape(item["label"])}</div>
                    <div class="spotlight-value">{escape(item["value"])}</div>
                    <div class="spotlight-body">{escape(item["body"])}</div>
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="spotlight-grid">{"".join(cards_html)}</div>')


def render_snapshot_strip(items: list[dict[str, str]]) -> None:
    cards_html: list[str] = []
    for item in items:
        delta = item.get("delta", "")
        delta_class = "snapshot-delta"
        if isinstance(delta, str) and delta.startswith("-"):
            delta_class += " negative"
        if not delta:
            delta_class += " neutral"
            delta = item.get("hint", "")
        tone = item.get("tone") or infer_dashboard_tone(item.get("label", ""), item.get("value", ""), delta)
        cards_html.append(
            dedent(
                f"""
                <div class="snapshot-card {escape(str(tone))}">
                    <div class="snapshot-label">{escape(item["label"])}</div>
                    <div class="snapshot-value">{escape(item["value"])}</div>
                    <div class="{delta_class}">{escape(delta)}</div>
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="snapshot-strip">{"".join(cards_html)}</div>')


def render_data_quality_report(report, current_user: dict[str, str]) -> None:
    if report.status == "blocked":
        st.error("Файл пока нельзя сохранять: найдены критичные ошибки качества данных.")
    elif report.status == "warning":
        st.warning("Файл можно сохранить, но перед этим стоит проверить предупреждения.")
    else:
        st.success("Качество данных выглядит хорошо. Файл готов к сохранению и анализу.")

    new_products = int(report.catalog_metrics.get("new_products", 0))
    matched_products = int(report.catalog_metrics.get("matched_catalog_products", 0))
    catalog_products = int(report.catalog_metrics.get("catalog_products", 0))
    render_snapshot_strip(
        [
            {
                "label": "Индекс качества",
                "value": f"{report.score}/100",
                "hint": "Готовность файла к сохранению",
                "tone": "success" if report.score >= 90 else ("warning" if report.score >= 70 else "danger"),
            },
            {
                "label": "Строк принято",
                "value": format_number(report.prepared_rows),
                "hint": f"из {format_number(report.raw_rows)} строк файла",
                "tone": "info",
            },
            {
                "label": "Исключено строк",
                "value": format_number(report.dropped_rows),
                "hint": "Пустая дата, товар или выручка",
                "tone": "success" if report.dropped_rows == 0 else "warning",
            },
            {
                "label": "Новые SKU",
                "value": format_number(new_products),
                "hint": f"совпало со справочником: {format_number(matched_products)} / {format_number(catalog_products)}",
                "tone": "success" if new_products == 0 else "warning",
            },
        ]
    )

    if report.issues:
        issue_rows = [
            {
                "Уровень": {
                    "blocker": "Критично",
                    "warning": "Проверить",
                    "info": "Инфо",
                }.get(issue.severity, issue.severity),
                "Проблема": issue.title,
                "Строк": issue.count,
                "Что значит": issue.detail,
                "Что сделать": issue.action,
            }
            for issue in report.issues
        ]
        st.dataframe(pd.DataFrame(issue_rows), use_container_width=True, hide_index=True, height=min(280, 62 + len(issue_rows) * 42))

    if not report.problem_rows.empty:
        with st.expander("Показать проблемные строки", expanded=False):
            st.dataframe(
                format_display_frame_for_role(report.problem_rows, current_user),
                use_container_width=True,
                hide_index=True,
                height=280,
            )


SCREEN_NAV_LABELS = {
    "Обзор": "Главная",
    "ABC-анализ": "Ассортимент · ABC",
    "Карточка SKU": "Ассортимент · Карточка SKU",
    "Финансы": "Продажи · Финансы",
    "Аналитика": "Продажи · Аналитика",
    "Поставщики": "Закупки · Поставщики",
    "Закупки": "Закупки · Планирование",
    "План / факт": "Планирование · План / факт",
    "Данные": "Данные и отчёты",
    "Управление": "Настройки и доступ",
    "Маржинальность": "Маржинальность",
    "Сравнение месяцев": "Сравнение месяцев",
    "Расширенный": "Расширенный анализ",
}


def sync_session_value(source_key: str, target_key: str) -> None:
    value = st.session_state.get(source_key)
    if value is not None:
        st.session_state[target_key] = value


def initialize_synchronized_state(
    primary_key: str,
    mobile_key: str,
    options: list[str],
) -> str:
    if not options:
        return ""

    current_value = st.session_state.get(primary_key)
    if current_value not in options:
        current_value = st.session_state.get(mobile_key)
    if current_value not in options:
        current_value = options[0]

    st.session_state[primary_key] = current_value
    st.session_state[mobile_key] = current_value
    return str(current_value)


def render_screen_switcher(
    title: str,
    options: list[str],
    *,
    key: str,
    description: str = "",
    sync_key: str | None = None,
) -> str:
    if not options:
        return ""
    if key in st.session_state and st.session_state[key] not in options:
        st.session_state[key] = options[0]

    st.markdown(f'<div class="sidebar-section-label">{escape(title)}</div>', unsafe_allow_html=True)
    current_value = st.session_state.get(key, options[0])
    widget_index = None if key in st.session_state else options.index(current_value)

    return str(
        st.radio(
            title,
            options=options,
            index=widget_index,
            format_func=lambda option: SCREEN_NAV_LABELS.get(str(option), str(option)),
            horizontal=False,
            label_visibility="collapsed",
            key=key,
            on_change=sync_session_value if sync_key else None,
            args=(key, sync_key) if sync_key else None,
        )
    )


def render_mobile_screen_navigation(
    screen_options: list[str],
    analytics_options: list[str],
    *,
    margin_visible: bool,
) -> tuple[str, str, str, str]:
    active_screen = initialize_synchronized_state(
        "primary_screen_nav",
        "mobile_primary_screen_nav",
        screen_options,
    )
    active_analytics_screen = ""
    active_advanced_screen = ""
    abc_metric = "revenue"

    with st.container(key="mobile_screen_navigation"):
        st.markdown(
            '<div class="mobile-nav-heading"><div class="mobile-nav-title">Текущий раздел</div>'
            '<div class="mobile-nav-hint">☰ режим и настройки</div></div>',
            unsafe_allow_html=True,
        )
        active_screen = str(
            st.selectbox(
                "Текущий раздел",
                options=screen_options,
                format_func=lambda option: SCREEN_NAV_LABELS.get(str(option), str(option)),
                key="mobile_primary_screen_nav",
                label_visibility="collapsed",
                on_change=sync_session_value,
                args=("mobile_primary_screen_nav", "primary_screen_nav"),
            )
        )

        if active_screen == "Аналитика":
            initialize_synchronized_state(
                "analytics_screen_nav",
                "mobile_analytics_screen_nav",
                analytics_options,
            )
            active_analytics_screen = str(
                st.selectbox(
                    "Раздел аналитики",
                    options=analytics_options,
                    format_func=lambda option: SCREEN_NAV_LABELS.get(str(option), str(option)),
                    key="mobile_analytics_screen_nav",
                    on_change=sync_session_value,
                    args=("mobile_analytics_screen_nav", "analytics_screen_nav"),
                )
            )

            if active_analytics_screen == "Расширенный":
                advanced_options = ["RFM-анализ", "Тепловая карта", "Прогноз", "Возвраты"]
                initialize_synchronized_state(
                    "advanced_screen_nav",
                    "mobile_advanced_screen_nav",
                    advanced_options,
                )
                active_advanced_screen = str(
                    st.selectbox(
                        "Глубокий разбор",
                        options=advanced_options,
                        key="mobile_advanced_screen_nav",
                        on_change=sync_session_value,
                        args=("mobile_advanced_screen_nav", "advanced_screen_nav"),
                    )
                )

        if active_screen == "ABC-анализ":
            abc_metric_options = {"revenue": "Выручка", "quantity": "Количество"}
            if margin_visible:
                abc_metric_options["margin"] = "Маржа"
            abc_metric_keys = list(abc_metric_options)
            initialize_synchronized_state(
                "abc_metric_sidebar",
                "mobile_abc_metric",
                abc_metric_keys,
            )
            abc_metric = str(
                st.selectbox(
                    "Метрика ABC",
                    options=abc_metric_keys,
                    format_func=abc_metric_options.get,
                    key="mobile_abc_metric",
                    on_change=sync_session_value,
                    args=("mobile_abc_metric", "abc_metric_sidebar"),
                )
            )

    return active_screen, active_analytics_screen, active_advanced_screen, abc_metric


MOBILE_BOTTOM_NAV_ITEMS = (
    ("Обзор", "Главная", ":material/home:"),
    ("ABC-анализ", "ABC", ":material/donut_large:"),
    ("Закупки", "Закупки", ":material/inventory_2:"),
    ("Данные", "Данные", ":material/upload_file:"),
)


def set_mobile_navigation_target(target_screen: str) -> None:
    st.session_state["primary_screen_nav"] = target_screen
    st.session_state["mobile_primary_screen_nav"] = target_screen


def render_mobile_bottom_navigation(active_screen: str, screen_options: list[str]) -> None:
    available_items = [item for item in MOBILE_BOTTOM_NAV_ITEMS if item[0] in screen_options]
    if not available_items:
        return

    with st.container(key="mobile_bottom_navigation"):
        columns = st.columns(len(available_items), gap="small")
        for column, (target_screen, label, icon) in zip(columns, available_items):
            with column:
                st.button(
                    label,
                    key=f"mobile_bottom_{target_screen}",
                    type="primary" if active_screen == target_screen else "tertiary",
                    icon=icon,
                    on_click=set_mobile_navigation_target,
                    args=(target_screen,),
                    use_container_width=True,
                )


ALLOWED_UPLOAD_EXTENSIONS = {".csv", ".xls", ".xlsx"}
DEFAULT_MAX_UPLOAD_SIZE_MB = 200


def get_max_upload_size_mb() -> int:
    raw_value = os.getenv("APP_MAX_UPLOAD_SIZE_MB") or os.getenv("STREAMLIT_SERVER_MAX_UPLOAD_SIZE")
    if not raw_value:
        return DEFAULT_MAX_UPLOAD_SIZE_MB

    try:
        parsed_value = int(float(raw_value))
    except (TypeError, ValueError):
        return DEFAULT_MAX_UPLOAD_SIZE_MB

    return max(1, parsed_value)


MAX_UPLOAD_SIZE_MB = get_max_upload_size_mb()
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024


def validate_uploaded_file(file_bytes: bytes, filename: str) -> str:
    normalized_name = Path(filename or "").name.strip()
    if not normalized_name:
        raise ValueError("Укажите файл с корректным именем.")
    if len(normalized_name) > 255:
        raise ValueError("Имя файла слишком длинное. Используйте имя короче 255 символов.")
    extension = Path(normalized_name).suffix.casefold()
    if extension not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError("Поддерживаются только файлы CSV, XLS и XLSX.")
    if not file_bytes:
        raise ValueError("Файл пустой. Загрузите выгрузку с данными.")
    if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
        raise ValueError(f"Файл слишком большой. Максимальный размер загрузки: {MAX_UPLOAD_SIZE_MB} МБ.")

    header = file_bytes[:8]
    if extension == ".xlsx" and not header.startswith(b"PK"):
        raise ValueError("Файл XLSX повреждён или не похож на Excel Open XML.")
    if extension == ".xls" and not header.startswith(b"\xD0\xCF\x11\xE0"):
        raise ValueError("Файл XLS повреждён или не похож на классический Excel.")
    if extension == ".csv" and b"\x00" in file_bytes[:4096]:
        raise ValueError("CSV содержит бинарные данные. Проверьте формат файла.")

    return normalized_name


def save_upload_with_feedback(
    *,
    file_bytes: bytes,
    filename: str,
    salon_name: str,
    report_date: date,
    mapping: dict[str, str | None],
    csv_separator: str,
    csv_encoding: str,
    sheet_name: str | int | None,
    replace_existing: bool,
    actor_username: str = "",
) -> None:
    save_salon(salon_name)
    save_result = register_upload(
        file_bytes=file_bytes,
        filename=filename,
        salon=salon_name,
        report_date=report_date,
        mapping=mapping,
        csv_separator=csv_separator,
        csv_encoding=csv_encoding,
        sheet_name=sheet_name,
        replace_existing=replace_existing,
    )
    st.cache_data.clear()
    replaced_text = ""
    if save_result["replaced"]:
        replaced_text = f" Заменено файлов за дату: {save_result['replaced']}."
    st.session_state["upload_flash_message"] = f"Файл «{filename}» сохранён в архив салона «{salon_name}».{replaced_text}"
    st.session_state["upload_last_saved"] = {
        "salon": salon_name,
        "report_date": report_date.isoformat(),
        "filename": filename,
    }
    audit_event(
        action="upload.save",
        user_id=actor_username,
        details={
            "salon": salon_name,
            "report_date": report_date.isoformat(),
            "filename": filename,
            "replace_existing": replace_existing,
            "replaced_count": int(save_result["replaced"]),
            "upload_id": str(save_result["record"].get("upload_id", "")),
        },
    )


def _query_param_text(value: object) -> str:
    if isinstance(value, list):
        return str(value[0]).strip() if value else ""
    return str(value).strip() if value is not None else ""


def get_request_ip() -> str:
    try:
        headers = getattr(st.context, "headers", None)
    except Exception:
        return ""
    if not headers:
        return ""

    forwarded_for = _query_param_text(headers.get("X-Forwarded-For", "") or headers.get("x-forwarded-for", ""))
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()

    for header_name in ("X-Real-IP", "x-real-ip", "X-Forwarded-Client-Ip", "x-forwarded-client-ip"):
        header_value = _query_param_text(headers.get(header_name, ""))
        if header_value:
            return header_value.strip()
    return ""


def audit_event(*, action: str, user_id: str, details: dict[str, object] | None = None) -> None:
    normalized_user_id = user_id.strip()
    if not normalized_user_id:
        return
    try:
        log_audit_event(
            user_id=normalized_user_id,
            action=action,
            ip=get_request_ip(),
            details=details or {},
        )
    except Exception as error:
        print(f"audit log error [{action}]: {error}")


def get_persistent_auth_token() -> str:
    return _query_param_text(st.query_params.get("auth", ""))


def set_persistent_auth_token(token: str) -> None:
    token = token.strip()
    if token:
        st.query_params["auth"] = token


def clear_persistent_auth_token() -> None:
    if "auth" in st.query_params:
        del st.query_params["auth"]


def build_safe_marker_size(series: pd.Series, *, absolute: bool = False) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce").fillna(0.0)
    if absolute:
        return numeric.abs()
    return numeric.clip(lower=0)


def format_plan_scope_label(salon_name: str) -> str:
    return "Вся сеть" if not str(salon_name).strip() else str(salon_name).strip()


def sum_plan_metric(series: pd.Series) -> float:
    numeric = pd.to_numeric(series, errors="coerce")
    return float(numeric.sum(min_count=1)) if numeric.notna().any() else float("nan")


def build_scope_plan_summary(
    plans_frame: pd.DataFrame,
    scope_salons: list[str],
    *,
    allow_network_fallback: bool = False,
) -> pd.DataFrame:
    if plans_frame.empty:
        return pd.DataFrame(columns=["month", "month_label", "revenue_plan", "margin_plan", "quantity_plan"])

    working = plans_frame.copy()
    working["plan_month"] = pd.to_datetime(working["plan_month"], errors="coerce").dt.normalize()
    working = working.dropna(subset=["plan_month"])
    working["salon"] = working["salon"].fillna("").astype(str).str.strip()

    normalized_scope = sorted({str(salon).strip() for salon in scope_salons if str(salon).strip()})
    scoped = working[working["salon"].isin(normalized_scope)].copy() if normalized_scope else pd.DataFrame()

    if scoped.empty and allow_network_fallback:
        scoped = working[working["salon"] == ""].copy()
        if scoped.empty:
            return pd.DataFrame(columns=["month", "month_label", "revenue_plan", "margin_plan", "quantity_plan"])
        grouped = scoped[["plan_month", "revenue_plan", "margin_plan", "quantity_plan"]].rename(columns={"plan_month": "month"})
    else:
        if scoped.empty:
            return pd.DataFrame(columns=["month", "month_label", "revenue_plan", "margin_plan", "quantity_plan"])
        grouped = (
            scoped.groupby("plan_month", as_index=False)
            .agg(
                revenue_plan=("revenue_plan", sum_plan_metric),
                margin_plan=("margin_plan", sum_plan_metric),
                quantity_plan=("quantity_plan", sum_plan_metric),
            )
            .rename(columns={"plan_month": "month"})
        )

    grouped["month_label"] = pd.to_datetime(grouped["month"], errors="coerce").dt.strftime("%Y-%m")
    return grouped[["month", "month_label", "revenue_plan", "margin_plan", "quantity_plan"]].sort_values("month").reset_index(drop=True)


def build_scope_plan_records(plans_frame: pd.DataFrame, month_label: str, scope_salons: list[str]) -> pd.DataFrame:
    if plans_frame.empty:
        return pd.DataFrame(columns=["salon", "revenue_plan", "margin_plan", "quantity_plan"])

    working = plans_frame.copy()
    working["plan_month"] = pd.to_datetime(working["plan_month"], errors="coerce").dt.strftime("%Y-%m")
    working["salon"] = working["salon"].fillna("").astype(str).str.strip()
    normalized_scope = sorted({str(salon).strip() for salon in scope_salons if str(salon).strip()})
    scoped = working[(working["plan_month"] == month_label) & (working["salon"].isin(normalized_scope))].copy()
    if scoped.empty:
        return pd.DataFrame(columns=["salon", "revenue_plan", "margin_plan", "quantity_plan"])
    return scoped[["salon", "revenue_plan", "margin_plan", "quantity_plan"]].reset_index(drop=True)


def get_plan_record(plans_frame: pd.DataFrame, month_label: str, salon_name: str) -> dict[str, object] | None:
    if plans_frame.empty:
        return None
    working = plans_frame.copy()
    working["plan_month_label"] = pd.to_datetime(working["plan_month"], errors="coerce").dt.strftime("%Y-%m")
    normalized_salon = str(salon_name).strip()
    mask = (
        working["plan_month_label"].astype(str) == str(month_label).strip()
    ) & (working["salon"].fillna("").astype(str).str.strip() == normalized_salon)
    if not mask.any():
        return None
    return working.loc[mask].iloc[-1].to_dict()


def render_user_strip(current_user: dict[str, str]) -> None:
    salon_text = current_user.get("salon") or "Вся сеть"
    contact = current_user.get("email") or current_user.get("phone") or current_user["username"]
    render_html_block(
        f"""
        <div class="user-strip">
            <div>
                <div class="user-name">{escape(current_user['display_name'])}</div>
                <div class="user-meta">Контакт: {escape(contact)} | Контур: {escape(salon_text)}</div>
            </div>
            <div class="role-pill {escape(current_user['role'])}">{escape(role_label(current_user['role']))}</div>
        </div>
        """
    )


def render_sidebar_navigation(current_user: dict[str, str], work_mode: str) -> None:
    if current_user["role"] == "salon":
        items = [
            ("Мой архив", work_mode == "Архив салона"),
            ("Ежедневная загрузка", work_mode == "Новая выгрузка"),
            ("Локальная аналитика", work_mode in {"Архив салона", "Новая выгрузка"}),
        ]
    elif current_user["role"] == "admin":
        items = [
            ("Сеть салонов", work_mode == "Сводка по сети"),
            ("Загрузка по салонам", work_mode == "Загрузка салона"),
            ("Разовый анализ", work_mode == "Разовая загрузка"),
            ("Управление", True),
        ]
    else:
        items = [
            ("Сеть салонов", work_mode == "Сводка по сети"),
            ("Загрузка по салонам", work_mode == "Загрузка салона"),
            ("Разовый анализ", work_mode == "Разовая загрузка"),
            ("Сводный контроль", True),
        ]

    items_html = "".join(
        f'<div class="nav-item{" active" if active else ""}">{escape(label)}</div>' for label, active in items
    )
    render_html_block(
        f"""
        <div class="nav-shell">
            <div class="nav-title">Рабочие зоны</div>
            {items_html}
        </div>
        """
    )


def generate_temp_password(length: int = 10) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def render_sidebar_admin_quick_actions(current_user: dict[str, str], registered_salons: list[str]) -> None:
    apply_pending_widget_resets()

    created_credentials = st.session_state.get("sidebar_created_credentials")
    if created_credentials:
        st.success("Пользователь создан. Эти данные можно сразу отправить сотруднику.")
        credentials_text = "\n".join(
            [
                f"Роль: {created_credentials['role_label']}",
                f"Салон: {created_credentials['salon'] or 'Без привязки'}",
                f"Логин: {created_credentials['username']}",
                f"Пароль: {created_credentials['password']}",
                f"Контакт: {created_credentials['contact'] or 'Не указан'}",
            ]
        )
        st.code(credentials_text, language="text")
        if st.button("Очистить карточку доступа", key="sidebar_clear_created_credentials", use_container_width=True):
            st.session_state.pop("sidebar_created_credentials", None)
            st.rerun()

    current_salons = load_salons()

    with st.expander("Добавить салон", expanded=False):
        new_salon_name = st.text_input(
            "Название салона",
            key="sidebar_new_salon_name",
            placeholder="Например: Бишкек | Азия Молл",
        )
        if st.button("Сохранить салон", key="sidebar_add_salon_button", use_container_width=True):
            normalized_salon_name = new_salon_name.strip()
            existing_salons = {salon.casefold() for salon in current_salons}

            if not normalized_salon_name:
                st.error("Введите название салона.")
            elif normalized_salon_name.casefold() in existing_salons:
                st.warning("Такой салон уже существует.")
            else:
                save_salon(normalized_salon_name)
                audit_event(
                    action="salon.create",
                    user_id=current_user["username"],
                    details={
                        "salon": normalized_salon_name,
                        "source": "sidebar_quick_action",
                    },
                )
                schedule_widget_reset({"sidebar_new_salon_name": ""})
                st.session_state["admin_flash_message"] = f"Салон «{normalized_salon_name}» добавлен."
                st.rerun()

    with st.expander("Создать пользователя", expanded=False):
        st.caption("Хотя бы один контакт обязателен: email или телефон.")

        role_choice = st.selectbox(
            "Роль",
            options=["salon", "manager", "admin"],
            format_func=role_label,
            key="sidebar_create_role",
        )
        username = st.text_input(
            "Логин",
            key="sidebar_create_username",
            placeholder="Например: salon_mega",
        )
        display_name = st.text_input(
            "Имя в системе",
            key="sidebar_create_display_name",
            placeholder="Например: Салон Мега",
        )
        email = st.text_input(
            "Email",
            key="sidebar_create_email",
            placeholder="user@company.ru",
        )
        phone = st.text_input(
            "Телефон",
            key="sidebar_create_phone",
            placeholder="+996 555 123 456",
        )

        selected_salon = ""
        if role_choice == "salon":
            salon_options = current_salons if current_salons else ["Сначала добавьте салон"]
            selected_salon = st.selectbox(
                "Салон",
                options=salon_options,
                key="sidebar_create_salon",
                disabled=not bool(current_salons),
            )
        else:
            st.caption("Для администратора и руководителя выбор салона не нужен.")

        generate_password_col, _ = st.columns([1, 1])
        with generate_password_col:
            if st.button("Сгенерировать пароль", key="sidebar_generate_password", use_container_width=True):
                generated_password = generate_temp_password()
                st.session_state["sidebar_create_password"] = generated_password
                st.session_state["sidebar_create_password_confirm"] = generated_password
                st.rerun()

        password = st.text_input("Пароль", type="password", key="sidebar_create_password")
        password_confirm = st.text_input("Повтор пароля", type="password", key="sidebar_create_password_confirm")

        create_user_disabled = role_choice == "salon" and not bool(current_salons)
        if create_user_disabled:
            st.warning("Сначала добавьте салон, потом можно создать пользователя салона.")

        if st.button(
            "Создать пользователя",
            key="sidebar_create_user_button",
            use_container_width=True,
            disabled=create_user_disabled,
        ):
            if password != password_confirm:
                st.error("Пароли не совпадают.")
            else:
                try:
                    create_user(
                        username=username,
                        password=password,
                        role=role_choice,
                        display_name=display_name,
                        email=email,
                        phone=phone,
                        salon=selected_salon,
                    )
                    st.session_state["sidebar_created_credentials"] = {
                        "role_label": role_label(role_choice),
                        "salon": selected_salon if role_choice == "salon" else "",
                        "username": username.strip(),
                        "password": password,
                        "contact": email.strip() or phone.strip(),
                    }
                    audit_event(
                        action="user.create",
                        user_id=current_user["username"],
                        details={
                            "username": username.strip(),
                            "role": role_choice,
                            "salon": selected_salon if role_choice == "salon" else "",
                            "has_email": bool(email.strip()),
                            "has_phone": bool(phone.strip()),
                            "source": "sidebar_quick_action",
                        },
                    )
                    schedule_widget_reset(
                        {
                            "sidebar_create_username": "",
                            "sidebar_create_display_name": "",
                            "sidebar_create_email": "",
                            "sidebar_create_phone": "",
                            "sidebar_create_password": "",
                            "sidebar_create_password_confirm": "",
                            "sidebar_create_salon": None,
                            "sidebar_create_role": None,
                        }
                    )
                    st.session_state["admin_flash_message"] = (
                        f"Пользователь «{display_name.strip() or username.strip()}» создан."
                    )
                    st.rerun()
                except Exception as error:
                    st.error(str(error))


def render_auth_gate() -> dict[str, str]:
    if "auth_user" in st.session_state:
        session_user = st.session_state["auth_user"]
        actual_user = find_user(str(session_user.get("username", "")))
        if actual_user and bool(actual_user.get("is_active", True)):
            refreshed_user = {
                "username": str(actual_user.get("username", "")),
                "display_name": str(actual_user.get("display_name", "") or actual_user.get("username", "")),
                "role": str(actual_user.get("role", "")),
                "salon": str(actual_user.get("salon", "") or ""),
                "email": str(actual_user.get("email", "") or ""),
                "phone": str(actual_user.get("phone", "") or ""),
                "is_active": bool(actual_user.get("is_active", True)),
                "created_at": str(actual_user.get("created_at", "") or ""),
            }
            st.session_state["auth_user"] = refreshed_user
            if not get_persistent_auth_token():
                session_token = create_auth_session(refreshed_user["username"])
                set_persistent_auth_token(session_token)
            return refreshed_user

        st.session_state.pop("auth_user", None)
        clear_persistent_auth_token()
        st.session_state["auth_notice"] = "Эта учётная запись удалена или отключена. Войдите под другим пользователем."

    persistent_token = get_persistent_auth_token()
    if persistent_token:
        persistent_user = authenticate_session(persistent_token)
        if persistent_user:
            st.session_state["auth_user"] = persistent_user
            return persistent_user
        clear_persistent_auth_token()
        st.session_state["auth_notice"] = "Сессия входа закончилась или стала недействительной. Войдите снова."

    left_col, right_col = st.columns([1.05, 0.95], gap="medium")

    with left_col:
        st.markdown(
            """
            <div class="login-shell">
                <div class="auth-kicker">Retail Intelligence</div>
                <h1 class="auth-title">Система продаж для сети салонов</h1>
                <p class="auth-copy">
                    Каждый салон видит только свой контур продаж и ежедневные выгрузки.
                    Руководитель получает общую картину по сети, сравнение салонов и единую управленческую панель.
                </p>
                <div class="feature-list">
                    <div class="feature-chip">Ежедневная загрузка файлов из 1С по салонам</div>
                    <div class="feature-chip">Роли доступа: салон и руководитель</div>
                    <div class="feature-chip">Сводка по сети, ABC, маржинальность и динамика</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with right_col:
        auth_notice = st.session_state.pop("auth_notice", "")
        if auth_notice:
            st.warning(auth_notice)

        if not has_users():
            st.subheader("Первичная настройка")
            st.caption("Создайте первого пользователя-руководителя. Для входа потребуется email или телефон.")
            with st.form("bootstrap_admin_form"):
                username = st.text_input("Логин руководителя")
                display_name = st.text_input("Имя")
                email = st.text_input("Email")
                phone = st.text_input("Телефон")
                password = st.text_input("Пароль", type="password")
                password_confirm = st.text_input("Подтверждение пароля", type="password")
                submitted = st.form_submit_button("Создать руководителя", use_container_width=True)

            if submitted:
                if password != password_confirm:
                    st.error("Пароли не совпадают.")
                else:
                    try:
                        user = bootstrap_first_admin(
                            username=username,
                            password=password,
                            display_name=display_name,
                            email=email,
                            phone=phone,
                        )
                        st.session_state["auth_user"] = user
                        set_persistent_auth_token(create_auth_session(user["username"]))
                        audit_event(
                            action="auth.bootstrap_admin",
                            user_id=user["username"],
                            details={
                                "role": user["role"],
                                "has_email": bool(user.get("email")),
                                "has_phone": bool(user.get("phone")),
                            },
                        )
                        audit_event(
                            action="auth.login",
                            user_id=user["username"],
                            details={
                                "method": "bootstrap",
                                "role": user["role"],
                            },
                        )
                        st.success("Руководитель создан. Выполняю вход.")
                        st.rerun()
                    except Exception as error:
                        st.error(str(error))
            st.stop()

        st.subheader("Вход в систему")
        with st.form("login_form"):
            identifier = st.text_input("Логин, email или телефон")
            password = st.text_input("Пароль", type="password")
            submitted = st.form_submit_button("Войти", use_container_width=True)

        if submitted:
            user = authenticate_user(identifier, password)
            if not user:
                st.error("Неверный логин, email, телефон или пароль.")
            else:
                st.session_state["auth_user"] = user
                set_persistent_auth_token(create_auth_session(user["username"]))
                audit_event(
                    action="auth.login",
                    user_id=user["username"],
                    details={
                        "method": "password",
                        "role": user["role"],
                    },
                )
                st.rerun()

    st.stop()


@st.cache_data(show_spinner=False)
def cached_load_input_file(
    file_bytes: bytes,
    filename: str,
    csv_separator: str,
    csv_encoding: str,
    sheet_name: str | int | None,
) -> pd.DataFrame:
    return load_input_file(
        file_bytes,
        filename,
        csv_separator=csv_separator,
        csv_encoding=csv_encoding,
        sheet_name=sheet_name,
    )


@st.cache_data(show_spinner=False)
def cached_prepare_sales_data(
    frame: pd.DataFrame,
    mapping_items: tuple[tuple[str, str | None], ...],
    supplier_rule_items: tuple[tuple[str, str], ...] = (),
):
    mapping = dict(mapping_items)
    return prepare_sales_data(frame, mapping, supplier_rules=supplier_rule_items)


@st.cache_data(show_spinner=False)
def cached_prepare_inventory_data(
    frame: pd.DataFrame,
    mapping_items: tuple[tuple[str, str | None], ...],
):
    mapping = dict(mapping_items)
    return prepare_inventory_data(frame, mapping)


@st.cache_data(show_spinner=False)
def cached_load_inventory_input_file(
    file_bytes: bytes,
    filename: str,
    csv_separator: str,
    csv_encoding: str,
    sheet_name: str | int | None,
) -> pd.DataFrame:
    return load_inventory_input_file(
        file_bytes,
        filename,
        csv_separator=csv_separator,
        csv_encoding=csv_encoding,
        sheet_name=sheet_name,
    )


@st.cache_data(show_spinner=False)
def cached_load_archive_data(
    selected_salons: tuple[str, ...],
    supplier_rule_items: tuple[tuple[str, str], ...] = (),
):
    return load_archive_data(salons=list(selected_salons), supplier_rules=supplier_rule_items)


@st.cache_data(show_spinner=False)
def cached_load_supplier_keyword_rules(include_inactive: bool = False) -> pd.DataFrame:
    return load_supplier_keyword_rules(include_inactive=include_inactive)


@st.cache_data(show_spinner=False)
def cached_load_supplier_product_assignments() -> pd.DataFrame:
    return load_supplier_product_assignments()


def supplier_rule_items_from_frame(frame: pd.DataFrame) -> tuple[tuple[str, str], ...]:
    if frame.empty or not {"supplier", "keyword"}.issubset(frame.columns):
        return ()

    if "is_active" in frame.columns:
        active_frame = frame[frame["is_active"].fillna(True).astype(bool)].copy()
    else:
        active_frame = frame.copy()

    items = []
    seen_keywords: set[str] = set()
    for row in active_frame.to_dict(orient="records"):
        supplier = str(row.get("supplier", "")).strip()
        keyword = str(row.get("keyword", "")).strip()
        keyword_key = keyword.casefold()
        if supplier and keyword and keyword_key not in seen_keywords:
            seen_keywords.add(keyword_key)
            items.append((supplier, keyword))
    return tuple(items)


def enrich_sales_with_supplier(
    data: pd.DataFrame,
    procurement_items: pd.DataFrame,
    supplier_rule_items: tuple[tuple[str, str], ...] = (),
    supplier_product_assignments: pd.DataFrame | None = None,
) -> pd.DataFrame:
    if data.empty:
        return data

    enriched = data.copy()
    if "supplier" not in enriched.columns:
        enriched["supplier"] = ""

    supplier_text = enriched["supplier"].fillna("").astype(str).str.strip()
    missing_supplier = supplier_text.eq("") | supplier_text.str.casefold().eq("не назначен")
    enriched["supplier"] = supplier_text

    if not procurement_items.empty and {"product", "supplier"}.issubset(procurement_items.columns):
        supplier_lookup_source = procurement_items.copy()
        supplier_lookup_source["product_lookup_key"] = (
            supplier_lookup_source["product"].fillna("").astype(str).str.strip().str.casefold()
        )
        supplier_lookup_source["supplier"] = supplier_lookup_source["supplier"].fillna("").astype(str).str.strip()
        supplier_lookup_source = supplier_lookup_source[
            supplier_lookup_source["product_lookup_key"].ne("") & supplier_lookup_source["supplier"].ne("")
        ]

        if not supplier_lookup_source.empty:
            supplier_lookup = (
                supplier_lookup_source.drop_duplicates("product_lookup_key", keep="last")
                .set_index("product_lookup_key")["supplier"]
                .to_dict()
            )
            product_lookup = enriched["product"].fillna("").astype(str).str.strip().str.casefold()
            mapped_supplier = product_lookup.map(supplier_lookup)

            if "product_key" in enriched.columns:
                product_key_lookup = enriched["product_key"].fillna("").astype(str).str.strip().str.casefold()
                mapped_supplier = mapped_supplier.fillna(product_key_lookup.map(supplier_lookup))

            fill_mask = missing_supplier & mapped_supplier.fillna("").astype(str).str.strip().ne("")
            enriched.loc[fill_mask, "supplier"] = mapped_supplier.loc[fill_mask].astype(str).str.strip()

    supplier_text = enriched["supplier"].fillna("").astype(str).str.strip()
    missing_supplier = supplier_text.eq("") | supplier_text.str.casefold().eq("не назначен")
    supplier_text_columns = [
        column for column in ("product", "product_key", "item_code", "category") if column in enriched.columns
    ]
    if supplier_text_columns and missing_supplier.any():
        combined_supplier_text = (
            enriched.loc[missing_supplier, supplier_text_columns]
            .fillna("")
            .astype(str)
            .agg(" ".join, axis=1)
        )
        inferred_supplier = combined_supplier_text.map(lambda value: infer_supplier_from_text(value, supplier_rule_items))
        inferred_mask = inferred_supplier.fillna("").astype(str).str.strip().ne("")
        if inferred_mask.any():
            enriched.loc[inferred_supplier.index[inferred_mask], "supplier"] = (
                inferred_supplier.loc[inferred_mask].astype(str).str.strip()
            )

    if (
        supplier_product_assignments is not None
        and not supplier_product_assignments.empty
        and {"product_key", "supplier"}.issubset(supplier_product_assignments.columns)
    ):
        assignment_source = supplier_product_assignments.copy()
        assignment_source["assignment_key"] = (
            assignment_source["product_key"].fillna("").astype(str).str.strip().str.casefold()
        )
        assignment_source["supplier"] = assignment_source["supplier"].fillna("").astype(str).str.strip()
        assignment_source = assignment_source[
            assignment_source["assignment_key"].ne("") & assignment_source["supplier"].ne("")
        ]
        if not assignment_source.empty:
            assignment_lookup = (
                assignment_source.drop_duplicates("assignment_key", keep="last")
                .set_index("assignment_key")["supplier"]
                .to_dict()
            )
            if "product_key" in enriched.columns:
                product_key_lookup = enriched["product_key"].fillna("").astype(str).str.strip().str.casefold()
            else:
                product_key_lookup = enriched["product"].fillna("").astype(str).str.strip().str.casefold()
            assigned_supplier = product_key_lookup.map(assignment_lookup)

            product_lookup = enriched["product"].fillna("").astype(str).str.strip().str.casefold()
            assigned_supplier = assigned_supplier.fillna(product_lookup.map(assignment_lookup))
            assigned_mask = assigned_supplier.fillna("").astype(str).str.strip().ne("")
            if assigned_mask.any():
                enriched.loc[assigned_mask, "supplier"] = assigned_supplier.loc[assigned_mask].astype(str).str.strip()

    enriched["supplier"] = enriched["supplier"].fillna("").astype(str).str.strip().replace("", "Не назначен")
    return enriched


def enrich_sales_with_brand(
    data: pd.DataFrame,
    procurement_items: pd.DataFrame,
) -> pd.DataFrame:
    if data.empty:
        return data

    enriched = data.copy()
    enriched["brand"] = (
        enriched.get("brand", pd.Series("", index=enriched.index))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    if procurement_items.empty or not {"product", "brand"}.issubset(procurement_items.columns):
        enriched["brand"] = enriched["brand"].replace("", "Не назначен")
        return enriched

    brand_lookup_source = procurement_items[["product", "brand"]].copy()
    brand_lookup_source["product_lookup_key"] = (
        brand_lookup_source["product"].fillna("").astype(str).str.strip().str.casefold()
    )
    brand_lookup_source["brand"] = (
        brand_lookup_source["brand"].fillna("").astype(str).str.strip()
    )
    brand_lookup_source = brand_lookup_source[
        brand_lookup_source["product_lookup_key"].ne("")
        & brand_lookup_source["brand"].ne("")
    ]
    if not brand_lookup_source.empty:
        brand_lookup = (
            brand_lookup_source.drop_duplicates("product_lookup_key", keep="last")
            .set_index("product_lookup_key")["brand"]
            .to_dict()
        )
        product_lookup = enriched["product"].fillna("").astype(str).str.strip().str.casefold()
        mapped_brand = product_lookup.map(brand_lookup)
        if "product_key" in enriched.columns:
            product_key_lookup = (
                enriched["product_key"].fillna("").astype(str).str.strip().str.casefold()
            )
            mapped_brand = mapped_brand.fillna(product_key_lookup.map(brand_lookup))

        missing_brand = enriched["brand"].eq("") | enriched["brand"].str.casefold().eq("не назначен")
        fill_mask = missing_brand & mapped_brand.fillna("").astype(str).str.strip().ne("")
        enriched.loc[fill_mask, "brand"] = mapped_brand.loc[fill_mask].astype(str).str.strip()

    enriched["brand"] = enriched["brand"].fillna("").astype(str).str.strip().replace("", "Не назначен")
    return enriched


def build_missing_supplier_assignment_candidates(data: pd.DataFrame) -> pd.DataFrame:
    columns = ["product_key", "product", "category", "revenue", "quantity", "line_count", "supplier"]
    if data.empty or "supplier" not in data.columns:
        return pd.DataFrame(columns=columns)

    supplier_text = data["supplier"].fillna("").astype(str).str.strip()
    missing_mask = supplier_text.eq("") | supplier_text.str.casefold().eq("не назначен")
    missing_data = data.loc[missing_mask].copy()
    if missing_data.empty:
        return pd.DataFrame(columns=columns)

    if "product_key" not in missing_data.columns:
        missing_data["product_key"] = missing_data["product"].fillna("").astype(str).str.strip()
    if "category" not in missing_data.columns:
        missing_data["category"] = ""

    summary = (
        missing_data.groupby("product_key", dropna=False)
        .agg(
            product=("product", lambda series: next((str(value).strip() for value in series if str(value).strip()), "")),
            category=("category", lambda series: next((str(value).strip() for value in series if str(value).strip()), "")),
            revenue=("revenue", "sum"),
            quantity=("quantity", "sum"),
            line_count=("product", "size"),
        )
        .reset_index()
    )
    summary["supplier"] = ""
    summary = summary.sort_values("revenue", ascending=False).reset_index(drop=True)
    return summary[columns]


def _first_non_empty_text(series: pd.Series) -> str:
    for value in series:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _most_common_non_empty_text(series: pd.Series, default: str = "Не назначен") -> str:
    values = series.fillna("").astype(str).str.strip()
    values = values[values.ne("")]
    if values.empty:
        return default
    return str(values.value_counts().index[0]).strip() or default


def build_supplier_assignment_catalog(
    data: pd.DataFrame,
    supplier_product_assignments: pd.DataFrame | None = None,
) -> pd.DataFrame:
    columns = [
        "assign",
        "product_key",
        "product",
        "category",
        "current_supplier",
        "manual_supplier",
        "new_supplier",
        "revenue",
        "quantity",
        "line_count",
        "assignment_status",
    ]
    if data.empty or "product" not in data.columns:
        return pd.DataFrame(columns=columns)

    working = data.copy()
    if "product_key" not in working.columns:
        working["product_key"] = working["product"].fillna("").astype(str).str.strip()
    if "category" not in working.columns:
        working["category"] = ""
    if "supplier" not in working.columns:
        working["supplier"] = "Не назначен"
    if "revenue" not in working.columns:
        working["revenue"] = 0.0
    if "quantity" not in working.columns:
        working["quantity"] = 0.0

    working["product_key"] = working["product_key"].fillna("").astype(str).str.strip()
    working = working[working["product_key"].ne("")]
    if working.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        working.groupby("product_key", dropna=False)
        .agg(
            product=("product", _first_non_empty_text),
            category=("category", _first_non_empty_text),
            current_supplier=("supplier", _most_common_non_empty_text),
            revenue=("revenue", "sum"),
            quantity=("quantity", "sum"),
            line_count=("product", "size"),
        )
        .reset_index()
    )

    summary["manual_supplier"] = ""
    if (
        supplier_product_assignments is not None
        and not supplier_product_assignments.empty
        and {"product_key", "supplier"}.issubset(supplier_product_assignments.columns)
    ):
        assignment_source = supplier_product_assignments.copy()
        assignment_source["assignment_key"] = (
            assignment_source["product_key"].fillna("").astype(str).str.strip().str.casefold()
        )
        assignment_source["supplier"] = assignment_source["supplier"].fillna("").astype(str).str.strip()
        assignment_source = assignment_source[
            assignment_source["assignment_key"].ne("") & assignment_source["supplier"].ne("")
        ]
        if not assignment_source.empty:
            assignment_lookup = (
                assignment_source.drop_duplicates("assignment_key", keep="last")
                .set_index("assignment_key")["supplier"]
                .to_dict()
            )
            summary["manual_supplier"] = (
                summary["product_key"].fillna("").astype(str).str.strip().str.casefold().map(assignment_lookup)
            ).fillna("")

    current_supplier = summary["current_supplier"].fillna("").astype(str).str.strip()
    manual_supplier = summary["manual_supplier"].fillna("").astype(str).str.strip()
    missing_supplier = current_supplier.eq("") | current_supplier.str.casefold().eq("не назначен")
    summary["assignment_status"] = "Есть поставщик"
    summary.loc[missing_supplier, "assignment_status"] = "Без поставщика"
    summary.loc[manual_supplier.ne(""), "assignment_status"] = "Назначен вручную"
    summary["assign"] = False
    summary["new_supplier"] = ""
    summary = summary.sort_values(
        by=["assignment_status", "revenue"],
        ascending=[True, False],
        key=lambda series: series.map({"Без поставщика": 0, "Есть поставщик": 1, "Назначен вручную": 2}).fillna(series)
        if series.name == "assignment_status"
        else series,
    ).reset_index(drop=True)
    return summary[columns]


def build_supplier_name_options(*frames: pd.DataFrame) -> list[str]:
    supplier_names: set[str] = set()
    for frame in frames:
        if frame is None or frame.empty or "supplier" not in frame.columns:
            continue
        for value in frame["supplier"].dropna().astype(str):
            supplier = value.strip()
            if supplier and supplier.casefold() != "не назначен":
                supplier_names.add(supplier)
    return sorted(supplier_names, key=str.casefold)


def _catalog_text(value: object) -> str:
    if is_missing(value):
        return ""
    return str(value).strip()


def _first_catalog_text(series: pd.Series) -> str:
    for value in series:
        text = _catalog_text(value)
        if text:
            return text
    return ""


def _most_common_catalog_text(series: pd.Series, default: str = "") -> str:
    values = series.fillna("").astype(str).str.strip()
    values = values[values.ne("")]
    if values.empty:
        return default
    return str(values.value_counts().index[0]).strip() or default


def _catalog_missing_text(value: object, *, empty_labels: set[str] | None = None) -> bool:
    text = _catalog_text(value)
    if not text:
        return True
    normalized = text.casefold()
    labels = empty_labels or set()
    return normalized in {label.casefold() for label in labels}


def build_sku_cleanup_queue(
    data: pd.DataFrame,
    procurement_items: pd.DataFrame,
    supplier_product_assignments: pd.DataFrame,
    *,
    require_item_code: bool = False,
) -> pd.DataFrame:
    columns = [
        "issue_count",
        "issues",
        "suggested_action",
        "product_key",
        "product",
        "item_code",
        "category",
        "effective_supplier",
        "sales_supplier",
        "manual_supplier",
        "stock_supplier",
        "revenue",
        "quantity",
        "line_count",
        "stock_on_hand",
        "stock_in_transit",
    ]
    if data.empty or "product" not in data.columns:
        return pd.DataFrame(columns=columns)

    working = data.copy()
    if "product_key" not in working.columns:
        working["product_key"] = working["product"].fillna("").astype(str).str.strip()
    for column in ("item_code", "category", "supplier"):
        if column not in working.columns:
            working[column] = ""
    for column in ("revenue", "quantity"):
        if column not in working.columns:
            working[column] = 0.0
        working[column] = pd.to_numeric(working[column], errors="coerce").fillna(0.0)

    working["product_key"] = working["product_key"].fillna("").astype(str).str.strip()
    working = working[working["product_key"].ne("")]
    if working.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        working.groupby("product_key", dropna=False)
        .agg(
            product=("product", _first_catalog_text),
            item_code=("item_code", _first_catalog_text),
            category=("category", _most_common_catalog_text),
            sales_supplier=("supplier", _most_common_catalog_text),
            revenue=("revenue", "sum"),
            quantity=("quantity", "sum"),
            line_count=("product", "size"),
        )
        .reset_index()
    )

    summary["manual_supplier"] = ""
    if (
        supplier_product_assignments is not None
        and not supplier_product_assignments.empty
        and {"product_key", "supplier"}.issubset(supplier_product_assignments.columns)
    ):
        assignment_source = supplier_product_assignments.copy()
        assignment_source["supplier"] = assignment_source["supplier"].fillna("").astype(str).str.strip()
        assignment_lookup: dict[str, str] = {}
        for row in assignment_source.to_dict(orient="records"):
            supplier = _catalog_text(row.get("supplier"))
            if not supplier:
                continue
            for key_field in ("product_key", "product"):
                key = _catalog_text(row.get(key_field)).casefold()
                if key:
                    assignment_lookup[key] = supplier
        summary["manual_supplier"] = (
            summary["product_key"].fillna("").astype(str).str.strip().str.casefold().map(assignment_lookup)
        ).fillna("")
        product_assignment_supplier = (
            summary["product"].fillna("").astype(str).str.strip().str.casefold().map(assignment_lookup)
        ).fillna("")
        summary["manual_supplier"] = summary["manual_supplier"].where(
            summary["manual_supplier"].ne(""),
            product_assignment_supplier,
        )

    stock_columns = ["stock_supplier", "stock_on_hand", "stock_in_transit", "has_stock_record"]
    for column in stock_columns:
        summary[column] = "" if column == "stock_supplier" else 0.0

    if procurement_items is not None and not procurement_items.empty and "product" in procurement_items.columns:
        stock_source = procurement_items.copy()
        stock_source["stock_product_key"] = stock_source["product"].fillna("").astype(str).str.strip().str.casefold()
        stock_source = stock_source[stock_source["stock_product_key"].ne("")]
        if not stock_source.empty:
            if "supplier" not in stock_source.columns:
                stock_source["supplier"] = ""
            for column in ("stock_on_hand", "stock_in_transit"):
                if column not in stock_source.columns:
                    stock_source[column] = 0.0
                stock_source[column] = pd.to_numeric(stock_source[column], errors="coerce").fillna(0.0)
            stock_source["supplier"] = stock_source["supplier"].fillna("").astype(str).str.strip()
            stock_lookup = (
                stock_source.drop_duplicates("stock_product_key", keep="last")
                .set_index("stock_product_key")[["supplier", "stock_on_hand", "stock_in_transit"]]
                .rename(columns={"supplier": "stock_supplier"})
            )
            summary["stock_lookup_key"] = summary["product"].fillna("").astype(str).str.strip().str.casefold()
            summary = summary.merge(
                stock_lookup,
                left_on="stock_lookup_key",
                right_index=True,
                how="left",
                suffixes=("", "_stock"),
            )
            summary["has_stock_record"] = summary["stock_supplier_stock"].notna() if "stock_supplier_stock" in summary.columns else summary["stock_supplier"].notna()
            for column in ("stock_supplier", "stock_on_hand", "stock_in_transit"):
                stock_column = f"{column}_stock"
                if stock_column in summary.columns:
                    summary[column] = summary[stock_column]
                    summary = summary.drop(columns=[stock_column])
            summary = summary.drop(columns=["stock_lookup_key"], errors="ignore")

    for column in ("sales_supplier", "manual_supplier", "stock_supplier", "category", "item_code", "product"):
        summary[column] = summary[column].fillna("").astype(str).str.strip()
    for column in ("stock_on_hand", "stock_in_transit"):
        summary[column] = pd.to_numeric(summary[column], errors="coerce").fillna(0.0)
    summary["has_stock_record"] = summary["has_stock_record"].fillna(False).astype(bool)

    sales_supplier = summary["sales_supplier"].where(
        ~summary["sales_supplier"].str.casefold().eq("не назначен"),
        "",
    )
    summary["effective_supplier"] = sales_supplier
    summary["effective_supplier"] = summary["effective_supplier"].where(
        summary["effective_supplier"].ne(""),
        summary["manual_supplier"],
    )
    summary["effective_supplier"] = summary["effective_supplier"].where(
        summary["effective_supplier"].ne(""),
        summary["stock_supplier"],
    )

    def _row_issues(row: pd.Series) -> tuple[str, str, int]:
        issues: list[str] = []
        actions: list[str] = []
        if _catalog_missing_text(row.get("effective_supplier"), empty_labels={"Не назначен"}):
            issues.append("Без поставщика")
            actions.append("Назначить поставщика")
        if _catalog_missing_text(row.get("category"), empty_labels={"Без категории", "Не указана"}):
            issues.append("Без категории")
            actions.append("Уточнить категорию")
        if require_item_code and _catalog_missing_text(row.get("item_code")):
            issues.append("Без кода товара")
            actions.append("Добавить артикул/код")
        if not bool(row.get("has_stock_record", False)):
            issues.append("Нет в остатках")
            actions.append("Загрузить или сверить остатки")
        if float(row.get("stock_on_hand", 0.0) or 0.0) <= 0 and bool(row.get("has_stock_record", False)):
            issues.append("Нулевой остаток")
            actions.append("Проверить потребность закупки")

        unique_actions = list(dict.fromkeys(actions))
        return " · ".join(issues), "; ".join(unique_actions), len(issues)

    issue_payload = summary.apply(_row_issues, axis=1)
    summary["issues"] = issue_payload.map(lambda payload: payload[0])
    summary["suggested_action"] = issue_payload.map(lambda payload: payload[1])
    summary["issue_count"] = issue_payload.map(lambda payload: payload[2])
    summary = summary[summary["issue_count"] > 0].copy()
    if summary.empty:
        return pd.DataFrame(columns=columns)

    summary = summary.sort_values(["issue_count", "revenue"], ascending=[False, False]).reset_index(drop=True)
    return summary[columns]


def default_supplier_rules_frame() -> pd.DataFrame:
    rows = [
        {"supplier": supplier, "keyword": keyword}
        for supplier, keywords in SUPPLIER_KEYWORD_PATTERNS
        for keyword in keywords
    ]
    return pd.DataFrame(rows, columns=["supplier", "keyword"])


@st.cache_data(show_spinner=False, max_entries=24)
def cached_build_overview_metrics(frame: pd.DataFrame) -> dict[str, float]:
    return build_overview_metrics(frame)


@st.cache_data(show_spinner=False, max_entries=48)
def cached_build_product_summary(frame: pd.DataFrame, group_column: str = "product") -> pd.DataFrame:
    return build_product_summary(frame, group_column)


@st.cache_data(show_spinner=False, max_entries=48)
def cached_build_monthly_summary(frame: pd.DataFrame) -> pd.DataFrame:
    return build_monthly_summary(frame)


@st.cache_data(show_spinner=False, max_entries=48)
def cached_build_returns_overview(frame: pd.DataFrame) -> dict[str, float]:
    return build_returns_overview(frame)


@st.cache_data(show_spinner=False, max_entries=24)
def cached_build_plan_fact_summary(
    monthly_summary: pd.DataFrame,
    plan_summary: pd.DataFrame,
) -> pd.DataFrame:
    return build_plan_fact_summary(monthly_summary, plan_summary)


@st.cache_data(show_spinner=False, max_entries=24)
def cached_build_procurement_forecast(
    frame: pd.DataFrame,
    *,
    history_months: int,
    coverage_days: int,
    lead_time_days: int,
    safety_days: int,
    min_active_months: int,
    procurement_items: pd.DataFrame,
    inbound_orders: pd.DataFrame,
) -> pd.DataFrame:
    return build_procurement_forecast(
        frame,
        history_months=history_months,
        coverage_days=coverage_days,
        lead_time_days=lead_time_days,
        safety_days=safety_days,
        min_active_months=min_active_months,
        procurement_items=procurement_items,
        inbound_orders=inbound_orders,
    )


@st.cache_data(show_spinner=False, max_entries=24)
def cached_build_procurement_overview(procurement_frame: pd.DataFrame) -> dict[str, float]:
    return build_procurement_overview(procurement_frame)


@st.cache_data(show_spinner=False, max_entries=24)
def cached_build_procurement_supplier_summary(procurement_frame: pd.DataFrame) -> pd.DataFrame:
    return build_procurement_supplier_summary(procurement_frame)


def select_column(
    label: str,
    columns: list[str],
    default_value: str | None,
    *,
    key: str | None = None,
    help_text: str | None = None,
) -> str | None:
    options = ["Не использовать", *columns]
    default_index = options.index(default_value) if default_value in options else 0
    selection = st.selectbox(label, options=options, index=default_index, help=help_text, key=key)
    return None if selection == "Не использовать" else selection


def build_download_frame(selected_mapping: dict[str, str | None]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Поле в приложении": DISPLAY_NAMES[key],
                "Колонка в файле": value or "Не используется",
            }
            for key, value in selected_mapping.items()
        ]
    )


DISPLAY_COLUMN_NAMES = {
    "date": "Дата",
    "Дата": "Дата",
    "metric": "Показатель",
    "amount": "Сумма",
    "comment": "Комментарий",
    "item_code": "Код товара",
    "product_key": "Ключ товара",
    "group_key": "Ключ товара",
    "product": "Товар",
    "product_name": "Название товара",
    "Номенклатура": "Товар",
    "group_name": "Позиция",
    "category": "Категория",
    "Категория": "Категория",
    "manager": "Менеджер",
    "salon": "Салон",
    "revenue": "Выручка",
    "Доход": "Выручка",
    "cost": "Себестоимость",
    "Себестоимость": "Себестоимость",
    "margin": "Маржа",
    "Прибыль": "Маржа",
    "margin_pct": "Маржа, %",
    "quantity": "Количество",
    "Количество": "Количество",
    "month": "Месяц",
    "month_label": "Месяц",
    "product_count": "Товаров",
    "sales_lines": "Строк продаж",
    "line_count": "Строк",
    "supplier": "Поставщик",
    "stock_on_hand": "Остаток",
    "available_stock_qty": "Доступный остаток",
    "stock_coverage_days": "Покрытие, дней",
    "recommended_order_qty": "К заказу",
    "avg_unit_cost": "Средняя себестоимость",
    "stock_value": "Стоимость остатка",
    "available_stock_value": "Стоимость доступного остатка",
    "monthly_cogs_forecast": "Прогноз себестоимости/мес.",
    "abc_basis": "База ABC",
    "abc_class": "Класс ABC",
    "share_pct": "Доля, %",
    "cum_share_pct": "Накопительная доля, %",
    "revenue_change_pct": "Изменение выручки, %",
    "margin_change_pct": "Изменение маржи, %",
    "report_date": "Дата отчёта",
    "source_filename": "Файл",
    "uploaded_at": "Загружен",
    "username": "Логин",
    "display_name": "Имя",
    "role": "Роль",
    "phone": "Телефон",
    "email": "Email",
    "created_at": "Создан",
    "Группа": "Группа",
    "Путь категории": "Путь категории",
    "Скидка": "Скидка",
    "Сумма НДС": "НДС",
    "Сумма НСП": "НСП",
    "Всего": "Итого",
}


def _normalized_label(value: object) -> str:
    return str(value).strip().casefold().replace("_", " ")


def _strip_metric_prefix(label: str) -> str:
    for prefix in ("план ", "факт ", "отклонение ", "плановая ", "фактическая "):
        if label.startswith(prefix):
            return label[len(prefix) :]
    return label


def _is_money_label(label: str) -> bool:
    normalized = _strip_metric_prefix(_normalized_label(label))
    return normalized.startswith(
        (
            "выручка",
            "выручки",
            "валовая выручка",
            "чистая выручка",
            "себестоимость",
            "себестоимости",
            "стоимость",
            "стоимости",
            "маржа",
            "маржи",
            "средняя себестоимость",
            "доход",
            "прибыль",
            "прибыли",
            "сумма",
            "суммы",
            "деньги",
            "прогноз себестоимости",
            "изменение выручки",
            "изменение маржи",
            "ндс",
            "нсп",
            "итого",
        )
    )


def _is_percent_label(label: str) -> bool:
    normalized = _normalized_label(label)
    return "%" in label or normalized.startswith(
        (
            "доля",
            "накопительная доля",
            "маржа, %",
            "изменение выручки, %",
            "изменение маржи, %",
            "изменение количества, %",
        )
    )


def _is_count_label(label: str) -> bool:
    normalized = _strip_metric_prefix(_normalized_label(label))
    if "/" in normalized:
        return False
    if normalized == "sku":
        return True
    return normalized.startswith(
        (
            "количество",
            "количества",
            "остаток",
            "остатка",
            "доступный остаток",
            "покрытие",
            "к заказу",
            "изменение количества",
            "товаров",
            "строк",
            "sku",
        )
    )


def _is_date_label(label: str) -> bool:
    normalized = _normalized_label(label)
    return normalized in {"дата", "дата отчёта", "создан", "загружен"}


def _is_month_label(label: str) -> bool:
    return _normalized_label(label) == "месяц"


def _format_datetime_value(value: object, *, include_time: bool) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return str(value)
    return parsed.strftime("%d.%m.%Y %H:%M" if include_time else "%d.%m.%Y")


def format_date_value(value: object, *, fallback: str = "—") -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return fallback
    return parsed.strftime("%d.%m.%Y")


def format_date_range_values(start_value: object, end_value: object, *, fallback: str = "—") -> str:
    start_text = format_date_value(start_value, fallback="")
    end_text = format_date_value(end_value, fallback="")
    if start_text and end_text:
        return f"{start_text} - {end_text}"
    if start_text:
        return start_text
    if end_text:
        return end_text
    return fallback


def _coerce_float_for_display(value: object) -> float:
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)

    text = str(value).strip()
    if not text:
        raise ValueError("Пустое значение нельзя преобразовать в число.")

    text = (
        text.replace("₽", "")
        .replace("сом.", "")
        .replace("сом", "")
        .replace("KGS", "")
        .replace("kgs", "")
        .replace("%", "")
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    numeric = pd.to_numeric(text, errors="coerce")
    if pd.isna(numeric):
        raise ValueError(f"Не удалось преобразовать значение в число: {value}")

    return float(numeric)


def _format_numeric_value_for_display(value: object, formatter) -> object:
    if is_missing(value):
        return "вЂ”"
    try:
        return formatter(_coerce_float_for_display(value))
    except (TypeError, ValueError):
        return str(value)


def _make_unique_label(label: str, source_column: str, used_labels: dict[str, int]) -> str:
    if label not in used_labels:
        used_labels[label] = 1
        return label

    fallback_label = DISPLAY_COLUMN_NAMES.get(source_column, str(source_column))
    if fallback_label not in used_labels:
        used_labels[fallback_label] = 1
        return fallback_label

    suffix = used_labels[label] + 1
    unique_label = f"{label} ({suffix})"
    while unique_label in used_labels:
        suffix += 1
        unique_label = f"{label} ({suffix})"

    used_labels[label] = suffix
    used_labels[unique_label] = 1
    return unique_label


def format_display_frame(
    frame: pd.DataFrame,
    rename_map: dict[str, str] | None = None,
    *,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    view = frame.copy()
    if columns is not None:
        available_columns = [column for column in columns if column in view.columns]
        view = view[available_columns]

    formatted_columns: dict[str, pd.Series] = {}
    used_labels: dict[str, int] = {}

    for source_column in view.columns:
        if rename_map and source_column in rename_map:
            preferred_label = rename_map[source_column]
        else:
            preferred_label = DISPLAY_COLUMN_NAMES.get(source_column, str(source_column))

        final_label = _make_unique_label(preferred_label, source_column, used_labels)
        series = view[source_column]

        if _is_date_label(final_label):
            include_time = _normalized_label(final_label) in {"создан", "загружен"}
            formatted_columns[final_label] = series.map(
                lambda value: "—" if is_missing(value) else _format_datetime_value(value, include_time=include_time)
            )
            continue

        if _is_month_label(final_label):
            formatted_columns[final_label] = series.map(
                lambda value: "—"
                if is_missing(value)
                else (
                    pd.to_datetime(value, errors="coerce").strftime("%Y-%m")
                    if pd.notna(pd.to_datetime(value, errors="coerce"))
                    else str(value)
                )
            )
            continue

        if _is_percent_label(final_label):
            formatted_columns[final_label] = series.map(
                lambda value: _format_numeric_value_for_display(value, format_percent)
            )
            continue

        if _is_money_label(final_label):
            formatted_columns[final_label] = series.map(
                lambda value: _format_numeric_value_for_display(value, format_money)
            )
            continue

        if _is_count_label(final_label):
            formatted_columns[final_label] = series.map(
                lambda value: _format_numeric_value_for_display(value, format_number)
            )
            continue

        formatted_columns[final_label] = series.map(lambda value: "—" if is_missing(value) else value)

    return pd.DataFrame(formatted_columns, index=view.index)


def to_display_table(frame: pd.DataFrame, rename_map: dict[str, str]) -> pd.DataFrame:
    visible_columns = [column for column in rename_map if column in frame.columns]
    if not visible_columns:
        return format_display_frame(frame)
    return format_display_frame(frame, rename_map, columns=visible_columns)


def format_display_frame_for_role(
    frame: pd.DataFrame,
    current_user: dict[str, str],
    rename_map: dict[str, str] | None = None,
    *,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    safe_rename_map = margin_safe_rename_map(rename_map, current_user)
    safe_columns = margin_safe_columns(columns, current_user) if columns is not None else None
    return format_display_frame(
        margin_safe_frame(frame, current_user),
        safe_rename_map,
        columns=safe_columns,
    )


def render_mobile_table_cards(
    frame: pd.DataFrame,
    *,
    title_column: str | None = None,
    field_columns: list[str] | None = None,
    max_rows: int = 6,
) -> None:
    if frame.empty or not len(frame.columns):
        return

    available_columns = [str(column) for column in frame.columns]
    resolved_title = title_column if title_column in available_columns else available_columns[0]
    if field_columns is None:
        resolved_fields = [column for column in available_columns if column != resolved_title][:4]
    else:
        resolved_fields = [
            column for column in field_columns if column in available_columns and column != resolved_title
        ][:4]

    cards: list[str] = []
    for _, row in frame.head(max_rows).iterrows():
        title_value = row.get(resolved_title, "—")
        safe_title = "—" if is_missing(title_value) else escape(str(title_value))
        fields = []
        for column in resolved_fields:
            value = row.get(column, "—")
            safe_value = "—" if is_missing(value) else escape(str(value))
            fields.append(
                '<div class="mobile-data-card-field">'
                f'<div class="mobile-data-card-label">{escape(column)}</div>'
                f'<div class="mobile-data-card-value">{safe_value}</div>'
                "</div>"
            )
        cards.append(
            '<article class="mobile-data-card">'
            f'<div class="mobile-data-card-title">{safe_title}</div>'
            f'<div class="mobile-data-card-fields">{"".join(fields)}</div>'
            "</article>"
        )

    note = ""
    if len(frame) > max_rows:
        note = (
            '<div class="mobile-table-card-note">'
            f'Показано {max_rows} из {len(frame)} строк. Полная таблица доступна ниже.'
            "</div>"
        )
    render_html_block(f'<div class="mobile-table-card-grid">{"".join(cards)}{note}</div>')


def _safe_excel_sheet_name(raw_name: str, used_names: set[str]) -> str:
    invalid_chars = "\\/*?:[]"
    safe_name = "".join("_" if char in invalid_chars else char for char in str(raw_name).strip())
    safe_name = (safe_name or "Лист")[:31]
    candidate = safe_name
    suffix = 2
    while candidate.casefold() in used_names:
        suffix_text = f" {suffix}"
        candidate = f"{safe_name[:31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate.casefold())
    return candidate


def _normalize_excel_cell(value: object) -> object:
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            return value.tz_convert(None)
        return value
    if isinstance(value, (list, tuple, set, dict)):
        return str(value)
    return value


def _normalize_excel_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    for column in normalized.columns:
        series = normalized[column]
        if isinstance(series.dtype, pd.DatetimeTZDtype):
            normalized[column] = pd.to_datetime(series, errors="coerce").dt.tz_convert(None)
        elif pd.api.types.is_datetime64_any_dtype(series):
            normalized[column] = pd.to_datetime(series, errors="coerce")
        elif series.dtype == "object":
            normalized[column] = series.map(_normalize_excel_cell)
    return normalized


def prepare_excel_report_frame(
    frame: pd.DataFrame,
    current_user: dict[str, str] | None = None,
    rename_map: dict[str, str] | None = None,
    *,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    view = frame.copy()
    safe_rename_map = rename_map
    safe_columns = columns
    if current_user is not None:
        view = margin_safe_frame(view, current_user)
        safe_rename_map = margin_safe_rename_map(rename_map, current_user)
        safe_columns = margin_safe_columns(columns, current_user) if columns is not None else None

    if safe_columns is not None:
        available_columns = [column for column in safe_columns if column in view.columns]
        view = view[available_columns]

    used_labels: dict[str, int] = {}
    labels: list[str] = []
    for source_column in view.columns:
        preferred_label = (
            safe_rename_map[source_column]
            if safe_rename_map and source_column in safe_rename_map
            else DISPLAY_COLUMN_NAMES.get(source_column, str(source_column))
        )
        labels.append(_make_unique_label(preferred_label, source_column, used_labels))

    export_frame = _normalize_excel_frame(view)
    export_frame.columns = labels
    return export_frame


def _excel_column_format(header: str) -> str | None:
    if _is_date_label(header):
        return "DD.MM.YYYY"
    if _is_percent_label(header):
        return '0.0"%"'
    if _is_money_label(header):
        return '#,##0" сом"'
    if _is_count_label(header):
        return "#,##0.00"
    return None


def _style_excel_worksheet(worksheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor=PRIMARY_COLOR.replace("#", ""))
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color=BORDER_COLOR.replace("#", "")))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        header = str(worksheet.cell(row=1, column=column_index).value or "")
        number_format = _excel_column_format(header)
        max_length = len(header)

        for cell in worksheet[column_letter][1:]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
            if number_format is not None:
                cell.number_format = number_format

        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 52)

    worksheet.row_dimensions[1].height = 32


@st.cache_data(show_spinner=False)
def to_excel_report_bytes(
    sheets: dict[str, pd.DataFrame],
    *,
    report_title: str | None = None,
) -> bytes:
    buffer = BytesIO()
    used_names: set[str] = set()
    export_sheets: dict[str, pd.DataFrame] = {}

    if report_title:
        export_sheets["Обзор"] = pd.DataFrame(
            [
                {"Показатель": "Отчет", "Значение": report_title},
                {"Показатель": "Сформировано", "Значение": pd.Timestamp.now().strftime("%d.%m.%Y %H:%M")},
            ]
        )

    export_sheets.update(sheets)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for raw_sheet_name, frame in export_sheets.items():
            sheet_name = _safe_excel_sheet_name(raw_sheet_name, used_names)
            export_frame = frame.copy() if not frame.empty else pd.DataFrame({"Сообщение": ["Нет данных"]})
            export_frame = _normalize_excel_frame(export_frame)
            export_frame.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_excel_worksheet(writer.sheets[sheet_name])

    return buffer.getvalue()


def render_access_tab(registered_salons: list[str]) -> None:
    users = pd.DataFrame(list_users())
    salons_table = pd.DataFrame({"Салон": registered_salons}) if registered_salons else pd.DataFrame(columns=["Салон"])

    left_col, right_col = st.columns([1.15, 1], gap="medium")

    with left_col:
        with st.container(border=True):
            st.markdown('<div class="panel-title">Пользователи</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="panel-caption">Список логинов и назначенных ролей в системе.</div>',
                unsafe_allow_html=True,
            )
            if users.empty:
                st.info("Пользователей пока нет.")
            else:
                display_users = users.rename(
                    columns={
                        "username": "Логин",
                        "display_name": "Имя",
                        "role": "Роль",
                        "salon": "Салон",
                        "email": "Email",
                        "phone": "Телефон",
                        "created_at": "Создан",
                    }
                )
                display_users["Роль"] = display_users["Роль"].map(role_label)
                st.dataframe(display_users, use_container_width=True, hide_index=True)

        if False:
            st.markdown('<div class="panel-title">Салоны</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="panel-caption">Реестр салонов, доступных для загрузок и назначения пользователям.</div>',
                unsafe_allow_html=True,
            )
            if salons_table.empty:
                st.info("Салоны ещё не зарегистрированы.")
            else:
                st.dataframe(salons_table, use_container_width=True, hide_index=True)

    with right_col:
        with st.container(border=True):
            st.markdown('<div class="panel-title">Создать пользователя</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="panel-caption">Создайте отдельный логин для салона или ещё одного руководителя.</div>',
                unsafe_allow_html=True,
            )
            with st.form("create_user_form"):
                username = st.text_input("Логин")
                display_name = st.text_input("Имя пользователя")
                role_choice = st.selectbox("Роль", options=["salon", "manager"], format_func=role_label)
                email = st.text_input("Email")
                phone = st.text_input("Телефон")
                salon_name = st.text_input("Салон", disabled=role_choice != "salon")
                password = st.text_input("Пароль", type="password")
                password_confirm = st.text_input("Подтверждение пароля", type="password")
                submitted = st.form_submit_button("Создать пользователя", use_container_width=True)

            if submitted:
                if password != password_confirm:
                    st.error("Пароли не совпадают.")
                else:
                    try:
                        if role_choice == "salon":
                            save_salon(salon_name)
                        create_user(
                            username=username,
                            password=password,
                            role=role_choice,
                            display_name=display_name,
                            email=email,
                            phone=phone,
                            salon=salon_name,
                        )
                        st.success("Пользователь создан.")
                        st.rerun()
                    except Exception as error:
                        st.error(str(error))

        with st.container(border=True):
            st.markdown('<div class="panel-title">Сброс пароля</div>', unsafe_allow_html=True)
            st.markdown(
                '<div class="panel-caption">Быстрый сброс пароля для действующего пользователя.</div>',
                unsafe_allow_html=True,
            )
            user_options = users["username"].tolist() if not users.empty else []
            with st.form("reset_password_form"):
                selected_username = st.selectbox("Пользователь", options=user_options) if user_options else st.selectbox("Пользователь", options=["Нет пользователей"], disabled=True)
                new_password = st.text_input("Новый пароль", type="password")
                new_password_confirm = st.text_input("Подтверждение нового пароля", type="password")
                submitted = st.form_submit_button("Сменить пароль", use_container_width=True, disabled=not user_options)

            if submitted:
                if new_password != new_password_confirm:
                    st.error("Пароли не совпадают.")
                else:
                    try:
                        set_user_password(selected_username, new_password)
                        st.success("Пароль обновлён.")
                    except Exception as error:
                        st.error(str(error))


def render_admin_tab(current_user: dict[str, str], registered_salons: list[str]) -> None:
    apply_pending_widget_resets()
    users = pd.DataFrame(list_users())
    salons_table = pd.DataFrame({"Салон": registered_salons}) if registered_salons else pd.DataFrame(columns=["Салон"])

    if users.empty:
        users = pd.DataFrame(columns=["username", "display_name", "role", "salon", "email", "phone", "created_at"])

    admin_count = int((users["role"] == "admin").sum()) if "role" in users else 0
    manager_count = int((users["role"] == "manager").sum()) if "role" in users else 0
    salon_user_count = int((users["role"] == "salon").sum()) if "role" in users else 0

    display_users = users.rename(
        columns={
            "username": "Логин",
            "display_name": "Имя",
            "role": "Роль",
            "salon": "Салон",
            "email": "Email",
            "phone": "Телефон",
            "created_at": "Создан",
        }
    )
    if not display_users.empty:
        display_users["Роль"] = display_users["Роль"].map(role_label)

    main_col = st.container()
    side_control_col = main_col

    with main_col:
        render_section_intro(
            "Управление системой",
            "Панель администратора для управления пользователями, салонами и доступами. Настройте структуру вашей сети и права сотрудников здесь.",
        )
        render_html_block(
            f"""
            <div class="admin-stat-grid">
                <div class="admin-stat">
                    <div class="admin-stat-label">Салоны</div>
                    <div class="admin-stat-value">{len(registered_salons)}</div>
                </div>
                <div class="admin-stat">
                    <div class="admin-stat-label">Администраторы</div>
                    <div class="admin-stat-value">{admin_count}</div>
                </div>
                <div class="admin-stat">
                    <div class="admin-stat-label">Руководители</div>
                    <div class="admin-stat-value">{manager_count}</div>
                </div>
                <div class="admin-stat">
                    <div class="admin-stat-label">Пользователи салонов</div>
                    <div class="admin-stat-value">{salon_user_count}</div>
                </div>
            </div>
            """
        )

        st.info(
            "Как работать: 1. Добавьте салон. 2. Перейдите в «Пользователи». "
            "3. Выберите роль и создайте аккаунт. Для входа пользователю достаточно логина, email или телефона."
        )
        flash_message = st.session_state.pop("admin_flash_message", "")
        if flash_message:
            st.success(flash_message)

    with side_control_col:
        with st.container(border=True):
            st.markdown('<div class="nav-shell" style="padding: 0; border: none; background: transparent; margin: 0;">', unsafe_allow_html=True)
            st.markdown('<div class="nav-title">Раздел управления</div>', unsafe_allow_html=True)
            section = st.radio(
                "Раздел управления",
                options=["Салоны", "Пользователи", "Пароли"],
                key="admin_section_switch",
                label_visibility="collapsed",
            )
            st.markdown('</div>', unsafe_allow_html=True)
            st.divider()
            st.caption(f"Вы вошли как: {current_user['display_name']}")
            st.caption(f"Роль: {role_label(current_user['role'])}")

    if section == "Салоны":
        with main_col:
            salons_left, salons_right = st.columns([1, 1], gap="medium")

            with salons_left:
                with st.container(border=True):
                    render_panel_header(
                        "Добавить салон",
                        "Введите название салона так, как оно должно отображаться в аналитике и фильтрах.",
                    )
                    new_salon_name = st.text_input(
                        "Название салона",
                        key="admin_new_salon_name",
                        placeholder="Например: Омск | Мега",
                        help="После создания этот салон появится в выборе при создании пользователя и загрузке данных.",
                    )
                    if st.button("Добавить салон", key="admin_add_salon_button", use_container_width=True):
                        normalized_salon_name = new_salon_name.strip()
                        existing_salons = {salon.casefold() for salon in registered_salons}

                        if not normalized_salon_name:
                            st.error("Введите название салона.")
                        elif normalized_salon_name.casefold() in existing_salons:
                            st.warning("Такой салон уже существует.")
                        else:
                            save_salon(normalized_salon_name)
                            audit_event(
                                action="salon.create",
                                user_id=current_user["username"],
                                details={
                                    "salon": normalized_salon_name,
                                    "source": "admin_tab",
                                },
                            )
                            schedule_widget_reset({"admin_new_salon_name": ""})
                            st.session_state["admin_flash_message"] = f"Салон «{normalized_salon_name}» добавлен."
                            st.rerun()

            with salons_right:
                with st.container(border=True):
                    render_panel_header(
                        "Список салонов",
                        "Все салоны, доступные для выгрузок и назначения пользователей.",
                    )
                    if salons_table.empty:
                        st.info("Салоны пока не зарегистрированы.")
                    else:
                        st.dataframe(salons_table, use_container_width=True, hide_index=True, height=200)

                with st.container(border=True):
                    render_panel_header(
                        "Удалить салон",
                        "Удаление салона можно делать безопасно: либо только пустой салон, либо полностью с архивом.",
                    )

                    salon_to_delete = (
                        st.selectbox("Какой салон удалить", options=registered_salons, key="admin_delete_salon_name")
                        if registered_salons
                        else st.selectbox("Какой салон удалить", options=["Нет салонов"], disabled=True, key="admin_delete_salon_name_empty")
                    )

                    if registered_salons:
                        related_users_count = int(
                            (
                                users["salon"].fillna("").astype(str).str.casefold() == str(salon_to_delete).strip().casefold()
                            ).sum()
                        ) if not users.empty else 0
                        related_uploads_count = count_uploads_for_salon(str(salon_to_delete))
                        st.caption(
                            f"Пользователей: {related_users_count} | Выгрузок: {related_uploads_count}"
                        )

                        delete_salon_users = st.checkbox(
                            "Удалить пользователей салона",
                            key="admin_delete_salon_users",
                        )
                        delete_salon_uploads = st.checkbox(
                            "Удалить архив выгрузок",
                            key="admin_delete_salon_uploads",
                        )
                        salon_confirm_text = st.text_input(
                            "Название для подтверждения",
                            key="admin_delete_salon_confirm",
                            placeholder=str(salon_to_delete),
                        )

                        if st.button(
                            "Удалить салон",
                            key="admin_delete_salon_button",
                            use_container_width=True,
                            type="secondary",
                        ):
                            if salon_confirm_text.strip() != str(salon_to_delete).strip():
                                st.error("Введите точное название салона.")
                            elif related_users_count and not delete_salon_users:
                                st.error("У салона есть пользователи.")
                            else:
                                try:
                                    deleted_users = delete_users_by_salon(str(salon_to_delete)) if delete_salon_users else 0
                                    delete_result = delete_salon(
                                        str(salon_to_delete),
                                        remove_uploads=delete_salon_uploads,
                                    )
                                    audit_event(
                                        action="salon.delete",
                                        user_id=current_user["username"],
                                        details={
                                            "salon": str(salon_to_delete),
                                            "deleted_users": int(deleted_users),
                                            "deleted_uploads": int(delete_result["deleted_uploads"]),
                                            "source": "admin_tab",
                                        },
                                    )
                                    schedule_widget_reset(
                                        {
                                            "admin_delete_salon_confirm": "",
                                            "admin_delete_salon_users": False,
                                            "admin_delete_salon_uploads": False,
                                        }
                                    )
                                    st.session_state["admin_flash_message"] = f"Салон «{salon_to_delete}» удалён."
                                    st.rerun()
                                except Exception as error:
                                    st.error(str(error))

    elif section == "Пользователи":
        with main_col:
            users_left, users_right = st.columns([1, 1], gap="medium")

            with users_left:
                with st.container(border=True):
                    render_panel_header(
                        "Создать пользователя",
                        "Выберите роль, заполните данные и сохраните нового пользователя.",
                    )

                    role_choice = st.selectbox(
                        "Роль пользователя",
                        options=["manager", "salon", "admin"],
                        format_func=role_label,
                        key="admin_create_role",
                    )

                    identity_col, name_col = st.columns(2)
                    with identity_col:
                        username = st.text_input(
                            "Логин",
                            key="admin_create_username",
                            placeholder="Например: ivanov",
                        )
                    with name_col:
                        display_name = st.text_input(
                            "Имя в системе",
                            key="admin_create_display_name",
                            placeholder="Например: Иван Иванов",
                        )

                    contact_col, phone_col = st.columns(2)
                    with contact_col:
                        email = st.text_input(
                            "Email",
                            key="admin_create_email",
                            placeholder="user@company.ru",
                        )
                    with phone_col:
                        phone = st.text_input(
                            "Телефон",
                            key="admin_create_phone",
                            placeholder="+7 999 123-45-67",
                        )

                    selected_salon = ""
                    if role_choice == "salon":
                        selected_salon = st.selectbox(
                            "Какой салон будет видеть пользователь",
                            options=registered_salons if registered_salons else ["Сначала создайте салон"],
                            key="admin_create_salon",
                            disabled=not bool(registered_salons),
                        )
                    else:
                        st.caption("Для администратора и руководителя выбор салона не нужен.")

                password_col, confirm_col = st.columns(2)
                with password_col:
                    password = st.text_input("Пароль", type="password", key="admin_create_password")
                with confirm_col:
                    password_confirm = st.text_input("Повтор пароля", type="password", key="admin_create_password_confirm")

                create_user_disabled = role_choice == "salon" and not bool(registered_salons)
                if create_user_disabled:
                    st.warning("Сначала добавьте хотя бы один салон, потом можно будет создать пользователя салона.")

                if st.button(
                    "Создать пользователя",
                    key="admin_create_user_button",
                    use_container_width=True,
                    disabled=create_user_disabled,
                ):
                    if password != password_confirm:
                        st.error("Пароли не совпадают.")
                    else:
                        try:
                            create_user(
                                username=username,
                                password=password,
                                role=role_choice,
                                display_name=display_name,
                                email=email,
                                phone=phone,
                                salon=selected_salon,
                            )
                            audit_event(
                                action="user.create",
                                user_id=current_user["username"],
                                details={
                                    "username": username.strip(),
                                    "role": role_choice,
                                    "salon": selected_salon if role_choice == "salon" else "",
                                    "has_email": bool(email.strip()),
                                    "has_phone": bool(phone.strip()),
                                    "source": "admin_tab",
                                },
                            )
                            schedule_widget_reset(
                                {
                                    "admin_create_username": "",
                                    "admin_create_display_name": "",
                                    "admin_create_email": "",
                                    "admin_create_phone": "",
                                    "admin_create_password": "",
                                    "admin_create_password_confirm": "",
                                    "admin_create_salon": None,
                                    "admin_create_role": None,
                                }
                            )
                            st.session_state["admin_flash_message"] = (
                                f"Пользователь «{display_name or username}» с ролью «{role_label(role_choice)}» создан."
                            )
                            st.rerun()
                        except Exception as error:
                            st.error(str(error))

        with users_right:
            with st.container(border=True):
                st.markdown('<div class="panel-title">Реестр пользователей</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Проверяйте роли, контакты и привязку пользователей к салонам.</div>',
                    unsafe_allow_html=True,
                )
                if display_users.empty:
                    st.info("Пользователи пока не созданы.")
                else:
                    st.dataframe(display_users, use_container_width=True, hide_index=True)

            with st.container(border=True):
                st.markdown('<div class="panel-title">Перевести администраторов в руководители</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Массовое действие для случаев, когда нужно быстро убрать лишние права администратора и оставить управленческий доступ.</div>',
                    unsafe_allow_html=True,
                )

                admin_usernames = (
                    users.loc[users["role"].astype(str).str.lower() == "admin", "username"].astype(str).tolist()
                    if not users.empty
                    else []
                )
                eligible_admin_usernames = [
                    username
                    for username in admin_usernames
                    if username.strip().casefold() != current_user["username"].strip().casefold()
                ]

                if current_user["username"] in admin_usernames:
                    st.caption("Текущий администратор не включён в список, чтобы вы случайно не сняли права с собственной сессии.")

                selected_admins_to_demote = st.multiselect(
                    "Каких администраторов перевести",
                    options=eligible_admin_usernames,
                    key="admin_bulk_demote_usernames",
                    disabled=not bool(eligible_admin_usernames),
                    placeholder="Выберите одного или нескольких администраторов",
                )

                if eligible_admin_usernames:
                    remaining_admins_after_change = len(admin_usernames) - len(selected_admins_to_demote)
                    st.caption(
                        f"Сейчас администраторов: {len(admin_usernames)}. "
                        f"После перевода останется: {remaining_admins_after_change}."
                    )
                else:
                    st.info("Нет других администраторов, которых можно перевести в руководители.")

                if st.button(
                    "Перевести в руководители",
                    key="admin_bulk_demote_button",
                    use_container_width=True,
                    disabled=not bool(eligible_admin_usernames),
                ):
                    if not selected_admins_to_demote:
                        st.error("Выберите хотя бы одного администратора.")
                    elif len(admin_usernames) - len(selected_admins_to_demote) < 1:
                        st.error("В системе должен остаться хотя бы один администратор.")
                    else:
                        try:
                            changed_users: list[str] = []
                            for username_to_demote in selected_admins_to_demote:
                                updated_user = update_user_role(
                                    str(username_to_demote),
                                    "manager",
                                    actor_username=current_user["username"],
                                )
                                changed_users.append(str(updated_user.get("display_name", "") or updated_user.get("username", "")))
                                audit_event(
                                    action="user.role_update",
                                    user_id=current_user["username"],
                                    details={
                                        "username": str(updated_user.get("username", "")),
                                        "old_role": "admin",
                                        "new_role": "manager",
                                        "old_salon": "",
                                        "new_salon": "",
                                        "source": "admin_tab_bulk_demotion",
                                    },
                                )
                            schedule_widget_reset({"admin_bulk_demote_usernames": []})
                            st.session_state["admin_flash_message"] = (
                                "В роль руководителя переведены: " + ", ".join(changed_users) + "."
                            )
                            st.rerun()
                        except Exception as error:
                            st.error(str(error))

            with st.container(border=True):
                st.markdown('<div class="panel-title">Изменить роль пользователя</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Выберите пользователя, задайте новую роль и при необходимости привяжите его к салону.</div>',
                    unsafe_allow_html=True,
                )

                role_usernames = users["username"].tolist() if not users.empty else []
                role_target_username = (
                    st.selectbox(
                        "Какого пользователя изменить",
                        options=role_usernames,
                        key="admin_edit_role_username",
                    )
                    if role_usernames
                    else st.selectbox(
                        "Какого пользователя изменить",
                        options=["Нет пользователей"],
                        disabled=True,
                        key="admin_edit_role_username_empty",
                    )
                )

                if role_usernames:
                    role_target_row = users.loc[users["username"] == role_target_username].iloc[0]
                    current_role_value = str(role_target_row.get("role", "")).strip().lower()
                    current_salon_value = str(role_target_row.get("salon", "") or "").strip()

                    st.caption(
                        f"Сейчас: {role_label(current_role_value)}"
                        + (
                            f" | Салон: {current_salon_value}"
                            if current_salon_value
                            else " | Салон не привязан"
                        )
                    )

                    role_options = ["admin", "manager", "salon"]
                    new_role_value = st.selectbox(
                        "Новая роль",
                        options=role_options,
                        index=role_options.index(current_role_value) if current_role_value in role_options else 0,
                        format_func=role_label,
                        key=f"admin_edit_role_value_{role_target_username}",
                    )

                    role_salon_value = current_salon_value if new_role_value == "salon" else ""
                    if new_role_value == "salon":
                        salon_options = registered_salons if registered_salons else ["Сначала создайте салон"]
                        default_salon_index = 0
                        if current_salon_value and current_salon_value in registered_salons:
                            default_salon_index = registered_salons.index(current_salon_value)
                        role_salon_value = st.selectbox(
                            "Салон для пользователя",
                            options=salon_options,
                            index=default_salon_index if registered_salons else 0,
                            key=f"admin_edit_role_salon_{role_target_username}",
                            disabled=not bool(registered_salons),
                        )
                        if not registered_salons:
                            st.warning("Сначала добавьте хотя бы один салон, потом можно назначить роль салона.")
                    else:
                        st.caption("Для администратора и руководителя салон не нужен.")

                    role_change_disabled = (
                        new_role_value == "salon" and not bool(registered_salons)
                    )
                    if st.button(
                        "Сохранить роль",
                        key="admin_update_role_button",
                        use_container_width=True,
                        disabled=role_change_disabled,
                    ):
                        proposed_salon = role_salon_value if new_role_value == "salon" else ""
                        if current_role_value == new_role_value and current_salon_value == proposed_salon:
                            st.warning("Изменений нет: у пользователя уже установлены такие доступы.")
                        else:
                            try:
                                updated_user = update_user_role(
                                    str(role_target_username),
                                    new_role_value,
                                    salon=proposed_salon,
                                    actor_username=current_user["username"],
                                )
                                audit_event(
                                    action="user.role_update",
                                    user_id=current_user["username"],
                                    details={
                                        "username": str(updated_user.get("username", "")),
                                        "old_role": current_role_value,
                                        "new_role": str(updated_user.get("role", "")),
                                        "old_salon": current_salon_value,
                                        "new_salon": str(updated_user.get("salon", "") or ""),
                                        "source": "admin_tab",
                                    },
                                )
                                st.session_state["admin_flash_message"] = (
                                    f"Роль пользователя «{updated_user['display_name']}» обновлена."
                                )
                                st.rerun()
                            except Exception as error:
                                st.error(str(error))

            with st.container(border=True):
                st.markdown('<div class="panel-title">Перевести пользователя в другой салон</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Отдельное действие для сотрудников салонов. Роль не меняется: пользователь остаётся в контуре салона, меняется только привязка к точке.</div>',
                    unsafe_allow_html=True,
                )

                salon_users_frame = (
                    users.loc[users["role"].astype(str).str.lower() == "salon"].copy()
                    if not users.empty and "role" in users
                    else pd.DataFrame()
                )
                salon_usernames = salon_users_frame["username"].astype(str).tolist() if not salon_users_frame.empty else []

                transfer_username = (
                    st.selectbox(
                        "Какого сотрудника перевести",
                        options=salon_usernames,
                        key="admin_reassign_salon_username",
                    )
                    if salon_usernames
                    else st.selectbox(
                        "Какого сотрудника перевести",
                        options=["Нет пользователей салонов"],
                        disabled=True,
                        key="admin_reassign_salon_username_empty",
                    )
                )

                if not registered_salons:
                    st.info("Сначала создайте салоны, потом можно будет переводить сотрудников между ними.")
                elif len(registered_salons) < 2:
                    st.info("Для перевода нужен минимум второй салон.")
                elif salon_usernames:
                    transfer_row = salon_users_frame.loc[salon_users_frame["username"] == transfer_username].iloc[0]
                    current_transfer_salon = str(transfer_row.get("salon", "") or "").strip()
                    transfer_salon_options = registered_salons
                    transfer_target_index = 0
                    for idx, salon_name in enumerate(transfer_salon_options):
                        if salon_name.strip().casefold() != current_transfer_salon.casefold():
                            transfer_target_index = idx
                            break

                    st.caption(
                        f"Сейчас сотрудник привязан к салону: {current_transfer_salon or 'Не привязан'}"
                    )
                    transfer_target_salon = st.selectbox(
                        "В какой салон перевести",
                        options=transfer_salon_options,
                        index=transfer_target_index,
                        key=f"admin_reassign_salon_target_{transfer_username}",
                    )

                    if st.button(
                        "Перевести в другой салон",
                        key="admin_reassign_salon_button",
                        use_container_width=True,
                    ):
                        if not current_transfer_salon:
                            st.error("У выбранного пользователя нет текущей привязки к салону.")
                        elif current_transfer_salon.strip().casefold() == transfer_target_salon.strip().casefold():
                            st.warning("Выберите другой салон для перевода.")
                        else:
                            try:
                                moved_user = reassign_salon_user(
                                    str(transfer_username),
                                    str(transfer_target_salon),
                                    actor_username=current_user["username"],
                                )
                                audit_event(
                                    action="user.salon_reassign",
                                    user_id=current_user["username"],
                                    details={
                                        "username": str(moved_user.get("username", "")),
                                        "old_salon": current_transfer_salon,
                                        "new_salon": str(moved_user.get("salon", "")),
                                        "source": "admin_tab",
                                    },
                                )
                                st.session_state["admin_flash_message"] = (
                                    f"Пользователь «{moved_user['display_name']}» переведён из салона "
                                    f"«{current_transfer_salon}» в «{moved_user['salon']}»."
                                )
                                st.rerun()
                            except Exception as error:
                                st.error(str(error))
                else:
                    st.info("Пока нет пользователей с ролью «Салон».")

            with st.container(border=True):
                st.markdown('<div class="panel-title">Удалить пользователя</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Удаление сразу закрывает доступ пользователю. Текущего администратора под собой удалить нельзя.</div>',
                    unsafe_allow_html=True,
                )

                available_usernames = users["username"].tolist() if not users.empty else []
                selected_user_to_delete = (
                    st.selectbox("Какого пользователя удалить", options=available_usernames, key="admin_delete_username")
                    if available_usernames
                    else st.selectbox("Какого пользователя удалить", options=["Нет пользователей"], disabled=True, key="admin_delete_username_empty")
                )

                if available_usernames:
                    selected_user_row = users.loc[users["username"] == selected_user_to_delete].iloc[0]
                    st.caption(
                        f"Роль: {role_label(str(selected_user_row['role']))} | "
                        f"Салон: {str(selected_user_row.get('salon', '') or 'Не привязан')}"
                    )
                    if str(selected_user_to_delete).strip().casefold() == current_user["username"].strip().casefold():
                        st.warning("Вы сейчас вошли под этим пользователем. Для безопасности самоуничтожение аккаунта отключено.")

                    user_confirm_text = st.text_input(
                        "Для подтверждения введите логин пользователя",
                        key="admin_delete_user_confirm",
                        placeholder=str(selected_user_to_delete),
                    )

                    if st.button(
                        "Удалить пользователя",
                        key="admin_delete_user_button",
                        use_container_width=True,
                        type="secondary",
                        disabled=str(selected_user_to_delete).strip().casefold() == current_user["username"].strip().casefold(),
                    ):
                        if user_confirm_text.strip() != str(selected_user_to_delete).strip():
                            st.error("Введите точный логин пользователя для подтверждения удаления.")
                        else:
                            try:
                                deleted_user = delete_user(
                                    str(selected_user_to_delete),
                                    actor_username=current_user["username"],
                                )
                                audit_event(
                                    action="user.delete",
                                    user_id=current_user["username"],
                                    details={
                                        "username": str(deleted_user.get("username", "")),
                                        "role": str(deleted_user.get("role", "")),
                                        "salon": str(deleted_user.get("salon", "")),
                                        "source": "admin_tab",
                                    },
                                )
                                schedule_widget_reset({"admin_delete_user_confirm": ""})
                                st.session_state["admin_flash_message"] = (
                                    f"Пользователь «{deleted_user['display_name']}» удалён."
                                )
                                st.rerun()
                            except Exception as error:
                                st.error(str(error))

    else:
        password_left, password_right = st.columns([0.95, 1.05], gap="medium")

        with password_left:
            with st.container(border=True):
                st.markdown('<div class="panel-title">Сброс пароля</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Выберите пользователя и задайте ему новый пароль.</div>',
                    unsafe_allow_html=True,
                )
                user_options = users["username"].tolist() if not users.empty else []
                selected_username = (
                    st.selectbox("Пользователь", options=user_options, key="admin_reset_username")
                    if user_options
                    else st.selectbox("Пользователь", options=["Нет пользователей"], disabled=True, key="admin_reset_username_empty")
                )
                password_col, confirm_col = st.columns(2)
                with password_col:
                    new_password = st.text_input("Новый пароль", type="password", key="admin_reset_password")
                with confirm_col:
                    new_password_confirm = st.text_input(
                        "Повтор нового пароля",
                        type="password",
                        key="admin_reset_password_confirm",
                    )

                if st.button(
                    "Сменить пароль",
                    key="admin_reset_password_button",
                    use_container_width=True,
                    disabled=not user_options,
                ):
                    if new_password != new_password_confirm:
                        st.error("Пароли не совпадают.")
                    else:
                        try:
                            set_user_password(selected_username, new_password)
                            audit_event(
                                action="user.password_reset",
                                user_id=current_user["username"],
                                details={
                                    "username": str(selected_username),
                                    "source": "admin_tab",
                                },
                            )
                            schedule_widget_reset(
                                {
                                    "admin_reset_password": "",
                                    "admin_reset_password_confirm": "",
                                }
                            )
                            st.session_state["admin_flash_message"] = f"Пароль для пользователя «{selected_username}» обновлён."
                            st.rerun()
                        except Exception as error:
                            st.error(str(error))

        with password_right:
            with st.container(border=True):
                st.markdown('<div class="panel-title">Что важно</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="panel-caption">Короткая памятка по созданию пользователей и доступам.</div>',
                    unsafe_allow_html=True,
                )
                st.write("1. Сначала создайте салон, если нужен пользователь салона.")
                st.write("2. Затем в разделе «Пользователи» выберите роль и заполните поля.")
                st.write("3. Нужно указать хотя бы один контакт: email или телефон.")
                st.write("4. Для входа можно использовать логин, email или телефон.")


def render_intro(current_user: dict[str, str]) -> None:
    salon_text = current_user.get("salon") or "Вся сеть"
    render_html_block(
        f"""
        <div class="app-shell-header">
            <div class="app-shell-brand">
                <div class="app-shell-name">ArtDB</div>
                <div class="app-shell-copy">Продажи, ассортимент и закупки в одном рабочем пространстве</div>
            </div>
            <div class="app-shell-user">
                <div>
                    <div class="app-shell-user-name">{escape(current_user['display_name'])}</div>
                    <div class="app-shell-user-meta">{escape(salon_text)}</div>
                </div>
                <div class="app-shell-role">{escape(role_label(current_user['role']))}</div>
            </div>
        </div>
        """
    )


def render_dataset_hero(
    filename: str,
    data: pd.DataFrame,
    overview: dict[str, float],
    selected_categories: list[str] | None,
    selected_managers: list[str] | None,
    current_user: dict[str, str],
    active_screen: str = "Обзор",
) -> None:
    min_date = format_date_value(data["date"].min())
    max_date = format_date_value(data["date"].max())

    chips = [
        f"Период: {min_date} - {max_date}",
        f"Строк: {format_number(overview['line_count'])}",
        f"Товаров: {format_number(overview['product_count'])}",
    ]
    chips_html = "".join(f'<span class="workspace-chip">{chip}</span>' for chip in chips)
    screen_title = SCREEN_NAV_LABELS.get(active_screen, active_screen)

    render_html_block(
        f"""
        <div class="workspace-header">
            <div>
                <div class="workspace-header-title">{escape(screen_title)}</div>
                <div class="workspace-header-meta">Источник: {escape(filename)} · {escape(role_label(current_user['role']))}</div>
            </div>
            <div class="workspace-context">
                {chips_html}
                <span class="workspace-chip">{escape(get_build_badge_label())}</span>
            </div>
        </div>
        """
    )


def build_insights(
    overview: dict[str, float],
    monthly_summary: pd.DataFrame,
    category_summary: pd.DataFrame,
    manager_summary: pd.DataFrame,
    product_summary: pd.DataFrame,
    abc_data: pd.DataFrame,
    returns_overview: dict[str, float] | None = None,
    salon_summary: pd.DataFrame | None = None,
    anomalies: list[tuple[str, str]] | None = None,
    *,
    allow_margin: bool = True,
) -> list[tuple[str, str]]:
    insights: list[tuple[str, str]] = []

    if len(monthly_summary) >= 2:
        latest_month = monthly_summary.iloc[-1]
        previous_month = monthly_summary.iloc[-2]
        insights.append(
            (
                f"Последний месяц: {latest_month['month_label']}",
                f"Выручка {format_money(latest_month['revenue'])}, изменение к {previous_month['month_label']} "
                f"составило {format_percent(latest_month['revenue_change_pct'])}.",
            )
        )

    if not category_summary.empty and overview["total_revenue"] > 0:
        top_category = category_summary.iloc[0]
        share = top_category["revenue"] / overview["total_revenue"] * 100
        insights.append(
            (
                f"Главная категория: {top_category['group_name']}",
                f"Даёт {format_money(top_category['revenue'])} и формирует {format_percent(share)} всей выручки.",
            )
        )

    if not manager_summary.empty:
        top_manager = manager_summary.iloc[0]
        insights.append(
            (
                f"Лидер по продажам: {top_manager['group_name']}",
                (
                    f"У менеджера {format_money(top_manager['revenue'])} выручки и {format_money(top_manager['margin'])} маржи."
                    if allow_margin
                    else f"У менеджера {format_money(top_manager['revenue'])} выручки в текущем срезе."
                ),
            )
        )

    if salon_summary is not None and not salon_summary.empty and len(salon_summary) > 1:
        top_salon = salon_summary.iloc[0]
        insights.append(
            (
                f"Сильнейший салон: {top_salon['group_name']}",
                (
                    f"Формирует {format_money(top_salon['revenue'])} выручки и {format_money(top_salon['margin'])} маржи."
                    if allow_margin
                    else f"Формирует {format_money(top_salon['revenue'])} выручки в текущем срезе."
                ),
            )
        )

    if allow_margin and not product_summary.empty and not product_summary["margin_pct"].isna().all():
        low_margin_count = int((product_summary["margin_pct"] < 20).fillna(False).sum())
        insights.append(
            (
                "Риски по марже",
                f"SKU с маржой ниже 20%: {low_margin_count}. Проверьте блок с зонами риска во вкладке `Маржинальность`.",
            )
        )

    if not abc_data.empty:
        a_share = abc_data.loc[abc_data["abc_class"] == "A", "share_pct"].sum()
        a_count = int((abc_data["abc_class"] == "A").sum())
        insights.append(
            (
                "Концентрация выручки",
                f"Класс A содержит {a_count} SKU и формирует {format_percent(a_share)} выбранной ABC-метрики.",
            )
        )

    if anomalies:
        for month_label, description in anomalies[:1]:
            insights.append((f"⚠ Аномалия: {month_label}", description))

    return insights[:6]


def render_insight_panel(insights: list[tuple[str, str]]) -> None:
    if not insights:
        return

    items_html = []
    for title, body in insights[:6]:
        tone = infer_dashboard_tone(title, body)
        items_html.append(
            dedent(
                f"""
                <div class="insight-compact-item {escape(tone)}">
                    <div class="insight-compact-title">{escape(title)}</div>
                    <div class="insight-compact-body">{escape(body)}</div>
                </div>
                """
            ).strip()
        )

    render_html_block(
        f"""
        <section class="overview-command-panel">
            <div class="overview-command-header">
                <div>
                    <div class="overview-command-kicker">Управленческая выжимка</div>
                    <div class="overview-command-title">Что важно сейчас</div>
                </div>
                <div class="overview-command-meta">{len(items_html)} сигналов</div>
            </div>
            <div class="insight-compact-list">{"".join(items_html)}</div>
        </section>
        """
    )


def render_overview_action_center(
    insights: list[tuple[str, str]],
    focus_cards: list[dict[str, str]],
) -> None:
    combined: list[dict[str, str]] = []
    for title, body in insights:
        combined.append(
            {
                "title": str(title),
                "body": str(body),
                "tone": infer_dashboard_tone(title, body),
            }
        )

    for card in focus_cards:
        combined.append(
            {
                "title": f"{card.get('label', '')}: {card.get('value', '')}".strip(": "),
                "body": str(card.get("body", "")),
                "tone": str(card.get("tone") or infer_dashboard_tone(card.get("label", ""), card.get("body", ""))),
            }
        )

    priority = {"danger": 0, "warning": 1, "success": 2, "info": 3}
    unique_items: list[dict[str, str]] = []
    seen_titles: set[str] = set()
    for item in sorted(combined, key=lambda value: priority.get(value.get("tone", "info"), 3)):
        normalized_title = item["title"].casefold().strip()
        if not normalized_title or normalized_title in seen_titles:
            continue
        seen_titles.add(normalized_title)
        unique_items.append(item)
        if len(unique_items) >= 5:
            break

    if not unique_items:
        return

    items_html = "".join(
        dedent(
            f"""
            <div class="overview-action-item {escape(item['tone'])}">
                <div class="overview-action-indicator"></div>
                <div>
                    <div class="overview-action-item-title">{escape(item['title'])}</div>
                    <div class="overview-action-item-body">{escape(item['body'])}</div>
                </div>
            </div>
            """
        ).strip()
        for item in unique_items
    )
    render_html_block(
        f"""
        <section class="overview-action-center">
            <div class="overview-action-header">
                <div class="overview-action-title">Что важно сейчас</div>
                <div class="overview-action-count">{len(unique_items)} сигналов</div>
            </div>
            <div class="overview-action-list">{items_html}</div>
        </section>
        """
    )


def build_overview_focus_cards(
    latest_month: pd.Series,
    latest_revenue_delta: float,
    latest_margin_delta: float,
    plan_fact_summary: pd.DataFrame,
    procurement_overview: dict[str, float],
    returns_overview: dict[str, float],
    *,
    allow_margin: bool = True,
) -> list[dict[str, str]]:
    month_label = str(latest_month.get("month_label", "Текущий период"))
    revenue_delta = percent_or_none(latest_revenue_delta)
    cards = [
        {
            "label": f"Итог {month_label}",
            "value": format_money(latest_month.get("revenue", 0)),
            "body": f"{revenue_delta} к предыдущему месяцу" if revenue_delta else "Выручка последнего доступного месяца",
            "tone": infer_dashboard_tone("выручка", revenue_delta or ""),
        }
    ]

    if allow_margin:
        margin_delta = percent_or_none(latest_margin_delta)
        cards.append(
            {
                "label": "Маржа месяца",
                "value": format_money(latest_month.get("margin", 0)),
                "body": f"{margin_delta} к предыдущему месяцу" if margin_delta else "Валовая маржа за последний месяц",
                "tone": infer_dashboard_tone("маржа", margin_delta or ""),
            }
        )

    if not plan_fact_summary.empty:
        latest_plan_row = plan_fact_summary.iloc[-1]
        execution_pct = latest_plan_row.get("revenue_execution_pct")
        cards.append(
            {
                "label": "План / факт",
                "value": format_percent(execution_pct) if not is_missing(execution_pct) else "план не задан",
                "body": f"Месяц: {latest_plan_row.get('month_label', month_label)}",
                "tone": "success" if not is_missing(execution_pct) and float(execution_pct) >= 100 else "warning",
            }
        )

    reorder_count = int(procurement_overview.get("reorder_sku_count", 0) or 0)
    critical_count = int(procurement_overview.get("critical_stock_count", 0) or 0)
    if reorder_count or critical_count:
        cards.append(
            {
                "label": "Закупки",
                "value": f"{format_number(reorder_count)} SKU",
                "body": f"К заказу сейчас, критичных позиций: {format_number(critical_count)}",
                "tone": "danger" if critical_count else "warning",
            }
        )
    else:
        cards.append(
            {
                "label": "Закупки",
                "value": "покрыто",
                "body": "Нет срочных рекомендаций к заказу по текущему срезу",
                "tone": "success",
            }
        )

    return_lines = int(returns_overview.get("return_lines", 0) or 0)
    cards.append(
        {
            "label": "Возвраты",
            "value": format_percent(returns_overview.get("return_share_pct", 0)),
            "body": f"Строк возврата: {format_number(return_lines)}, сумма: {format_money(returns_overview.get('return_revenue', 0))}",
            "tone": "warning" if return_lines else "success",
        }
    )

    return cards[:5]


def render_overview_focus_stack(cards: list[dict[str, str]]) -> None:
    if not cards:
        return

    cards_html = []
    for card in cards:
        tone = card.get("tone") or infer_dashboard_tone(card.get("label", ""), card.get("body", ""))
        cards_html.append(
            dedent(
                f"""
                <div class="overview-focus-card {escape(str(tone))}">
                    <div class="overview-focus-label">{escape(card["label"])}</div>
                    <div class="overview-focus-value">{escape(card["value"])}</div>
                    <div class="overview-focus-body">{escape(card.get("body", ""))}</div>
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="overview-focus-stack">{"".join(cards_html)}</div>')


def build_text_summary(
    overview: dict[str, float],
    monthly_summary: pd.DataFrame,
    returns_overview: dict[str, float],
    plan_fact_summary: pd.DataFrame,
    latest_revenue_delta: float,
    *,
    allow_margin: bool = True,
) -> list[str]:
    lines: list[str] = []
    if len(monthly_summary) >= 1:
        latest = monthly_summary.iloc[-1]
        line = f"\U0001f4ca {latest['month_label']}: выручка {format_money(overview['total_revenue'])}"
        if not is_missing(latest_revenue_delta):
            sign = "\u2191" if latest_revenue_delta >= 0 else "\u2193"
            line += f" ({sign} {abs(float(latest_revenue_delta)):.1f}% к пред. месяцу)"
        lines.append(line)
    if allow_margin and not is_missing(overview.get("margin_pct")):
        lines.append(f"\U0001f4b0 Маржа: {format_percent(overview['margin_pct'])}")
    if not plan_fact_summary.empty:
        lp = plan_fact_summary.iloc[-1]
        rev_exec = lp.get("revenue_execution_pct")
        if rev_exec is not None and not is_missing(rev_exec):
            pct = float(rev_exec)
            ico = "\u2705" if pct >= 100 else "\u26a0\ufe0f" if pct >= 80 else "\U0001f534"
            lines.append(f"{ico} План: {format_percent(pct)}")
    if returns_overview["return_lines"] > 0:
        lines.append(f"\u21a9\ufe0f Возвраты: {format_percent(returns_overview['return_share_pct'])}")
    return lines


def get_numeric_column(frame: pd.DataFrame, column: str, *, default: float = 0.0) -> pd.Series:
    if column not in frame.columns:
        return pd.Series(default, index=frame.index, dtype="float64")
    return pd.to_numeric(frame[column], errors="coerce")


def safe_pct(numerator: float, denominator: float) -> float:
    if is_missing(numerator) or is_missing(denominator) or abs(float(denominator)) < 1e-9:
        return float("nan")
    return float(numerator) / float(denominator) * 100.0


def build_financial_pnl_frame(frame: pd.DataFrame, returns_overview: dict[str, float]) -> pd.DataFrame:
    revenue = get_numeric_column(frame, "revenue").fillna(0.0)
    cost = get_numeric_column(frame, "cost", default=float("nan"))
    margin = get_numeric_column(frame, "margin", default=float("nan"))

    gross_revenue = float(revenue.clip(lower=0).sum())
    return_revenue = float(returns_overview.get("return_revenue", 0.0) or 0.0)
    net_revenue = float(revenue.sum())
    total_cost_raw = cost.sum(min_count=1)
    total_cost = float(total_cost_raw) if pd.notna(total_cost_raw) else float("nan")
    total_margin_raw = margin.sum(min_count=1)
    total_margin = float(total_margin_raw) if pd.notna(total_margin_raw) else float("nan")
    if is_missing(total_margin) and not is_missing(total_cost):
        total_margin = net_revenue - total_cost

    return pd.DataFrame(
        [
            {
                "metric": "Валовая выручка",
                "amount": gross_revenue,
                "share_pct": 100.0 if gross_revenue else float("nan"),
                "comment": "Продажи до учёта возвратов.",
            },
            {
                "metric": "Возвраты",
                "amount": -return_revenue,
                "share_pct": safe_pct(return_revenue, gross_revenue),
                "comment": "Сколько выручки вернулось клиентам.",
            },
            {
                "metric": "Чистая выручка",
                "amount": net_revenue,
                "share_pct": safe_pct(net_revenue, gross_revenue),
                "comment": "Итоговая выручка после возвратов.",
            },
            {
                "metric": "Себестоимость",
                "amount": -total_cost if not is_missing(total_cost) else float("nan"),
                "share_pct": safe_pct(total_cost, net_revenue),
                "comment": "Закупочная стоимость проданных товаров.",
            },
            {
                "metric": "Валовая маржа",
                "amount": total_margin,
                "share_pct": safe_pct(total_margin, net_revenue),
                "comment": "Управленческая валовая прибыль до операционных расходов.",
            },
        ]
    )


def build_financial_inventory_frame(procurement_forecast: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "product",
        "category",
        "supplier",
        "stock_on_hand",
        "available_stock_qty",
        "avg_unit_cost",
        "stock_value",
        "available_stock_value",
        "monthly_cogs_forecast",
        "stock_coverage_days",
        "recommended_order_qty",
        "priority",
        "stock_status",
    ]
    if procurement_forecast.empty:
        return pd.DataFrame(columns=columns)

    inventory = procurement_forecast.copy()
    for column in [
        "stock_on_hand",
        "available_stock_qty",
        "forecast_qty",
        "avg_revenue_per_unit",
        "avg_margin_per_unit",
        "stock_coverage_days",
        "recommended_order_qty",
    ]:
        inventory[column] = get_numeric_column(inventory, column)

    inventory["avg_unit_cost"] = inventory["avg_revenue_per_unit"] - inventory["avg_margin_per_unit"]
    inventory["avg_unit_cost"] = inventory["avg_unit_cost"].where(inventory["avg_unit_cost"] > 0)
    inventory["stock_value"] = inventory["stock_on_hand"].fillna(0.0) * inventory["avg_unit_cost"]
    inventory["available_stock_value"] = inventory["available_stock_qty"].fillna(0.0) * inventory["avg_unit_cost"]
    inventory["monthly_cogs_forecast"] = inventory["forecast_qty"].fillna(0.0) * inventory["avg_unit_cost"]
    inventory = inventory.replace([float("inf"), -float("inf")], pd.NA)

    for column in columns:
        if column not in inventory.columns:
            inventory[column] = "" if column in {"product", "category", "supplier", "priority", "stock_status"} else float("nan")

    return (
        inventory[columns]
        .sort_values(["stock_value", "available_stock_value"], ascending=[False, False], na_position="last")
        .reset_index(drop=True)
    )


def build_financial_alerts(
    overview: dict[str, float],
    category_summary: pd.DataFrame,
    product_summary: pd.DataFrame,
    returns_overview: dict[str, float],
    inventory_frame: pd.DataFrame,
    procurement_overview: dict[str, float],
    plan_fact_summary: pd.DataFrame,
) -> list[dict[str, str]]:
    alerts: list[dict[str, str]] = []

    margin_pct = overview.get("margin_pct", float("nan"))
    if not is_missing(margin_pct):
        if float(margin_pct) < 15:
            alerts.append(
                {
                    "title": "Маржа ниже безопасного уровня",
                    "body": f"Текущая маржинальность {format_percent(margin_pct)}. Нужна проверка цен, скидок и себестоимости.",
                    "tone": "danger",
                }
            )
        elif float(margin_pct) < 25:
            alerts.append(
                {
                    "title": "Маржа в зоне внимания",
                    "body": f"Маржинальность {format_percent(margin_pct)}. Стоит отдельно посмотреть низкомаржинальные SKU.",
                    "tone": "warning",
                }
            )

    if not product_summary.empty and "margin" in product_summary.columns:
        negative_margin_count = int((pd.to_numeric(product_summary["margin"], errors="coerce").fillna(0.0) < 0).sum())
        low_margin_count = int((pd.to_numeric(product_summary["margin_pct"], errors="coerce").fillna(9999) < 20).sum())
        if negative_margin_count:
            alerts.append(
                {
                    "title": "Есть продажи в минус",
                    "body": f"SKU с отрицательной маржей: {format_number(negative_margin_count)}. Это первый список для ревизии.",
                    "tone": "danger",
                }
            )
        elif low_margin_count:
            alerts.append(
                {
                    "title": "Низкомаржинальные позиции",
                    "body": f"SKU ниже 20% маржи: {format_number(low_margin_count)}. Проверьте скидки, закупку и переоценку.",
                    "tone": "warning",
                }
            )

    returns_share = returns_overview.get("return_share_pct", float("nan"))
    if not is_missing(returns_share) and float(returns_share) >= 2:
        alerts.append(
            {
                "title": "Возвраты влияют на результат",
                "body": f"Доля возвратов {format_percent(returns_share)} от валовой выручки. Нужна причина по товарам и салонам.",
                "tone": "warning",
            }
        )

    if not category_summary.empty and overview.get("total_revenue", 0):
        top_category = category_summary.iloc[0]
        top_share = safe_pct(float(top_category.get("revenue", 0.0) or 0.0), float(overview.get("total_revenue", 0.0) or 0.0))
        if not is_missing(top_share) and float(top_share) >= 65:
            alerts.append(
                {
                    "title": "Высокая зависимость от категории",
                    "body": f"{top_category.get('group_name', 'Категория')} даёт {format_percent(top_share)} выручки. Нужен контроль ассортимента-ядра.",
                    "tone": "info",
                }
            )

    stock_value = (
        pd.to_numeric(inventory_frame.get("stock_value", pd.Series(dtype="float64")), errors="coerce").sum(min_count=1)
        if not inventory_frame.empty
        else float("nan")
    )
    if pd.notna(stock_value) and float(stock_value) > 0:
        alerts.append(
            {
                "title": "Деньги в остатках",
                "body": f"Оценка замороженных денег: {format_money(float(stock_value))}. Смотрите товары с большим остатком и низким спросом.",
                "tone": "info",
            }
        )

    critical_stock_count = int(procurement_overview.get("critical_stock_count", 0) or 0)
    if critical_stock_count:
        alerts.append(
            {
                "title": "Риск потери продаж",
                "body": f"Критичных позиций по остаткам: {format_number(critical_stock_count)}. Их лучше разобрать до следующей закупки.",
                "tone": "danger",
            }
        )

    if not plan_fact_summary.empty:
        latest_plan_row = plan_fact_summary.iloc[-1]
        execution_pct = latest_plan_row.get("revenue_execution_pct")
        if execution_pct is not None and not is_missing(execution_pct) and float(execution_pct) < 90:
            alerts.append(
                {
                    "title": "План продаж проседает",
                    "body": f"Выполнение плана выручки за {latest_plan_row.get('month_label', 'последний месяц')}: {format_percent(execution_pct)}.",
                    "tone": "warning",
                }
            )

    if not alerts:
        alerts.append(
            {
                "title": "Критичных финансовых сигналов нет",
                "body": "По текущему срезу маржа, возвраты и остатки выглядят без явных красных флагов.",
                "tone": "success",
            }
        )

    return alerts[:6]


def render_financial_alerts(alerts: list[dict[str, str]]) -> None:
    items_html = []
    for alert in alerts[:6]:
        items_html.append(
            dedent(
                f"""
                <div class="insight-compact-item {escape(alert.get("tone", "info"))}">
                    <div class="insight-compact-title">{escape(alert["title"])}</div>
                    <div class="insight-compact-body">{escape(alert["body"])}</div>
                </div>
                """
            ).strip()
        )

    render_html_block(f'<div class="insight-compact-list">{"".join(items_html)}</div>')


def safe_display_text(value: object, fallback: str = "Не указан") -> str:
    if is_missing(value):
        return fallback
    text = str(value).strip()
    if not text or text.casefold() in {"nan", "none", "nat"}:
        return fallback
    return text


def first_display_text(*values: object, fallback: str = "Не указан") -> str:
    for value in values:
        text = safe_display_text(value, "")
        if text:
            return text
    return fallback


def safe_row_float(row: pd.Series, column: str, default: float = float("nan")) -> float:
    if row.empty or column not in row.index:
        return default
    value = pd.to_numeric(pd.Series([row.get(column)]), errors="coerce").iloc[0]
    return float(value) if pd.notna(value) else default


def format_optional_date(value: object) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "н/д"
    return parsed.strftime("%d.%m.%Y")


def most_common_text(frame: pd.DataFrame, column: str, fallback: str = "Не указан") -> str:
    if frame.empty or column not in frame.columns:
        return fallback
    series = frame[column].dropna().astype(str).str.strip()
    series = series[(series != "") & (~series.str.casefold().isin({"nan", "none", "nat"}))]
    if series.empty:
        return fallback
    return safe_display_text(series.value_counts().idxmax(), fallback)


def build_sku_monthly_detail(product_monthly: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "month",
        "month_label",
        "revenue",
        "cost",
        "margin",
        "quantity",
        "avg_price",
        "avg_margin_per_unit",
        "revenue_change_pct",
        "quantity_change_pct",
        "avg_price_change_pct",
        "margin_change_pct",
        "moving_avg_revenue_3m",
        "moving_avg_quantity_3m",
        "moving_avg_price_3m",
        "moving_avg_margin_3m",
    ]
    if product_monthly.empty:
        return pd.DataFrame(columns=columns)

    detail = product_monthly.copy()
    for metric_column in ("revenue", "cost", "margin", "quantity"):
        if metric_column not in detail.columns:
            detail[metric_column] = 0.0
        detail[metric_column] = pd.to_numeric(detail[metric_column], errors="coerce").fillna(0.0)

    quantity_non_zero = detail["quantity"].where(detail["quantity"].abs() > 1e-9)
    detail["avg_price"] = detail["revenue"].divide(quantity_non_zero)
    detail["avg_margin_per_unit"] = detail["margin"].divide(quantity_non_zero)

    if "month" in detail.columns:
        detail["month"] = pd.to_datetime(detail["month"], errors="coerce")
        detail = detail.sort_values("month").reset_index(drop=True)
    if "month_label" not in detail.columns:
        detail["month_label"] = detail["month"].dt.strftime("%Y-%m") if "month" in detail.columns else ""

    detail["revenue_change_pct"] = detail["revenue"].pct_change() * 100
    detail["quantity_change_pct"] = detail["quantity"].pct_change() * 100
    detail["avg_price_change_pct"] = detail["avg_price"].pct_change() * 100
    detail["margin_change_pct"] = detail["margin"].pct_change() * 100

    detail["moving_avg_revenue_3m"] = detail["revenue"].rolling(3, min_periods=1).mean()
    detail["moving_avg_quantity_3m"] = detail["quantity"].rolling(3, min_periods=1).mean()
    detail["moving_avg_price_3m"] = detail["avg_price"].rolling(3, min_periods=1).mean()
    detail["moving_avg_margin_3m"] = detail["margin"].rolling(3, min_periods=1).mean()

    return detail.reindex(columns=columns)


def format_sku_trend_value(metric: str, value: float) -> str:
    if metric in {"revenue", "margin", "avg_price"}:
        return format_money(value)
    return format_number(value)


def sku_trend_tone(change_pct: float) -> str:
    if is_missing(change_pct):
        return "info"
    if change_pct >= 10:
        return "success"
    if change_pct <= -10:
        return "warning"
    return "info"


def build_product_detail_options(product_summary: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "option_id",
        "option_label",
        "group_key",
        "group_name",
        "product_name",
        "item_code",
        "revenue",
        "cost",
        "margin",
        "margin_pct",
        "quantity",
        "sales_lines",
    ]
    if product_summary.empty:
        return pd.DataFrame(columns=columns)

    options = product_summary.copy().reset_index(drop=True)
    if "group_name" not in options.columns:
        options["group_name"] = "Не указан"
    if "group_key" not in options.columns:
        options["group_key"] = options["group_name"].astype(str)
    if "product_name" not in options.columns:
        options["product_name"] = options["group_name"]
    if "item_code" not in options.columns:
        options["item_code"] = ""

    for numeric_column in ["revenue", "cost", "margin", "margin_pct", "quantity", "sales_lines"]:
        if numeric_column not in options.columns:
            options[numeric_column] = float("nan")
        options[numeric_column] = pd.to_numeric(options[numeric_column], errors="coerce")

    options["option_id"] = options.index.astype(str)

    def _label(row: pd.Series) -> str:
        code = safe_display_text(row.get("item_code"), "")
        name = first_display_text(row.get("product_name"), row.get("group_name"))
        prefix = f"{code} · " if code else ""
        revenue = row.get("revenue")
        revenue_hint = format_money(float(revenue)) if pd.notna(revenue) else "выручка н/д"
        return f"{prefix}{name} · {revenue_hint}"

    options["option_label"] = options.apply(_label, axis=1)
    return options[columns].sort_values("revenue", ascending=False, na_position="last").reset_index(drop=True)


def build_partner_sku_portfolio(
    sales_frame: pd.DataFrame,
    procurement_frame: pd.DataFrame,
    *,
    dimension: str,
    selected_value: str,
    product_analysis_column: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    portfolio_columns = [
        "portfolio_key",
        "item_code",
        "product",
        "category",
        "brand",
        "supplier",
        "revenue",
        "cost",
        "margin",
        "margin_pct",
        "quantity",
        "avg_price",
        "sales_lines",
        "return_revenue",
        "return_lines",
        "last_sale_date",
        "abc_class",
        "xyz_class",
        "stock_on_hand",
        "stock_value",
        "available_stock_qty",
        "stock_coverage_days",
        "forecast_qty",
        "forecast_revenue",
        "recommended_order_qty",
        "priority",
        "stock_status",
    ]
    monthly_columns = [
        "month",
        "month_label",
        "revenue",
        "cost",
        "margin",
        "quantity",
        "product_count",
        "revenue_change_pct",
        "margin_change_pct",
    ]
    selected_normalized = str(selected_value).strip().casefold()
    if not selected_normalized:
        return pd.DataFrame(columns=portfolio_columns), pd.DataFrame(columns=monthly_columns)

    sales_scope = sales_frame.copy()
    if dimension not in sales_scope.columns:
        sales_scope = sales_scope.iloc[0:0].copy()
    else:
        dimension_values = (
            sales_scope[dimension]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Не назначен")
            .str.casefold()
        )
        sales_scope = sales_scope.loc[dimension_values.eq(selected_normalized)].copy()

    procurement_scope = procurement_frame.copy()
    if dimension not in procurement_scope.columns:
        procurement_scope = procurement_scope.iloc[0:0].copy()
    else:
        procurement_values = (
            procurement_scope[dimension]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Не назначен")
            .str.casefold()
        )
        procurement_scope = procurement_scope.loc[procurement_values.eq(selected_normalized)].copy()

    sales_summary = pd.DataFrame()
    supplier_monthly = pd.DataFrame(columns=monthly_columns)
    if not sales_scope.empty and "product" in sales_scope.columns:
        portfolio_identity_column = (
            product_analysis_column
            if product_analysis_column in sales_scope.columns
            else "product"
        )
        sales_scope["_portfolio_key"] = (
            sales_scope[portfolio_identity_column]
            .fillna("")
            .astype(str)
            .str.strip()
        )
        sales_scope = sales_scope[sales_scope["_portfolio_key"].ne("")].copy()
        if not sales_scope.empty:
            for numeric_column in ("revenue", "cost", "margin", "quantity"):
                if numeric_column not in sales_scope.columns:
                    sales_scope[numeric_column] = 0.0
                sales_scope[numeric_column] = pd.to_numeric(
                    sales_scope[numeric_column],
                    errors="coerce",
                ).fillna(0.0)
            if "date" not in sales_scope.columns:
                sales_scope["date"] = pd.NaT
            sales_scope["date"] = pd.to_datetime(sales_scope["date"], errors="coerce")
            for text_column in ("item_code", "category", "brand", "supplier"):
                if text_column not in sales_scope.columns:
                    sales_scope[text_column] = ""

            sales_summary = (
                sales_scope.groupby("_portfolio_key", dropna=False)
                .agg(
                    product=("product", _first_catalog_text),
                    item_code=("item_code", _first_catalog_text),
                    category=("category", _first_catalog_text),
                    brand=("brand", _first_catalog_text),
                    supplier=("supplier", _first_catalog_text),
                    revenue=("revenue", "sum"),
                    cost=("cost", "sum"),
                    margin=("margin", "sum"),
                    quantity=("quantity", "sum"),
                    sales_lines=("product", "size"),
                    last_sale_date=("date", "max"),
                )
                .reset_index()
                .rename(columns={"_portfolio_key": "portfolio_key"})
            )
            sales_summary["margin_pct"] = (
                sales_summary["margin"] / sales_summary["revenue"]
            ).where(sales_summary["revenue"].ne(0)) * 100
            sales_summary["avg_price"] = (
                sales_summary["revenue"] / sales_summary["quantity"]
            ).where(sales_summary["quantity"].ne(0))

            abc_source = sales_summary[
                [
                    "portfolio_key",
                    "product",
                    "revenue",
                    "cost",
                    "margin",
                    "margin_pct",
                    "quantity",
                    "sales_lines",
                ]
            ].copy()
            abc_source = build_abc_analysis(abc_source, "revenue")
            sales_summary = sales_summary.merge(
                abc_source[["portfolio_key", "abc_class"]],
                on="portfolio_key",
                how="left",
            )

            return_rows = extract_return_rows(sales_scope)
            if not return_rows.empty:
                return_rows["_portfolio_key"] = (
                    return_rows[portfolio_identity_column]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                return_summary = (
                    return_rows.groupby("_portfolio_key", dropna=False)
                    .agg(
                        return_revenue=("revenue", lambda values: pd.to_numeric(values, errors="coerce").abs().sum()),
                        return_lines=("product", "size"),
                    )
                    .reset_index()
                    .rename(columns={"_portfolio_key": "portfolio_key"})
                )
                sales_summary = sales_summary.merge(
                    return_summary,
                    on="portfolio_key",
                    how="left",
                )
            supplier_monthly = build_monthly_summary(sales_scope)

    if sales_summary.empty:
        sales_summary = pd.DataFrame(
            columns=[
                "portfolio_key",
                "product",
                "item_code",
                "category",
                "brand",
                "supplier",
                "revenue",
                "cost",
                "margin",
                "margin_pct",
                "quantity",
                "avg_price",
                "sales_lines",
                "last_sale_date",
                "abc_class",
                "return_revenue",
                "return_lines",
            ]
        )

    sales_summary["_join_key"] = (
        sales_summary["product"].fillna("").astype(str).str.strip().str.casefold()
    )
    sales_summary = sales_summary[sales_summary["_join_key"].ne("")].copy()

    procurement_columns = [
        "product",
        "category",
        "brand",
        "supplier",
        "xyz_class",
        "stock_on_hand",
        "stock_value",
        "available_stock_qty",
        "stock_coverage_days",
        "forecast_qty",
        "forecast_revenue",
        "recommended_order_qty",
        "priority",
        "stock_status",
    ]
    for column in procurement_columns:
        if column not in procurement_scope.columns:
            procurement_scope[column] = ""
    procurement_scope = procurement_scope[procurement_columns].copy()
    procurement_scope["_join_key"] = (
        procurement_scope["product"].fillna("").astype(str).str.strip().str.casefold()
    )
    procurement_scope = (
        procurement_scope[procurement_scope["_join_key"].ne("")]
        .drop_duplicates("_join_key", keep="last")
        .rename(
            columns={
                "product": "inventory_product",
                "category": "inventory_category",
                "brand": "inventory_brand",
                "supplier": "inventory_supplier",
            }
        )
    )

    portfolio = sales_summary.merge(procurement_scope, on="_join_key", how="outer")
    if portfolio.empty:
        return pd.DataFrame(columns=portfolio_columns), supplier_monthly

    for target_column, inventory_column in (
        ("product", "inventory_product"),
        ("category", "inventory_category"),
        ("brand", "inventory_brand"),
        ("supplier", "inventory_supplier"),
    ):
        if target_column not in portfolio.columns:
            portfolio[target_column] = ""
        portfolio[target_column] = portfolio[target_column].fillna("").astype(str).str.strip()
        inventory_values = portfolio.get(
            inventory_column,
            pd.Series("", index=portfolio.index, dtype="object"),
        ).fillna("").astype(str).str.strip()
        portfolio[target_column] = portfolio[target_column].where(
            portfolio[target_column].ne(""),
            inventory_values,
        )

    portfolio["portfolio_key"] = (
        portfolio.get("portfolio_key", pd.Series("", index=portfolio.index, dtype="object"))
        .fillna("")
        .astype(str)
        .str.strip()
    )
    portfolio["portfolio_key"] = portfolio["portfolio_key"].where(
        portfolio["portfolio_key"].ne(""),
        portfolio["product"],
    )
    for text_column in (
        "item_code",
        "category",
        "brand",
        "supplier",
        "abc_class",
        "xyz_class",
        "priority",
        "stock_status",
    ):
        if text_column not in portfolio.columns:
            portfolio[text_column] = ""
        portfolio[text_column] = portfolio[text_column].fillna("").astype(str).str.strip()
    portfolio["abc_class"] = portfolio["abc_class"].replace("", "C")
    portfolio["xyz_class"] = portfolio["xyz_class"].replace("", "Z")
    portfolio["brand"] = portfolio["brand"].replace("", "Не назначен")
    portfolio["supplier"] = portfolio["supplier"].replace("", "Не назначен")
    portfolio["stock_status"] = portfolio["stock_status"].replace("", "Нет данных")

    numeric_columns = [
        "revenue",
        "cost",
        "margin",
        "margin_pct",
        "quantity",
        "avg_price",
        "sales_lines",
        "return_revenue",
        "return_lines",
        "stock_on_hand",
        "stock_value",
        "available_stock_qty",
        "stock_coverage_days",
        "forecast_qty",
        "forecast_revenue",
        "recommended_order_qty",
    ]
    for numeric_column in numeric_columns:
        if numeric_column not in portfolio.columns:
            portfolio[numeric_column] = 0.0
        portfolio[numeric_column] = pd.to_numeric(
            portfolio[numeric_column],
            errors="coerce",
        )
    for zero_fill_column in (
        "revenue",
        "cost",
        "margin",
        "quantity",
        "sales_lines",
        "return_revenue",
        "return_lines",
        "stock_on_hand",
        "stock_value",
        "available_stock_qty",
        "forecast_qty",
        "forecast_revenue",
        "recommended_order_qty",
    ):
        portfolio[zero_fill_column] = portfolio[zero_fill_column].fillna(0.0)
    portfolio["margin_pct"] = (
        portfolio["margin"] / portfolio["revenue"]
    ).where(portfolio["revenue"].ne(0)) * 100
    portfolio["avg_price"] = (
        portfolio["revenue"] / portfolio["quantity"]
    ).where(portfolio["quantity"].ne(0))
    portfolio["last_sale_date"] = pd.to_datetime(
        portfolio.get("last_sale_date"),
        errors="coerce",
    )

    portfolio = portfolio.sort_values(
        ["revenue", "stock_value", "recommended_order_qty"],
        ascending=[False, False, False],
        ignore_index=True,
    )
    return portfolio[portfolio_columns], supplier_monthly


def filter_product_detail_frame(
    frame: pd.DataFrame,
    selected_product: pd.Series,
    product_analysis_column: str,
) -> pd.DataFrame:
    if frame.empty or selected_product.empty:
        return frame.iloc[0:0].copy()

    if product_analysis_column in frame.columns and "group_key" in selected_product.index:
        selected_key = safe_display_text(selected_product.get("group_key"), "")
        if selected_key:
            mask = frame[product_analysis_column].astype(str) == selected_key
            matched = frame.loc[mask].copy()
            if not matched.empty:
                return matched

    candidate_names = [
        selected_product.get("product_name"),
        selected_product.get("group_name"),
    ]
    if "product" in frame.columns:
        product_names = frame["product"].astype(str).str.strip().str.casefold()
        for candidate in candidate_names:
            candidate_text = safe_display_text(candidate, "")
            if not candidate_text:
                continue
            matched = frame.loc[product_names == candidate_text.casefold()].copy()
            if not matched.empty:
                return matched

    return frame.iloc[0:0].copy()


def find_procurement_row_for_product(
    procurement_forecast: pd.DataFrame,
    product_frame: pd.DataFrame,
    selected_product: pd.Series,
) -> pd.Series:
    if procurement_forecast.empty or "product" not in procurement_forecast.columns:
        return pd.Series(dtype="object")

    candidates: list[str] = []
    if not product_frame.empty and "product" in product_frame.columns:
        candidates.extend(
            product_frame["product"]
            .dropna()
            .astype(str)
            .str.strip()
            .loc[lambda values: values != ""]
            .value_counts()
            .index
            .tolist()
        )
    candidates.extend(
        [
            safe_display_text(selected_product.get("product_name"), ""),
            safe_display_text(selected_product.get("group_name"), ""),
        ]
    )

    seen: set[str] = set()
    procurement_names = procurement_forecast["product"].astype(str).str.strip().str.casefold()
    for candidate in candidates:
        normalized = candidate.casefold().strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        match = procurement_forecast.loc[procurement_names == normalized]
        if not match.empty:
            return match.iloc[0]

    return pd.Series(dtype="object")


def find_product_abc_row(abc_frame: pd.DataFrame, selected_product: pd.Series) -> pd.Series:
    if abc_frame.empty or selected_product.empty:
        return pd.Series(dtype="object")

    for column, selected_column in [("group_key", "group_key"), ("group_name", "group_name")]:
        if column not in abc_frame.columns or selected_column not in selected_product.index:
            continue
        selected_value = safe_display_text(selected_product.get(selected_column), "")
        if not selected_value:
            continue
        match = abc_frame.loc[abc_frame[column].astype(str) == selected_value]
        if not match.empty:
            return match.iloc[0]

    return pd.Series(dtype="object")


def build_product_action_cards(
    selected_product: pd.Series,
    product_frame: pd.DataFrame,
    procurement_row: pd.Series,
    product_returns: pd.DataFrame,
    *,
    allow_margin: bool,
) -> list[dict[str, str]]:
    actions: list[dict[str, str]] = []
    margin_pct = safe_row_float(selected_product, "margin_pct")
    revenue = safe_row_float(selected_product, "revenue", 0.0)
    quantity = safe_row_float(selected_product, "quantity", 0.0)
    return_revenue = float(pd.to_numeric(product_returns.get("revenue", pd.Series(dtype="float64")), errors="coerce").abs().sum())
    recommended_order_qty = safe_row_float(procurement_row, "recommended_order_qty", 0.0)
    stock_on_hand = safe_row_float(procurement_row, "stock_on_hand")
    coverage_days = safe_row_float(procurement_row, "stock_coverage_days")
    priority = safe_display_text(procurement_row.get("priority"), "") if not procurement_row.empty else ""
    stock_status = safe_display_text(procurement_row.get("stock_status"), "") if not procurement_row.empty else ""

    if allow_margin and not is_missing(margin_pct):
        if margin_pct < 0:
            actions.append(
                {
                    "title": "Маржа отрицательная",
                    "body": "Проверьте себестоимость, скидки и корректность загрузки. Этот SKU продаётся в минус.",
                    "tone": "danger",
                }
            )
        elif margin_pct < 20:
            actions.append(
                {
                    "title": "Низкая маржа",
                    "body": f"Маржинальность {format_percent(margin_pct)}. Стоит проверить цену, закупку и условия скидок.",
                    "tone": "warning",
                }
            )
        elif margin_pct >= 35 and revenue > 0:
            actions.append(
                {
                    "title": "Хороший маржинальный SKU",
                    "body": f"Маржинальность {format_percent(margin_pct)}. Можно держать товар в фокусе продаж и наличия.",
                    "tone": "success",
                }
            )

    if recommended_order_qty > 0:
        coverage_text = f", покрытие {coverage_days:.0f} дней" if not is_missing(coverage_days) else ""
        actions.append(
            {
                "title": f"К закупке: {format_number(recommended_order_qty)}",
                "body": f"Прогноз показывает потребность к заказу{coverage_text}. Приоритет: {priority or 'не указан'}.",
                "tone": "danger" if "крит" in stock_status.casefold() else "warning",
            }
        )
    elif not is_missing(stock_on_hand) and stock_on_hand > 0 and not is_missing(coverage_days) and coverage_days > 90:
        actions.append(
            {
                "title": "Риск излишка",
                "body": f"Остаток {format_number(stock_on_hand)}, покрытие около {coverage_days:.0f} дней. Лучше не дозакупать без причины.",
                "tone": "warning",
            }
        )

    if not product_returns.empty:
        actions.append(
            {
                "title": "Есть возвраты",
                "body": f"Возвратных строк: {format_number(len(product_returns))}, сумма возвратов {format_money(return_revenue)}. Проверьте причины.",
                "tone": "warning",
            }
        )

    if quantity <= 0 and revenue <= 0:
        actions.append(
            {
                "title": "Нет положительных продаж",
                "body": "По текущему срезу товар не даёт положительной выручки. Возможно, это возврат или спящая позиция.",
                "tone": "info",
            }
        )

    if not actions:
        actions.append(
            {
                "title": "SKU без явных красных флагов",
                "body": "По текущему срезу продажи, возвраты и остатки не требуют срочного вмешательства.",
                "tone": "success",
            }
        )

    return actions[:6]


def build_movement_chart(comparison: pd.DataFrame, left_month: str, right_month: str) -> go.Figure:
    movement = comparison.copy()
    movement["abs_revenue_delta"] = movement["revenue_delta"].abs()
    movement = movement.nlargest(12, "abs_revenue_delta").sort_values("revenue_delta")
    movement_scale = [
        [0.0, CHART_DANGER_COLOR],
        [0.48, "#F3D78C"],
        [1.0, PRIMARY_COLOR],
    ]

    figure = px.bar(
        movement,
        x="revenue_delta",
        y="group_name",
        orientation="h",
        color="revenue_delta",
        color_continuous_scale=movement_scale,
        labels={"group_name": "Товар", "revenue_delta": "Изменение выручки"},
        title=f"Главные движения: {right_month} к {left_month}",
    )
    figure.add_vline(x=0, line_width=1.2, line_color="rgba(15, 23, 42, 0.28)")
    figure.update_layout(coloraxis_showscale=False, bargap=0.22)
    return polish_figure(figure, height=460)


promoted_admin = promote_first_manager_to_admin()
if promoted_admin and st.session_state.get("auth_user", {}).get("username", "").casefold() == promoted_admin["username"].casefold():
    st.session_state["auth_user"] = promoted_admin


current_user = render_auth_gate()

uploaded_file = None
raw_data: pd.DataFrame | None = None
selected_mapping: dict[str, str | None] = {}
prepared_result = None
quality_report = None
archive_result = None
manifest_view = load_manifest()
registered_salons = load_salons()
analysis_source = ""
source_label = ""
selected_salon_name = ""
selected_categories: list[str] | None = None
selected_managers: list[str] | None = None
selected_salons_filter: list[str] | None = None
selected_salons_for_archive: list[str] = []
current_file_report_date = date.today()
auto_detected_report_date: date | None = None
sheet_name: str | int | None = 0
csv_separator = ";"
csv_encoding = "utf-8"
replace_existing_upload = True
data = pd.DataFrame()
plan_fact_source_data = pd.DataFrame()
procurement_source_data = pd.DataFrame()
procurement_uses_manager_unfiltered_scope = False
supplier_keyword_rules = cached_load_supplier_keyword_rules(include_inactive=True)
active_supplier_keyword_rules = supplier_keyword_rules[
    supplier_keyword_rules["is_active"].fillna(True).astype(bool)
].copy() if not supplier_keyword_rules.empty else pd.DataFrame()
supplier_rule_items = supplier_rule_items_from_frame(active_supplier_keyword_rules)
supplier_product_assignments = cached_load_supplier_product_assignments()

if current_user["role"] == "salon":
    if not current_user.get("salon"):
        st.error("Для пользователя салона не назначен салон. Зайдите под руководителем и привяжите пользователя.")
        st.stop()
    selected_salon_name = current_user["salon"]
    manifest_view = manifest_view[manifest_view["salon"].astype(str) == selected_salon_name].copy()
else:
    manifest_view = manifest_view.copy()

if current_user["role"] == "salon" and selected_salon_name not in registered_salons:
    registered_salons = sorted([*registered_salons, selected_salon_name])

# --- Sidebar Configuration ---
with st.sidebar:
    st.markdown(f"**Пользователь:** {current_user['display_name']}")
    st.caption(f"Роль: {role_label(current_user['role'])}")
    if current_user["role"] == "salon":
        st.caption(f"Салон: {selected_salon_name}")

    st.markdown("---")
    st.header("Режим работы")

    if current_user["role"] == "salon":
        work_mode = st.radio("Режим салона", ["Архив салона", "Новая выгрузка"], index=0, key="main_work_mode")
        if work_mode == "Новая выгрузка":
            current_file_report_date = st.date_input("Дата отчёта", value=date.today(), key="main_report_date")
            replace_existing_upload = st.checkbox("Заменять файл за эту дату", value=True, key="main_replace_upload")
    else:
        work_mode = st.radio(
            "Режим сети",
            ["Сводка по сети", "Загрузка салона", "Разовая загрузка"],
            index=0,
            key="main_work_mode",
        )

        if work_mode == "Сводка по сети":
            selected_salons_for_archive = st.multiselect(
                "Салоны в отчёте",
                registered_salons,
                default=registered_salons,
                key="main_archive_salons",
            )
        elif work_mode == "Загрузка салона":
            salon_options = [*registered_salons, "Новый салон"] if registered_salons else ["Новый салон"]
            chosen_salon = st.selectbox("Салон для загрузки", options=salon_options, key="main_upload_salon")
            if chosen_salon == "Новый салон":
                selected_salon_name = st.text_input("Название нового салона", key="main_new_upload_salon").strip()
            else:
                selected_salon_name = chosen_salon

            current_file_report_date = st.date_input("Дата отчёта", value=date.today(), key="main_report_date")
            replace_existing_upload = st.checkbox("Заменять файл за эту дату", value=True, key="main_replace_upload")

    st.markdown("---")
    render_sidebar_navigation(current_user, work_mode)

    if can_manage_access(current_user):
        with st.expander("Быстрое создание", expanded=False):
            render_sidebar_admin_quick_actions(current_user, registered_salons)

    if st.button("Выйти", key="main_logout_button", use_container_width=True):
        revoke_auth_session(get_persistent_auth_token())
        clear_persistent_auth_token()
        st.session_state.pop("auth_user", None)
        st.rerun()

# --- Main Area ---
render_intro(current_user)

upload_modes = {"Новая выгрузка", "Загрузка салона", "Разовая загрузка"}
upload_flash_message = st.session_state.pop("upload_flash_message", "")
if upload_flash_message:
    st.success(upload_flash_message)

main_col = st.container()
mobile_navigation_slot = main_col.empty()
mobile_bottom_navigation_slot = main_col.empty()
control_col = main_col

if work_mode in upload_modes:
    upload_shell = st.container(border=True)
    with upload_shell:
        st.markdown('<div class="panel-title">Загрузка файла</div>', unsafe_allow_html=True)
        if work_mode == "Разовая загрузка":
            st.markdown('<div class="panel-caption">Загрузите файл для быстрого анализа без сохранения в архив.</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="panel-caption">Загрузка выгрузки для салона <b>{selected_salon_name}</b>.</div>', unsafe_allow_html=True)

        upload_left, upload_right = st.columns([1, 1], gap="large")
        with upload_left:
            st.markdown('<div class="panel-title">Шаг 1: Выбор файла</div>', unsafe_allow_html=True)
            uploaded_file = st.file_uploader("Файл из 1С", type=["xlsx", "xls", "csv"], key=f"main_uploader_{work_mode}", label_visibility="collapsed")
            if not uploaded_file:
                st.info("Пожалуйста, выберите файл Excel или CSV.")

        with upload_right:
            upload_scope = selected_salon_name or "Без привязки к салону"
            if work_mode in {"Новая выгрузка", "Загрузка салона"} and selected_salon_name:
                current_manifest = load_manifest()
                salon_manifest = current_manifest[current_manifest["salon"].astype(str) == selected_salon_name].copy()
                uploads_count = len(salon_manifest)
                latest_report_date = ""
                if not salon_manifest.empty:
                    latest_report_date = str(salon_manifest["report_date"].astype(str).max())

                st.markdown(f"""
                <div style="background: {BG_COLOR}; border: 1px solid {BORDER_COLOR}; border-radius: 12px; padding: 1rem;">
                    <div style="color: {TEXT_MUTED}; font-size: 0.75rem; text-transform: uppercase; font-weight: 700; margin-bottom: 0.4rem;">Контекст загрузки</div>
                    <div style="font-weight: 800; color: {PRIMARY_COLOR}; margin-bottom: 0.2rem; font-family: 'Manrope', sans-serif;">{upload_scope}</div>
                    <div style="color: {TEXT_SECONDARY}; font-size: 0.85rem;">Режим: {work_mode}</div>
                    <div style="color: {TEXT_SECONDARY}; font-size: 0.85rem;">В архиве: {uploads_count} файлов</div>
                    {f'<div style="color: {TEXT_SECONDARY}; font-size: 0.85rem;">Последний: {latest_report_date}</div>' if latest_report_date else ""}
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div style="background: {BG_COLOR}; border: 1px solid {BORDER_COLOR}; border-radius: 12px; padding: 1.2rem;">
                    <div style="color: {SECONDARY_COLOR}; font-size: 0.75rem; text-transform: uppercase; font-weight: 700; margin-bottom: 0.5rem;">Контекст загрузки</div>
                    <div style="font-weight: 800; color: {PRIMARY_COLOR}; margin-bottom: 0.3rem; font-family: 'Manrope', sans-serif;">{upload_scope}</div>
                    <div style="color: {TEXT_SECONDARY}; font-size: 0.88rem;">Режим: {work_mode}</div>
                </div>
                """, unsafe_allow_html=True)

    if uploaded_file is not None:
        # Step 2: Configuration & Preview (Wizard-like flow)
        file_bytes = uploaded_file.getvalue()
        try:
            filename = validate_uploaded_file(file_bytes, uploaded_file.name)
        except ValueError as error:
            st.error(str(error))
            st.stop()

        with upload_shell:
            st.divider()
            st.markdown('<div class="panel-title" style="margin-top: 1rem;">Шаг 2: Настройка и сопоставление</div>', unsafe_allow_html=True)
            conf_col, map_col = st.columns([1, 1.5], gap="medium")

            with conf_col:
                with st.container(border=True):
                    st.markdown("**Настройка чтения**")
                    if filename.lower().endswith(".csv"):
                        csv_separator = st.selectbox("Разделитель", options=[";", ",", "\t"], index=0, key="upload_csv_separator")
                        csv_encoding = st.selectbox("Кодировка", options=["utf-8", "cp1251", "utf-8-sig"], index=0, key="upload_csv_encoding")
                        sheet_name = None
                    else:
                        sheet_names = list_excel_sheets(file_bytes)
                        sheet_name = st.selectbox("Лист Excel", options=sheet_names, index=0, key="upload_sheet_name")
                        csv_separator = ";"
                        csv_encoding = "utf-8"

            try:
                raw_data = cached_load_input_file(file_bytes, filename, csv_separator, csv_encoding, sheet_name)
            except Exception as error:
                st.error(f"Не удалось прочитать файл: {error}")
                st.stop()

            if raw_data.empty:
                st.warning("Файл прочитан, но в нём нет строк.")
                st.stop()

            columns = raw_data.columns.astype(str).tolist()
            guesses = guess_column_mapping(columns)

            with map_col:
                with st.container(border=True):
                    st.markdown("**Шаг 3: Сопоставление колонок**")
                    st.caption("Выберите соответствующие поля из вашего файла.")
                    m1, m2 = st.columns(2)
                    with m1:
                        col_date = select_column("Дата (обязательно)", columns, guesses.get("date"), key="map_date")
                        col_product = select_column("Товар (обязательно)", columns, guesses.get("product"), key="map_product")
                        col_item_code = select_column("Код / артикул товара", columns, guesses.get("item_code"), key="map_item_code")
                        col_cat = select_column("Категория", columns, guesses.get("category"), key="map_cat")
                        col_supplier = select_column("Поставщик", columns, guesses.get("supplier"), key="map_supplier")
                        col_manager = select_column("Менеджер", columns, guesses.get("manager"), key="map_man")
                    with m2:
                        col_rev = select_column("Выручка", columns, guesses.get("revenue"), key="map_rev")
                        col_cost = select_column("Себестоимость", columns, guesses.get("cost"), key="map_cost")
                        col_qty = select_column("Количество", columns, guesses.get("quantity"), key="map_qty")
                        col_price = select_column("Цена ед.", columns, guesses.get("unit_price"), key="map_price")

                    selected_mapping = {
                        "date": col_date, "product": col_product, "revenue": col_rev,
                        "cost": col_cost, "margin": None, "quantity": col_qty,
                        "unit_price": col_price, "unit_cost": None, "category": col_cat,
                        "supplier": col_supplier, "manager": col_manager, "item_code": col_item_code
                    }

            st.divider()
            st.markdown('<div class="panel-title">Шаг 4: Проверка и сохранение</div>', unsafe_allow_html=True)
            mapping_items = tuple(sorted(selected_mapping.items()))
            try:
                prepared_result = cached_prepare_sales_data(raw_data, mapping_items, supplier_rule_items)
            except Exception as error:
                st.error(str(error))
                st.stop()

            auto_detected_report_date = None
            detected_dates = prepared_result.data["date"].dropna().dt.date.unique().tolist()
            if len(detected_dates) == 1:
                auto_detected_report_date = detected_dates[0]
                if work_mode in {"Новая выгрузка", "Загрузка салона"} and current_file_report_date == date.today():
                    current_file_report_date = auto_detected_report_date

            quality_expected_date = current_file_report_date if work_mode in {"Новая выгрузка", "Загрузка салона"} else None
            quality_report = analyze_sales_quality(
                raw_data,
                prepared_result.data,
                selected_mapping,
                procurement_items=load_procurement_items(),
                expected_report_date=quality_expected_date,
                include_margin_checks=can_view_margin(current_user),
            )
            with st.container(border=True):
                render_panel_header(
                    "Контроль качества данных",
                    "Проверяем даты, товары, выручку, количество, дубли и связь со справочником перед сохранением.",
                )
                render_data_quality_report(quality_report, current_user)

            if work_mode in {"Новая выгрузка", "Загрузка салона"}:
                render_panel_header(
                    "Сохранение в архив",
                    "Все параметры сохранения собраны прямо в этом окне.",
                )
                if auto_detected_report_date is not None:
                    st.markdown(f"**Найдена дата в файле:** `{auto_detected_report_date.strftime('%d.%m.%Y')}`")

                if not selected_salon_name:
                    st.warning("Сначала выберите салон.")
                else:
                    settings_left, settings_right = st.columns([1.25, 0.95], gap="medium")
                    with settings_left:
                        target_date = st.date_input("Дата отчёта", value=current_file_report_date, key="final_save_date")
                    with settings_right:
                        replace_check = st.checkbox(
                            "Заменить существующий файл за эту дату",
                            value=replace_existing_upload,
                            key="final_save_replace",
                        )

                    save_meta_left, save_meta_right = st.columns([1.15, 1], gap="medium")
                    with save_meta_left:
                        st.caption(f"Файл: {filename}")
                        st.caption(f"Салон: {selected_salon_name}")
                    with save_meta_right:
                        st.caption(f"Строк после подготовки: {format_number(len(prepared_result.data))}")
                        st.caption(f"Распознано полей: {sum(1 for value in selected_mapping.values() if value)}")

                    save_disabled = quality_report is not None and not quality_report.can_save
                    if save_disabled:
                        st.error("Сохранение отключено до исправления критичных ошибок качества данных.")

                    if st.button(
                        "🚀 Сохранить в архив",
                        key="main_save_upload_button",
                        use_container_width=True,
                        type="primary",
                        disabled=save_disabled,
                    ):
                        save_upload_with_feedback(
                            file_bytes=file_bytes,
                            filename=filename,
                            salon_name=selected_salon_name,
                            report_date=target_date,
                            mapping=selected_mapping,
                            csv_separator=csv_separator,
                            csv_encoding=csv_encoding,
                            sheet_name=sheet_name,
                            replace_existing=replace_check,
                            actor_username=current_user["username"],
                        )
                        st.rerun()

    if work_mode == "Разовая загрузка":
        if prepared_result is None:
            st.info("Загрузите файл, чтобы открыть обзор.")
            st.stop()
        data = prepared_result.data.copy()
        source_label = filename
    elif work_mode in {"Новая выгрузка", "Загрузка салона"}:
        if prepared_result is None:
            st.info("Загрузите файл, чтобы провести анализ.")
            st.stop()
        data = prepared_result.data.copy()
        if selected_salon_name:
            data["salon"] = selected_salon_name
        source_label = f"Текущая выгрузка: {filename}"
    else:
        selected_archive_salons = []
        if current_user["role"] == "salon":
            selected_archive_salons = [selected_salon_name]
        else:
            selected_archive_salons = selected_salons_for_archive if selected_salons_for_archive else registered_salons

        archive_result = cached_load_archive_data(tuple(selected_archive_salons), supplier_rule_items)
        manifest_view = archive_result.manifest.copy()
        data = archive_result.data.copy()

        for warning in archive_result.warnings:
            st.warning(warning)

        if data.empty:
            if current_user["role"] == "salon":
                st.info("У этого салона пока нет сохранённых выгрузок. Загрузите первый файл и сохраните его в архив.")
            else:
                st.info("В архиве пока нет данных по выбранным салонам.")
            st.stop()

        source_label = "Архив сети" if is_network_role(current_user["role"]) else f"Архив салона: {selected_salon_name}"

if prepared_result is not None and work_mode in {"Разовая загрузка", "Новая выгрузка", "Загрузка салона"}:
    for warning in prepared_result.warnings:
        st.warning(warning)

if work_mode not in upload_modes:
    selected_archive_salons = []
    if current_user["role"] == "salon":
        selected_archive_salons = [selected_salon_name]
    else:
        selected_archive_salons = selected_salons_for_archive if selected_salons_for_archive else registered_salons

    archive_result = cached_load_archive_data(tuple(selected_archive_salons), supplier_rule_items)
    manifest_view = archive_result.manifest.copy()
    data = archive_result.data.copy()

    for warning in archive_result.warnings:
        st.warning(warning)

    if data.empty:
        if current_user["role"] == "salon":
            st.info("У этого салона пока нет сохранённых выгрузок. Загрузите первый файл и сохраните его в архив.")
        else:
            st.info("В архиве пока нет данных по выбранным салонам.")
        st.stop()

    source_label = "Архив сети" if is_network_role(current_user["role"]) else f"Архив салона: {selected_salon_name}"

if data.empty:
    st.warning("После применения фильтров не осталось данных.")
    st.stop()

procurement_items = load_procurement_items()
data = enrich_sales_with_supplier(
    data,
    procurement_items,
    supplier_rule_items=supplier_rule_items,
    supplier_product_assignments=supplier_product_assignments,
)
data = enrich_sales_with_brand(data, procurement_items)

margin_visible = can_view_margin(current_user)
available_history_months = (
    int(data["date"].dropna().dt.to_period("M").nunique())
    if "date" in data.columns and not data.empty
    else 6
)

screen_options = ["Обзор", "ABC-анализ", "Карточка SKU"]
if margin_visible:
    screen_options.append("Финансы")
screen_options.append("Аналитика")
if is_network_role(current_user["role"]):
    screen_options.append("Поставщики")
screen_options.extend(["Закупки", "План / факт", "Данные"])
if can_manage_access(current_user):
    screen_options.append("Управление")

active_screen = screen_options[0]
active_analytics_screen = ""
active_advanced_screen = ""
abc_metric = "revenue"
procurement_history_months = min(
    max(available_history_months, 3),
    6,
)
procurement_coverage_days = 30
procurement_lead_time_days = 14
procurement_safety_days = 7
procurement_min_active_months = 2
selected_categories = []
selected_suppliers = []
selected_brands = []
selected_managers = []
selected_salons_filter = []
analytics_screen_options = []
if margin_visible:
    analytics_screen_options.append("Маржинальность")
analytics_screen_options.extend(["Сравнение месяцев", "Расширенный"])

if (
    st.session_state.get("primary_screen_nav") == "Аналитика"
    and st.session_state.get("analytics_screen_nav") == "ABC-анализ"
):
    st.session_state["primary_screen_nav"] = "ABC-анализ"

with mobile_navigation_slot.container():
    (
        active_screen,
        active_analytics_screen,
        active_advanced_screen,
        abc_metric,
    ) = render_mobile_screen_navigation(
        screen_options,
        analytics_screen_options,
        margin_visible=margin_visible,
    )

with mobile_bottom_navigation_slot.container():
    render_mobile_bottom_navigation(active_screen, screen_options)

with st.sidebar:
    st.markdown('<div class="sidebar-section-label">Рабочее пространство</div>', unsafe_allow_html=True)

    active_screen = render_screen_switcher(
        "Основной экран",
        screen_options,
        key="primary_screen_nav",
        sync_key="mobile_primary_screen_nav",
    )

    if active_screen == "Аналитика":
        active_analytics_screen = render_screen_switcher(
            "Раздел аналитики",
            analytics_screen_options,
            key="analytics_screen_nav",
            sync_key="mobile_analytics_screen_nav",
        )
    else:
        active_analytics_screen = ""

    if active_screen == "Аналитика" and active_analytics_screen == "Расширенный":
        active_advanced_screen = render_screen_switcher(
            "Глубокий разбор",
            ["RFM-анализ", "Тепловая карта", "Прогноз", "Возвраты"],
            key="advanced_screen_nav",
            sync_key="mobile_advanced_screen_nav",
        )
    else:
        active_advanced_screen = ""

    if active_screen == "ABC-анализ":
        abc_metric_options = {"revenue": "Выручка", "quantity": "Количество"}
        if margin_visible:
            abc_metric_options["margin"] = "Маржа"
        abc_metric = st.selectbox(
            "Метрика ABC",
            options=list(abc_metric_options.keys()),
            format_func=abc_metric_options.get,
            index=0,
            key="abc_metric_sidebar",
            on_change=sync_session_value,
            args=("abc_metric_sidebar", "mobile_abc_metric"),
        )

    if active_screen == "Закупки":
        procurement_month_cap = max(3, min(12, int(data["date"].dt.to_period("M").nunique()) if not data.empty else 6))
        with st.expander("Параметры закупки", expanded=False):
            if procurement_month_cap <= 3:
                procurement_history_months = 3
                st.caption("История для прогноза: 3 мес. Чтобы менять период, нужны данные за 4+ месяца.")
            else:
                procurement_history_months = st.slider(
                    "История для прогноза, мес.",
                    min_value=3,
                    max_value=procurement_month_cap,
                    value=min(procurement_history_months, procurement_month_cap),
                    key="procurement_history_months",
                )
            procurement_coverage_days = st.slider(
                "Период покрытия, дней",
                min_value=7,
                max_value=90,
                value=procurement_coverage_days,
                step=1,
                key="procurement_coverage_days",
            )
            procurement_lead_time_days = st.slider(
                "Срок поставки, дней",
                min_value=0,
                max_value=60,
                value=procurement_lead_time_days,
                step=1,
                key="procurement_lead_time_days",
            )
            procurement_safety_days = st.slider(
                "Страховой запас, дней",
                min_value=0,
                max_value=45,
                value=procurement_safety_days,
                step=1,
                key="procurement_safety_days",
            )
            procurement_min_active_months = st.slider(
                "Мин. активных месяцев",
                min_value=1,
                max_value=min(6, procurement_month_cap),
                value=min(procurement_min_active_months, min(6, procurement_month_cap)),
                step=1,
                key="procurement_min_active_months",
            )

    if is_network_role(current_user["role"]):
        with st.expander("Telegram", expanded=False):
            st.caption("Отправляет управленческую сводку, риски, CSV-отчёты и Excel-заказ поставщикам в настроенный чат.")
            telegram_ready = telegram_is_configured()
            if not telegram_ready:
                st.warning("Заполните TG_BOT_TOKEN и TG_CHAT_ID в переменных окружения.")
            if st.button(
                "Отправить отчёт в Telegram",
                key="telegram_send_report_button",
                use_container_width=True,
                disabled=not telegram_ready,
            ):
                try:
                    with st.spinner("Отправляю отчёт в Telegram..."):
                        sent_files = send_telegram_report_pack(with_files=True)
                    audit_event(
                        action="telegram.report_send",
                        user_id=current_user["username"],
                        details={"sent_files": int(sent_files)},
                    )
                    st.success(f"Отчёт отправлен. Файлов: {sent_files}.")
                except Exception as error:
                    st.error(f"Не удалось отправить отчёт: {error}")

filter_source_data = data.copy()
valid_filter_dates = filter_source_data["date"].dropna()
if valid_filter_dates.empty:
    st.error("В данных нет корректных дат для построения аналитики.")
    st.stop()

min_date = valid_filter_dates.min().date()
max_date = valid_filter_dates.max().date()
all_salons = (
    sorted(filter_source_data["salon"].dropna().astype(str).unique().tolist())
    if is_network_role(current_user["role"])
    and "salon" in filter_source_data.columns
    and filter_source_data["salon"].nunique() > 1
    else []
)
all_categories = (
    sorted(filter_source_data["category"].dropna().astype(str).unique().tolist())
    if "category" in filter_source_data.columns and filter_source_data["category"].nunique() > 1
    else []
)
all_suppliers = (
    sorted(filter_source_data["supplier"].dropna().astype(str).unique().tolist())
    if "supplier" in filter_source_data.columns and filter_source_data["supplier"].nunique() > 1
    else []
)
brand_scope_values: set[str] = set()
if "brand" in filter_source_data.columns:
    brand_scope_values.update(
        value
        for value in filter_source_data["brand"].dropna().astype(str).str.strip().tolist()
        if value
    )
if "brand" in procurement_items.columns:
    brand_scope_values.update(
        value
        for value in procurement_items["brand"].dropna().astype(str).str.strip().tolist()
        if value
    )
all_brands = sorted(brand_scope_values) if len(brand_scope_values) > 1 else []
all_managers = (
    sorted(filter_source_data["manager"].dropna().astype(str).unique().tolist())
    if "manager" in filter_source_data.columns and filter_source_data["manager"].nunique() > 1
    else []
)

for state_key, available_values in (
    ("main_selected_salons_filter", all_salons),
    ("main_selected_categories", all_categories),
    ("main_selected_suppliers", all_suppliers),
    ("main_selected_brands", all_brands),
    ("main_selected_managers", all_managers),
):
    if state_key in st.session_state:
        saved_values = st.session_state.get(state_key) or []
        st.session_state[state_key] = [
            value for value in saved_values if value in available_values
        ]

saved_date_range = st.session_state.get("main_selected_dates")
if saved_date_range:
    saved_dates = list(saved_date_range) if isinstance(saved_date_range, (list, tuple)) else [saved_date_range]
    if len(saved_dates) != 2 or saved_dates[0] < min_date or saved_dates[1] > max_date:
        st.session_state.pop("main_selected_dates", None)

with main_col:
    filter_shell = st.container(border=True)
    with filter_shell:
        render_sticky_header_marker()
        filter_title_col, filter_reset_col = st.columns([4.6, 1], gap="small")
        with filter_title_col:
            st.markdown(
                '<div class="filter-toolbar-heading"><div><div class="filter-toolbar-title">Фильтры и срез данных</div>'
                '<div class="filter-toolbar-meta">Один контекст для всех показателей и отчётов</div></div></div>',
                unsafe_allow_html=True,
            )
        with filter_reset_col:
            if st.button("Сбросить", key="reset_global_filters", use_container_width=True):
                for filter_key in (
                    "main_selected_dates",
                    "main_selected_salons_filter",
                    "main_selected_categories",
                    "main_selected_suppliers",
                    "main_selected_brands",
                    "main_selected_managers",
                ):
                    st.session_state.pop(filter_key, None)
                st.rerun()

        with st.expander("Настроить фильтры", expanded=False):
            date_col, salon_col, category_col = st.columns([1.15, 1, 1], gap="small")
            with date_col:
                selected_dates = st.date_input(
                    "Период продаж",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                    key="main_selected_dates",
                )
            with salon_col:
                if all_salons:
                    selected_salons_filter = st.multiselect(
                        "Салоны",
                        all_salons,
                        default=all_salons,
                        key="main_selected_salons_filter",
                    )
                else:
                    st.text_input("Салоны", value="Текущий контур", disabled=True, key="main_salon_scope_display")
            with category_col:
                if all_categories:
                    selected_categories = st.multiselect(
                        "Категории",
                        all_categories,
                        default=all_categories,
                        key="main_selected_categories",
                    )
                else:
                    st.text_input("Категории", value="Все", disabled=True, key="main_category_scope_display")

            supplier_col, brand_col, manager_col = st.columns(3, gap="small")
            with supplier_col:
                if all_suppliers:
                    selected_suppliers = st.multiselect(
                        "Поставщики",
                        all_suppliers,
                        default=all_suppliers,
                        key="main_selected_suppliers",
                    )
                else:
                    st.text_input("Поставщики", value="Все", disabled=True, key="main_supplier_scope_display")
            with brand_col:
                if all_brands:
                    selected_brands = st.multiselect(
                        "Бренды",
                        all_brands,
                        default=all_brands,
                        key="main_selected_brands",
                    )
                else:
                    st.text_input("Бренды", value="Все", disabled=True, key="main_brand_scope_display")
            with manager_col:
                if all_managers:
                    selected_managers = st.multiselect(
                        "Менеджеры",
                        all_managers,
                        default=all_managers,
                        key="main_selected_managers",
                    )
                else:
                    st.text_input("Менеджеры", value="Все", disabled=True, key="main_manager_scope_display")

filter_badges = []
if len(selected_dates) == 2:
    date_from, date_to = selected_dates
    data = data[(data["date"].dt.date >= date_from) & (data["date"].dt.date <= date_to)]
    if date_from != min_date or date_to != max_date:
        filter_badges.append(f"Период: {date_from:%d.%m.%Y} - {date_to:%d.%m.%Y}")

if all_salons:
    data = data[data["salon"].astype(str).isin(selected_salons_filter)]
    if len(selected_salons_filter) != len(all_salons):
        filter_badges.append(f"Салоны: {len(selected_salons_filter)} из {len(all_salons)}")

plan_fact_source_data = data.copy()

if all_categories:
    data = data[data["category"].astype(str).isin(selected_categories)]
    if len(selected_categories) != len(all_categories):
        filter_badges.append(f"Категории: {len(selected_categories)} из {len(all_categories)}")

if all_suppliers:
    data = data[data["supplier"].astype(str).isin(selected_suppliers)]
    if len(selected_suppliers) != len(all_suppliers):
        filter_badges.append(f"Поставщики: {len(selected_suppliers)} из {len(all_suppliers)}")

procurement_source_data = data.copy()

if all_brands:
    data = data[data["brand"].astype(str).isin(selected_brands)]
    if len(selected_brands) != len(all_brands):
        filter_badges.append(f"Бренды: {len(selected_brands)} из {len(all_brands)}")

procurement_uses_manager_unfiltered_scope = False

if all_managers:
    manager_filtered_data = data[data["manager"].astype(str).isin(selected_managers)]
    procurement_uses_manager_unfiltered_scope = len(manager_filtered_data) != len(data)
    data = manager_filtered_data
    if len(selected_managers) != len(all_managers):
        filter_badges.append(f"Менеджеры: {len(selected_managers)} из {len(all_managers)}")

if filter_badges:
    with main_col:
        render_html_block(
            '<div class="filter-summary">'
            + "".join(f"<span>{escape(badge)}</span>" for badge in filter_badges)
            + "</div>"
        )

if data.empty:
    st.warning("После применения фильтров не осталось данных.")
    st.stop()

has_item_codes = (
    "item_code" in data.columns
    and data["item_code"].fillna("").astype(str).str.strip().ne("").any()
)
product_analysis_column = "product_key" if has_item_codes and "product_key" in data.columns else "product"
overview = cached_build_overview_metrics(data)

needs_product_summary = active_screen in {"Обзор", "Карточка SKU", "Финансы", "ABC-анализ"} or (
    active_screen == "Аналитика" and active_analytics_screen == "Маржинальность"
)
needs_breakdown_summaries = active_screen in {"Обзор", "Финансы"}
needs_monthly_summary = active_screen in {"Обзор", "Данные"} or (
    active_screen == "Аналитика"
    and (
        active_analytics_screen == "Сравнение месяцев"
        or (
            active_analytics_screen == "Расширенный"
            and active_advanced_screen == "Прогноз"
        )
    )
)
needs_returns_overview = active_screen in {"Обзор", "Финансы"} or (
    active_screen == "Аналитика"
    and active_analytics_screen == "Расширенный"
    and active_advanced_screen == "Возвраты"
)
needs_plan_analysis = active_screen in {"Обзор", "Финансы", "План / факт"}
needs_order_analysis = active_screen in {"Обзор", "Карточка SKU", "Финансы", "Закупки"}

product_summary = (
    cached_build_product_summary(data, product_analysis_column)
    if needs_product_summary
    else pd.DataFrame()
)
category_summary = (
    cached_build_product_summary(data, "category")
    if needs_breakdown_summaries
    else pd.DataFrame()
)
manager_summary = (
    cached_build_product_summary(data, "manager")
    if active_screen == "Обзор"
    else pd.DataFrame()
)
salon_summary = (
    cached_build_product_summary(data, "salon")
    if needs_breakdown_summaries and "salon" in data.columns
    else pd.DataFrame()
)
supplier_sales_summary = (
    cached_build_product_summary(data, "supplier")
    if active_screen == "Обзор" and "supplier" in data.columns
    else pd.DataFrame()
)
monthly_summary = (
    cached_build_monthly_summary(data)
    if needs_monthly_summary
    else pd.DataFrame()
)
returns_overview = cached_build_returns_overview(data) if needs_returns_overview else {}

plan_monthly_summary = pd.DataFrame()
monthly_plans = pd.DataFrame()
plan_scope_salons: list[str] = []
scope_plan_summary = pd.DataFrame()
plan_fact_summary = pd.DataFrame()
if needs_plan_analysis:
    plan_monthly_summary = cached_build_monthly_summary(plan_fact_source_data)
    monthly_plans = load_monthly_plans()
    plan_scope_salons = (
        sorted(plan_fact_source_data["salon"].dropna().astype(str).unique().tolist())
        if "salon" in plan_fact_source_data.columns
        else ([current_user.get("salon", "")] if current_user.get("salon") else [])
    )
    allow_network_plan_fallback = bool(
        is_network_role(current_user["role"])
        and registered_salons
        and sorted({str(salon).strip() for salon in plan_scope_salons if str(salon).strip()})
        == sorted({str(salon).strip() for salon in registered_salons})
    )
    scope_plan_summary = build_scope_plan_summary(
        monthly_plans,
        plan_scope_salons,
        allow_network_fallback=allow_network_plan_fallback,
    )
    plan_fact_summary = cached_build_plan_fact_summary(plan_monthly_summary, scope_plan_summary)

procurement_orders = pd.DataFrame()
procurement_order_items = pd.DataFrame()
open_procurement_orders = pd.DataFrame()
procurement_order_overview = pd.DataFrame()
if needs_order_analysis:
    procurement_orders = load_procurement_orders()
    procurement_order_items = load_procurement_order_items()
    open_procurement_orders = build_open_order_summary(procurement_orders, procurement_order_items)
    if active_screen == "Закупки":
        procurement_order_overview = build_procurement_order_overview(
            procurement_orders,
            procurement_order_items,
        )

plan_fact_uses_unfiltered_scope = len(data) != len(plan_fact_source_data)
latest_revenue_delta = monthly_summary.iloc[-1]["revenue_change_pct"] if len(monthly_summary) >= 2 else float("nan")
latest_margin_delta = monthly_summary.iloc[-1]["margin_change_pct"] if len(monthly_summary) >= 2 else float("nan")
procurement_items_for_forecast = procurement_items.copy()
if all_suppliers and "supplier" in procurement_items_for_forecast.columns:
    procurement_supplier_text = (
        procurement_items_for_forecast["supplier"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", "Не назначен")
    )
    procurement_items_for_forecast = procurement_items_for_forecast[
        procurement_supplier_text.isin(set(selected_suppliers))
    ].copy()
procurement_forecast = pd.DataFrame()
procurement_overview = cached_build_procurement_overview(procurement_forecast)
procurement_supplier_summary = cached_build_procurement_supplier_summary(procurement_forecast)
needs_procurement_analysis = needs_order_analysis
if needs_procurement_analysis:
    procurement_forecast = cached_build_procurement_forecast(
        procurement_source_data,
        history_months=procurement_history_months,
        coverage_days=procurement_coverage_days,
        lead_time_days=procurement_lead_time_days,
        safety_days=procurement_safety_days,
        min_active_months=procurement_min_active_months,
        procurement_items=procurement_items_for_forecast,
        inbound_orders=open_procurement_orders,
    )
    if all_brands and "brand" in procurement_forecast.columns:
        procurement_brand_text = (
            procurement_forecast["brand"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Не назначен")
        )
        procurement_forecast = procurement_forecast[
            procurement_brand_text.isin(set(selected_brands))
        ].copy()
    procurement_overview = cached_build_procurement_overview(procurement_forecast)
    procurement_supplier_summary = cached_build_procurement_supplier_summary(procurement_forecast)

with main_col:
    render_dataset_hero(
        source_label,
        data,
        overview,
        selected_categories,
        selected_managers,
        current_user,
        active_screen=active_screen,
    )

    if active_screen == "Обзор":
        _rev_delta_str = (
            f"{format_change_percent(latest_revenue_delta)} к предыдущему месяцу"
            if not is_missing(latest_revenue_delta)
            else ""
        )
        _mar_delta_str = (
            f"{format_change_percent(latest_margin_delta)} к предыдущему месяцу"
            if not is_missing(latest_margin_delta)
            else ""
        )
        overview_cards = [
            {
                "label": "Выручка",
                "value": format_money(overview["total_revenue"]),
                "delta": _rev_delta_str,
            },
        ]
        if margin_visible:
            overview_cards.extend(
                [
                    {
                        "label": "Валовая маржа",
                        "value": format_money(overview["total_margin"]),
                        "delta": _mar_delta_str,
                    },
                    {
                        "label": "Маржинальность",
                        "value": format_percent(overview["margin_pct"]),
                        "delta": "",
                    },
                ]
            )

        if not plan_fact_summary.empty:
            latest_plan_execution = plan_fact_summary.iloc[-1].get("revenue_execution_pct")
            overview_cards.append(
                {
                    "label": "Выполнение плана",
                    "value": (
                        format_percent(latest_plan_execution)
                        if not is_missing(latest_plan_execution)
                        else "Не задан"
                    ),
                    "delta": str(plan_fact_summary.iloc[-1].get("month_label", "")),
                    "tone": (
                        "success"
                        if not is_missing(latest_plan_execution) and float(latest_plan_execution) >= 100
                        else "warning"
                    ),
                }
            )

        overview_cards.append(
            {
                "label": "К закупке",
                "value": f"{format_number(procurement_overview.get('reorder_sku_count', 0))} SKU",
                "delta": f"Критичных: {format_number(procurement_overview.get('critical_stock_count', 0))}",
                "tone": "danger" if procurement_overview.get("critical_stock_count", 0) else "success",
            }
        )
        overview_cards.append(
            {
                "label": "Продано единиц",
                "value": format_number(overview["total_quantity"]),
                "delta": f"Позиций: {format_number(overview['product_count'])}",
            }
        )
        render_metric_cards(overview_cards)

if active_screen == "Карточка SKU":
    with main_col:
        render_section_intro(
            "Карточка SKU",
            "Быстрый drill-down по товару: паспорт, продажи, возвраты, остатки и рекомендация к закупке в одном месте.",
        )

    pending_product_id = st.session_state.pop("sku_portfolio_pending_product_id", "")
    if pending_product_id:
        st.session_state["sku_card_mode"] = "Один SKU"
        st.session_state["product_detail_selector"] = str(pending_product_id)

    with main_col:
        sku_card_mode = st.radio(
            "Режим анализа",
            ["Один SKU", "Портфель поставщика"],
            horizontal=True,
            key="sku_card_mode",
        )

    if sku_card_mode == "Портфель поставщика":
        with main_col:
            portfolio_dimension_label = st.radio(
                "Разрез портфеля",
                ["Поставщик", "Бренд"],
                horizontal=True,
                key="sku_portfolio_dimension",
                help=(
                    "Поставщик — юридический контрагент, у которого закупается товар. "
                    "Бренд — торговая марка или производитель."
                ),
            )
            portfolio_dimension = (
                "supplier"
                if portfolio_dimension_label == "Поставщик"
                else "brand"
            )
            portfolio_values: set[str] = set()
            for portfolio_source in (data, procurement_forecast):
                if portfolio_dimension not in portfolio_source.columns:
                    continue
                portfolio_values.update(
                    value
                    for value in (
                        portfolio_source[portfolio_dimension]
                        .fillna("")
                        .astype(str)
                        .str.strip()
                        .replace("", "Не назначен")
                        .tolist()
                    )
                    if value
                )
            portfolio_options = sorted(
                portfolio_values,
                key=lambda value: (value == "Не назначен", value.casefold()),
            )

            if not portfolio_options:
                st.info(
                    f"В текущем срезе пока нет назначений: {portfolio_dimension_label.lower()}. "
                    "Для уже перенесённых Hettich, Slotex и Nuomi выберите разрез «Бренд»."
                )
                st.stop()

            portfolio_selector_col, portfolio_hint_col = st.columns(
                [1.25, 0.75],
                gap="medium",
            )
            with portfolio_selector_col:
                selected_portfolio_value = st.selectbox(
                    portfolio_dimension_label,
                    portfolio_options,
                    key=f"sku_portfolio_{portfolio_dimension}_selector",
                )
            with portfolio_hint_col:
                st.caption(
                    "Портфель учитывает текущий период, салоны и остальные общие фильтры."
                )

            partner_portfolio, partner_monthly = build_partner_sku_portfolio(
                data,
                procurement_forecast,
                dimension=portfolio_dimension,
                selected_value=selected_portfolio_value,
                product_analysis_column=product_analysis_column,
            )
            if partner_portfolio.empty:
                st.warning(
                    "Для выбранного значения не нашлось SKU в текущем срезе. "
                    "Проверьте общие фильтры и назначения номенклатуры."
                )
                st.stop()

            portfolio_revenue = float(partner_portfolio["revenue"].sum())
            portfolio_margin = float(partner_portfolio["margin"].sum())
            portfolio_margin_pct = (
                portfolio_margin / portfolio_revenue * 100
                if portfolio_revenue
                else float("nan")
            )
            portfolio_stock_value = float(partner_portfolio["stock_value"].sum())
            portfolio_order_qty = float(
                partner_portfolio["recommended_order_qty"].sum()
            )
            portfolio_return_revenue = float(
                partner_portfolio["return_revenue"].sum()
            )
            portfolio_reorder_count = int(
                partner_portfolio["recommended_order_qty"].gt(0).sum()
            )
            portfolio_critical_count = int(
                partner_portfolio["stock_status"]
                .isin(["Нет остатка", "Риск дефицита"])
                .sum()
            )

            st.markdown(
                f"### {portfolio_dimension_label}: {selected_portfolio_value}"
            )
            portfolio_metric_cards = [
                {
                    "label": "Выручка портфеля",
                    "value": format_money(portfolio_revenue),
                    "delta": f"{format_number(len(partner_portfolio))} SKU",
                    "tone": "info",
                },
                {
                    "label": "Продано",
                    "value": format_number(partner_portfolio["quantity"].sum()),
                    "delta": f"{format_number(partner_portfolio['sales_lines'].sum())} строк продаж",
                    "tone": "info",
                },
                {
                    "label": "Сумма склада",
                    "value": format_money(portfolio_stock_value),
                    "delta": "по загруженным остаткам",
                    "tone": "info",
                },
                {
                    "label": "К заказу",
                    "value": format_number(portfolio_order_qty),
                    "delta": f"{format_number(portfolio_reorder_count)} SKU",
                    "tone": "danger" if portfolio_reorder_count else "success",
                },
                {
                    "label": "Дефицит",
                    "value": f"{format_number(portfolio_critical_count)} SKU",
                    "delta": "нет остатка или риск дефицита",
                    "tone": "danger" if portfolio_critical_count else "success",
                },
                {
                    "label": "Возвраты",
                    "value": format_money(portfolio_return_revenue),
                    "delta": f"{format_number(partner_portfolio['return_lines'].sum())} строк",
                    "tone": "warning" if portfolio_return_revenue else "success",
                },
            ]
            if margin_visible:
                portfolio_metric_cards.insert(
                    1,
                    {
                        "label": "Валовая маржа",
                        "value": format_money(portfolio_margin),
                        "delta": (
                            "н/д"
                            if is_missing(portfolio_margin_pct)
                            else format_percent(portfolio_margin_pct)
                        ),
                        "tone": (
                            "success"
                            if not is_missing(portfolio_margin_pct)
                            and portfolio_margin_pct >= 25
                            else "warning"
                        ),
                    },
                )
            render_metric_cards(portfolio_metric_cards)

            with st.container(border=True):
                render_panel_header(
                    "Динамика портфеля",
                    "Продажи всей номенклатуры выбранного поставщика или бренда по месяцам.",
                )
                if (
                    partner_monthly.empty
                    or partner_monthly[["revenue", "quantity"]].abs().sum().sum() == 0
                ):
                    st.info("Недостаточно данных для построения месячной динамики.")
                else:
                    portfolio_trend_chart = go.Figure()
                    portfolio_trend_chart.add_trace(
                        go.Bar(
                            x=partner_monthly["month_label"],
                            y=partner_monthly["revenue"],
                            name="Выручка",
                            marker=dict(
                                color=partner_monthly["revenue"],
                                colorscale=[
                                    [0, CHART_SOFT_BLUE],
                                    [1, PRIMARY_COLOR],
                                ],
                                showscale=False,
                                line=dict(color="rgba(255,255,255,0.5)", width=1),
                            ),
                            hovertemplate="%{x}<br>Выручка: %{y:,.0f}<extra></extra>",
                        )
                    )
                    portfolio_trend_chart.add_trace(
                        go.Scatter(
                            x=partner_monthly["month_label"],
                            y=partner_monthly["quantity"],
                            name="Количество",
                            mode="lines+markers",
                            line=dict(color=SECONDARY_COLOR, width=3, shape="spline"),
                            marker=dict(size=8, color=SECONDARY_COLOR),
                            yaxis="y2",
                            hovertemplate="%{x}<br>Количество: %{y:,.2f}<extra></extra>",
                        )
                    )
                    portfolio_trend_chart.update_layout(
                        xaxis_title="",
                        yaxis_title="Выручка",
                        yaxis2=dict(
                            title="Количество",
                            overlaying="y",
                            side="right",
                            showgrid=False,
                        ),
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=1.02,
                            xanchor="left",
                            x=0,
                        ),
                    )
                    polish_figure(portfolio_trend_chart, height=360)
                    st.plotly_chart(
                        portfolio_trend_chart,
                        width="stretch",
                    )

            portfolio_structure_col, portfolio_stock_col = st.columns(
                2,
                gap="medium",
            )
            with portfolio_structure_col:
                with st.container(border=True):
                    render_panel_header(
                        "Структура ABC",
                        "Ценность SKU внутри выбранного портфеля по выручке.",
                    )
                    portfolio_abc_summary = (
                        partner_portfolio.groupby("abc_class", as_index=False)
                        .agg(
                            sku_count=("product", "nunique"),
                            revenue=("revenue", "sum"),
                        )
                        .sort_values("abc_class")
                    )
                    portfolio_abc_chart = px.pie(
                        portfolio_abc_summary,
                        names="abc_class",
                        values="revenue",
                        hole=0.62,
                        color="abc_class",
                        color_discrete_map={
                            "A": PRIMARY_COLOR,
                            "B": SECONDARY_COLOR,
                            "C": CHART_SOFT_AMBER,
                        },
                        custom_data=["sku_count"],
                    )
                    portfolio_abc_chart.update_traces(
                        textposition="inside",
                        textinfo="label+percent",
                        hovertemplate=(
                            "Класс %{label}<br>Выручка: %{value:,.0f}"
                            "<br>SKU: %{customdata[0]:.0f}<extra></extra>"
                        ),
                    )
                    portfolio_abc_chart.update_layout(showlegend=False)
                    polish_figure(portfolio_abc_chart, height=300)
                    st.plotly_chart(
                        portfolio_abc_chart,
                        width="stretch",
                    )

            with portfolio_stock_col:
                with st.container(border=True):
                    render_panel_header(
                        "Состояние запасов",
                        "Количество SKU по текущему статусу остатка.",
                    )
                    portfolio_stock_summary = (
                        partner_portfolio.groupby("stock_status", as_index=False)
                        .agg(sku_count=("product", "nunique"))
                        .sort_values("sku_count", ascending=True)
                    )
                    portfolio_stock_chart = px.bar(
                        portfolio_stock_summary,
                        x="sku_count",
                        y="stock_status",
                        orientation="h",
                        color="stock_status",
                        color_discrete_map={
                            "Нет остатка": CHART_DANGER_COLOR,
                            "Риск дефицита": CHART_ACCENT_COLOR,
                            "Требуется пополнение": CHART_SOFT_AMBER,
                            "Запас достаточен": PRIMARY_COLOR,
                            "Избыточный запас": CHART_INFO_COLOR,
                            "Нет данных": TEXT_MUTED,
                        },
                        text="sku_count",
                    )
                    portfolio_stock_chart.update_traces(
                        texttemplate="%{text:.0f}",
                        textposition="outside",
                        hovertemplate="%{y}<br>SKU: %{x:.0f}<extra></extra>",
                    )
                    portfolio_stock_chart.update_layout(
                        showlegend=False,
                        xaxis_title="SKU",
                        yaxis_title="",
                    )
                    polish_figure(portfolio_stock_chart, height=300)
                    st.plotly_chart(
                        portfolio_stock_chart,
                        width="stretch",
                    )

            with st.container(border=True):
                render_panel_header(
                    "Полный анализ номенклатуры",
                    "Продажи, маржа, остатки, прогноз и заказ по каждому SKU выбранного портфеля.",
                )
                portfolio_filter_col, portfolio_status_col, portfolio_abc_col = st.columns(
                    [1.1, 0.8, 0.55],
                    gap="small",
                )
                with portfolio_filter_col:
                    portfolio_search = st.text_input(
                        "Поиск по SKU",
                        key=f"sku_portfolio_search_{portfolio_dimension}",
                        placeholder="Код, товар, категория или бренд",
                    ).strip()
                with portfolio_status_col:
                    portfolio_status_options = sorted(
                        partner_portfolio["stock_status"]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    )
                    selected_portfolio_statuses = st.multiselect(
                        "Статус остатка",
                        portfolio_status_options,
                        default=portfolio_status_options,
                        key=f"sku_portfolio_status_{portfolio_dimension}",
                    )
                with portfolio_abc_col:
                    portfolio_abc_options = sorted(
                        partner_portfolio["abc_class"]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    )
                    selected_portfolio_abc = st.multiselect(
                        "ABC",
                        portfolio_abc_options,
                        default=portfolio_abc_options,
                        key=f"sku_portfolio_abc_{portfolio_dimension}",
                    )

                portfolio_view = partner_portfolio.copy()
                if selected_portfolio_statuses:
                    portfolio_view = portfolio_view[
                        portfolio_view["stock_status"].isin(
                            selected_portfolio_statuses
                        )
                    ]
                else:
                    portfolio_view = portfolio_view.iloc[0:0].copy()
                if selected_portfolio_abc:
                    portfolio_view = portfolio_view[
                        portfolio_view["abc_class"].isin(selected_portfolio_abc)
                    ]
                else:
                    portfolio_view = portfolio_view.iloc[0:0].copy()
                if portfolio_search:
                    portfolio_search_text = portfolio_search.casefold()
                    portfolio_search_source = (
                        portfolio_view[
                            [
                                "item_code",
                                "product",
                                "category",
                                "brand",
                                "supplier",
                            ]
                        ]
                        .fillna("")
                        .astype(str)
                        .agg(" ".join, axis=1)
                        .str.casefold()
                    )
                    portfolio_view = portfolio_view[
                        portfolio_search_source.str.contains(
                            portfolio_search_text,
                            regex=False,
                            na=False,
                        )
                    ]

                portfolio_table_columns = [
                    "item_code",
                    "product",
                    "category",
                    "brand",
                    "supplier",
                    "abc_class",
                    "xyz_class",
                    "revenue",
                    "cost",
                    "margin",
                    "margin_pct",
                    "quantity",
                    "avg_price",
                    "return_revenue",
                    "stock_on_hand",
                    "stock_value",
                    "stock_coverage_days",
                    "forecast_qty",
                    "recommended_order_qty",
                    "priority",
                    "stock_status",
                    "last_sale_date",
                ]
                portfolio_table_rename = {
                    "item_code": "Код",
                    "product": "Номенклатура",
                    "category": "Категория",
                    "brand": "Бренд",
                    "supplier": "Поставщик",
                    "abc_class": "ABC",
                    "xyz_class": "XYZ",
                    "revenue": "Выручка",
                    "cost": "Себестоимость",
                    "margin": "Маржа",
                    "margin_pct": "Маржа, %",
                    "quantity": "Продано",
                    "avg_price": "Средняя цена",
                    "return_revenue": "Возвраты",
                    "stock_on_hand": "Остаток",
                    "stock_value": "Сумма остатка",
                    "stock_coverage_days": "Покрытие, дней",
                    "forecast_qty": "Прогноз спроса",
                    "recommended_order_qty": "К заказу",
                    "priority": "Приоритет",
                    "stock_status": "Статус остатка",
                    "last_sale_date": "Последняя продажа",
                }
                st.caption(
                    f"Показано {format_number(len(portfolio_view))} из "
                    f"{format_number(len(partner_portfolio))} SKU."
                )
                st.dataframe(
                    format_display_frame_for_role(
                        portfolio_view,
                        current_user,
                        rename_map=portfolio_table_rename,
                        columns=portfolio_table_columns,
                    ),
                    width="stretch",
                    hide_index=True,
                    height=min(640, max(260, 80 + len(portfolio_view) * 34)),
                )

            with st.container(border=True):
                render_panel_header(
                    "Открыть детальную карточку",
                    "Выберите SKU из портфеля, чтобы перейти к его полной индивидуальной карточке.",
                )
                portfolio_jump_source = (
                    portfolio_view
                    if not portfolio_view.empty
                    else partner_portfolio
                )
                portfolio_jump_lookup = {
                    str(row["portfolio_key"]): (
                        f"{row['item_code']} · {row['product']}"
                        if str(row.get("item_code", "")).strip()
                        else str(row["product"])
                    )
                    for row in portfolio_jump_source.to_dict(orient="records")
                }
                selected_portfolio_sku_key = st.selectbox(
                    "SKU",
                    list(portfolio_jump_lookup),
                    format_func=lambda key: portfolio_jump_lookup.get(key, key),
                    key=f"sku_portfolio_jump_{portfolio_dimension}",
                )
                full_product_options = build_product_detail_options(product_summary)
                matching_product_option = pd.DataFrame()
                if not full_product_options.empty:
                    selected_portfolio_row = portfolio_jump_source[
                        portfolio_jump_source["portfolio_key"]
                        .astype(str)
                        .eq(str(selected_portfolio_sku_key))
                    ].iloc[0]
                    selected_portfolio_product = str(
                        selected_portfolio_row.get("product", "")
                    ).strip()
                    matching_product_option = full_product_options[
                        full_product_options["group_key"]
                        .fillna("")
                        .astype(str)
                        .str.casefold()
                        .eq(str(selected_portfolio_sku_key).casefold())
                    ]
                    if matching_product_option.empty:
                        matching_product_option = full_product_options[
                            full_product_options["product_name"]
                            .fillna("")
                            .astype(str)
                            .str.casefold()
                            .eq(selected_portfolio_product.casefold())
                        ]
                if st.button(
                    "Открыть карточку выбранного SKU",
                    type="primary",
                    width="stretch",
                    disabled=matching_product_option.empty,
                    key=f"sku_portfolio_open_{portfolio_dimension}",
                ):
                    st.session_state["sku_portfolio_pending_product_id"] = str(
                        matching_product_option.iloc[0]["option_id"]
                    )
                    st.rerun()

            portfolio_summary_export = pd.DataFrame(
                [
                    {"Показатель": portfolio_dimension_label, "Значение": selected_portfolio_value},
                    {"Показатель": "SKU", "Значение": len(partner_portfolio)},
                    {"Показатель": "Выручка", "Значение": portfolio_revenue},
                    {"Показатель": "Маржа", "Значение": portfolio_margin},
                    {"Показатель": "Маржа, %", "Значение": portfolio_margin_pct},
                    {"Показатель": "Сумма склада", "Значение": portfolio_stock_value},
                    {"Показатель": "SKU к заказу", "Значение": portfolio_reorder_count},
                    {"Показатель": "Количество к заказу", "Значение": portfolio_order_qty},
                    {"Показатель": "Критичных SKU", "Значение": portfolio_critical_count},
                    {"Показатель": "Возвраты", "Значение": portfolio_return_revenue},
                ]
            )
            safe_portfolio_token = "".join(
                character
                if character.isalnum() or character in {"-", "_"}
                else "_"
                for character in selected_portfolio_value
            ).strip("_")[:60] or "portfolio"
            st.download_button(
                f"Скачать портфель «{selected_portfolio_value}»",
                data=to_excel_report_bytes(
                    {
                        "Сводка": prepare_excel_report_frame(
                            portfolio_summary_export,
                            current_user,
                        ),
                        "Номенклатура": prepare_excel_report_frame(
                            partner_portfolio,
                            current_user,
                            portfolio_table_rename,
                            columns=portfolio_table_columns,
                        ),
                        "Динамика": prepare_excel_report_frame(
                            partner_monthly,
                            current_user,
                        ),
                    },
                    report_title=(
                        f"Портфель {portfolio_dimension_label.lower()}: "
                        f"{selected_portfolio_value}"
                    ),
                ),
                file_name=(
                    f"{portfolio_dimension}_portfolio_{safe_portfolio_token}.xlsx"
                ),
                mime=EXCEL_MIME,
                width="stretch",
            )
        st.stop()

    product_detail_options = build_product_detail_options(product_summary)
    if product_detail_options.empty:
        with main_col:
            st.info("Загрузите продажи, чтобы появилась карточка товара.")
    else:
        product_labels = product_detail_options.set_index("option_id")["option_label"].to_dict()

        with main_col:
            selector_col, selector_meta_col = st.columns([1.55, 0.45], gap="medium")
            with selector_col:
                selected_product_id = st.selectbox(
                    "Товар / SKU",
                    options=product_detail_options["option_id"].tolist(),
                    format_func=lambda option_id: product_labels.get(option_id, option_id),
                    key="product_detail_selector",
                )
            selected_product = product_detail_options.loc[
                product_detail_options["option_id"] == selected_product_id
            ].iloc[0]
            product_frame = filter_product_detail_frame(data, selected_product, product_analysis_column)
            product_returns = extract_return_rows(product_frame)
            product_monthly = build_monthly_summary(product_frame)
            product_monthly_detail = build_sku_monthly_detail(product_monthly)
            product_abc_row = find_product_abc_row(build_abc_analysis(product_summary, "revenue"), selected_product)
            procurement_row = find_procurement_row_for_product(
                procurement_forecast,
                product_frame,
                selected_product,
            )

            product_name = first_display_text(selected_product.get("product_name"), selected_product.get("group_name"))
            product_code = safe_display_text(selected_product.get("item_code"), "")
            product_title = f"{product_code} · {product_name}" if product_code else product_name
            selected_revenue = safe_row_float(selected_product, "revenue", 0.0)
            selected_cost = safe_row_float(selected_product, "cost")
            selected_margin = safe_row_float(selected_product, "margin")
            selected_margin_pct = safe_row_float(selected_product, "margin_pct")
            selected_quantity = safe_row_float(selected_product, "quantity", 0.0)
            selected_lines = safe_row_float(selected_product, "sales_lines", float(len(product_frame)))
            return_revenue = float(
                pd.to_numeric(product_returns.get("revenue", pd.Series(dtype="float64")), errors="coerce")
                .abs()
                .sum()
            )
            stock_on_hand = safe_row_float(procurement_row, "stock_on_hand")
            recommended_order_qty = safe_row_float(procurement_row, "recommended_order_qty", 0.0)
            stock_coverage_days = safe_row_float(procurement_row, "stock_coverage_days")

            with selector_meta_col:
                st.caption("В списке товары отсортированы по выручке.")
                st.caption(f"Всего SKU в текущем срезе: {format_number(len(product_detail_options))}.")

            st.markdown(f"### {product_title}")
            metric_cards = [
                {
                    "label": "Выручка SKU",
                    "value": format_money(selected_revenue),
                    "delta": f"{format_number(selected_lines)} строк продаж",
                    "tone": "info",
                },
                {
                    "label": "Количество",
                    "value": format_number(selected_quantity),
                    "delta": "продано в выбранном срезе",
                    "tone": "info",
                },
                {
                    "label": "Возвраты",
                    "value": format_money(return_revenue),
                    "delta": f"{format_number(len(product_returns))} строк",
                    "tone": "warning" if return_revenue > 0 else "success",
                },
                {
                    "label": "Остаток",
                    "value": "н/д" if is_missing(stock_on_hand) else format_number(stock_on_hand),
                    "delta": "из файла остатков",
                    "tone": "info",
                },
                {
                    "label": "К заказу",
                    "value": format_number(recommended_order_qty),
                    "delta": "по прогнозу потребности",
                    "tone": "danger" if recommended_order_qty > 0 else "success",
                },
            ]
            if margin_visible:
                metric_cards.insert(
                    1,
                    {
                        "label": "Маржа SKU",
                        "value": "н/д" if is_missing(selected_margin) else format_money(selected_margin),
                        "delta": "н/д" if is_missing(selected_margin_pct) else format_percent(selected_margin_pct),
                        "tone": "success" if not is_missing(selected_margin_pct) and selected_margin_pct >= 25 else "warning",
                    },
                )
            render_metric_cards(metric_cards)

            latest_sale_value = (
                procurement_row.get("last_sale_date")
                if not procurement_row.empty and not is_missing(procurement_row.get("last_sale_date"))
                else product_frame["date"].max()
                if not product_frame.empty and "date" in product_frame.columns
                else None
            )
            abc_class = safe_display_text(
                product_abc_row.get("abc_class") if not product_abc_row.empty else procurement_row.get("abc_class"),
                "н/д",
            )
            xyz_class = safe_display_text(procurement_row.get("xyz_class") if not procurement_row.empty else None, "н/д")
            render_snapshot_strip(
                [
                    {
                        "label": "Код товара",
                        "value": product_code or "н/д",
                        "delta": "стабильный ключ анализа",
                        "tone": "info",
                    },
                    {
                        "label": "Категория",
                        "value": most_common_text(product_frame, "category"),
                        "delta": "по строкам продаж",
                        "tone": "info",
                    },
                    {
                        "label": "Поставщик",
                        "value": most_common_text(product_frame, "supplier"),
                        "delta": "из продаж/остатков",
                        "tone": "info",
                    },
                    {
                        "label": "ABC / XYZ",
                        "value": f"{abc_class} / {xyz_class}",
                        "delta": "ценность и стабильность спроса",
                        "tone": "success" if abc_class == "A" else "info",
                    },
                    {
                        "label": "Статус остатка",
                        "value": safe_display_text(procurement_row.get("stock_status") if not procurement_row.empty else None, "н/д"),
                        "delta": "по прогнозу закупки",
                        "tone": "warning" if recommended_order_qty > 0 else "info",
                    },
                    {
                        "label": "Последняя продажа",
                        "value": format_optional_date(latest_sale_value),
                        "delta": "по выбранному срезу",
                        "tone": "info",
                    },
                ]
            )

            with st.container(border=True):
                render_panel_header(
                    "Что делать с этим SKU",
                    "Система собирает сигналы по марже, возвратам, остаткам и прогнозу закупки.",
                )
                render_financial_alerts(
                    build_product_action_cards(
                        selected_product,
                        product_frame,
                        procurement_row,
                        product_returns,
                        allow_margin=margin_visible,
                    )
                )

            detail_chart_col, procurement_col = st.columns([1.25, 0.75], gap="medium")
            with detail_chart_col:
                with st.container(border=True):
                    render_panel_header(
                        "Динамика SKU",
                        "Помесячная динамика выбранной номенклатуры: количество, выручка, средняя цена и маржа для руководителей.",
                    )
                    if product_monthly_detail.empty or product_monthly_detail[["revenue", "quantity"]].abs().sum().sum() == 0:
                        st.info("По этому SKU недостаточно месячной динамики для графика.")
                    else:
                        trend_metric_options = {
                            "quantity": "Количество",
                            "revenue": "Выручка",
                            "avg_price": "Средняя цена",
                        }
                        if margin_visible:
                            trend_metric_options["margin"] = "Маржа"
                        selected_trend_metric = st.radio(
                            "Что показать на графике",
                            options=list(trend_metric_options.keys()),
                            format_func=trend_metric_options.get,
                            horizontal=True,
                            key=f"sku_trend_metric_{selected_product_id}",
                        )
                        trend_series = pd.to_numeric(
                            product_monthly_detail[selected_trend_metric],
                            errors="coerce",
                        ).fillna(0.0)
                        trend_rolling_column = {
                            "quantity": "moving_avg_quantity_3m",
                            "revenue": "moving_avg_revenue_3m",
                            "avg_price": "moving_avg_price_3m",
                            "margin": "moving_avg_margin_3m",
                        }.get(selected_trend_metric, "moving_avg_quantity_3m")
                        trend_rolling = pd.to_numeric(
                            product_monthly_detail[trend_rolling_column],
                            errors="coerce",
                        )
                        latest_trend_value = float(trend_series.iloc[-1]) if len(trend_series) else float("nan")
                        previous_trend_value = float(trend_series.iloc[-2]) if len(trend_series) >= 2 else float("nan")
                        trend_change_pct = calculate_change_pct(latest_trend_value, previous_trend_value)
                        recent_avg_value = float(trend_series.tail(3).mean()) if len(trend_series) else float("nan")
                        active_month_count = int((pd.to_numeric(product_monthly_detail["quantity"], errors="coerce").fillna(0.0).abs() > 0).sum())

                        render_snapshot_strip(
                            [
                                {
                                    "label": f"Последний месяц: {trend_metric_options[selected_trend_metric]}",
                                    "value": format_sku_trend_value(selected_trend_metric, latest_trend_value),
                                    "hint": safe_display_text(product_monthly_detail["month_label"].iloc[-1], "период н/д"),
                                    "tone": "info",
                                },
                                {
                                    "label": "Изменение к прошлому месяцу",
                                    "value": format_change_percent(trend_change_pct),
                                    "hint": "если прошлый месяц есть в данных",
                                    "tone": sku_trend_tone(trend_change_pct),
                                },
                                {
                                    "label": "Среднее за 3 месяца",
                                    "value": format_sku_trend_value(selected_trend_metric, recent_avg_value),
                                    "hint": f"активных месяцев: {format_number(active_month_count)}",
                                    "tone": "info",
                                },
                            ]
                        )
                        sku_trend_chart = go.Figure()
                        sku_trend_chart.add_trace(
                            go.Bar(
                                x=product_monthly_detail["month_label"],
                                y=trend_series,
                                name=trend_metric_options[selected_trend_metric],
                                marker_color=PRIMARY_COLOR,
                            )
                        )
                        sku_trend_chart.add_trace(
                            go.Scatter(
                                x=product_monthly_detail["month_label"],
                                y=trend_rolling,
                                name="Среднее 3 мес.",
                                mode="lines+markers",
                                line=dict(color=SECONDARY_COLOR, width=3),
                            )
                        )
                        forecast_metric_value = float("nan")
                        if selected_trend_metric == "quantity":
                            forecast_metric_value = safe_row_float(procurement_row, "forecast_qty")
                        elif selected_trend_metric == "revenue":
                            forecast_metric_value = safe_row_float(procurement_row, "forecast_revenue")
                        elif selected_trend_metric == "margin" and margin_visible:
                            forecast_metric_value = safe_row_float(procurement_row, "forecast_margin")
                        if not is_missing(forecast_metric_value):
                            sku_trend_chart.add_trace(
                                go.Scatter(
                                    x=["Прогноз"],
                                    y=[forecast_metric_value],
                                    name="Прогноз",
                                    mode="markers+text",
                                    text=["прогноз"],
                                    textposition="top center",
                                    marker=dict(color=CHART_ACCENT_COLOR, size=12, symbol="diamond"),
                                )
                            )
                        sku_trend_chart.update_layout(
                            legend_title="",
                            xaxis_title="",
                            yaxis_title=trend_metric_options[selected_trend_metric],
                        )
                        polish_figure(sku_trend_chart, height=360)
                        st.plotly_chart(sku_trend_chart, use_container_width=True)

                        monthly_detail_columns = [
                            "month_label",
                            "quantity",
                            "revenue",
                            "avg_price",
                            "quantity_change_pct",
                            "revenue_change_pct",
                            "moving_avg_quantity_3m",
                        ]
                        monthly_detail_rename = {
                            "month_label": "Месяц",
                            "quantity": "Количество",
                            "revenue": "Выручка",
                            "avg_price": "Средняя цена",
                            "margin": "Маржа",
                            "avg_margin_per_unit": "Маржа на ед.",
                            "quantity_change_pct": "Изм. количества, %",
                            "revenue_change_pct": "Изм. выручки, %",
                            "margin_change_pct": "Изм. маржи, %",
                            "moving_avg_quantity_3m": "Среднее кол-во 3 мес.",
                            "moving_avg_revenue_3m": "Средняя выручка 3 мес.",
                        }
                        if margin_visible:
                            monthly_detail_columns.insert(3, "margin")
                            monthly_detail_columns.insert(5, "avg_margin_per_unit")
                            monthly_detail_columns.append("margin_change_pct")
                        st.dataframe(
                            format_display_frame_for_role(
                                product_monthly_detail.tail(12),
                                current_user,
                                rename_map=monthly_detail_rename,
                                columns=monthly_detail_columns,
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=260,
                        )

            with procurement_col:
                with st.container(border=True):
                    render_panel_header(
                        "Остатки и закупка",
                        "Короткий статус наличия и рекомендуемого заказа.",
                    )
                    if procurement_row.empty:
                        st.info("Для этого товара пока нет строки в прогнозе закупки. Загрузите остатки или проверьте совпадение названия.")
                    else:
                        procurement_detail = pd.DataFrame([procurement_row.to_dict()])
                        st.dataframe(
                            format_display_frame(
                                procurement_detail,
                                {
                                    "stock_on_hand": "Остаток, шт",
                                    "available_stock_qty": "Доступно, шт",
                                    "stock_coverage_days": "Покрытие, дней",
                                    "forecast_qty": "Прогноз спроса",
                                    "recommended_order_qty": "К заказу, шт",
                                    "priority": "Приоритет",
                                    "stock_status": "Статус",
                                    "notes": "Комментарий",
                                },
                                columns=[
                                    "stock_on_hand",
                                    "available_stock_qty",
                                    "stock_coverage_days",
                                    "forecast_qty",
                                    "recommended_order_qty",
                                    "priority",
                                    "stock_status",
                                    "notes",
                                ],
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=225,
                        )
                        if not is_missing(stock_coverage_days):
                            st.caption(f"Покрытие запаса: примерно {stock_coverage_days:.0f} дней.")

            with st.container(border=True):
                render_panel_header(
                    "Последние строки продаж",
                    "Фактические операции по выбранному SKU. Салоны не видят себестоимость и маржу.",
                )
                if product_frame.empty:
                    st.info("Не нашёл строки продаж по выбранному товару.")
                    product_sales_view = product_frame.copy()
                else:
                    product_sales_view = product_frame.copy()
                    if "date" in product_sales_view.columns:
                        product_sales_view["_sort_date"] = pd.to_datetime(product_sales_view["date"], errors="coerce")
                        product_sales_view = product_sales_view.sort_values("_sort_date", ascending=False)
                    sales_columns = [
                        "date",
                        "item_code",
                        "product",
                        "category",
                        "supplier",
                        "manager",
                        "salon",
                        "quantity",
                        "revenue",
                        "cost",
                        "margin",
                        "margin_pct",
                    ]
                    st.dataframe(
                        format_display_frame_for_role(
                            product_sales_view.head(80).drop(columns=["_sort_date"], errors="ignore"),
                            current_user,
                            {
                                "date": "Дата",
                                "item_code": "Код",
                                "product": "Товар",
                                "category": "Категория",
                                "supplier": "Поставщик",
                                "manager": "Менеджер",
                                "salon": "Салон",
                                "quantity": "Количество",
                                "revenue": "Выручка",
                                "cost": "Себестоимость",
                                "margin": "Маржа",
                                "margin_pct": "Маржа, %",
                            },
                            columns=sales_columns,
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=380,
                    )

            if not product_returns.empty:
                with st.container(border=True):
                    render_panel_header(
                        "Возвраты по SKU",
                        "Отдельно вынесены строки с отрицательной выручкой или количеством.",
                    )
                    st.dataframe(
                        format_display_frame_for_role(
                            product_returns.sort_values("date", ascending=False) if "date" in product_returns.columns else product_returns,
                            current_user,
                            {
                                "date": "Дата",
                                "item_code": "Код",
                                "product": "Товар",
                                "category": "Категория",
                                "manager": "Менеджер",
                                "salon": "Салон",
                                "quantity": "Количество",
                                "revenue": "Сумма возврата",
                                "margin": "Маржа возврата",
                            },
                            columns=[
                                "date",
                                "item_code",
                                "product",
                                "category",
                                "manager",
                                "salon",
                                "quantity",
                                "revenue",
                                "margin",
                            ],
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=260,
                    )

            passport_rows = [
                {"Показатель": "Товар", "Значение": product_name},
                {"Показатель": "Код", "Значение": product_code or "н/д"},
                {"Показатель": "Категория", "Значение": most_common_text(product_frame, "category")},
                {"Показатель": "Поставщик", "Значение": most_common_text(product_frame, "supplier")},
                {"Показатель": "ABC / XYZ", "Значение": f"{abc_class} / {xyz_class}"},
                {"Показатель": "Выручка", "Значение": selected_revenue},
                {"Показатель": "Количество", "Значение": selected_quantity},
                {"Показатель": "Возвраты", "Значение": return_revenue},
                {"Показатель": "Остаток", "Значение": stock_on_hand},
                {"Показатель": "К заказу", "Значение": recommended_order_qty},
                {"Показатель": "Покрытие, дней", "Значение": stock_coverage_days},
            ]
            if margin_visible:
                passport_rows[6:6] = [
                    {"Показатель": "Себестоимость", "Значение": selected_cost},
                    {"Показатель": "Маржа", "Значение": selected_margin},
                    {"Показатель": "Маржа, %", "Значение": selected_margin_pct},
                ]
            passport_export = pd.DataFrame(passport_rows)
            procurement_export = pd.DataFrame([procurement_row.to_dict()]) if not procurement_row.empty else pd.DataFrame()
            sku_export_sheets = {
                "Паспорт SKU": prepare_excel_report_frame(passport_export, current_user),
                "Месячная динамика": prepare_excel_report_frame(
                    product_monthly_detail,
                    current_user,
                    {
                        "month_label": "Месяц",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "quantity": "Количество",
                        "avg_price": "Средняя цена",
                        "avg_margin_per_unit": "Маржа на ед.",
                        "revenue_change_pct": "Изм. выручки, %",
                        "quantity_change_pct": "Изм. количества, %",
                        "avg_price_change_pct": "Изм. средней цены, %",
                        "margin_change_pct": "Изм. маржи, %",
                        "moving_avg_revenue_3m": "Средняя выручка 3 мес.",
                        "moving_avg_quantity_3m": "Среднее количество 3 мес.",
                        "moving_avg_price_3m": "Средняя цена 3 мес.",
                        "moving_avg_margin_3m": "Средняя маржа 3 мес.",
                    },
                    columns=[
                        "month_label",
                        "revenue",
                        "cost",
                        "margin",
                        "quantity",
                        "avg_price",
                        "avg_margin_per_unit",
                        "revenue_change_pct",
                        "quantity_change_pct",
                        "avg_price_change_pct",
                        "margin_change_pct",
                        "moving_avg_revenue_3m",
                        "moving_avg_quantity_3m",
                        "moving_avg_price_3m",
                        "moving_avg_margin_3m",
                    ],
                ),
                "Продажи": prepare_excel_report_frame(
                    product_sales_view.drop(columns=["_sort_date"], errors="ignore"),
                    current_user,
                    {
                        "date": "Дата",
                        "item_code": "Код",
                        "product": "Товар",
                        "category": "Категория",
                        "supplier": "Поставщик",
                        "manager": "Менеджер",
                        "salon": "Салон",
                        "quantity": "Количество",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "margin_pct": "Маржа, %",
                    },
                    columns=[
                        "date",
                        "item_code",
                        "product",
                        "category",
                        "supplier",
                        "manager",
                        "salon",
                        "quantity",
                        "revenue",
                        "cost",
                        "margin",
                        "margin_pct",
                    ],
                ),
                "Закупка": prepare_excel_report_frame(
                    procurement_export,
                    current_user,
                    {
                        "stock_on_hand": "Остаток, шт",
                        "available_stock_qty": "Доступно, шт",
                        "stock_coverage_days": "Покрытие, дней",
                        "forecast_qty": "Прогноз спроса",
                        "recommended_order_qty": "К заказу, шт",
                        "priority": "Приоритет",
                        "stock_status": "Статус",
                        "notes": "Комментарий",
                    },
                    columns=[
                        "stock_on_hand",
                        "available_stock_qty",
                        "stock_coverage_days",
                        "forecast_qty",
                        "recommended_order_qty",
                        "priority",
                        "stock_status",
                        "notes",
                    ],
                ),
            }
            if not product_returns.empty:
                sku_export_sheets["Возвраты"] = prepare_excel_report_frame(product_returns, current_user)

            safe_file_token = "".join(
                char if char.isalnum() or char in {"-", "_"} else "_"
                for char in (product_code or product_name)
            ).strip("_")[:60] or "sku"
            st.download_button(
                "Скачать карточку SKU Excel",
                data=to_excel_report_bytes(sku_export_sheets, report_title=f"Карточка SKU: {product_title}"),
                file_name=f"sku_card_{safe_file_token}.xlsx",
                mime=EXCEL_MIME,
                use_container_width=True,
            )

if active_screen == "Финансы":
    with main_col:
        render_section_intro(
            "Финансовый контроль",
            "Управленческий P&L, вклад в валовую прибыль, деньги в остатках и сигналы, которые помогают быстро понять финансовое здоровье текущего среза.",
        )

    if not margin_visible:
        with main_col:
            st.warning("Финансовый экран доступен только администратору и руководителю.")
    else:
        financial_pnl = build_financial_pnl_frame(data, returns_overview)
        financial_inventory = build_financial_inventory_frame(procurement_forecast)
        latest_inventory_snapshot = load_latest_inventory_snapshot()
        financial_alerts = build_financial_alerts(
            overview,
            category_summary,
            product_summary,
            returns_overview,
            financial_inventory,
            procurement_overview,
            plan_fact_summary,
        )

        pnl_amounts = financial_pnl.set_index("metric")["amount"].to_dict()
        gross_revenue = float(pnl_amounts.get("Валовая выручка", 0.0) or 0.0)
        return_revenue = abs(float(pnl_amounts.get("Возвраты", 0.0) or 0.0))
        net_revenue = float(pnl_amounts.get("Чистая выручка", 0.0) or 0.0)
        total_cost = abs(float(pnl_amounts.get("Себестоимость", 0.0) or 0.0)) if not is_missing(pnl_amounts.get("Себестоимость")) else float("nan")
        total_margin = float(pnl_amounts.get("Валовая маржа", float("nan")) or 0.0)

        estimated_stock_value_total_raw = (
            pd.to_numeric(financial_inventory["stock_value"], errors="coerce").sum(min_count=1)
            if not financial_inventory.empty
            else float("nan")
        )
        estimated_stock_value_total = (
            float(estimated_stock_value_total_raw)
            if pd.notna(estimated_stock_value_total_raw)
            else float("nan")
        )
        uploaded_stock_value = (
            float(latest_inventory_snapshot.get("total_value", 0.0))
            if latest_inventory_snapshot
            else float("nan")
        )
        stock_value_total = (
            uploaded_stock_value
            if not is_missing(uploaded_stock_value)
            else estimated_stock_value_total
        )
        if latest_inventory_snapshot:
            snapshot_date = format_date_value(latest_inventory_snapshot.get("uploaded_at"))
            snapshot_item_count = int(latest_inventory_snapshot.get("item_count", 0) or 0)
            stock_value_delta = (
                f"файл остатков от {snapshot_date} · "
                f"{format_number(snapshot_item_count)} SKU"
            )
        else:
            stock_value_delta = "оценка: остаток × средняя себестоимость"
        monthly_cogs_forecast_raw = (
            pd.to_numeric(financial_inventory["monthly_cogs_forecast"], errors="coerce").sum(min_count=1)
            if not financial_inventory.empty
            else float("nan")
        )
        monthly_cogs_forecast = float(monthly_cogs_forecast_raw) if pd.notna(monthly_cogs_forecast_raw) else float("nan")
        inventory_turnover = (
            monthly_cogs_forecast / stock_value_total
            if not is_missing(monthly_cogs_forecast)
            and not is_missing(stock_value_total)
            and stock_value_total > 0
            else float("nan")
        )
        inventory_coverage_days = (
            stock_value_total / (monthly_cogs_forecast / 30.4)
            if not is_missing(monthly_cogs_forecast)
            and not is_missing(stock_value_total)
            and monthly_cogs_forecast > 0
            else float("nan")
        )

        with main_col:
            render_metric_cards(
                [
                    {
                        "label": "Чистая выручка",
                        "value": format_money(net_revenue),
                        "delta": f"возвраты: {format_money(return_revenue)}",
                        "tone": "info",
                    },
                    {
                        "label": "Валовая маржа",
                        "value": format_money(total_margin),
                        "delta": format_percent(safe_pct(total_margin, net_revenue)),
                        "tone": "success" if safe_pct(total_margin, net_revenue) >= 25 else "warning",
                    },
                    {
                        "label": "Себестоимость",
                        "value": format_money(total_cost),
                        "delta": format_percent(safe_pct(total_cost, net_revenue)),
                        "tone": "info",
                    },
                    {
                        "label": "Сумма склада",
                        "value": format_money(stock_value_total),
                        "delta": stock_value_delta,
                        "tone": "warning" if not is_missing(stock_value_total) and stock_value_total > total_margin else "info",
                    },
                    {
                        "label": "Оборачиваемость",
                        "value": "н/д" if is_missing(inventory_turnover) else f"{inventory_turnover:.1f}x/мес",
                        "delta": "прогноз COGS / остаток",
                        "tone": "success" if not is_missing(inventory_turnover) and inventory_turnover >= 1 else "warning",
                    },
                    {
                        "label": "Покрытие запасов",
                        "value": "н/д" if is_missing(inventory_coverage_days) else f"{inventory_coverage_days:.0f} дней",
                        "delta": "по текущему прогнозу спроса",
                        "tone": "warning" if not is_missing(inventory_coverage_days) and inventory_coverage_days > 90 else "info",
                    },
                ]
            )

            with st.container(border=True):
                render_panel_header(
                    "Финансовые сигналы",
                    "Короткий список того, что требует внимания в текущем срезе.",
                )
                render_financial_alerts(financial_alerts)

            pnl_left, pnl_right = st.columns([1, 1.15], gap="medium")
            with pnl_left:
                with st.container(border=True):
                    render_panel_header(
                        "Управленческий P&L",
                        "От валовой выручки до валовой маржи. Это не бухгалтерская форма, а быстрый управленческий разбор.",
                    )
                    st.dataframe(
                        format_display_frame(
                            financial_pnl,
                            {
                                "metric": "Показатель",
                                "amount": "Сумма",
                                "share_pct": "Доля, %",
                                "comment": "Комментарий",
                            },
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=270,
                    )

            with pnl_right:
                with st.container(border=True):
                    render_panel_header(
                        "Водопад прибыли",
                        "Как продажи, возвраты и себестоимость приводят к валовой марже.",
                    )
                    waterfall_y = [
                        gross_revenue,
                        -return_revenue,
                        0,
                        -total_cost if not is_missing(total_cost) else 0,
                        0,
                    ]
                    waterfall_chart = go.Figure(
                        go.Waterfall(
                            measure=["relative", "relative", "total", "relative", "total"],
                            x=["Валовая выручка", "Возвраты", "Чистая выручка", "Себестоимость", "Валовая маржа"],
                            y=waterfall_y,
                            text=[
                                format_money(gross_revenue),
                                f"-{format_money(return_revenue)}",
                                format_money(net_revenue),
                                f"-{format_money(total_cost)}",
                                format_money(total_margin),
                            ],
                            textposition="outside",
                            connector={"line": {"color": "rgba(100,116,139,0.45)"}},
                            increasing={"marker": {"color": PRIMARY_COLOR}},
                            decreasing={"marker": {"color": "#dc2626"}},
                            totals={"marker": {"color": SECONDARY_COLOR}},
                        )
                    )
                    waterfall_chart.update_layout(yaxis_title="Сумма", xaxis_title="", showlegend=False)
                    polish_figure(waterfall_chart, height=330)
                    st.plotly_chart(waterfall_chart, use_container_width=True)

            finance_left, finance_right = st.columns(2, gap="medium")
            with finance_left:
                with st.container(border=True):
                    render_panel_header(
                        "Вклад категорий в прибыль",
                        "Где зарабатываем валовую маржу и где прибыльность проседает.",
                    )
                    category_finance = category_summary.copy()
                    if not category_finance.empty:
                        category_finance["revenue_share_pct"] = category_finance["revenue"].map(
                            lambda value: safe_pct(float(value or 0.0), net_revenue)
                        )
                        category_finance["margin_share_pct"] = category_finance["margin"].map(
                            lambda value: safe_pct(float(value or 0.0), total_margin)
                        )
                        category_chart_data = category_finance.head(12).sort_values("margin", ascending=True)
                        category_finance_chart = px.bar(
                            category_chart_data,
                            x="margin",
                            y="group_name",
                            orientation="h",
                            color="margin_pct",
                            color_continuous_scale=[CHART_DANGER_COLOR, CHART_SOFT_AMBER, SECONDARY_COLOR],
                            labels={"group_name": "Категория", "margin": "Маржа", "margin_pct": "Маржа, %"},
                        )
                        category_finance_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                        polish_figure(category_finance_chart, height=340)
                        st.plotly_chart(category_finance_chart, use_container_width=True)
                    else:
                        st.info("Категории не найдены в текущем наборе данных.")

            with finance_right:
                with st.container(border=True):
                    render_panel_header(
                        "Деньги в остатках",
                        "Оценка товарного запаса по средней себестоимости из продаж и прогноза закупок.",
                    )
                    stock_value_frame = financial_inventory.dropna(subset=["stock_value"]).head(12).sort_values("stock_value", ascending=True)
                    if stock_value_frame.empty:
                        st.info("Чтобы увидеть деньги в остатках, загрузите остатки и убедитесь, что в продажах есть себестоимость/маржа.")
                    else:
                        stock_value_chart = px.bar(
                            stock_value_frame,
                            x="stock_value",
                            y="product",
                            orientation="h",
                            color="stock_coverage_days",
                            color_continuous_scale=[SECONDARY_COLOR, CHART_SOFT_AMBER, CHART_DANGER_COLOR],
                            labels={"product": "Товар", "stock_value": "Стоимость остатка", "stock_coverage_days": "Покрытие, дней"},
                        )
                        stock_value_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                        polish_figure(stock_value_chart, height=340)
                        st.plotly_chart(stock_value_chart, use_container_width=True)

            detail_left, detail_right = st.columns(2, gap="medium")
            with detail_left:
                with st.container(border=True):
                    render_panel_header(
                        "Категории: финансовая таблица",
                        "Выручка, себестоимость, маржа и доли по категориям.",
                    )
                    category_table_columns = [
                        "group_name",
                        "revenue",
                        "cost",
                        "margin",
                        "margin_pct",
                        "revenue_share_pct",
                        "margin_share_pct",
                        "quantity",
                        "sales_lines",
                    ]
                    if "category_finance" in locals() and not category_finance.empty:
                        st.dataframe(
                            format_display_frame(
                                category_finance,
                                {
                                    "group_name": "Категория",
                                    "revenue": "Выручка",
                                    "cost": "Себестоимость",
                                    "margin": "Маржа",
                                    "margin_pct": "Маржа, %",
                                    "revenue_share_pct": "Доля выручки, %",
                                    "margin_share_pct": "Доля маржи, %",
                                    "quantity": "Количество",
                                    "sales_lines": "Строк продаж",
                                },
                                columns=category_table_columns,
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=360,
                        )
                    else:
                        st.info("Нет данных по категориям.")

            with detail_right:
                with st.container(border=True):
                    render_panel_header(
                        "Салоны: вклад в результат",
                        "Сравнение точек по выручке, себестоимости и валовой марже.",
                    )
                    if not salon_summary.empty and len(salon_summary) > 1:
                        salon_finance = salon_summary.copy()
                        salon_finance["revenue_share_pct"] = salon_finance["revenue"].map(
                            lambda value: safe_pct(float(value or 0.0), net_revenue)
                        )
                        salon_finance["margin_share_pct"] = salon_finance["margin"].map(
                            lambda value: safe_pct(float(value or 0.0), total_margin)
                        )
                        st.dataframe(
                            format_display_frame(
                                salon_finance,
                                {
                                    "group_name": "Салон",
                                    "revenue": "Выручка",
                                    "cost": "Себестоимость",
                                    "margin": "Маржа",
                                    "margin_pct": "Маржа, %",
                                    "revenue_share_pct": "Доля выручки, %",
                                    "margin_share_pct": "Доля маржи, %",
                                    "quantity": "Количество",
                                    "sales_lines": "Строк продаж",
                                },
                                columns=[
                                    "group_name",
                                    "revenue",
                                    "cost",
                                    "margin",
                                    "margin_pct",
                                    "revenue_share_pct",
                                    "margin_share_pct",
                                    "quantity",
                                    "sales_lines",
                                ],
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=360,
                        )
                    else:
                        st.info("Для сравнения салонов нужно выбрать сеть с несколькими салонами.")

            with st.container(border=True):
                render_panel_header(
                    "Товары с замороженными деньгами",
                    "Позиции, где остаток в деньгах самый большой. Здесь удобно искать неликвиды, излишки и риск закупки сверх спроса.",
                )
                if financial_inventory.empty:
                    st.info("Нет данных по остаткам. Загрузите файл остатков в разделе `Закупки`.")
                else:
                    st.dataframe(
                        format_display_frame(
                            financial_inventory.head(40),
                            {
                                "product": "Товар",
                                "category": "Категория",
                                "supplier": "Поставщик",
                                "stock_on_hand": "Остаток, шт",
                                "available_stock_qty": "Доступно, шт",
                                "avg_unit_cost": "Средняя себестоимость",
                                "stock_value": "Стоимость остатка",
                                "available_stock_value": "Стоимость доступного остатка",
                                "monthly_cogs_forecast": "Прогноз себестоимости/мес.",
                                "stock_coverage_days": "Покрытие, дней",
                                "recommended_order_qty": "К заказу, шт",
                                "priority": "Приоритет",
                                "stock_status": "Статус остатка",
                            },
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=430,
                    )

            finance_export_sheets = {
                "P&L": prepare_excel_report_frame(
                    financial_pnl,
                    current_user,
                    {
                        "metric": "Показатель",
                        "amount": "Сумма",
                        "share_pct": "Доля, %",
                        "comment": "Комментарий",
                    },
                ),
                "Категории": prepare_excel_report_frame(
                    category_finance if "category_finance" in locals() else category_summary,
                    current_user,
                    {
                        "group_name": "Категория",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "margin_pct": "Маржа, %",
                        "revenue_share_pct": "Доля выручки, %",
                        "margin_share_pct": "Доля маржи, %",
                        "quantity": "Количество",
                        "sales_lines": "Строк продаж",
                    },
                ),
                "Товары": prepare_excel_report_frame(
                    product_summary,
                    current_user,
                    {
                        "group_name": "Товар",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "margin_pct": "Маржа, %",
                        "quantity": "Количество",
                        "sales_lines": "Строк продаж",
                    },
                ),
                "Остатки в деньгах": prepare_excel_report_frame(
                    financial_inventory,
                    current_user,
                    {
                        "product": "Товар",
                        "category": "Категория",
                        "supplier": "Поставщик",
                        "stock_on_hand": "Остаток, шт",
                        "available_stock_qty": "Доступно, шт",
                        "avg_unit_cost": "Средняя себестоимость",
                        "stock_value": "Стоимость остатка",
                        "available_stock_value": "Стоимость доступного остатка",
                        "monthly_cogs_forecast": "Прогноз себестоимости/мес.",
                        "stock_coverage_days": "Покрытие, дней",
                        "recommended_order_qty": "К заказу, шт",
                        "priority": "Приоритет",
                        "stock_status": "Статус остатка",
                    },
                ),
            }
            if not salon_summary.empty and len(salon_summary) > 1 and "salon_finance" in locals():
                finance_export_sheets["Салоны"] = prepare_excel_report_frame(
                    salon_finance,
                    current_user,
                    {
                        "group_name": "Салон",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "margin_pct": "Маржа, %",
                        "revenue_share_pct": "Доля выручки, %",
                        "margin_share_pct": "Доля маржи, %",
                        "quantity": "Количество",
                        "sales_lines": "Строк продаж",
                    },
                )

            st.download_button(
                "Скачать финансовый отчёт Excel",
                data=to_excel_report_bytes(finance_export_sheets, report_title="Финансовый контроль"),
                file_name="financial_control_report.xlsx",
                mime=EXCEL_MIME,
                use_container_width=True,
            )

if active_screen == "Обзор":
    abc_data = build_abc_analysis(product_summary, "revenue")
    forecast_data = build_forecast(monthly_summary)
    revenue_anomalies = detect_anomalies(monthly_summary)
    insights = build_insights(
        overview,
        monthly_summary,
        category_summary,
        manager_summary,
        product_summary,
        abc_data,
        returns_overview=returns_overview,
        salon_summary=salon_summary,
        anomalies=revenue_anomalies,
        allow_margin=margin_visible,
    )
    if returns_overview["return_lines"] > 0:
        insights.append(
            (
                "Возвраты в текущем контуре",
                f"Возвратных строк: {format_number(returns_overview['return_lines'])}, "
                f"сумма возвратов {format_money(returns_overview['return_revenue'])}, "
                f"доля от валовой выручки {format_percent(returns_overview['return_share_pct'])}.",
            )
        )
    insights = insights[:6]
    latest_month = monthly_summary.iloc[-1]
    top_products = product_summary.head(7)
    overview_focus_cards = build_overview_focus_cards(
        latest_month,
        latest_revenue_delta,
        latest_margin_delta,
        plan_fact_summary,
        procurement_overview,
        returns_overview,
        allow_margin=margin_visible,
    )

    with main_col:
        overview_left, overview_right = st.columns([1.55, 0.75], gap="medium")

        with overview_left:
            with st.container(border=True):
                render_panel_header(
                    "Динамика продаж",
                    "Факт, маржа и ближайший прогноз по месяцам.",
                )
                trend_chart = go.Figure()
                trend_revenue_colors = [CHART_SOFT_BLUE] * len(monthly_summary)
                if trend_revenue_colors:
                    trend_revenue_colors[-1] = PRIMARY_COLOR
                trend_chart.add_trace(
                    go.Bar(
                        x=monthly_summary["month_label"],
                        y=monthly_summary["revenue"],
                        name="Выручка",
                        marker=dict(
                            color=trend_revenue_colors,
                            line=dict(color="rgba(255,255,255,0.72)", width=0.8),
                        ),
                    )
                )
                if margin_visible:
                    trend_chart.add_trace(
                        go.Scatter(
                            x=monthly_summary["month_label"],
                            y=monthly_summary["margin"],
                            name="Маржа",
                            mode="lines+markers",
                            line=dict(color=CHART_ACCENT_COLOR, width=3.4, shape="spline", smoothing=0.45),
                            marker=dict(size=9, color=CHART_ACCENT_COLOR, symbol="circle"),
                        )
                    )
                if not forecast_data.empty:
                    trend_chart.add_trace(
                        go.Bar(
                            x=forecast_data["month_label"],
                            y=forecast_data["revenue"],
                            name="Прогноз выручки",
                            marker=dict(
                                color="#CBD5E1",
                                line=dict(color="rgba(255,255,255,0.7)", width=0.8),
                            ),
                            opacity=0.68,
                        )
                    )
                    if margin_visible and not forecast_data["margin"].isna().all():
                        trend_chart.add_trace(
                            go.Scatter(
                                x=forecast_data["month_label"],
                                y=forecast_data["margin"],
                                name="Прогноз маржи",
                                mode="lines+markers",
                                line=dict(color=CHART_ACCENT_COLOR, width=2.6, dash="dot", shape="spline", smoothing=0.45),
                                marker=dict(symbol="diamond", size=8, color=CHART_ACCENT_COLOR),
                            )
                        )
                trend_chart.update_layout(legend_title="", xaxis_title="", yaxis_title="", bargap=0.24)
                polish_figure(trend_chart, height=340)
                st.plotly_chart(trend_chart, use_container_width=True)
                if not forecast_data.empty:
                    st.caption(f"Прогноз на {len(forecast_data)} мес. по линейному тренду.")

        with overview_right:
            render_overview_action_center(insights, overview_focus_cards)

    if not salon_summary.empty and len(salon_summary) > 1:
        with main_col:
            with st.container(border=True):
                render_panel_header("Салоны сети", "Выручка по точкам, а для руководителя ещё и маржа.")
                salon_chart = go.Figure()
                salon_chart.add_trace(
                    go.Bar(
                        x=salon_summary["group_name"],
                        y=salon_summary["revenue"],
                        name="Выручка",
                        marker=dict(
                            color=gradient_colors(len(salon_summary), CHART_SOFT_BLUE, PRIMARY_COLOR),
                            line=dict(color="rgba(255,255,255,0.72)", width=0.8),
                        ),
                    )
                )
                if margin_visible:
                    salon_chart.add_trace(
                        go.Scatter(
                            x=salon_summary["group_name"],
                            y=salon_summary["margin"],
                            name="Маржа",
                            mode="lines+markers",
                            line=dict(color=CHART_ACCENT_COLOR, width=3.2, shape="spline", smoothing=0.35),
                            marker=dict(size=8, color=CHART_ACCENT_COLOR),
                        )
                    )
                salon_chart.update_layout(xaxis_tickangle=-20, bargap=0.26)
                polish_figure(salon_chart, height=360)
                st.plotly_chart(salon_chart, use_container_width=True)

    if not supplier_sales_summary.empty and len(supplier_sales_summary) > 1:
        with main_col:
            supplier_chart_data = supplier_sales_summary.head(12).sort_values("revenue", ascending=True)
            supplier_table = supplier_sales_summary.copy()
            if overview["total_revenue"]:
                supplier_table["revenue_share_pct"] = supplier_table["revenue"] / overview["total_revenue"] * 100
            else:
                supplier_table["revenue_share_pct"] = 0.0
            with st.container(border=True):
                render_panel_header(
                    "Поставщики",
                    "Продажи, маржа и доля выручки по поставщикам в текущем срезе.",
                )
                supplier_chart_col, supplier_table_col = st.columns([1.05, 0.95], gap="medium")
                with supplier_chart_col:
                    if margin_visible:
                        supplier_sales_chart = px.bar(
                            supplier_chart_data,
                            x="revenue",
                            y="group_name",
                            orientation="h",
                            color="margin_pct",
                            color_continuous_scale=[CHART_DANGER_COLOR, CHART_SOFT_AMBER, SECONDARY_COLOR],
                            labels={
                                "group_name": "Поставщик",
                                "revenue": "Выручка",
                                "margin_pct": "Маржа, %",
                            },
                        )
                        supplier_sales_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                    else:
                        supplier_sales_chart = px.bar(
                            supplier_chart_data,
                            x="revenue",
                            y="group_name",
                            orientation="h",
                            color_discrete_sequence=[PRIMARY_COLOR],
                            labels={"group_name": "Поставщик", "revenue": "Выручка"},
                        )
                        supplier_sales_chart.update_layout(yaxis_title="")
                    supplier_sales_chart.update_yaxes(automargin=True)
                    polish_figure(supplier_sales_chart, height=compact_bar_chart_height(max(len(supplier_chart_data), 3)))
                    st.plotly_chart(supplier_sales_chart, use_container_width=True)
                with supplier_table_col:
                    st.dataframe(
                        format_display_frame_for_role(
                            supplier_table,
                            current_user,
                            {
                                "group_name": "Поставщик",
                                "revenue": "Выручка",
                                "cost": "Себестоимость",
                                "margin": "Маржа",
                                "margin_pct": "Маржа, %",
                                "revenue_share_pct": "Доля выручки, %",
                                "quantity": "Количество",
                                "sales_lines": "Строк продаж",
                            },
                            columns=[
                                "group_name",
                                "revenue",
                                "cost",
                                "margin",
                                "margin_pct",
                                "revenue_share_pct",
                                "quantity",
                                "sales_lines",
                            ],
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=max(260, min(420, 86 + len(supplier_table) * 34)),
                    )

    with main_col:
        second_left, second_right = st.columns(2, gap="medium")
        category_chart_data = category_summary.head(10).sort_values("revenue", ascending=True)
        manager_chart_data = manager_summary.head(10).sort_values("revenue", ascending=True)
        overview_breakdown_height = compact_bar_chart_height(max(len(category_chart_data), len(manager_chart_data), 3))

        with second_left:
            with st.container(border=True):
                render_panel_header("Категории", "Вклад категорий в выручку.")
                if margin_visible:
                    category_chart = px.bar(
                        category_chart_data, x="revenue", y="group_name", orientation="h",
                        color="margin_pct", color_continuous_scale=[CHART_SOFT_BLUE, CHART_INFO_COLOR, PRIMARY_COLOR],
                        labels={"group_name": "Категория", "revenue": "Выручка", "margin_pct": "Маржа %"},
                    )
                    category_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                else:
                    category_chart = px.bar(
                        category_chart_data,
                        x="revenue",
                        y="group_name",
                        orientation="h",
                        color_discrete_sequence=[PRIMARY_COLOR],
                        labels={"group_name": "Категория", "revenue": "Выручка"},
                    )
                    category_chart.update_layout(yaxis_title="")
                category_chart.update_yaxes(automargin=True)
                polish_figure(category_chart, height=overview_breakdown_height)
                st.plotly_chart(category_chart, use_container_width=True)

        with second_right:
            with st.container(border=True):
                render_panel_header("Менеджеры", "Результаты по команде продаж.")
                if margin_visible:
                    manager_chart = px.bar(
                        manager_chart_data, x="revenue", y="group_name", orientation="h",
                        color="margin_pct", color_continuous_scale=[CHART_SOFT_TEAL, CHART_INFO_COLOR, PRIMARY_COLOR],
                        labels={"group_name": "Менеджер", "revenue": "Выручка", "margin_pct": "Маржа %"},
                    )
                    manager_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                else:
                    manager_chart = px.bar(
                        manager_chart_data,
                        x="revenue",
                        y="group_name",
                        orientation="h",
                        color_discrete_sequence=[CHART_INFO_COLOR],
                        labels={"group_name": "Менеджер", "revenue": "Выручка"},
                    )
                    manager_chart.update_layout(yaxis_title="")
                manager_chart.update_yaxes(automargin=True)
                polish_figure(manager_chart, height=overview_breakdown_height)
                st.plotly_chart(manager_chart, use_container_width=True)

        with st.container(border=True):
            render_panel_header("Топ товаров", "Лидеры по выручке.")
            top_products_display = format_display_frame_for_role(
                top_products,
                current_user,
                {"group_name": "Товар", "revenue": "Выручка", "margin": "Маржа", "quantity": "Количество", "margin_pct": "Маржа, %"},
            )
            render_mobile_table_cards(
                top_products_display,
                title_column="Товар",
                field_columns=["Выручка", "Маржа", "Количество", "Маржа, %"],
                max_rows=7,
            )
            st.dataframe(
                top_products_display,
                use_container_width=True,
                hide_index=True,
                height=300,
            )

if active_screen == "ABC-анализ":
    with main_col:
        render_section_intro(
            "ABC-анализ",
            "Показывает, какие товары формируют основную долю выручки, а для руководителя ещё и маржи или объема продаж.",
        )

    abc_metric_options = {"revenue": "Выручка", "quantity": "Количество"}
    if margin_visible:
        abc_metric_options["margin"] = "Маржа"

    abc_tab_data = build_abc_analysis(product_summary, abc_metric)
    abc_classes = (
        abc_tab_data.groupby("abc_class", as_index=False)
        .agg(items=("group_name", "count"), metric_total=("abc_basis", "sum"))
        .sort_values("abc_class")
    )
    abc_share_map = abc_tab_data.groupby("abc_class")["share_pct"].sum().to_dict()
    abc_count_map = abc_tab_data.groupby("abc_class")["group_name"].count().to_dict()

    with main_col:
        render_section_marker(
            "Картина ассортимента",
            "Как распределён результат по классам",
            "Сначала смотрите, какую часть метрики удерживают классы A, B и C, затем переходите к лидерам, структуре и кривой Парето.",
        )
        render_snapshot_strip(
            [
                {
                    "label": "Класс A",
                    "value": format_percent(abc_share_map.get("A", 0)),
                    "hint": f"{format_number(abc_count_map.get('A', 0))} товаров",
                },
                {
                    "label": "Класс B",
                    "value": format_percent(abc_share_map.get("B", 0)),
                    "hint": f"{format_number(abc_count_map.get('B', 0))} товаров",
                },
                {
                    "label": "Класс C",
                    "value": format_percent(abc_share_map.get("C", 0)),
                    "hint": f"{format_number(abc_count_map.get('C', 0))} товаров",
                },
            ]
        )

        abc_chart_l, abc_chart_r = st.columns(2, gap="medium")

        with abc_chart_l:
            with st.container(border=True):
                render_panel_header(
                    "Лидеры по метрике",
                    "Товары с наибольшим вкладом.",
                )
                fig_abc = px.bar(
                    abc_tab_data.head(15).sort_values("abc_basis", ascending=True),
                    x="abc_basis",
                    y="group_name",
                    orientation="h",
                    color="abc_class",
                    color_discrete_map={"A": PRIMARY_COLOR, "B": CHART_ACCENT_COLOR, "C": "#7B8796"},
                    labels={"group_name": "Товар", "abc_basis": abc_metric_options[abc_metric]},
                )
                fig_abc.update_layout(yaxis_title="", coloraxis_showscale=False)
                polish_figure(fig_abc, height=450)
                st.plotly_chart(fig_abc, use_container_width=True)

        with abc_chart_r:
            with st.container(border=True):
                render_panel_header(
                    "Структура ABC",
                    "Распределение всего ассортимента.",
                )
                fig_treemap = px.treemap(
                    abc_tab_data,
                    path=["abc_class", "group_name"],
                    values="abc_basis",
                    color="abc_class",
                    color_discrete_map={"A": PRIMARY_COLOR, "B": CHART_ACCENT_COLOR, "C": "#7B8796"},
                )
                polish_figure(fig_treemap, height=450)
                st.plotly_chart(fig_treemap, use_container_width=True)

        with st.container(border=True):
            render_panel_header(
                "Кривая Парето",
                "Накопленная доля метрики. Линия 80% — ядро ассортимента.",
            )
            pareto_chart_data = abc_tab_data.copy().reset_index(drop=True)
            pareto_chart_data["rank"] = range(1, len(pareto_chart_data) + 1)
            pareto_chart_data["abc_label"] = pareto_chart_data["abc_class"].map(
                {"A": "Класс A", "B": "Класс B", "C": "Класс C"}
            )
            pareto_class_colors = {"A": PRIMARY_COLOR, "B": CHART_ACCENT_COLOR, "C": "#7B8796"}
            pareto_hover = pareto_chart_data[["group_name", "cum_share_pct", "abc_label"]]
            pareto_tick_step = max(1, len(pareto_chart_data) // 12)
            cutoff_80 = pareto_chart_data[pareto_chart_data["cum_share_pct"] >= 80].head(1)
            pareto_fig = go.Figure()
            pareto_fig.add_trace(
                go.Bar(
                    x=pareto_chart_data["rank"],
                    y=pareto_chart_data["share_pct"],
                    name=abc_metric_options[abc_metric],
                    marker=dict(
                        color=pareto_chart_data["abc_class"].map(pareto_class_colors),
                        line=dict(color="rgba(15, 23, 42, 0.14)", width=0.6),
                    ),
                    customdata=pareto_hover.to_numpy(),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Ранг: %{x}<br>"
                        "Доля: %{y:.1f}%<br>"
                        "Накопленная доля: %{customdata[1]:.1f}%<br>"
                        "%{customdata[2]}<extra></extra>"
                    ),
                    showlegend=False,
                )
            )
            pareto_fig.add_trace(
                go.Scatter(
                    x=pareto_chart_data["rank"],
                    y=pareto_chart_data["cum_share_pct"],
                    name="Накопленная доля, %",
                    mode="lines+markers",
                    line=dict(color=CHART_ACCENT_COLOR, width=3.4, shape="spline", smoothing=0.45),
                    marker=dict(size=7.5, color=CHART_ACCENT_COLOR),
                    customdata=pareto_hover.to_numpy(),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "Ранг: %{x}<br>"
                        "Накопленная доля: %{y:.1f}%<br>"
                        "%{customdata[2]}<extra></extra>"
                    ),
                    yaxis="y2",
                )
            )
            pareto_fig.add_hrect(
                y0=80,
                y1=100,
                yref="y2",
                line_width=0,
                fillcolor="rgba(216, 154, 43, 0.08)",
                layer="below",
            )
            pareto_fig.add_shape(
                type="line",
                x0=0.5,
                x1=float(pareto_chart_data["rank"].max()) + 0.5,
                y0=80,
                y1=80,
                xref="x",
                yref="y2",
                line=dict(color=CHART_ACCENT_COLOR, dash="dash", width=1.8),
            )
            pareto_fig.add_annotation(
                x=1,
                y=80,
                xref="x",
                yref="y2",
                text="Порог 80%",
                showarrow=False,
                yshift=14,
                font=dict(size=12, color=TEXT_PRIMARY),
                bgcolor="rgba(255, 255, 255, 0.84)",
            )
            if not cutoff_80.empty:
                cutoff_rank = int(cutoff_80["rank"].iloc[0])
                pareto_fig.add_vline(
                    x=cutoff_rank,
                    line_dash="dot",
                    line_color=PRIMARY_COLOR,
                    line_width=1.6,
                )
                st.caption(
                    f"80% накопленной доли достигается на ранге {cutoff_rank}: "
                    f"{cutoff_80['group_name'].iloc[0]}"
                )
            pareto_fig.update_layout(
                xaxis=dict(
                    title="Ранг товара",
                    tickmode="array",
                    tickvals=pareto_chart_data["rank"][::pareto_tick_step],
                    showgrid=False,
                    zeroline=False,
                ),
                yaxis=dict(
                    title="Доля в метрике, %",
                    rangemode="tozero",
                    gridcolor="rgba(148, 163, 184, 0.18)",
                    zeroline=False,
                ),
                yaxis2=dict(
                    title="Накопленная доля, %",
                    overlaying="y",
                    side="right",
                    range=[0, 105],
                    showgrid=False,
                    zeroline=False,
                ),
                hovermode="x unified",
                bargap=0.18,
            )
            polish_figure(pareto_fig, height=390)
            st.plotly_chart(pareto_fig, width="stretch")

        with st.container(border=True):
            render_panel_header(
                "Детальная таблица",
                "Расшифровка по каждой позиции.",
            )
            abc_display = format_display_frame_for_role(
                abc_tab_data,
                current_user,
                {
                    "group_name": "Товар",
                    "abc_basis": abc_metric_options[abc_metric],
                    "share_pct": "Доля, %",
                    "abc_class": "Класс ABC",
                    "revenue": "Выручка",
                    "margin": "Маржа",
                    "quantity": "Кол-во",
                    "margin_pct": "Маржа, %",
                },
            )
            render_mobile_table_cards(
                abc_display,
                title_column="Товар",
                field_columns=["Класс ABC", abc_metric_options[abc_metric], "Доля, %", "Кол-во"],
                max_rows=8,
            )
            st.dataframe(
                abc_display,
                use_container_width=True,
                hide_index=True,
                height=400,
            )
            st.download_button(
                "Скачать ABC-анализ",
                data=to_excel_report_bytes(
                    {
                        "ABC-анализ": prepare_excel_report_frame(
                            abc_tab_data,
                            current_user,
                            {
                                "group_name": "Товар",
                                "abc_basis": abc_metric_options[abc_metric],
                                "share_pct": "Доля, %",
                                "cum_share_pct": "Накопительная доля, %",
                                "abc_class": "Класс ABC",
                                "revenue": "Выручка",
                                "margin": "Маржа",
                                "quantity": "Количество",
                                "margin_pct": "Маржа, %",
                            },
                        )
                    },
                    report_title="ABC-анализ",
                ),
                file_name=f"abc_analysis_{abc_metric}.xlsx",
                mime=EXCEL_MIME,
            )

if active_screen == "Аналитика" and active_analytics_screen == "Маржинальность" and margin_visible:
    main_col = st.container()
    control_col = main_col

    with main_col:
        render_section_intro(
            "Маржинальность",
            "Показывает, какие товары приносят больше валовой прибыли, а какие создают риск по марже.",
        )

    if data["margin"].isna().all():
        with main_col:
            st.info("Для маржинальности не хватает колонок `Себестоимость` или `Маржа` в исходном файле.")
    else:
        margin_sorted = product_summary.sort_values("margin", ascending=False).copy()
        low_margin = product_summary.sort_values("margin_pct", ascending=True).head(15).copy()
        margin_pct_series = product_summary["margin_pct"].dropna()
        risk_threshold = 20.0
        risk_count = int((product_summary["margin_pct"].fillna(9999) < risk_threshold).sum())
        top_margin_product = margin_sorted.iloc[0] if not margin_sorted.empty else None

        with main_col:
            render_metric_cards(
                [
                    {
                        "label": "Общая маржа",
                        "value": format_money(overview["total_margin"]),
                        "delta": format_percent(overview["margin_pct"]),
                    },
                    {
                        "label": "Средняя маржа по товарам",
                        "value": format_percent(margin_pct_series.mean()) if not margin_pct_series.empty else "н/д",
                        "delta": "Средний процент по текущему ассортименту",
                    },
                    {
                        "label": "Позиции в зоне риска",
                        "value": format_number(risk_count),
                        "delta": f"Ниже {int(risk_threshold)}% по марже",
                    },
                    {
                        "label": "Лидер по валовой марже",
                        "value": str(top_margin_product['group_name']) if top_margin_product is not None else "н/д",
                        "delta": format_money(top_margin_product["margin"]) if top_margin_product is not None else "",
                    },
                ]
            )
            render_section_marker(
                "Прибыль и риск",
                "Что создаёт маржу, а что её съедает",
                "Сначала смотрите лидеров и зону риска, затем переходите к карте распределения и детальной таблице. Такой порядок помогает не потеряться в цифрах и сразу выделить товары для действия.",
            )

        with main_col:
             with st.container(border=True):
                render_panel_header(
                    "Управление видом",
                    "Настройте детализацию маржинального анализа.",
                )
                # Here we could add specific filters for Margin if needed
                st.info("Используйте глобальные фильтры в боковой панели для смены периода или выбора салона.")

        with main_col:
            leaders_left, leaders_right = st.columns(2, gap="medium")

            with leaders_left:
                with st.container(border=True):
                    render_panel_header(
                        "Лидеры по валовой марже",
                        "Крупный список товаров, которые приносят наибольшую сумму маржи.",
                    )
                    fig_margin = px.bar(
                        margin_sorted.head(20).sort_values("margin", ascending=True),
                        x="margin",
                        y="group_name",
                        orientation="h",
                        color="margin_pct",
                        color_continuous_scale=[CHART_SOFT_AMBER, CHART_ACCENT_COLOR, PRIMARY_COLOR],
                        title="",
                        labels={"group_name": "Товар", "margin": "Маржа", "margin_pct": "Маржа %"},
                    )
                    fig_margin.update_layout(coloraxis_showscale=False, yaxis_title="")
                    polish_figure(fig_margin, height=520)
                    st.plotly_chart(fig_margin, use_container_width=True)

            with leaders_right:
                with st.container(border=True):
                    render_panel_header(
                        "Зоны риска по марже",
                        "Показывает товары с самой низкой маржой в процентах.",
                    )
                    fig_low = px.bar(
                        low_margin.sort_values("margin_pct", ascending=True),
                        x="margin_pct",
                        y="group_name",
                        orientation="h",
                        color="margin_pct",
                        color_continuous_scale=[CHART_DANGER_COLOR, "#E9785F", CHART_ACCENT_COLOR],
                        title="",
                        labels={"group_name": "Товар", "margin_pct": "Маржа %"},
                    )
                    fig_low.update_layout(coloraxis_showscale=False, yaxis_title="")
                    polish_figure(fig_low, height=520)
                    st.plotly_chart(fig_low, use_container_width=True)

            with st.container(border=True):
                render_panel_header(
                    "Выручка vs Маржинальность",
                    "Каждая точка — отдельный товар. Правый верхний угол показывает сильные позиции, а левый нижний — слабые.",
                )
                scatter_data = product_summary[product_summary["margin_pct"].notna()].copy()
                if not scatter_data.empty:
                    scatter_data["quantity_bubble"] = build_safe_marker_size(scatter_data["quantity"], absolute=True)
                    scatter_fig = px.scatter(
                        scatter_data,
                        x="revenue",
                        y="margin_pct",
                        size="quantity_bubble",
                        color="margin_pct",
                        color_continuous_scale=[CHART_DANGER_COLOR, CHART_ACCENT_COLOR, PRIMARY_COLOR],
                        hover_name="group_name",
                        labels={"revenue": "Выручка", "margin_pct": "Маржа, %", "quantity_bubble": "Объём продаж"},
                        title="",
                        size_max=60,
                    )
                    scatter_fig.add_hline(y=20, line_dash="dash", line_color=CHART_DANGER_COLOR, annotation_text="Порог 20%")
                    polish_figure(scatter_fig, height=560)
                    st.plotly_chart(scatter_fig, use_container_width=True)
                else:
                    st.info("Недостаточно данных для построения скаттер-плота.")

            with st.container(border=True):
                render_panel_header(
                    "Таблица по маржинальности товаров",
                    "Подробная расшифровка по выручке, себестоимости, марже и количеству.",
                )
                st.dataframe(
                    format_display_frame(
                        margin_sorted,
                        {
                            "group_name": "Товар",
                            "revenue": "Выручка",
                            "cost": "Себестоимость",
                            "margin": "Маржа",
                            "quantity": "Количество",
                            "sales_lines": "Строк продаж",
                            "margin_pct": "Маржа, %",
                        },
                    ),
                    use_container_width=True,
                    hide_index=True,
                    height=640,
                )
                st.download_button(
                    "Скачать маржинальность по товарам",
                    data=to_excel_report_bytes(
                        {
                            "Маржинальность": prepare_excel_report_frame(
                                margin_sorted,
                                current_user,
                                {
                                    "group_name": "Товар",
                                    "revenue": "Выручка",
                                    "cost": "Себестоимость",
                                    "margin": "Маржа",
                                    "quantity": "Количество",
                                    "sales_lines": "Строк продаж",
                                    "margin_pct": "Маржа, %",
                                },
                            )
                        },
                        report_title="Маржинальность по товарам",
                    ),
                    file_name="margin_by_product.xlsx",
                    mime=EXCEL_MIME,
                    key="dl_margin_csv"
                )

if active_screen == "Аналитика" and active_analytics_screen == "Сравнение месяцев":
    main_col = st.container()
    control_col = main_col
    yoy_data = build_yoy_comparison(data)

    with main_col:
        render_section_intro(
            "Сравнение месяцев",
            "Показывает, как меняются продажи и количество между двумя периодами, а для руководителя ещё и маржа.",
        )
        render_section_marker(
            "Сценарий сравнения",
            "Сначала выберите два периода, затем ищите причину изменения",
            "Эта вкладка читается в три шага: сначала смотрите общий результат двух месяцев, потом разбирайте движение товаров и вклад в выручку.",
        )

    available_months = monthly_summary["month_label"].tolist()

    if len(available_months) < 2:
        with main_col:
            st.info("Для сравнения нужно минимум два месяца в данных.")
    else:
        default_left_index = max(len(available_months) - 2, 0)
        default_right_index = len(available_months) - 1

        with main_col:
            with st.container(border=True):
                render_panel_header(
                    "Выбор периодов",
                    "Выберите месяцы для сравнения.",
                )
                selected_left_month = st.selectbox("Базовый месяц", options=available_months, index=default_left_index)
                selected_right_month = st.selectbox(
                    "Месяц сравнения",
                    options=available_months,
                    index=default_right_index,
                )

        month_comparison = build_month_comparison(
            data,
            selected_left_month,
            selected_right_month,
            group_column=product_analysis_column,
        )

        if month_comparison.empty:
            with main_col:
                st.warning("Не получилось собрать сравнение для выбранных месяцев.")
        else:
            selected_left_row = monthly_summary.loc[monthly_summary["month_label"] == selected_left_month].iloc[-1]
            selected_right_row = monthly_summary.loc[monthly_summary["month_label"] == selected_right_month].iloc[-1]
            revenue_change_selected = calculate_change_pct(selected_right_row["revenue"], selected_left_row["revenue"])
            margin_change_selected = calculate_change_pct(selected_right_row["margin"], selected_left_row["margin"])
            quantity_change_selected = calculate_change_pct(selected_right_row["quantity"], selected_left_row["quantity"])

            with main_col:
                comparison_snapshot_items = [
                    {
                        "label": f"Выручка {selected_left_month}",
                        "value": format_money(selected_left_row["revenue"]),
                        "hint": "Базовый месяц",
                    },
                    {
                        "label": f"Выручка {selected_right_month}",
                        "value": format_money(selected_right_row["revenue"]),
                        "delta": format_change_percent(revenue_change_selected),
                        "hint": "К базовому месяцу",
                    },
                    {
                        "label": f"Количество {selected_right_month}",
                        "value": format_number(selected_right_row["quantity"]),
                        "delta": format_change_percent(quantity_change_selected),
                        "hint": "К базовому месяцу",
                    },
                ]
                if margin_visible:
                    comparison_snapshot_items.insert(
                        2,
                        {
                            "label": f"Маржа {selected_right_month}",
                            "value": format_money(selected_right_row["margin"]),
                            "delta": format_change_percent(margin_change_selected),
                            "hint": "К базовому месяцу",
                        },
                    )
                render_snapshot_strip(comparison_snapshot_items)

                chart_left, chart_right = st.columns(2, gap="medium")

                with chart_left:
                    with st.container(border=True):
                        render_panel_header(
                            "Движение товаров",
                            "Какие позиции выросли или просели между периодами.",
                        )
                        movement_chart = build_movement_chart(month_comparison, selected_left_month, selected_right_month)
                        st.plotly_chart(movement_chart, use_container_width=True)

                with chart_right:
                    with st.container(border=True):
                        render_panel_header(
                            "Вклад в изменение выручки",
                            "За счет каких товаров общая выручка выросла или снизилась.",
                        )
                        top_movers = month_comparison.nlargest(5, "revenue_delta")
                        bottom_movers = month_comparison.nsmallest(5, "revenue_delta")
                        left_total = month_comparison[f"revenue_{selected_left_month}"].sum()
                        right_total = month_comparison[f"revenue_{selected_right_month}"].sum()

                        wf_x = (
                            [selected_left_month]
                            + top_movers["group_name"].tolist()
                            + bottom_movers["group_name"].tolist()
                            + [selected_right_month]
                        )
                        wf_y = (
                            [left_total]
                            + top_movers["revenue_delta"].tolist()
                            + bottom_movers["revenue_delta"].tolist()
                            + [right_total]
                        )
                        wf_measure = (
                            ["absolute"]
                            + ["relative"] * len(top_movers)
                            + ["relative"] * len(bottom_movers)
                            + ["total"]
                        )
                        wf_fig = go.Figure(
                            go.Waterfall(
                                x=wf_x,
                                y=wf_y,
                                measure=wf_measure,
                                connector=dict(line=dict(color=PRIMARY_COLOR, width=1)),
                                increasing=dict(marker_color=PRIMARY_COLOR),
                                decreasing=dict(marker_color=CHART_DANGER_COLOR),
                                totals=dict(marker_color=CHART_ACCENT_COLOR),
                                name="Изменение выручки",
                            )
                        )
                        wf_fig.update_layout(title="")
                        polish_figure(wf_fig, height=460)
                        st.plotly_chart(wf_fig, use_container_width=True)

                with st.container(border=True):
                    render_panel_header(
                        "Таблица сравнения месяцев",
                        "Подробная расшифровка по товарам.",
                    )
                    rename_map = {
                        "group_name": "Товар",
                        f"revenue_{selected_left_month}": f"Выручка {selected_left_month}",
                        f"revenue_{selected_right_month}": f"Выручка {selected_right_month}",
                        "revenue_delta": "Изменение выручки",
                        "revenue_delta_pct": "Изменение выручки %",
                        f"margin_{selected_left_month}": f"Маржа {selected_left_month}",
                        f"margin_{selected_right_month}": f"Маржа {selected_right_month}",
                        "margin_delta": "Изменение маржи",
                        "margin_delta_pct": "Изменение маржи %",
                        f"quantity_{selected_left_month}": f"Количество {selected_left_month}",
                        f"quantity_{selected_right_month}": f"Количество {selected_right_month}",
                        "quantity_delta": "Изменение количества",
                        "quantity_delta_pct": "Изменение количества %",
                    }
                    st.dataframe(
                        format_display_frame_for_role(month_comparison, current_user, rename_map),
                        use_container_width=True,
                        hide_index=True,
                        height=520,
                    )
                    st.download_button(
                        "Скачать сравнение месяцев",
                        data=to_excel_report_bytes(
                            {
                                "Сравнение": prepare_excel_report_frame(
                                    month_comparison,
                                    current_user,
                                    rename_map,
                                )
                            },
                            report_title=f"Сравнение месяцев {selected_left_month} vs {selected_right_month}",
                        ),
                        file_name=f"month_comparison_{selected_left_month}_vs_{selected_right_month}.xlsx",
                        mime=EXCEL_MIME,
                        key="dl_month_comp_csv"
                    )

    if not yoy_data.empty and yoy_data["year"].nunique() >= 2:
        with st.container(border=True):
            render_panel_header(
                "Год к году (YoY)",
                "Сравнение одинаковых месяцев разных лет. Этот блок помогает отделить сезонность от реального роста или просадки бизнеса.",
            )
            yoy_metric_options = ["revenue", "quantity"]
            if margin_visible:
                yoy_metric_options.insert(1, "margin")
            yoy_metric = st.radio(
                "Метрика YoY",
                options=yoy_metric_options,
                format_func={"revenue": "Выручка", "margin": "Маржа", "quantity": "Количество"}.get,
                horizontal=True,
                key="yoy_metric_radio",
            )
            yoy_fig = px.bar(
                yoy_data,
                x="month_num",
                y=yoy_metric,
                color="year",
                barmode="group",
                color_discrete_sequence=[PRIMARY_COLOR, SECONDARY_COLOR, CHART_INFO_COLOR, CHART_DANGER_COLOR],
                labels={"month_num": "Месяц", yoy_metric: {"revenue": "Выручка", "margin": "Маржа", "quantity": "Количество"}[yoy_metric], "year": "Год"},
                title="",
            )
            yoy_fig.update_xaxes(tickmode="array", tickvals=list(range(1, 13)), ticktext=["Янв","Фев","Мар","Апр","Май","Июн","Июл","Авг","Сен","Окт","Ноя","Дек"])
            polish_figure(yoy_fig, height=420)
            st.plotly_chart(yoy_fig, use_container_width=True)

if active_screen == "Поставщики":
    render_section_intro(
        "Поставщики",
        "Справочник правил и ручных назначений помогает распределить старые продажи по поставщикам без повторной загрузки отчетов.",
    )

    can_edit_suppliers = can_manage_procurement(current_user)
    missing_supplier_candidates = build_missing_supplier_assignment_candidates(data)
    default_rules = default_supplier_rules_frame()
    supplier_assignment_catalog = build_supplier_assignment_catalog(data, supplier_product_assignments)
    supplier_name_options = build_supplier_name_options(
        data,
        supplier_product_assignments,
        supplier_keyword_rules,
        default_rules,
        procurement_items,
    )
    active_custom_rule_count = len(active_supplier_keyword_rules) if not active_supplier_keyword_rules.empty else 0
    assignment_count = len(supplier_product_assignments) if not supplier_product_assignments.empty else 0

    render_snapshot_strip(
        [
            {
                "label": "Авто-правил",
                "value": format_number(float(len(default_rules) + active_custom_rule_count)),
                "hint": f"встроенных: {format_number(float(len(default_rules)))}, ваших: {format_number(float(active_custom_rule_count))}",
                "tone": "info",
            },
            {
                "label": "Ручных назначений",
                "value": format_number(float(assignment_count)),
                "hint": "товары, где поставщик закреплен вручную",
                "tone": "success" if assignment_count else "info",
            },
            {
                "label": "Без поставщика",
                "value": format_number(float(len(missing_supplier_candidates))),
                "hint": "SKU в текущем срезе, которые еще нужно разобрать",
                "tone": "success" if missing_supplier_candidates.empty else "warning",
            },
        ]
    )

    if not can_edit_suppliers:
        st.info("Редактировать справочник поставщиков может только администратор. Руководитель видит текущие правила и проблемные позиции.")

    with st.container(border=True):
        render_panel_header(
            "Правила авто-распознавания",
            "Если поставщик встречается в названии товара или отдельной строкой-группой в старом Excel, правило назначит поставщика автоматически.",
        )
        rules_left, rules_right = st.columns([1.15, 0.85], gap="medium")
        with rules_left:
            keyword_rules_editor_source = supplier_keyword_rules.copy()
            if keyword_rules_editor_source.empty:
                keyword_rules_editor_source = pd.DataFrame(columns=KEYWORD_RULE_COLUMNS)
            visible_rule_columns = ["supplier", "keyword", "is_active", "updated_by", "updated_at"]
            for column in visible_rule_columns:
                if column not in keyword_rules_editor_source.columns:
                    keyword_rules_editor_source[column] = "" if column != "is_active" else True

            edited_keyword_rules = st.data_editor(
                keyword_rules_editor_source[visible_rule_columns],
                use_container_width=True,
                hide_index=True,
                height=320,
                num_rows="dynamic" if can_edit_suppliers else "fixed",
                disabled=[] if can_edit_suppliers else visible_rule_columns,
                key="supplier_keyword_rules_editor",
                column_config={
                    "supplier": st.column_config.TextColumn("Поставщик", required=True),
                    "keyword": st.column_config.TextColumn("Ключевое слово", required=True),
                    "is_active": st.column_config.CheckboxColumn("Активно", default=True),
                    "updated_by": st.column_config.TextColumn("Кто обновил", disabled=True),
                    "updated_at": st.column_config.TextColumn("Когда", disabled=True),
                },
            )
            if can_edit_suppliers and st.button(
                "Сохранить правила поставщиков",
                key="supplier_keyword_rules_save_button",
                use_container_width=True,
                type="primary",
            ):
                saved_count = replace_supplier_keyword_rules(
                    edited_keyword_rules,
                    updated_by=current_user["username"],
                )
                st.cache_data.clear()
                audit_event(
                    action="supplier.keyword_rules_replace",
                    user_id=current_user["username"],
                    details={"saved_count": int(saved_count)},
                )
                st.session_state["supplier_flash_message"] = f"Правила поставщиков сохранены: {saved_count}."
                st.rerun()

        with rules_right:
            st.markdown("**Встроенные правила**")
            st.caption("Эти правила работают всегда. Ваши правила из таблицы слева имеют приоритет выше встроенных.")
            st.dataframe(
                default_rules.rename(columns={"supplier": "Поставщик", "keyword": "Ключевое слово"}),
                use_container_width=True,
                hide_index=True,
                height=320,
            )

    supplier_flash_message = st.session_state.pop("supplier_flash_message", "")
    if supplier_flash_message:
        st.success(supplier_flash_message)

    with st.container(border=True):
        render_panel_header(
            "Справочник номенклатуры",
            "Выберите товар из текущего среза и закрепите поставщика один раз. Это назначение будет применяться к старым и новым продажам.",
        )

        if supplier_assignment_catalog.empty:
            st.info("Загрузите или выберите продажи, чтобы появился список номенклатуры для привязки поставщиков.")
        else:
            quick_col_product, quick_col_supplier, quick_col_action = st.columns([1.35, 0.85, 0.45], gap="medium")
            product_label_lookup = {
                row["product_key"]: (
                    f"{row['product']} · сейчас: {row['current_supplier']} · выручка: {format_money(row['revenue'])}"
                )
                for row in supplier_assignment_catalog.to_dict(orient="records")
            }
            product_key_options = supplier_assignment_catalog["product_key"].astype(str).tolist()

            with quick_col_product:
                selected_supplier_product_key = st.selectbox(
                    "Номенклатура",
                    options=product_key_options,
                    format_func=lambda key: product_label_lookup.get(key, key),
                    key="supplier_quick_product_key",
                    disabled=not can_edit_suppliers,
                )

            selected_supplier_product_row = supplier_assignment_catalog[
                supplier_assignment_catalog["product_key"].astype(str).eq(str(selected_supplier_product_key))
            ].iloc[0]

            custom_supplier_option = "Ввести нового поставщика"
            supplier_select_options = [custom_supplier_option, *supplier_name_options]
            current_supplier_hint = str(
                selected_supplier_product_row.get("manual_supplier")
                or selected_supplier_product_row.get("current_supplier")
                or ""
            ).strip()
            supplier_option_index = (
                supplier_select_options.index(current_supplier_hint)
                if current_supplier_hint in supplier_select_options
                else 0
            )

            with quick_col_supplier:
                selected_supplier_choice = st.selectbox(
                    "Поставщик",
                    options=supplier_select_options,
                    index=supplier_option_index,
                    key="supplier_quick_supplier_choice",
                    disabled=not can_edit_suppliers,
                )
                if selected_supplier_choice == custom_supplier_option:
                    quick_supplier_value = st.text_input(
                        "Название поставщика",
                        key="supplier_quick_supplier_text",
                        placeholder="Например: Slotex",
                        disabled=not can_edit_suppliers,
                    ).strip()
                else:
                    quick_supplier_value = selected_supplier_choice.strip()

            with quick_col_action:
                st.write("")
                st.write("")
                if st.button(
                    "Назначить",
                    key="supplier_quick_assignment_save_button",
                    use_container_width=True,
                    type="primary",
                    disabled=not can_edit_suppliers,
                ):
                    if not quick_supplier_value:
                        st.warning("Укажите поставщика перед сохранением.")
                    else:
                        assignment_row = pd.DataFrame(
                            [
                                {
                                    "product_key": selected_supplier_product_row["product_key"],
                                    "product": selected_supplier_product_row["product"],
                                    "supplier": quick_supplier_value,
                                }
                            ]
                        )
                        saved_count = upsert_supplier_product_assignments(
                            assignment_row,
                            updated_by=current_user["username"],
                        )
                        st.cache_data.clear()
                        audit_event(
                            action="supplier.product_assignments_upsert",
                            user_id=current_user["username"],
                            details={
                                "saved_count": int(saved_count),
                                "source": "quick_product_assignment",
                                "product_key": str(selected_supplier_product_row["product_key"]),
                                "supplier": quick_supplier_value,
                            },
                        )
                        st.session_state["supplier_flash_message"] = f"Поставщик назначен: {quick_supplier_value}."
                        st.rerun()

            with st.expander("Массовая разборка текущего среза", expanded=not missing_supplier_candidates.empty):
                filter_col, search_col = st.columns([0.45, 1.0], gap="medium")
                with filter_col:
                    supplier_catalog_filter = st.selectbox(
                        "Показать",
                        [
                            "Без поставщика",
                            "Все товары",
                            "С ручным назначением",
                            "Без ручного назначения",
                        ],
                        key="supplier_catalog_status_filter",
                    )
                with search_col:
                    supplier_catalog_search = st.text_input(
                        "Поиск по товару, ключу или категории",
                        key="supplier_catalog_search",
                        placeholder="Например: кромка, Hettich, ЛДСП",
                    ).strip()

                supplier_catalog_view = supplier_assignment_catalog.copy()
                if supplier_catalog_filter == "Без поставщика":
                    supplier_catalog_view = supplier_catalog_view[
                        supplier_catalog_view["assignment_status"].eq("Без поставщика")
                    ]
                elif supplier_catalog_filter == "С ручным назначением":
                    supplier_catalog_view = supplier_catalog_view[
                        supplier_catalog_view["manual_supplier"].fillna("").astype(str).str.strip().ne("")
                    ]
                elif supplier_catalog_filter == "Без ручного назначения":
                    supplier_catalog_view = supplier_catalog_view[
                        supplier_catalog_view["manual_supplier"].fillna("").astype(str).str.strip().eq("")
                    ]

                if supplier_catalog_search:
                    search_text = supplier_catalog_search.casefold()
                    search_source = (
                        supplier_catalog_view[["product_key", "product", "category", "current_supplier", "manual_supplier"]]
                        .fillna("")
                        .astype(str)
                        .agg(" ".join, axis=1)
                        .str.casefold()
                    )
                    supplier_catalog_view = supplier_catalog_view[search_source.str.contains(search_text, na=False)]

                visible_supplier_catalog = supplier_catalog_view.head(300).copy()
                st.caption(
                    f"Показано {format_number(float(len(visible_supplier_catalog)))} из "
                    f"{format_number(float(len(supplier_catalog_view)))} позиций. "
                    "Отметьте строки и впишите нового поставщика."
                )
                if len(supplier_catalog_view) > len(visible_supplier_catalog):
                    st.info("Список ограничен первыми 300 позициями. Используйте поиск, чтобы сузить выборку.")

                assign_all_visible = st.checkbox(
                    "Отметить все показанные строки",
                    key="supplier_catalog_assign_all_visible",
                    disabled=not can_edit_suppliers or visible_supplier_catalog.empty,
                )
                visible_supplier_catalog["assign"] = bool(assign_all_visible)
                supplier_catalog_editor_source = visible_supplier_catalog.drop(columns=["new_supplier"], errors="ignore")

                with st.form("supplier_catalog_assignment_form", clear_on_submit=False):
                    supplier_form_col, helper_form_col = st.columns([0.55, 1.0], gap="medium")
                    with supplier_form_col:
                        bulk_supplier_options = [*supplier_name_options, custom_supplier_option]
                        default_bulk_supplier_index = 0 if supplier_name_options else len(bulk_supplier_options) - 1
                        bulk_supplier_choice = st.selectbox(
                            "Поставщик для отмеченных строк",
                            options=bulk_supplier_options,
                            index=default_bulk_supplier_index,
                            key="supplier_catalog_bulk_supplier_choice",
                            disabled=not can_edit_suppliers,
                        )
                        bulk_supplier_text = st.text_input(
                            "Новый поставщик",
                            key="supplier_catalog_bulk_supplier_text",
                            placeholder="Заполните, если поставщика нет в списке",
                            disabled=not can_edit_suppliers,
                        ).strip()
                    with helper_form_col:
                        st.caption(
                            "Галочки внутри этой формы не запускают пересчёт страницы. "
                            "Отметьте товары, выберите поставщика и нажмите кнопку сохранения."
                        )

                    edited_supplier_catalog = st.data_editor(
                        supplier_catalog_editor_source,
                        use_container_width=True,
                        hide_index=True,
                        height=min(560, max(260, 92 + len(supplier_catalog_editor_source) * 34)),
                        disabled=[
                            column
                            for column in supplier_catalog_editor_source.columns
                            if column != "assign"
                        ]
                        if can_edit_suppliers
                        else list(supplier_catalog_editor_source.columns),
                        key="supplier_catalog_assignment_editor",
                        column_config={
                            "assign": st.column_config.CheckboxColumn("Назначить", default=False, width="small"),
                            "product_key": st.column_config.TextColumn("Ключ товара", width="medium"),
                            "product": st.column_config.TextColumn("Товар", width="large"),
                            "category": st.column_config.TextColumn("Категория", width="medium"),
                            "current_supplier": st.column_config.TextColumn("Поставщик сейчас", width="medium"),
                            "manual_supplier": st.column_config.TextColumn("Ручное назначение", width="medium"),
                            "revenue": st.column_config.NumberColumn("Выручка", format="%.2f"),
                            "quantity": st.column_config.NumberColumn("Количество", format="%.2f"),
                            "line_count": st.column_config.NumberColumn("Строк", format="%d"),
                            "assignment_status": st.column_config.TextColumn("Статус", width="medium"),
                        },
                    )

                    submitted_supplier_catalog = st.form_submit_button(
                        "Назначить поставщика отмеченным строкам",
                        use_container_width=True,
                        type="primary",
                        disabled=not can_edit_suppliers,
                    )

                if submitted_supplier_catalog:
                    selected_supplier_value = (
                        bulk_supplier_text
                        or ("" if bulk_supplier_choice == custom_supplier_option else str(bulk_supplier_choice).strip())
                    )
                    selected_assignment_rows = edited_supplier_catalog[
                        edited_supplier_catalog["assign"].fillna(False).astype(bool)
                    ].copy()
                    if selected_assignment_rows.empty:
                        st.warning("Отметьте хотя бы одну строку перед сохранением.")
                    elif not selected_supplier_value:
                        st.warning("Выберите поставщика или впишите нового.")
                    else:
                        selected_assignment_rows["supplier"] = selected_supplier_value
                        assignment_rows = selected_assignment_rows[["product_key", "product", "supplier"]]
                        saved_count = upsert_supplier_product_assignments(
                            assignment_rows,
                            updated_by=current_user["username"],
                        )
                        st.cache_data.clear()
                        audit_event(
                            action="supplier.product_assignments_upsert",
                            user_id=current_user["username"],
                            details={"saved_count": int(saved_count), "source": "supplier_catalog_bulk_assignment"},
                        )
                        st.session_state["supplier_flash_message"] = (
                            f"Поставщик {selected_supplier_value} назначен позициям: {saved_count}."
                        )
                        st.rerun()

    with st.container(border=True):
        render_panel_header(
            "Неразобранные позиции",
            "Товары, где поставщик пока не найден автоматически. Можно заполнить поставщика прямо в таблице.",
        )

        if missing_supplier_candidates.empty:
            st.success("В текущем срезе нет товаров без поставщика. Если сменить период или салон, здесь появятся только новые неразобранные позиции.")
        else:
            st.caption("Заполните поставщика в нужных строках. Лучше начинать сверху: таблица отсортирована по выручке.")
            with st.form("supplier_missing_assignment_form", clear_on_submit=False):
                edited_missing_suppliers = st.data_editor(
                    missing_supplier_candidates,
                    use_container_width=True,
                    hide_index=True,
                    height=min(520, max(260, 96 + len(missing_supplier_candidates) * 34)),
                    disabled=["product_key", "product", "category", "revenue", "quantity", "line_count"] if can_edit_suppliers else list(missing_supplier_candidates.columns),
                    key="supplier_missing_assignment_editor",
                    column_config={
                        "product_key": st.column_config.TextColumn("Ключ товара", width="medium"),
                        "product": st.column_config.TextColumn("Товар", width="large"),
                        "category": st.column_config.TextColumn("Категория", width="medium"),
                        "revenue": st.column_config.NumberColumn("Выручка", format="%.2f"),
                        "quantity": st.column_config.NumberColumn("Количество", format="%.2f"),
                        "line_count": st.column_config.NumberColumn("Строк", format="%d"),
                        "supplier": st.column_config.TextColumn("Поставщик"),
                    },
                )
                submitted_missing_suppliers = st.form_submit_button(
                    "Сохранить ручные назначения",
                    use_container_width=True,
                    type="primary",
                    disabled=not can_edit_suppliers,
                )

            if submitted_missing_suppliers:
                assignment_rows = edited_missing_suppliers[
                    edited_missing_suppliers["supplier"].fillna("").astype(str).str.strip().ne("")
                ][["product_key", "product", "supplier"]]
                saved_count = upsert_supplier_product_assignments(
                    assignment_rows,
                    updated_by=current_user["username"],
                )
                st.cache_data.clear()
                audit_event(
                    action="supplier.product_assignments_upsert",
                    user_id=current_user["username"],
                    details={"saved_count": int(saved_count), "source": "missing_supplier_candidates"},
                )
                st.session_state["supplier_flash_message"] = f"Ручные назначения сохранены: {saved_count}."
                st.rerun()

    with st.container(border=True):
        render_panel_header(
            "Исправить назначение поставщика",
            "Если товар закрепили не за тем поставщиком, найдите его здесь и выберите правильного. "
            "Изменение перезапишет только ручную привязку этой номенклатуры.",
        )

        assignment_fix_source = supplier_product_assignments.copy()
        if assignment_fix_source.empty:
            st.info("Пока нет ручных назначений. Ошибочные привязки появятся здесь после первого сохранения поставщика.")
        else:
            assignment_fix_columns = ["product_key", "product", "supplier", "updated_by", "updated_at"]
            for column in assignment_fix_columns:
                if column not in assignment_fix_source.columns:
                    assignment_fix_source[column] = ""

            assignment_fix_search = st.text_input(
                "Поиск по номенклатуре или поставщику",
                key="supplier_assignment_fix_search",
                placeholder="Например: кромка, Slotex, Hettich",
            ).strip()

            assignment_fix_view = assignment_fix_source[assignment_fix_columns].copy()
            if assignment_fix_search:
                assignment_fix_query = assignment_fix_search.casefold()
                assignment_fix_mask = (
                    assignment_fix_view[["product_key", "product", "supplier"]]
                    .fillna("")
                    .astype(str)
                    .agg(" ".join, axis=1)
                    .str.casefold()
                    .str.contains(assignment_fix_query, na=False)
                )
                assignment_fix_view = assignment_fix_view[assignment_fix_mask]

            if assignment_fix_view.empty:
                st.warning("По этому поиску ручных назначений не найдено.")
            else:
                assignment_fix_view = assignment_fix_view.sort_values(
                    by=["supplier", "product"],
                    key=lambda series: series.fillna("").astype(str).str.casefold(),
                )
                assignment_fix_labels = {
                    str(row["product_key"]): (
                        f"{row['product'] or row['product_key']} · сейчас: {row['supplier']}"
                    )
                    for row in assignment_fix_view.to_dict(orient="records")
                }
                assignment_fix_options = assignment_fix_view["product_key"].astype(str).tolist()
                selected_assignment_fix_key = st.selectbox(
                    "Номенклатура с ручным назначением",
                    options=assignment_fix_options,
                    format_func=lambda key: assignment_fix_labels.get(key, key),
                    key="supplier_assignment_fix_product_key",
                    disabled=not can_edit_suppliers,
                )
                selected_assignment_fix_row = assignment_fix_view[
                    assignment_fix_view["product_key"].astype(str).eq(str(selected_assignment_fix_key))
                ].iloc[0]

                fix_info_col, fix_supplier_col = st.columns([1.05, 0.95], gap="medium")
                current_assignment_supplier = str(selected_assignment_fix_row.get("supplier", "")).strip()
                current_assignment_product = str(
                    selected_assignment_fix_row.get("product", "") or selected_assignment_fix_key
                ).strip()

                with fix_info_col:
                    st.caption("Текущее ручное назначение")
                    st.markdown(f"**{current_assignment_product}**")
                    st.write(f"Поставщик сейчас: **{current_assignment_supplier or 'Не указан'}**")
                    st.caption(
                        "Если нажать «Снять ручное назначение», товар снова будет определяться "
                        "по авто-правилам и данным из файла."
                    )

                with fix_supplier_col:
                    fix_custom_supplier_option = "Ввести нового поставщика"
                    known_fix_suppliers = [
                        str(supplier).strip()
                        for supplier in supplier_name_options
                        if str(supplier).strip()
                    ]
                    if current_assignment_supplier:
                        known_fix_suppliers = [current_assignment_supplier, *known_fix_suppliers]
                    known_fix_suppliers = list(dict.fromkeys(known_fix_suppliers))
                    fix_supplier_options = [*known_fix_suppliers, fix_custom_supplier_option]
                    fix_supplier_index = (
                        fix_supplier_options.index(current_assignment_supplier)
                        if current_assignment_supplier in fix_supplier_options
                        else len(fix_supplier_options) - 1
                    )
                    selected_fix_supplier_choice = st.selectbox(
                        "Новый поставщик",
                        options=fix_supplier_options,
                        index=fix_supplier_index,
                        key="supplier_assignment_fix_supplier_choice",
                        disabled=not can_edit_suppliers,
                    )
                    if selected_fix_supplier_choice == fix_custom_supplier_option:
                        fix_supplier_value = st.text_input(
                            "Название нового поставщика",
                            key="supplier_assignment_fix_supplier_text",
                            placeholder="Например: Slotex",
                            disabled=not can_edit_suppliers,
                        ).strip()
                    else:
                        fix_supplier_value = selected_fix_supplier_choice.strip()

                    save_fix_col, remove_fix_col = st.columns(2, gap="small")
                    with save_fix_col:
                        if st.button(
                            "Изменить поставщика",
                            key="supplier_assignment_fix_save_button",
                            use_container_width=True,
                            type="primary",
                            disabled=not can_edit_suppliers,
                        ):
                            if not fix_supplier_value:
                                st.warning("Выберите поставщика или впишите нового.")
                            else:
                                saved_count = upsert_supplier_product_assignments(
                                    pd.DataFrame(
                                        [
                                            {
                                                "product_key": selected_assignment_fix_key,
                                                "product": current_assignment_product,
                                                "supplier": fix_supplier_value,
                                            }
                                        ]
                                    ),
                                    updated_by=current_user["username"],
                                )
                                st.cache_data.clear()
                                audit_event(
                                    action="supplier.product_assignment_update",
                                    user_id=current_user["username"],
                                    details={
                                        "saved_count": int(saved_count),
                                        "product_key": str(selected_assignment_fix_key),
                                        "old_supplier": current_assignment_supplier,
                                        "new_supplier": fix_supplier_value,
                                    },
                                )
                                st.session_state["supplier_flash_message"] = (
                                    f"Поставщик изменен: {current_assignment_supplier or 'Не указан'} -> {fix_supplier_value}."
                                )
                                st.rerun()
                    with remove_fix_col:
                        if st.button(
                            "Снять ручное назначение",
                            key="supplier_assignment_fix_remove_button",
                            use_container_width=True,
                            disabled=not can_edit_suppliers,
                        ):
                            deleted_count = delete_supplier_product_assignments([str(selected_assignment_fix_key)])
                            st.cache_data.clear()
                            audit_event(
                                action="supplier.product_assignment_delete",
                                user_id=current_user["username"],
                                details={
                                    "deleted_count": int(deleted_count),
                                    "product_key": str(selected_assignment_fix_key),
                                    "old_supplier": current_assignment_supplier,
                                },
                            )
                            st.session_state["supplier_flash_message"] = (
                                "Ручное назначение снято. Теперь поставщик будет определяться автоматически."
                            )
                            st.rerun()

    with st.expander("Существующие ручные назначения", expanded=False):
        existing_assignments = supplier_product_assignments.copy()
        if existing_assignments.empty:
            existing_assignments = pd.DataFrame(columns=PRODUCT_ASSIGNMENT_COLUMNS)
        visible_assignment_columns = ["product_key", "product", "supplier", "updated_by", "updated_at"]
        for column in visible_assignment_columns:
            if column not in existing_assignments.columns:
                existing_assignments[column] = ""
        edited_assignments = st.data_editor(
            existing_assignments[visible_assignment_columns],
            use_container_width=True,
            hide_index=True,
            height=320,
            num_rows="dynamic" if can_edit_suppliers else "fixed",
            disabled=["updated_by", "updated_at"] if can_edit_suppliers else visible_assignment_columns,
            key="supplier_existing_assignments_editor",
            column_config={
                "product_key": st.column_config.TextColumn("Ключ товара", required=True),
                "product": st.column_config.TextColumn("Товар"),
                "supplier": st.column_config.TextColumn("Поставщик", required=True),
                "updated_by": st.column_config.TextColumn("Кто обновил", disabled=True),
                "updated_at": st.column_config.TextColumn("Когда", disabled=True),
            },
        )
        if can_edit_suppliers and st.button(
            "Сохранить таблицу назначений",
            key="supplier_existing_assignments_save_button",
            use_container_width=True,
        ):
            assignment_rows = edited_assignments[
                edited_assignments["supplier"].fillna("").astype(str).str.strip().ne("")
            ][["product_key", "product", "supplier"]]
            saved_count = upsert_supplier_product_assignments(
                assignment_rows,
                updated_by=current_user["username"],
            )
            st.cache_data.clear()
            audit_event(
                action="supplier.product_assignments_upsert",
                user_id=current_user["username"],
                details={"saved_count": int(saved_count), "source": "existing_assignments_editor"},
            )
            st.session_state["supplier_flash_message"] = f"Таблица назначений сохранена: {saved_count}."
            st.rerun()

if active_screen == "Закупки":
    render_section_intro(
        "Закупки",
        "Прогнозирует потребность по SKU на основе последних продаж и помогает собрать приоритетный заказ на следующий цикл поставки.",
    )

    procurement_total_window_days = (
        procurement_coverage_days
        + procurement_lead_time_days
        + procurement_safety_days
    )
    st.caption(
        f"Окно расчёта: {procurement_history_months} мес. истории, "
        f"{procurement_coverage_days} дн. покрытия, "
        f"{procurement_lead_time_days} дн. поставки и "
        f"{procurement_safety_days} дн. страхового запаса."
    )

    if procurement_uses_manager_unfiltered_scope:
        st.info(
            "Прогноз закупки учитывает фильтры по периоду, салонам и категориям, "
            "но не сужается по менеджерам, чтобы персональные продажи не искажали потребность."
        )

    procurement_flash_message = st.session_state.pop("procurement_flash_message", "")
    if procurement_flash_message:
        st.success(procurement_flash_message)
    procurement_error_message = st.session_state.pop("procurement_error_message", "")
    if procurement_error_message:
        st.error(procurement_error_message)

    active_order_statuses = ["draft", *[status for status in PROCUREMENT_ORDER_STATUSES if status in ACTIVE_INBOUND_ORDER_STATUSES]]
    active_procurement_orders = procurement_order_overview[
        procurement_order_overview["status"].isin(active_order_statuses)
    ].copy()
    active_procurement_order_count = (
        int(active_procurement_orders["order_id"].nunique())
        if not active_procurement_orders.empty
        else 0
    )
    catalog_health = build_catalog_health(procurement_items, procurement_forecast)

    with st.container(border=True):
        render_panel_header(
            "Паспорт справочника товаров",
            "Показывает, насколько текущий ассортимент готов к точному прогнозу закупок: поставщики, сроки, правила заказа и связь с продажами.",
        )
        render_snapshot_strip(
            [
                {
                    "label": "Готовность",
                    "value": format_percent(float(catalog_health["readiness_pct"])),
                    "hint": "Чем выше, тем точнее закупочный автопилот",
                    "tone": "success"
                    if float(catalog_health["readiness_pct"]) >= 85
                    else ("warning" if float(catalog_health["readiness_pct"]) >= 60 else "danger"),
                },
                {
                    "label": "SKU в справочнике",
                    "value": format_number(float(catalog_health["catalog_products"])),
                    "hint": f"в прогнозе: {format_number(float(catalog_health['forecast_products']))}",
                    "tone": "info",
                },
                {
                    "label": "Новых вне справочника",
                    "value": format_number(float(catalog_health["missing_in_catalog"])),
                    "hint": "Товары из продаж без карточки закупки",
                    "tone": "success" if int(catalog_health["missing_in_catalog"]) == 0 else "warning",
                },
                {
                    "label": "Без поставщика",
                    "value": format_number(float(catalog_health["missing_supplier"])),
                    "hint": f"поставщиков: {format_number(float(catalog_health['supplier_count']))}",
                    "tone": "success" if int(catalog_health["missing_supplier"]) == 0 else "warning",
                },
                {
                    "label": "Без бренда",
                    "value": format_number(float(catalog_health["missing_brand"])),
                    "hint": f"брендов: {format_number(float(catalog_health['brand_count']))}",
                    "tone": "success" if int(catalog_health["missing_brand"]) == 0 else "warning",
                },
                {
                    "label": "Без срока поставки",
                    "value": format_number(float(catalog_health["missing_lead_time"])),
                    "hint": "Иначе берется общий срок из настроек",
                    "tone": "success" if int(catalog_health["missing_lead_time"]) == 0 else "warning",
                },
                {
                    "label": "Без правил заказа",
                    "value": format_number(float(catalog_health["missing_order_rules"])),
                    "hint": "MOQ или кратность не заданы",
                    "tone": "success" if int(catalog_health["missing_order_rules"]) == 0 else "info",
                },
            ]
        )

    if procurement_forecast.empty:
        st.info(
            "Для прогноза нужны положительные продажи по товарам хотя бы за один месяц "
            "в выбранном контуре."
        )
    else:
        if can_manage_procurement(current_user):
            with st.container(border=True):
                render_panel_header(
                    "Параметры пополнения",
                    "Заполните реальные остатки, поставщика и правила заказа. После сохранения прогноз сразу пересоберётся.",
                )
                st.caption("Редактируются SKU текущего контура. Удобно сначала сузить категории, а потом быстро внести параметры пополнения.")

                with st.expander("Загрузить остатки файлом", expanded=False):
                    st.caption(
                        "Файл может содержать остаток, товар в пути, поставщика и параметры заказа. "
                        "Пустые ячейки не затрут уже сохранённые настройки по SKU."
                    )
                    inventory_upload_file = st.file_uploader(
                        "Файл остатков",
                        type=["xlsx", "xls", "csv"],
                        key="procurement_stock_upload_file",
                    )

                    if inventory_upload_file is not None:
                        inventory_file_bytes = inventory_upload_file.getvalue()
                        try:
                            inventory_filename = validate_uploaded_file(inventory_file_bytes, inventory_upload_file.name)
                        except ValueError as error:
                            st.error(str(error))
                        else:
                            inventory_config_col, inventory_mapping_col = st.columns([1, 1.4], gap="medium")
                            with inventory_config_col:
                                with st.container(border=True):
                                    st.markdown("**Настройка чтения**")
                                    if inventory_filename.lower().endswith(".csv"):
                                        inventory_csv_separator = st.selectbox(
                                            "Разделитель",
                                            options=[";", ",", "\t"],
                                            index=0,
                                            key="procurement_stock_csv_separator",
                                        )
                                        inventory_csv_encoding = st.selectbox(
                                            "Кодировка",
                                            options=["utf-8", "cp1251", "utf-8-sig"],
                                            index=0,
                                            key="procurement_stock_csv_encoding",
                                        )
                                        inventory_sheet_name = None
                                    else:
                                        inventory_sheet_names = list_excel_sheets(inventory_file_bytes)
                                        inventory_sheet_name = st.selectbox(
                                            "Лист Excel",
                                            options=inventory_sheet_names,
                                            index=0,
                                            key="procurement_stock_sheet_name",
                                        )
                                        inventory_csv_separator = ";"
                                        inventory_csv_encoding = "utf-8"

                            try:
                                raw_inventory_data = cached_load_inventory_input_file(
                                    inventory_file_bytes,
                                    inventory_filename,
                                    inventory_csv_separator,
                                    inventory_csv_encoding,
                                    inventory_sheet_name,
                                )
                            except Exception as error:
                                st.error(f"Не удалось прочитать файл остатков: {error}")
                            else:
                                if raw_inventory_data.empty:
                                    st.warning("Файл прочитан, но в нём нет строк.")
                                else:
                                    inventory_columns = raw_inventory_data.columns.astype(str).tolist()
                                    inventory_guesses = guess_inventory_column_mapping(inventory_columns)

                                    with inventory_mapping_col:
                                        with st.container(border=True):
                                            st.markdown("**Сопоставление колонок**")
                                            inventory_map_left, inventory_map_right = st.columns(2)
                                            with inventory_map_left:
                                                inventory_product_col = select_column(
                                                    "Товар (обязательно)",
                                                    inventory_columns,
                                                    inventory_guesses.get("product"),
                                                    key="stock_map_product",
                                                )
                                                inventory_stock_col = select_column(
                                                    "Остаток (обязательно)",
                                                    inventory_columns,
                                                    inventory_guesses.get("stock_on_hand"),
                                                    key="stock_map_on_hand",
                                                )
                                                inventory_stock_value_col = select_column(
                                                    "Сумма склада",
                                                    inventory_columns,
                                                    inventory_guesses.get("stock_value"),
                                                    key="stock_map_value",
                                                )
                                                inventory_in_transit_col = select_column(
                                                    "В пути",
                                                    inventory_columns,
                                                    inventory_guesses.get("stock_in_transit"),
                                                    key="stock_map_in_transit",
                                                )
                                                inventory_supplier_col = select_column(
                                                    "Поставщик",
                                                    inventory_columns,
                                                    inventory_guesses.get("supplier"),
                                                    key="stock_map_supplier",
                                                )
                                            with inventory_map_right:
                                                inventory_brand_col = select_column(
                                                    "Бренд",
                                                    inventory_columns,
                                                    inventory_guesses.get("brand"),
                                                    key="stock_map_brand",
                                                )
                                                inventory_moq_col = select_column(
                                                    "MOQ",
                                                    inventory_columns,
                                                    inventory_guesses.get("min_order_qty"),
                                                    key="stock_map_moq",
                                                )
                                                inventory_multiple_col = select_column(
                                                    "Кратность",
                                                    inventory_columns,
                                                    inventory_guesses.get("order_multiple"),
                                                    key="stock_map_multiple",
                                                )
                                                inventory_lead_time_col = select_column(
                                                    "Срок поставки, дн.",
                                                    inventory_columns,
                                                    inventory_guesses.get("lead_time_days"),
                                                    key="stock_map_lead_time",
                                                )
                                                inventory_notes_col = select_column(
                                                    "Примечание",
                                                    inventory_columns,
                                                    inventory_guesses.get("notes"),
                                                    key="stock_map_notes",
                                                )

                                    inventory_mapping = {
                                        "product": inventory_product_col,
                                        "stock_on_hand": inventory_stock_col,
                                        "stock_value": inventory_stock_value_col,
                                        "stock_in_transit": inventory_in_transit_col,
                                        "supplier": inventory_supplier_col,
                                        "brand": inventory_brand_col,
                                        "min_order_qty": inventory_moq_col,
                                        "order_multiple": inventory_multiple_col,
                                        "lead_time_days": inventory_lead_time_col,
                                        "notes": inventory_notes_col,
                                    }

                                    try:
                                        prepared_inventory_result = cached_prepare_inventory_data(
                                            raw_inventory_data,
                                            tuple(sorted(inventory_mapping.items())),
                                        )
                                    except Exception as error:
                                        st.error(str(error))
                                    else:
                                        for warning in prepared_inventory_result.warnings:
                                            st.warning(warning)

                                        inventory_preview = prepared_inventory_result.data[
                                            [
                                                "product",
                                                "stock_on_hand",
                                                "stock_value",
                                                "stock_in_transit",
                                                "supplier",
                                                "brand",
                                                "min_order_qty",
                                                "order_multiple",
                                                "lead_time_days",
                                                "notes",
                                            ]
                                        ].rename(
                                            columns={
                                                "product": "Товар",
                                                "stock_on_hand": "Остаток",
                                                "stock_value": "Сумма склада",
                                                "stock_in_transit": "В пути",
                                                "supplier": "Поставщик",
                                                "brand": "Бренд",
                                                "min_order_qty": "MOQ",
                                                "order_multiple": "Кратность",
                                                "lead_time_days": "Срок поставки, дн.",
                                                "notes": "Примечание",
                                            }
                                        )
                                        st.dataframe(
                                            inventory_preview,
                                            use_container_width=True,
                                            hide_index=True,
                                            height=260,
                                        )

                                        override_fields = {
                                            field
                                            for field, column_name in inventory_mapping.items()
                                            if field != "product" and column_name
                                        }
                                        stock_value_series = pd.to_numeric(
                                            prepared_inventory_result.data.get(
                                                "stock_value",
                                                pd.Series(dtype="float64"),
                                            ),
                                            errors="coerce",
                                        )
                                        stock_value_total_raw = stock_value_series.sum(min_count=1)
                                        uploaded_stock_value_total = (
                                            float(stock_value_total_raw)
                                            if pd.notna(stock_value_total_raw)
                                            and inventory_stock_value_col
                                            else float("nan")
                                        )
                                        st.caption(
                                            f"К загрузке подготовлено SKU: {format_number(len(prepared_inventory_result.data))}. "
                                            "Обновятся только поля, которые вы сопоставили в файле."
                                        )
                                        if not is_missing(uploaded_stock_value_total):
                                            st.success(
                                                "Сумма склада по файлу: "
                                                f"{format_money(uploaded_stock_value_total)}"
                                            )
                                        else:
                                            st.info(
                                                "Чтобы обновить карточку «Сумма склада», "
                                                "сопоставьте колонку с общей суммой остатка."
                                            )
                                        if st.button(
                                            "Сохранить остатки из файла",
                                            key="procurement_stock_upload_save_button",
                                            use_container_width=True,
                                            type="primary",
                                        ):
                                            saved_count = merge_procurement_upload(
                                                prepared_inventory_result.data,
                                                updated_by=current_user["username"],
                                                override_fields=override_fields,
                                            )
                                            if not is_missing(uploaded_stock_value_total):
                                                save_inventory_snapshot(
                                                    total_value=uploaded_stock_value_total,
                                                    item_count=len(prepared_inventory_result.data),
                                                    filename=inventory_filename,
                                                    uploaded_by=current_user["username"],
                                                )
                                            st.cache_data.clear()
                                            audit_event(
                                                action="procurement.stock_upload",
                                                user_id=current_user["username"],
                                                details={
                                                    "filename": inventory_filename,
                                                    "saved_count": int(saved_count),
                                                    "stock_value_total": (
                                                        uploaded_stock_value_total
                                                        if not is_missing(uploaded_stock_value_total)
                                                        else None
                                                    ),
                                                    "override_fields": sorted(override_fields),
                                                    "categories": selected_categories or [],
                                                    "salons": selected_salons_filter or [],
                                                },
                                            )
                                            st.session_state["procurement_flash_message"] = (
                                                f"Остатки из файла обновлены для {saved_count} SKU."
                                                + (
                                                    f" Сумма склада: {format_money(uploaded_stock_value_total)}."
                                                    if not is_missing(uploaded_stock_value_total)
                                                    else ""
                                                )
                                            )
                                            st.rerun()

                with st.expander("Массово назначить бренд", expanded=False):
                    st.caption(
                        "Выберите несколько номенклатур и укажите один бренд. "
                        "Продажи, остатки и закупочные рекомендации обновятся после сохранения."
                    )
                    with st.form("procurement_bulk_brand_form"):
                        bulk_brand_products = st.multiselect(
                            "Номенклатуры",
                            procurement_forecast["product"].dropna().astype(str).drop_duplicates().tolist(),
                            key="procurement_bulk_brand_products",
                            placeholder="Начните вводить название товара",
                        )
                        bulk_brand_name = st.text_input(
                            "Бренд",
                            key="procurement_bulk_brand_name",
                            placeholder="Например: Hettich",
                        )
                        bulk_brand_submit = st.form_submit_button(
                            "Назначить бренд",
                            type="primary",
                            use_container_width=True,
                        )
                    if bulk_brand_submit:
                        normalized_brand_name = bulk_brand_name.strip()
                        if not bulk_brand_products:
                            st.error("Выберите хотя бы одну номенклатуру.")
                        elif not normalized_brand_name:
                            st.error("Укажите название бренда.")
                        else:
                            bulk_brand_rows = procurement_forecast[
                                procurement_forecast["product"].astype(str).isin(bulk_brand_products)
                            ].copy()
                            bulk_brand_rows["brand"] = normalized_brand_name
                            bulk_brand_rows = bulk_brand_rows.rename(
                                columns={"manual_stock_in_transit": "stock_in_transit"}
                            )[
                                [
                                    "product",
                                    "supplier",
                                    "brand",
                                    "stock_on_hand",
                                    "stock_value",
                                    "stock_in_transit",
                                    "min_order_qty",
                                    "order_multiple",
                                    "lead_time_days",
                                    "notes",
                                ]
                            ]
                            saved_count = upsert_procurement_items(
                                bulk_brand_rows,
                                updated_by=current_user["username"],
                            )
                            st.cache_data.clear()
                            audit_event(
                                action="procurement.brand_bulk_assign",
                                user_id=current_user["username"],
                                details={
                                    "brand": normalized_brand_name,
                                    "saved_count": int(saved_count),
                                    "products": bulk_brand_products,
                                },
                            )
                            st.session_state["procurement_flash_message"] = (
                                f"Бренд «{normalized_brand_name}» назначен для {saved_count} SKU."
                            )
                            st.rerun()

                procurement_editor_source = procurement_forecast[
                    [
                        "product",
                        "supplier",
                        "brand",
                        "stock_on_hand",
                        "stock_value",
                        "manual_stock_in_transit",
                        "ordered_in_transit_qty",
                        "min_order_qty",
                        "order_multiple",
                        "lead_time_days",
                        "notes",
                    ]
                ].copy()
                procurement_editor = st.data_editor(
                    procurement_editor_source,
                    use_container_width=True,
                    hide_index=True,
                    height=460,
                    key="procurement_editor_table",
                    disabled=["product", "ordered_in_transit_qty"],
                    column_config={
                        "product": st.column_config.TextColumn("SKU / Товар"),
                        "supplier": st.column_config.TextColumn("Поставщик"),
                        "brand": st.column_config.TextColumn("Бренд"),
                        "stock_on_hand": st.column_config.NumberColumn("Остаток", min_value=0.0, step=1.0, format="%.2f"),
                        "stock_value": st.column_config.NumberColumn("Сумма остатка", min_value=0.0, step=1.0, format="%.2f"),
                        "manual_stock_in_transit": st.column_config.NumberColumn("В пути (ручной)", min_value=0.0, step=1.0, format="%.2f"),
                        "ordered_in_transit_qty": st.column_config.NumberColumn("В пути по заказам", min_value=0.0, step=1.0, format="%.2f"),
                        "min_order_qty": st.column_config.NumberColumn("MOQ", min_value=0.0, step=1.0, format="%.2f"),
                        "order_multiple": st.column_config.NumberColumn("Кратность", min_value=0.0, step=1.0, format="%.2f"),
                        "lead_time_days": st.column_config.NumberColumn("Срок поставки, дн.", min_value=0, step=1),
                        "notes": st.column_config.TextColumn("Примечание"),
                    },
                )
                if st.button("Сохранить параметры закупки", key="procurement_save_button", use_container_width=True, type="primary"):
                    procurement_editor_to_save = procurement_editor.rename(
                        columns={"manual_stock_in_transit": "stock_in_transit"}
                    )[
                        [
                            "product",
                            "supplier",
                            "brand",
                            "stock_on_hand",
                            "stock_value",
                            "stock_in_transit",
                            "min_order_qty",
                            "order_multiple",
                            "lead_time_days",
                            "notes",
                        ]
                    ].copy()
                    saved_count = upsert_procurement_items(
                        procurement_editor_to_save,
                        updated_by=current_user["username"],
                    )
                    audit_event(
                        action="procurement.settings_upsert",
                        user_id=current_user["username"],
                        details={
                            "saved_count": int(saved_count),
                            "history_months": int(procurement_history_months),
                            "coverage_days": int(procurement_coverage_days),
                            "lead_time_days": int(procurement_lead_time_days),
                            "safety_days": int(procurement_safety_days),
                            "categories": selected_categories or [],
                            "salons": selected_salons_filter or [],
                        },
                    )
                    st.session_state["procurement_flash_message"] = f"Параметры закупки сохранены для {saved_count} SKU."
                    st.rerun()

        procurement_snapshot_items = [
            {
                "label": "SKU в расчёте",
                "value": format_number(procurement_overview["sku_count"]),
                "hint": "Все товары, попавшие в расчёт прогноза",
            },
            {
                "label": "Нужно заказать",
                "value": format_number(procurement_overview["reorder_sku_count"]),
                "hint": "SKU с ненулевой рекомендацией к заказу",
            },
            {
                "label": "Риск дефицита",
                "value": format_number(procurement_overview["critical_stock_count"]),
                "hint": "Позиции без остатка или с риском не дожить до поставки",
            },
            {
                "label": "Рекомендованный заказ",
                "value": format_number(procurement_overview["recommended_order_qty_total"]),
                "hint": f"Уже с учётом MOQ и кратности на окно {procurement_total_window_days} дней",
            },
            {
                "label": "Доступный остаток",
                "value": format_number(procurement_overview["available_stock_qty_total"]),
                "hint": "Сумма остатка и товара в пути по текущему контуру",
            },
            {
                "label": "В пути по заказам",
                "value": format_number(procurement_overview["ordered_in_transit_qty_total"]),
                "hint": "Количество, которое уже заказано и должно доехать",
            },
            {
                "label": "Поставщиков",
                "value": format_number(procurement_overview["supplier_count"]),
                "hint": "Сколько поставщиков уже привязано к SKU в этом срезе",
            },
            {
                "label": "Брендов",
                "value": format_number(procurement_overview["brand_count"]),
                "hint": "Сколько брендов назначено номенклатурам в текущем срезе",
            },
            {
                "label": "Активных заказов",
                "value": format_number(active_procurement_order_count),
                "hint": "Черновики, заказанные и заказы в пути",
            },
        ]
        render_snapshot_strip(procurement_snapshot_items)

        priority_color_map = {
            "Критичный": CHART_DANGER_COLOR,
            "Высокий": CHART_ACCENT_COLOR,
            "Средний": PRIMARY_COLOR,
            "Низкий": TEXT_MUTED,
            "Новый": SECONDARY_COLOR,
            "Спящий": "#CBD5E1",
        }
        stock_status_color_map = {
            "Нет остатка": CHART_DANGER_COLOR,
            "Риск дефицита": CHART_ACCENT_COLOR,
            "Нужно пополнение": PRIMARY_COLOR,
            "Покрыто остатком": SECONDARY_COLOR,
            "Нет спроса": "#CBD5E1",
        }

        stock_risk_frames = build_procurement_stock_risk_frames(
            procurement_forecast,
            total_window_days=procurement_total_window_days,
        )
        stock_risk_frame = stock_risk_frames["all"]
        shortage_risks = stock_risk_frames["shortage"]
        out_of_stock_risks = stock_risk_frames["out_of_stock"]
        reorder_risks = stock_risk_frames["reorder"]
        overstock_risks = stock_risk_frames["overstock"]
        dormant_stock_risks = stock_risk_frames["dormant"]
        stock_risk_report = stock_risk_frames["report"]
        overstock_days_threshold = stock_risk_frames["overstock_days_threshold"]
        stock_risk_rename_map = {
            "risk_type": "Тип риска",
            "product": "SKU / Товар",
            "category": "Категория",
            "supplier": "Поставщик",
            "brand": "Бренд",
            "abc_class": "ABC",
            "xyz_class": "XYZ",
            "priority": "Приоритет",
            "stock_status": "Статус остатка",
            "demand_state": "Состояние спроса",
            "forecast_qty": "Количество прогноза, шт/мес.",
            "avg_daily_qty": "Среднедневной спрос",
            "stock_on_hand": "Количество остатка",
            "stock_value": "Сумма остатка",
            "manual_stock_in_transit": "Количество в пути вручную",
            "ordered_in_transit_qty": "Количество в заказах",
            "stock_in_transit": "Количество в пути",
            "available_stock_qty": "Количество доступно",
            "stock_coverage_days": "Покрытие, дней",
            "gross_requirement_qty": "Брутто-потребность, шт",
            "net_requirement_qty": "Чистая потребность, шт",
            "recommended_order_qty": "Количество к заказу",
            "lead_time_days": "Срок поставки, дней",
            "last_sale_date": "Последняя продажа",
            "days_since_last_sale": "Дней без продаж",
            "notes": "Примечание",
        }

        def render_stock_risk_table(
            frame: pd.DataFrame,
            columns: list[str],
            *,
            empty_message: str,
            height: int = 330,
        ) -> None:
            if frame.empty:
                st.info(empty_message)
                return
            visible_columns = [column for column in columns if column in frame.columns]
            stock_risk_display = format_display_frame_for_role(
                frame[visible_columns].copy(),
                current_user,
                rename_map=stock_risk_rename_map,
            )
            preferred_title = "Товар" if "Товар" in stock_risk_display.columns else None
            render_mobile_table_cards(
                stock_risk_display,
                title_column=preferred_title,
                field_columns=["Бренд", "Поставщик", "Статус остатка", "Количество доступно", "Количество к заказу"],
                max_rows=6,
            )
            st.dataframe(
                stock_risk_display,
                use_container_width=True,
                hide_index=True,
                height=height,
            )

        brand_summary_source = stock_risk_frame.copy()
        brand_summary_source["brand"] = (
            brand_summary_source["brand"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "Не назначен")
        )
        brand_summary_source["deficit_sku"] = brand_summary_source["stock_status"].isin(
            ["Нет остатка", "Риск дефицита"]
        ).astype(int)
        brand_summary_source["overstock_sku"] = (
            (brand_summary_source["stock_on_hand"] > 0)
            & (brand_summary_source["stock_coverage_days"] > overstock_days_threshold)
        ).astype(int)
        brand_summary_source["order_sku"] = (
            brand_summary_source["recommended_order_qty"] > 0
        ).astype(int)
        brand_summary = (
            brand_summary_source.groupby("brand", as_index=False)
            .agg(
                sku_count=("product", "nunique"),
                forecast_revenue=("forecast_revenue", "sum"),
                stock_value=("stock_value", "sum"),
                stock_on_hand=("stock_on_hand", "sum"),
                deficit_sku=("deficit_sku", "sum"),
                overstock_sku=("overstock_sku", "sum"),
                order_sku=("order_sku", "sum"),
                recommended_order_qty=("recommended_order_qty", "sum"),
            )
            .sort_values(["forecast_revenue", "stock_value"], ascending=[False, False])
        )
        with st.container(border=True):
            render_panel_header(
                "Сводка по брендам",
                "Продажи, стоимость остатка и закупочные риски в одном разрезе. "
                "Выберите бренд в общих фильтрах, чтобы весь экран пересчитался только по нему.",
            )
            st.dataframe(
                format_display_frame_for_role(
                    brand_summary,
                    current_user,
                    rename_map={
                        "brand": "Бренд",
                        "sku_count": "SKU",
                        "forecast_revenue": "Прогноз выручки",
                        "stock_value": "Сумма склада",
                        "stock_on_hand": "Остаток, шт",
                        "deficit_sku": "Дефицит, SKU",
                        "overstock_sku": "Излишки, SKU",
                        "order_sku": "К заказу, SKU",
                        "recommended_order_qty": "К заказу, шт",
                    },
                ),
                use_container_width=True,
                hide_index=True,
                height=min(420, 46 + max(len(brand_summary), 1) * 35),
            )

        with st.container(border=True):
            render_panel_header(
                "Остатки и риски",
                f"Сверка текущих остатков с прогнозом спроса на окно {procurement_total_window_days} дней.",
            )
            render_snapshot_strip(
                [
                    {
                        "label": "Дефицит",
                        "value": format_number(len(shortage_risks)),
                        "hint": "Нет остатка или не хватает до поставки",
                    },
                    {
                        "label": "Без остатка",
                        "value": format_number(len(out_of_stock_risks)),
                        "hint": "Спрос есть, доступный остаток нулевой",
                    },
                    {
                        "label": "Излишки",
                        "value": format_number(len(overstock_risks)),
                        "hint": f"Покрытие больше {format_number(overstock_days_threshold)} дней",
                    },
                    {
                        "label": "Неликвид",
                        "value": format_number(len(dormant_stock_risks)),
                        "hint": "Остаток есть, текущего спроса нет",
                    },
                    {
                        "label": "Чистая потребность",
                        "value": format_number(stock_risk_frame["net_requirement_qty"].sum()),
                        "hint": "Потребность после вычета остатков",
                    },
                    {
                        "label": "К заказу",
                        "value": format_number(stock_risk_frame["recommended_order_qty"].sum()),
                        "hint": "С учётом MOQ и кратности",
                    },
                ]
            )

            stock_status_summary = (
                stock_risk_frame.assign(
                    stock_status=stock_risk_frame["stock_status"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", "Не указан")
                )
                .groupby("stock_status", as_index=False)
                .agg(
                    sku_count=("product", "nunique"),
                    recommended_order_qty=("recommended_order_qty", "sum"),
                    available_stock_qty=("available_stock_qty", "sum"),
                )
                .sort_values("sku_count", ascending=True)
            )
            status_col, urgent_col = st.columns([0.78, 1.22], gap="medium")
            with status_col:
                if stock_status_summary.empty:
                    st.info("Нет данных для сводки по статусам остатка.")
                else:
                    stock_status_chart = px.bar(
                        stock_status_summary,
                        x="sku_count",
                        y="stock_status",
                        orientation="h",
                        color="stock_status",
                        color_discrete_map=stock_status_color_map,
                        text="sku_count",
                        labels={
                            "sku_count": "SKU",
                            "stock_status": "",
                        },
                        title="",
                    )
                    stock_status_chart.update_traces(texttemplate="%{text:.0f}", textposition="outside")
                    stock_status_chart.update_layout(
                        xaxis_title="SKU",
                        yaxis_title="",
                        showlegend=False,
                    )
                    polish_figure(stock_status_chart, height=280)
                    st.plotly_chart(stock_status_chart, use_container_width=True)

            with urgent_col:
                st.markdown("**Самые срочные позиции**")
                render_stock_risk_table(
                    shortage_risks.head(8),
                    [
                        "product",
                        "category",
                        "supplier",
                        "brand",
                        "stock_status",
                        "available_stock_qty",
                        "stock_coverage_days",
                        "forecast_qty",
                        "recommended_order_qty",
                        "abc_class",
                        "xyz_class",
                    ],
                    empty_message="Сейчас нет критичных позиций по остаткам.",
                    height=280,
                )

            deficit_tab, overstock_tab, dormant_tab, reorder_tab = st.tabs(
                ["Дефицит", "Излишки", "Неликвид", "Заказ"]
            )
            with deficit_tab:
                render_stock_risk_table(
                    shortage_risks.head(30),
                    [
                        "product",
                        "category",
                        "supplier",
                        "brand",
                        "priority",
                        "stock_status",
                        "available_stock_qty",
                        "stock_coverage_days",
                        "gross_requirement_qty",
                        "net_requirement_qty",
                        "recommended_order_qty",
                    ],
                    empty_message="Дефицита в текущем контуре не найдено.",
                )
            with overstock_tab:
                render_stock_risk_table(
                    overstock_risks.head(30),
                    [
                        "product",
                        "category",
                        "supplier",
                        "brand",
                        "stock_status",
                        "available_stock_qty",
                        "stock_value",
                        "stock_coverage_days",
                        "forecast_qty",
                        "last_sale_date",
                        "days_since_last_sale",
                    ],
                    empty_message="Излишков по выбранному порогу покрытия не найдено.",
                )
            with dormant_tab:
                render_stock_risk_table(
                    dormant_stock_risks.head(30),
                    [
                        "product",
                        "category",
                        "supplier",
                        "brand",
                        "available_stock_qty",
                        "stock_on_hand",
                        "stock_value",
                        "stock_in_transit",
                        "demand_state",
                        "last_sale_date",
                        "days_since_last_sale",
                        "notes",
                    ],
                    empty_message="Неликвидного остатка в текущем контуре не найдено.",
                )
            with reorder_tab:
                render_stock_risk_table(
                    reorder_risks.head(30),
                    [
                        "product",
                        "category",
                        "supplier",
                        "brand",
                        "priority",
                        "stock_status",
                        "forecast_qty",
                        "available_stock_qty",
                        "net_requirement_qty",
                        "recommended_order_qty",
                        "lead_time_days",
                    ],
                    empty_message="Пока нет SKU с рекомендацией к заказу.",
                )

            if not stock_risk_report.empty:
                stock_risk_download_columns = [
                    "risk_type",
                    "product",
                    "category",
                    "supplier",
                    "brand",
                    "priority",
                    "stock_status",
                    "demand_state",
                    "abc_class",
                    "xyz_class",
                    "forecast_qty",
                    "stock_on_hand",
                    "stock_value",
                    "stock_in_transit",
                    "available_stock_qty",
                    "stock_coverage_days",
                    "net_requirement_qty",
                    "recommended_order_qty",
                    "last_sale_date",
                    "days_since_last_sale",
                    "notes",
                ]
                stock_risk_export_columns = [
                    column for column in stock_risk_download_columns if column in stock_risk_report.columns
                ]
                st.download_button(
                    "Скачать отчёт по остаткам и рискам",
                    data=to_excel_report_bytes(
                        {
                            "Остатки и риски": prepare_excel_report_frame(
                                stock_risk_report,
                                current_user,
                                stock_risk_rename_map,
                                columns=stock_risk_export_columns,
                            )
                        },
                        report_title="Остатки и риски",
                    ),
                    file_name="stock_risks_report.xlsx",
                    mime=EXCEL_MIME,
                    key="dl_stock_risks_report",
                )

            supplier_order_file = build_supplier_order_file(procurement_forecast)
            if supplier_order_file is not None:
                st.download_button(
                    "Скачать Excel-заказ поставщикам",
                    data=supplier_order_file.content,
                    file_name=supplier_order_file.filename,
                    mime=supplier_order_file.content_type,
                    key="dl_supplier_order_workbook",
                )

        top_procurement = procurement_forecast[
            procurement_forecast["recommended_order_qty"].fillna(0) > 0
        ].head(12).copy()
        top_procurement = top_procurement.sort_values("recommended_order_qty", ascending=True)

        scatter_data = procurement_forecast[
            procurement_forecast["forecast_qty"].fillna(0) > 0
        ].copy()
        scatter_data["stock_coverage_days"] = pd.to_numeric(
            scatter_data["stock_coverage_days"],
            errors="coerce",
        ).fillna(0.0)
        scatter_data["bubble_size"] = (
            build_safe_marker_size(
                scatter_data["recommended_order_qty"].where(
                    scatter_data["recommended_order_qty"].fillna(0) > 0,
                    scatter_data["forecast_qty"],
                ),
                absolute=True,
            )
            .clip(lower=1.0)
        )

        procurement_left, procurement_right = st.columns([1.18, 0.82], gap="medium")

        with procurement_left:
            with st.container(border=True):
                render_panel_header(
                    "Что заказывать сейчас",
                    "Позиции отсортированы по рекомендованному заказу после вычета остатка и товара в пути.",
                )
                if top_procurement.empty:
                    st.info("В текущем контуре нет SKU, которым уже нужно формировать заказ.")
                else:
                    procurement_bar = px.bar(
                        top_procurement,
                        x="recommended_order_qty",
                        y="product",
                        orientation="h",
                        color="priority",
                        color_discrete_map=priority_color_map,
                        hover_data={
                            "supplier": True,
                            "stock_status": True,
                            "available_stock_qty": ":.2f",
                            "net_requirement_qty": ":.2f",
                            "ordered_in_transit_qty": ":.2f",
                            "open_order_count": True,
                            "category": True,
                            "forecast_qty": ":.2f",
                            "gross_requirement_qty": ":.2f",
                            "min_order_qty": ":.2f",
                            "order_multiple": ":.2f",
                            "abc_class": True,
                            "xyz_class": True,
                        },
                        labels={
                            "recommended_order_qty": "Рекомендованный заказ, шт",
                            "product": "",
                            "priority": "Приоритет",
                        },
                        title="",
                    )
                    procurement_bar.update_layout(
                        xaxis_title="Рекомендованный заказ, шт",
                        yaxis_title="",
                        showlegend=False,
                    )
                    polish_figure(
                        procurement_bar,
                        height=compact_bar_chart_height(
                            len(top_procurement),
                            minimum=360,
                            maximum=520,
                            base=160,
                            row_step=28,
                        ),
                    )
                    st.plotly_chart(procurement_bar, use_container_width=True)

        with procurement_right:
            with st.container(border=True):
                render_panel_header(
                    "Карта спроса и покрытия",
                    "Помогает увидеть, где спрос уже есть, а остатка и покрытия в днях недостаточно.",
                )
                if scatter_data.empty:
                    st.info("Недостаточно данных для карты покрытия по закупкам.")
                else:
                    procurement_scatter = px.scatter(
                        scatter_data,
                        x="forecast_qty",
                        y="stock_coverage_days",
                        size="bubble_size",
                        size_max=34,
                        color="stock_status",
                        color_discrete_map=stock_status_color_map,
                        hover_name="product",
                        hover_data={
                            "supplier": True,
                            "category": True,
                            "abc_class": True,
                            "xyz_class": True,
                            "available_stock_qty": ":.2f",
                            "ordered_in_transit_qty": ":.2f",
                            "open_order_count": True,
                            "net_requirement_qty": ":.2f",
                            "recommended_order_qty": ":.2f",
                            "avg_daily_qty": ":.2f",
                            "days_since_last_sale": True,
                            "bubble_size": False,
                        },
                        labels={
                            "forecast_qty": "Прогноз потребности, шт / мес.",
                            "stock_coverage_days": "Покрытие остатком, дней",
                            "stock_status": "Статус остатка",
                        },
                        title="",
                    )
                    procurement_scatter.update_layout(
                        xaxis_title="Прогноз потребности, шт / мес.",
                        yaxis_title="Покрытие остатком, дней",
                    )
                    polish_figure(procurement_scatter, height=420)
                    st.plotly_chart(procurement_scatter, use_container_width=True)

        supplier_reorder_summary = procurement_supplier_summary[
            procurement_supplier_summary["recommended_order_qty"].fillna(0) > 0
        ].copy()
        supplier_chart_data = supplier_reorder_summary.head(10).sort_values(
            "recommended_order_qty",
            ascending=True,
        )
        supplier_summary_rename_map = {
            "supplier": "Поставщик",
            "sku_count": "SKU в контуре",
            "reorder_sku_count": "SKU к заказу",
            "critical_sku_count": "Критичных SKU",
            "recommended_order_qty": "Рекомендованный заказ, шт",
            "net_requirement_qty": "Чистая потребность, шт",
            "available_stock_qty": "Доступный остаток",
            "forecast_qty": "Прогноз спроса, шт",
            "forecast_revenue": "Прогноз выручки",
            "max_lead_time_days": "Макс. срок поставки, дн.",
        }

        supplier_summary_left, supplier_summary_right = st.columns([0.92, 1.08], gap="medium")

        with supplier_summary_left:
            with st.container(border=True):
                render_panel_header(
                    "Поставщики к заказу",
                    "Показывает, кому уже нужно отправлять заказ и где сосредоточен риск дефицита.",
                )
                if supplier_chart_data.empty:
                    st.info("Пока нет поставщиков с ненулевой рекомендацией к заказу.")
                else:
                    supplier_bar = px.bar(
                        supplier_chart_data,
                        x="recommended_order_qty",
                        y="supplier",
                        orientation="h",
                        text="recommended_order_qty",
                        color_discrete_sequence=[SECONDARY_COLOR],
                        hover_data={
                            "reorder_sku_count": True,
                            "critical_sku_count": True,
                            "sku_count": True,
                            "net_requirement_qty": ":.2f",
                            "available_stock_qty": ":.2f",
                            "max_lead_time_days": True,
                        },
                        labels={
                            "recommended_order_qty": "Рекомендованный заказ, шт",
                            "supplier": "",
                        },
                        title="",
                    )
                    supplier_bar.update_traces(texttemplate="%{text:.0f}", textposition="outside")
                    supplier_bar.update_layout(
                        xaxis_title="Рекомендованный заказ, шт",
                        yaxis_title="",
                        showlegend=False,
                    )
                    polish_figure(
                        supplier_bar,
                        height=compact_bar_chart_height(
                            len(supplier_chart_data),
                            minimum=340,
                            maximum=480,
                            base=160,
                            row_step=28,
                        ),
                    )
                    st.plotly_chart(supplier_bar, use_container_width=True)

        with supplier_summary_right:
            with st.container(border=True):
                render_panel_header(
                    "Сводка по поставщикам",
                    "Готовая таблица для переговоров и сборки ближайшей закупочной волны.",
                )
                st.dataframe(
                    supplier_reorder_summary.rename(columns=supplier_summary_rename_map),
                    use_container_width=True,
                    hide_index=True,
                    height=390,
                )
                st.download_button(
                    "Скачать сводку по поставщикам",
                    data=to_excel_report_bytes(
                        {
                            "Поставщики": prepare_excel_report_frame(
                                supplier_reorder_summary,
                                current_user,
                                supplier_summary_rename_map,
                            )
                        },
                        report_title="Сводка по поставщикам",
                    ),
                    file_name="procurement_suppliers.xlsx",
                    mime=EXCEL_MIME,
                )

        if can_manage_procurement(current_user):
            with st.container(border=True):
                render_panel_header(
                    "Создать заказ поставщику",
                    "Соберите заказ прямо из текущего прогноза. В расчёт попадут только строки с количеством больше нуля.",
                )
                order_candidates = procurement_forecast[
                    procurement_forecast["recommended_order_qty"].fillna(0) > 0
                ].copy()
                order_candidates["supplier_group"] = (
                    order_candidates["supplier"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", "Не назначен")
                )
                order_supplier_options = (
                    order_candidates["supplier_group"].dropna().astype(str).drop_duplicates().tolist()
                    if not order_candidates.empty
                    else []
                )
                if not order_supplier_options:
                    st.info("Сейчас нет поставщиков, по которым уже нужно формировать заказ.")
                else:
                    selected_order_supplier = st.selectbox(
                        "Поставщик для заказа",
                        options=order_supplier_options,
                        key="procurement_order_supplier",
                    )
                    selected_order_source = order_candidates[
                        order_candidates["supplier_group"] == selected_order_supplier
                    ][
                        [
                            "product",
                            "category",
                            "recommended_order_qty",
                            "available_stock_qty",
                            "lead_time_days",
                            "notes",
                        ]
                    ].copy()
                    selected_order_source["order_qty"] = selected_order_source["recommended_order_qty"]
                    selected_order_source["line_comment"] = ""

                    order_editor_col, order_meta_col = st.columns([1.24, 0.76], gap="medium")
                    with order_editor_col:
                        order_editor = st.data_editor(
                            selected_order_source,
                            use_container_width=True,
                            hide_index=True,
                            height=430,
                            key=f"procurement_order_editor_{selected_order_supplier}",
                            disabled=[
                                "product",
                                "category",
                                "recommended_order_qty",
                                "available_stock_qty",
                                "lead_time_days",
                                "notes",
                            ],
                            column_config={
                                "product": st.column_config.TextColumn("SKU / Товар"),
                                "category": st.column_config.TextColumn("Категория"),
                                "recommended_order_qty": st.column_config.NumberColumn("Рекомендовано", format="%.2f"),
                                "available_stock_qty": st.column_config.NumberColumn("Доступно сейчас", format="%.2f"),
                                "lead_time_days": st.column_config.NumberColumn("Срок поставки, дн.", format="%d"),
                                "notes": st.column_config.TextColumn("Примечание SKU"),
                                "order_qty": st.column_config.NumberColumn("Заказать", min_value=0.0, step=1.0, format="%.2f"),
                                "line_comment": st.column_config.TextColumn("Комментарий строки"),
                            },
                        )
                    with order_meta_col:
                        order_creation_status = st.selectbox(
                            "Статус при создании",
                            options=["draft", "ordered", "in_transit"],
                            index=1,
                            format_func=lambda value: PROCUREMENT_ORDER_STATUS_LABELS.get(value, value),
                            key="procurement_order_creation_status",
                        )
                        order_date_value = st.date_input(
                            "Дата заказа",
                            value=date.today(),
                            key="procurement_order_date",
                        )
                        order_expected_default = (
                            pd.Timestamp(date.today()) + pd.Timedelta(days=max(procurement_lead_time_days, 1))
                        ).date()
                        order_expected_date_value = st.date_input(
                            "Ожидаемая дата",
                            value=order_expected_default,
                            key="procurement_order_expected_date",
                        )
                        order_comment = st.text_area(
                            "Комментарий к заказу",
                            key="procurement_order_comment",
                            placeholder="Например: срочная поставка, частичная отгрузка, согласовано с поставщиком.",
                        )
                        if st.button(
                            "Создать заказ поставщику",
                            key="procurement_create_order_button",
                            use_container_width=True,
                            type="primary",
                        ):
                            try:
                                order_items_to_create = order_editor.rename(
                                    columns={"order_qty": "quantity", "line_comment": "comment"}
                                )[["product", "quantity", "comment"]].copy()
                                created_order = create_procurement_order(
                                    supplier=selected_order_supplier,
                                    items_frame=order_items_to_create,
                                    order_date=order_date_value,
                                    expected_date=order_expected_date_value,
                                    comment=order_comment,
                                    created_by=current_user["username"],
                                    status=order_creation_status,
                                )
                                created_item_count = int((order_items_to_create["quantity"] > 0).sum())
                                audit_event(
                                    action="procurement.order_create",
                                    user_id=current_user["username"],
                                    details={
                                        "order_id": created_order["order_id"],
                                        "supplier": selected_order_supplier,
                                        "status": order_creation_status,
                                        "item_count": created_item_count,
                                    },
                                )
                                st.session_state["procurement_flash_message"] = (
                                    f"Заказ {created_order['order_id']} создан для поставщика {selected_order_supplier}."
                                )
                            except ValueError as error:
                                st.session_state["procurement_error_message"] = str(error)
                            st.rerun()

        with st.container(border=True):
            render_panel_header(
                "Заказы в работе",
                "Следите за созданными заказами, обновляйте статусы и принимайте поставки прямо из системы.",
            )
            if procurement_order_overview.empty:
                st.info("Заказы ещё не создавались. Сформируйте первый заказ из прогноза выше.")
            else:
                order_filter_options = ["Активные", "Все", *[PROCUREMENT_ORDER_STATUS_LABELS[status] for status in PROCUREMENT_ORDER_STATUSES]]
                selected_order_filter = st.selectbox(
                    "Какие заказы показать",
                    options=order_filter_options,
                    index=0,
                    key="procurement_order_filter",
                )
                filtered_order_overview = procurement_order_overview.copy()
                if selected_order_filter == "Активные":
                    filtered_order_overview = filtered_order_overview[
                        filtered_order_overview["status"].isin(active_order_statuses)
                    ].copy()
                elif selected_order_filter != "Все":
                    matching_status = next(
                        (
                            status
                            for status, label in PROCUREMENT_ORDER_STATUS_LABELS.items()
                            if label == selected_order_filter
                        ),
                        "",
                    )
                    filtered_order_overview = filtered_order_overview[
                        filtered_order_overview["status"] == matching_status
                    ].copy()

                order_summary_col, order_detail_col = st.columns([1.08, 0.92], gap="medium")
                with order_summary_col:
                    order_table_rename_map = {
                        "order_id": "Заказ",
                        "supplier": "Поставщик",
                        "status_label": "Статус",
                        "order_date": "Дата заказа",
                        "expected_date": "Ожидаемая дата",
                        "sku_count": "SKU",
                        "total_quantity": "Количество",
                        "updated_by": "Последнее изменение",
                        "updated_at": "Обновлён",
                    }
                    order_table_for_display = filtered_order_overview[
                        [
                            "order_id",
                            "supplier",
                            "status_label",
                            "order_date",
                            "expected_date",
                            "sku_count",
                            "total_quantity",
                            "updated_by",
                            "updated_at",
                        ]
                    ].rename(columns=order_table_rename_map)
                    order_table_column_config = {
                        order_table_rename_map["supplier"]: st.column_config.TextColumn(width="medium"),
                        order_table_rename_map["status_label"]: st.column_config.TextColumn(width="medium"),
                        order_table_rename_map["updated_by"]: st.column_config.TextColumn(width="medium"),
                    }
                    st.dataframe(
                        order_table_for_display,
                        column_config=order_table_column_config,
                        use_container_width=True,
                        hide_index=True,
                        height=430,
                    )

                with order_detail_col:
                    if filtered_order_overview.empty:
                        st.info("Под выбранный фильтр пока нет заказов.")
                    else:
                        order_option_labels = {
                            row["order_id"]: (
                                f"{row['order_id']} • {row['supplier'] or 'Без поставщика'} • {row['status_label']}"
                            )
                            for _, row in filtered_order_overview.iterrows()
                        }
                        selected_order_id = st.selectbox(
                            "Карточка заказа",
                            options=filtered_order_overview["order_id"].tolist(),
                            format_func=lambda value: order_option_labels.get(value, value),
                            key="procurement_selected_order_id",
                        )
                        selected_order_row = filtered_order_overview[
                            filtered_order_overview["order_id"] == selected_order_id
                        ].iloc[0]
                        selected_order_lines = procurement_order_items[
                            procurement_order_items["order_id"] == selected_order_id
                        ][["product", "quantity", "unit_cost", "comment"]].copy()
                        order_lines_rename_map = {
                            "product": "SKU / Товар",
                            "quantity": "Количество",
                            "unit_cost": "Цена закупки",
                            "comment": "Комментарий",
                        }
                        selected_order_lines_display = selected_order_lines.rename(columns=order_lines_rename_map)

                        render_snapshot_strip(
                            [
                                {
                                    "label": "Статус",
                                    "value": str(selected_order_row["status_label"]),
                                    "hint": f"Поставщик: {selected_order_row['supplier'] or 'Без поставщика'}",
                                },
                                {
                                    "label": "SKU в заказе",
                                    "value": format_number(selected_order_row["sku_count"]),
                                    "hint": f"Всего единиц: {format_number(selected_order_row['total_quantity'])}",
                                },
                                {
                                    "label": "Ожидаемая дата",
                                    "value": (
                                        pd.to_datetime(selected_order_row["expected_date"], errors="coerce").strftime("%d.%m.%Y")
                                        if pd.notna(pd.to_datetime(selected_order_row["expected_date"], errors="coerce"))
                                        else "Не указана"
                                    ),
                                    "hint": f"Создал: {selected_order_row['created_by'] or 'н/д'}",
                                },
                            ]
                        )
                        st.dataframe(
                            selected_order_lines.rename(
                                columns={
                                    "product": "SKU / Товар",
                                    "quantity": "Количество",
                                    "unit_cost": "Цена закупки",
                                    "comment": "Комментарий",
                                }
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=300,
                        )

                        if can_manage_procurement(current_user):
                            current_status_value = str(selected_order_row["status"]).strip()
                            selected_order_comment = str(selected_order_row.get("comment", "") or "")
                            parsed_expected_date = pd.to_datetime(
                                selected_order_row["expected_date"],
                                errors="coerce",
                            )
                            expected_date_input_value = (
                                parsed_expected_date.date()
                                if pd.notna(parsed_expected_date)
                                else (pd.Timestamp(date.today()) + pd.Timedelta(days=max(procurement_lead_time_days, 1))).date()
                            )
                            updated_status_value = st.selectbox(
                                "Новый статус",
                                options=list(PROCUREMENT_ORDER_STATUSES),
                                index=list(PROCUREMENT_ORDER_STATUSES).index(current_status_value)
                                if current_status_value in PROCUREMENT_ORDER_STATUSES
                                else 0,
                                format_func=lambda value: PROCUREMENT_ORDER_STATUS_LABELS.get(value, value),
                                key=f"procurement_order_status_{selected_order_id}",
                            )
                            updated_expected_date = st.date_input(
                                "Обновить ожидаемую дату",
                                value=expected_date_input_value,
                                key=f"procurement_order_expected_{selected_order_id}",
                            )
                            updated_order_comment = st.text_area(
                                "Комментарий заказа",
                                value=selected_order_comment,
                                key=f"procurement_order_comment_{selected_order_id}",
                            )

                            action_label = (
                                "Принять заказ и зачислить на остаток"
                                if updated_status_value == "received"
                                else "Сохранить статус заказа"
                            )
                            if st.button(
                                action_label,
                                key=f"procurement_order_update_{selected_order_id}",
                                use_container_width=True,
                            ):
                                try:
                                    if updated_status_value == "received":
                                        apply_procurement_order_receipt(
                                            selected_order_id,
                                            updated_by=current_user["username"],
                                            expected_date=updated_expected_date,
                                            comment=updated_order_comment,
                                        )
                                        audit_event(
                                            action="procurement.order_receive",
                                            user_id=current_user["username"],
                                            details={"order_id": selected_order_id},
                                        )
                                        st.session_state["procurement_flash_message"] = (
                                            f"Заказ {selected_order_id} принят, остатки обновлены."
                                        )
                                    else:
                                        update_procurement_order(
                                            selected_order_id,
                                            status=updated_status_value,
                                            expected_date=updated_expected_date,
                                            comment=updated_order_comment,
                                            updated_by=current_user["username"],
                                        )
                                        audit_event(
                                            action="procurement.order_update",
                                            user_id=current_user["username"],
                                            details={
                                                "order_id": selected_order_id,
                                                "status": updated_status_value,
                                            },
                                        )
                                        st.session_state["procurement_flash_message"] = (
                                            f"Заказ {selected_order_id} обновлён."
                                        )
                                except ValueError as error:
                                    st.session_state["procurement_error_message"] = str(error)
                                st.rerun()

        procurement_table_columns = [
            "product",
            "category",
            "supplier",
            "brand",
            "abc_class",
            "xyz_class",
            "priority",
            "stock_status",
            "demand_state",
            "active_months",
            "history_months_used",
            "last_month_qty",
            "avg_monthly_qty",
            "recent_avg_qty",
            "trend_factor",
            "forecast_qty",
            "avg_daily_qty",
            "stock_on_hand",
            "stock_value",
            "manual_stock_in_transit",
            "ordered_in_transit_qty",
            "stock_in_transit",
            "available_stock_qty",
            "open_order_count",
            "stock_coverage_days",
            "lead_time_days",
            "lead_time_requirement_qty",
            "coverage_requirement_qty",
            "safety_stock_qty",
            "gross_requirement_qty",
            "net_requirement_qty",
            "min_order_qty",
            "order_multiple",
            "recommended_order_qty",
            "demand_variability_pct",
            "days_since_last_sale",
            "forecast_revenue",
            "notes",
        ]
        procurement_rename_map = {
            "product": "SKU / Товар",
            "category": "Категория",
            "supplier": "Поставщик",
            "brand": "Бренд",
            "abc_class": "ABC",
            "xyz_class": "XYZ",
            "priority": "Приоритет",
            "stock_status": "Статус остатка",
            "demand_state": "Состояние спроса",
            "active_months": "Активных месяцев",
            "history_months_used": "Месяцев в истории",
            "last_month_qty": "Последний месяц, шт",
            "avg_monthly_qty": "Средний спрос, шт",
            "recent_avg_qty": "Последние месяцы, шт",
            "trend_factor": "Тренд",
            "forecast_qty": "Прогноз потребности, шт",
            "avg_daily_qty": "Среднедневной спрос",
            "stock_on_hand": "Остаток",
            "stock_value": "Сумма остатка",
            "manual_stock_in_transit": "В пути (ручной)",
            "ordered_in_transit_qty": "В пути по заказам",
            "stock_in_transit": "В пути",
            "available_stock_qty": "Доступно сейчас",
            "open_order_count": "Активных заказов",
            "stock_coverage_days": "Покрытие остатком, дней",
            "lead_time_days": "Срок поставки, дн.",
            "lead_time_requirement_qty": "Потребность до поставки, шт",
            "coverage_requirement_qty": "Покрытие окна, шт",
            "safety_stock_qty": "Страховой запас, шт",
            "gross_requirement_qty": "Брутто-потребность, шт",
            "net_requirement_qty": "Чистая потребность, шт",
            "min_order_qty": "MOQ",
            "order_multiple": "Кратность",
            "recommended_order_qty": "Рекомендованный заказ, шт",
            "demand_variability_pct": "Вариативность спроса, %",
            "days_since_last_sale": "Дней с последней продажи",
            "forecast_revenue": "Прогноз выручки",
            "notes": "Примечание",
        }
        if margin_visible:
            procurement_table_columns.extend(["avg_margin_per_unit", "forecast_margin"])
            procurement_rename_map["avg_margin_per_unit"] = "Маржа на единицу"
            procurement_rename_map["forecast_margin"] = "Прогноз маржи"

        procurement_table = procurement_forecast[procurement_table_columns].copy()

        with st.container(border=True):
            render_panel_header(
                "Таблица закупок",
                "Полный список товаров с прогнозом потребности, стабильностью спроса и приоритетом к закупке.",
            )
            st.dataframe(
                format_display_frame_for_role(
                    procurement_table,
                    current_user,
                    rename_map=procurement_rename_map,
                ),
                use_container_width=True,
                hide_index=True,
                height=680,
            )
            st.download_button(
                "Скачать прогноз закупки",
                data=to_excel_report_bytes(
                    {
                        "Прогноз закупки": prepare_excel_report_frame(
                            procurement_table,
                            current_user,
                            procurement_rename_map,
                        )
                    },
                    report_title="Прогноз закупки",
                ),
                file_name="procurement_forecast.xlsx",
                mime=EXCEL_MIME,
            )

if active_screen == "План / факт":
    render_section_intro(
        "План / факт",
        "Сравнивает помесячный план с фактическими продажами по выручке и количеству, а для руководителя ещё и по марже.",
    )
    render_section_marker(
        "Управление целями",
        "Сначала задайте план, потом контролируйте выполнение",
        "Рабочий порядок здесь такой: сначала выбираете контур и месяц плана, затем задаёте цифры по выручке, марже и количеству. После этого смотрите выполнение по месяцам и, если нужно, разбираете отставание по салонам.",
    )

    if plan_fact_uses_unfiltered_scope:
        st.info("План / факт считается по выбранному периоду и салонам, но не сужается фильтрами по категориям и менеджерам. Это сделано специально, чтобы план не искажался товарными срезами.")

    if plan_fact_summary.empty:
        st.info("Для контроля плана нужен хотя бы один месяц данных или сохранённый план на будущий месяц.")
    else:
        plan_month_options = plan_fact_summary["month_label"].dropna().astype(str).tolist()
        latest_plan_row = plan_fact_summary.iloc[-1]

        plan_snapshot_items = [
            {
                "label": f"План выручки {latest_plan_row['month_label']}",
                "value": format_money(latest_plan_row["revenue_plan"]) if not is_missing(latest_plan_row["revenue_plan"]) else "н/д",
                "hint": "Плановая выручка на выбранный месяц",
            },
            {
                "label": f"Факт выручки {latest_plan_row['month_label']}",
                "value": format_money(latest_plan_row["revenue"]),
                "delta": format_percent(latest_plan_row["revenue_execution_pct"]) if not is_missing(latest_plan_row["revenue_execution_pct"]) else "",
                "hint": "Процент выполнения плана по выручке",
            },
            {
                "label": f"План количества {latest_plan_row['month_label']}",
                "value": format_number(latest_plan_row["quantity_plan"]) if not is_missing(latest_plan_row["quantity_plan"]) else "н/д",
                "hint": "План по количеству проданных единиц",
            },
            {
                "label": f"Факт количества {latest_plan_row['month_label']}",
                "value": format_number(latest_plan_row["quantity"]),
                "delta": format_percent(latest_plan_row["quantity_execution_pct"]) if not is_missing(latest_plan_row["quantity_execution_pct"]) else "",
                "hint": "Процент выполнения плана по количеству",
            },
        ]
        if margin_visible:
            plan_snapshot_items[2:2] = [
                {
                    "label": f"План маржи {latest_plan_row['month_label']}",
                    "value": format_money(latest_plan_row["margin_plan"]) if not is_missing(latest_plan_row["margin_plan"]) else "н/д",
                    "hint": "Плановая валовая маржа на месяц",
                },
                {
                    "label": f"Факт маржи {latest_plan_row['month_label']}",
                    "value": format_money(latest_plan_row["margin"]),
                    "delta": format_percent(latest_plan_row["margin_execution_pct"]) if not is_missing(latest_plan_row["margin_execution_pct"]) else "",
                    "hint": "Процент выполнения плана по марже",
                },
            ]
        render_snapshot_strip(plan_snapshot_items)

        plan_chart_left, plan_chart_right = st.columns([1.35, 0.95], gap="medium")

        with plan_chart_left:
            with st.container(border=True):
                render_panel_header(
                    "Динамика плана и факта по выручке",
                    "Главный график вкладки: показывает, как месячный факт идёт относительно плана. Используйте его как первую точку контроля, чтобы сразу увидеть месяцы с недовыполнением и месяцы, где план уже закрыт.",
                )
                plan_chart = go.Figure()
                plan_chart.add_trace(
                    go.Bar(
                        x=plan_fact_summary["month_label"],
                        y=plan_fact_summary["revenue"],
                        name="Факт: выручка",
                        marker_color=PRIMARY_COLOR,
                    )
                )
                if plan_fact_summary["revenue_plan"].notna().any():
                    plan_chart.add_trace(
                        go.Scatter(
                            x=plan_fact_summary["month_label"],
                            y=plan_fact_summary["revenue_plan"],
                            name="План: выручка",
                            mode="lines+markers",
                            line=dict(color=SECONDARY_COLOR, width=3),
                            marker=dict(size=8),
                        )
                    )
                plan_chart.update_layout(xaxis_title="Месяц", yaxis_title="Сумма", legend_title="")
                polish_figure(plan_chart, height=440)
                st.plotly_chart(plan_chart, use_container_width=True)

        with plan_chart_right:
            with st.container(border=True):
                render_panel_header(
                    "Отклонение последнего месяца",
                    "Короткая контрольная панель по самому свежему месяцу в текущем контуре. Здесь видно, сколько не добрали или перевыполнили по ключевым показателям.",
                )
                latest_gap_cards = [
                    {
                        "label": "Отклонение выручки",
                        "value": format_money(latest_plan_row["revenue_gap"]) if not is_missing(latest_plan_row["revenue_gap"]) else "н/д",
                        "delta": format_percent(latest_plan_row["revenue_execution_pct"]) if not is_missing(latest_plan_row["revenue_execution_pct"]) else "",
                    },
                    {
                        "label": "Отклонение количества",
                        "value": format_number(latest_plan_row["quantity_gap"]) if not is_missing(latest_plan_row["quantity_gap"]) else "н/д",
                        "delta": format_percent(latest_plan_row["quantity_execution_pct"]) if not is_missing(latest_plan_row["quantity_execution_pct"]) else "",
                    },
                ]
                if margin_visible:
                    latest_gap_cards.insert(
                        1,
                        {
                            "label": "Отклонение маржи",
                            "value": format_money(latest_plan_row["margin_gap"]) if not is_missing(latest_plan_row["margin_gap"]) else "н/д",
                            "delta": format_percent(latest_plan_row["margin_execution_pct"]) if not is_missing(latest_plan_row["margin_execution_pct"]) else "",
                        },
                    )
                render_metric_cards(latest_gap_cards)

        if can_manage_plans(current_user):
            with st.container(border=True):
                render_panel_header(
                    "Редактирование планов",
                    "Здесь можно задать цель для всей сети или отдельного салона. Пустое поле означает, что план по этой метрике не задан, а не равен нулю.",
                )

                existing_month_labels = set(monthly_plans["plan_month"].dt.strftime("%Y-%m").dropna().tolist()) if not monthly_plans.empty else set()
                existing_month_labels.update(plan_month_options)
                if plan_month_options:
                    latest_editor_month = pd.to_datetime(f"{plan_month_options[-1]}-01", errors="coerce")
                else:
                    latest_editor_month = pd.Timestamp(date.today().replace(day=1))
                for offset in range(4):
                    existing_month_labels.add((latest_editor_month + pd.DateOffset(months=offset)).strftime("%Y-%m"))
                editor_month_options = sorted(existing_month_labels)

                scope_options = ["Вся сеть", *registered_salons] if is_network_role(current_user["role"]) else [current_user.get("salon", "Текущий салон")]
                plan_form_left, plan_form_right = st.columns([1, 1.2], gap="medium")
                with plan_form_left:
                    selected_plan_scope_label = st.selectbox(
                        "Контур плана",
                        options=scope_options,
                        key="plan_editor_scope",
                    )
                    selected_plan_month = st.selectbox(
                        "Месяц плана",
                        options=editor_month_options,
                        index=max(len(editor_month_options) - 1, 0),
                        key="plan_editor_month",
                    )

                selected_plan_scope = "" if selected_plan_scope_label == "Вся сеть" else selected_plan_scope_label
                existing_plan_record = get_plan_record(monthly_plans, selected_plan_month, selected_plan_scope)
                current_month_fact_row = plan_monthly_summary.loc[plan_monthly_summary["month_label"] == selected_plan_month]
                fact_hint = current_month_fact_row.iloc[-1].to_dict() if not current_month_fact_row.empty else {}

                def _plan_prefill(field_name: str) -> str:
                    if not existing_plan_record:
                        return ""
                    value = existing_plan_record.get(field_name)
                    return "" if is_missing(value) else str(int(value) if float(value).is_integer() else round(float(value), 2))

                with plan_form_right:
                    revenue_plan_text = st.text_input(
                        "План по выручке",
                        value=_plan_prefill("revenue_plan"),
                        key=f"plan_revenue_{selected_plan_scope}_{selected_plan_month}",
                        placeholder="Например: 1500000",
                    )
                    margin_plan_text = st.text_input(
                        "План по марже",
                        value=_plan_prefill("margin_plan"),
                        key=f"plan_margin_{selected_plan_scope}_{selected_plan_month}",
                        placeholder="Например: 320000",
                    )
                    quantity_plan_text = st.text_input(
                        "План по количеству",
                        value=_plan_prefill("quantity_plan"),
                        key=f"plan_quantity_{selected_plan_scope}_{selected_plan_month}",
                        placeholder="Например: 1800",
                    )

                if fact_hint:
                    st.caption(
                        f"Факт за {selected_plan_month} в текущем контуре: "
                        f"выручка {format_money(fact_hint.get('revenue'))}, "
                        f"маржа {format_money(fact_hint.get('margin'))}, "
                        f"количество {format_number(fact_hint.get('quantity'))}."
                    )

                save_plan_col, delete_plan_col = st.columns([1, 1], gap="medium")
                with save_plan_col:
                    if st.button("Сохранить план", key="plan_save_button", use_container_width=True):
                        try:
                            revenue_plan_value = parse_plan_input(revenue_plan_text)
                            margin_plan_value = parse_plan_input(margin_plan_text)
                            quantity_plan_value = parse_plan_input(quantity_plan_text)
                            record = upsert_monthly_plan(
                                plan_month=normalize_plan_month(f"{selected_plan_month}-01"),
                                salon=selected_plan_scope,
                                revenue_plan=revenue_plan_value,
                                margin_plan=margin_plan_value,
                                quantity_plan=quantity_plan_value,
                                updated_by=current_user["username"],
                            )
                            audit_event(
                                action="plan.upsert",
                                user_id=current_user["username"],
                                details={
                                    "scope": selected_plan_scope,
                                    "scope_label": selected_plan_scope_label,
                                    "plan_month": selected_plan_month,
                                    "revenue_plan": record.get("revenue_plan"),
                                    "margin_plan": record.get("margin_plan"),
                                    "quantity_plan": record.get("quantity_plan"),
                                },
                            )
                            st.session_state["plan_flash_message"] = f"План для «{selected_plan_scope_label}» на {selected_plan_month} сохранён."
                            st.rerun()
                        except Exception as error:
                            st.error(str(error))
                with delete_plan_col:
                    if st.button(
                        "Удалить план",
                        key="plan_delete_button",
                        use_container_width=True,
                        type="secondary",
                        disabled=existing_plan_record is None,
                    ):
                        deleted = delete_monthly_plan(
                            plan_month=normalize_plan_month(f"{selected_plan_month}-01"),
                            salon=selected_plan_scope,
                        )
                        if deleted:
                            audit_event(
                                action="plan.delete",
                                user_id=current_user["username"],
                                details={
                                    "scope": selected_plan_scope,
                                    "scope_label": selected_plan_scope_label,
                                    "plan_month": selected_plan_month,
                                },
                            )
                            st.session_state["plan_flash_message"] = f"План для «{selected_plan_scope_label}» на {selected_plan_month} удалён."
                            st.rerun()
                        st.warning("Для выбранного месяца и контура сохранённого плана не было.")

        plan_flash_message = st.session_state.pop("plan_flash_message", "")
        if plan_flash_message:
            st.success(plan_flash_message)

        with st.container(border=True):
            render_panel_header(
                "Таблица план / факт по месяцам",
                "Полная управленческая таблица по месяцам: план, факт, отклонение и процент выполнения. Подходит для ежемесячных разборов и выгрузки.",
            )
            plan_table = plan_fact_summary[
                [
                    "month_label",
                    "revenue_plan",
                    "revenue",
                    "revenue_gap",
                    "revenue_execution_pct",
                    "margin_plan",
                    "margin",
                    "margin_gap",
                    "margin_execution_pct",
                    "quantity_plan",
                    "quantity",
                    "quantity_gap",
                    "quantity_execution_pct",
                ]
            ].copy()
            st.dataframe(
                format_display_frame_for_role(
                    plan_table,
                    current_user,
                    rename_map={
                        "month_label": "Месяц",
                        "revenue_plan": "План выручки",
                        "revenue": "Факт выручки",
                        "revenue_gap": "Отклонение выручки",
                        "revenue_execution_pct": "Выполнение выручки, %",
                        "margin_plan": "План маржи",
                        "margin": "Факт маржи",
                        "margin_gap": "Отклонение маржи",
                        "margin_execution_pct": "Выполнение маржи, %",
                        "quantity_plan": "План количества",
                        "quantity": "Факт количества",
                        "quantity_gap": "Отклонение количества",
                        "quantity_execution_pct": "Выполнение количества, %",
                    },
                ),
                use_container_width=True,
                hide_index=True,
                height=480,
            )
            st.download_button(
                "Скачать план / факт",
                data=to_excel_report_bytes(
                    {
                        "План факт": prepare_excel_report_frame(
                            plan_table,
                            current_user,
                            {
                                "month_label": "Месяц",
                                "revenue_plan": "План выручки",
                                "revenue": "Факт выручки",
                                "revenue_gap": "Отклонение выручки",
                                "revenue_execution_pct": "Выполнение выручки, %",
                                "margin_plan": "План маржи",
                                "margin": "Факт маржи",
                                "margin_gap": "Отклонение маржи",
                                "margin_execution_pct": "Выполнение маржи, %",
                                "quantity_plan": "План количества",
                                "quantity": "Факт количества",
                                "quantity_gap": "Отклонение количества",
                                "quantity_execution_pct": "Выполнение количества, %",
                            },
                        )
                    },
                    report_title="План / факт",
                ),
                file_name="plan_fact_summary.xlsx",
                mime=EXCEL_MIME,
            )

        if is_network_role(current_user["role"]) and len(plan_scope_salons) > 1 and plan_month_options:
            selected_plan_month_label = st.selectbox(
                "Месяц для контроля выполнения по салонам",
                options=plan_month_options,
                index=max(len(plan_month_options) - 1, 0),
                key="plan_fact_salon_month",
            )
            salon_plan_records = build_scope_plan_records(monthly_plans, selected_plan_month_label, plan_scope_salons)
            salon_plan_fact = build_plan_fact_by_salon(plan_fact_source_data, salon_plan_records, selected_plan_month_label)

            with st.container(border=True):
                render_panel_header(
                    "Выполнение плана по салонам",
                    "Показывает, какие точки закрывают свой план, а какие уже отстают в выбранном месяце. Это рабочая таблица для еженедельного контроля сети.",
                )
                if salon_plan_records.empty:
                    st.info("Для выбранного месяца пока нет салонных планов. Сначала задайте план хотя бы для одного салона.")
                else:
                    st.dataframe(
                        format_display_frame_for_role(
                            salon_plan_fact,
                            current_user,
                            rename_map={
                                "scope_name": "Салон",
                                "revenue_plan": "План выручки",
                                "revenue": "Факт выручки",
                                "revenue_gap": "Отклонение выручки",
                                "revenue_execution_pct": "Выполнение выручки, %",
                                "margin_plan": "План маржи",
                                "margin": "Факт маржи",
                                "margin_gap": "Отклонение маржи",
                                "margin_execution_pct": "Выполнение маржи, %",
                                "quantity_plan": "План количества",
                                "quantity": "Факт количества",
                                "quantity_gap": "Отклонение количества",
                                "quantity_execution_pct": "Выполнение количества, %",
                            },
                        ),
                        use_container_width=True,
                        hide_index=True,
                        height=420,
                    )

if active_screen == "Аналитика" and active_analytics_screen == "Расширенный":
    main_col = st.container()
    control_col = main_col

    with main_col:
        render_section_intro(
            "Расширенный анализ",
            "Дополнительные блоки для более глубокого разбора продаж: сегментация по активности, поиск ритма продаж по дням и простой прогноз.",
        )

    if active_advanced_screen == "RFM-анализ":
        with main_col:
            with st.container(border=True):
                render_panel_header(
                    "RFM-анализ",
                    "Показывает, кто продает чаще, кто давно не активен и кто приносит наибольшую выручку.",
                )
                rfm_group = st.radio(
                    "Группировать по",
                    options=["manager", "category"],
                    format_func={"manager": "Менеджерам", "category": "Категориям"}.get,
                    horizontal=True,
                    key="rfm_group_radio",
                )

        rfm_data = build_rfm_summary(data, group_column=rfm_group)

        if rfm_data.empty:
            with main_col:
                st.info("Для RFM-анализа нужны данные по менеджерам или категориям.")
        else:
            rfm_data["monetary_bubble"] = build_safe_marker_size(rfm_data["monetary"], absolute=False)

            with main_col:
                rfm_left, rfm_right = st.columns(2, gap="medium")

                with rfm_left:
                    with st.container(border=True):
                        render_panel_header(
                            "Карта RFM-сегментов",
                            "Точки показывают активность и вклад в выручку.",
                        )
                        rfm_scatter = px.scatter(
                            rfm_data,
                            x="recency",
                            y="frequency",
                            size="monetary_bubble",
                            color="segment",
                            hover_name="group_name",
                            color_discrete_map={
                                "Лидер": PRIMARY_COLOR,
                                "Активный": CHART_ACCENT_COLOR,
                                "Стабильный": CHART_INFO_COLOR,
                                "Слабый": CHART_DANGER_COLOR,
                            },
                            labels={
                                "recency": "Давность (дней)",
                                "frequency": "Активность (месяцев)",
                                "monetary_bubble": "Выручка",
                                "segment": "Сегмент",
                            },
                            title="",
                            size_max=60,
                        )
                        rfm_scatter.update_xaxes(autorange="reversed")
                        polish_figure(rfm_scatter, height=460)
                        st.plotly_chart(rfm_scatter, use_container_width=True)

                with rfm_right:
                    with st.container(border=True):
                        render_panel_header(
                            "Распределение по сегментам",
                            "Доля каждой группы в общем количестве.",
                        )
                        seg_counts = rfm_data["segment"].value_counts().reset_index()
                        seg_counts.columns = ["segment", "count"]
                        seg_pie = px.pie(
                            seg_counts,
                            names="segment",
                            values="count",
                            color="segment",
                            color_discrete_map={
                                "Лидер": PRIMARY_COLOR,
                                "Активный": CHART_ACCENT_COLOR,
                                "Стабильный": CHART_INFO_COLOR,
                                "Слабый": CHART_DANGER_COLOR,
                            },
                            title="",
                        )
                        polish_figure(seg_pie, height=340)
                        st.plotly_chart(seg_pie, use_container_width=True)

                rfm_display = rfm_data[["group_name", "recency", "frequency", "monetary", "rfm_score", "segment"]].copy()
                rfm_display.columns = ["Позиция", "Давность (дней)", "Активность (мес.)", "Выручка", "RFM-балл", "Сегмент"]
                with st.container(border=True):
                    render_panel_header(
                        "Таблица RFM",
                        "Подробный список сегментов.",
                    )
                    st.dataframe(rfm_display, use_container_width=True, hide_index=True, height=420)

    if active_advanced_screen == "Тепловая карта":
        heatmap_df = build_heatmap_data(data)
        if heatmap_df.empty:
            with main_col:
                st.info("Нет данных для тепловой карты.")
        else:
            dow_labels = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
            all_weeks = sorted(heatmap_df["week_label"].unique())
            all_dows = list(range(7))

            pivot = heatmap_df.pivot(index="dow", columns="week_label", values="revenue").reindex(
                index=all_dows, columns=all_weeks
            ).fillna(0)

            with main_col:
                with st.container(border=True):
                    render_panel_header(
                        "Тепловая карта продаж",
                        "Ритм продаж по дням недели и неделям периода.",
                    )
                    heatmap_fig = go.Figure(
                        go.Heatmap(
                            z=pivot.values,
                            x=all_weeks,
                            y=dow_labels,
                            colorscale=[[0, "#F0FDF4"], [0.5, PRIMARY_COLOR], [1, "#022C22"]],
                            hoverongaps=False,
                            hovertemplate="Неделя: %{x}<br>День: %{y}<br>Выручка: %{z:,.0f}<extra></extra>",
                        )
                    )
                    heatmap_fig.update_layout(
                        title="",
                        xaxis=dict(showticklabels=len(all_weeks) <= 52),
                        yaxis=dict(title=""),
                    )
                    polish_figure(heatmap_fig, height=320)
                    st.plotly_chart(heatmap_fig, use_container_width=True)
                    st.caption("Чем темнее ячейка, тем выше выручка в этот день и неделю.")

    if active_advanced_screen == "Прогноз":
        forecast_data = build_forecast(monthly_summary)
        if forecast_data.empty:
            with main_col:
                st.info("Для прогноза нужно минимум 3 месяца данных.")
        else:
            combined = pd.concat(
                [
                    monthly_summary[["month_label", "revenue", "margin"]].assign(is_forecast=False),
                    forecast_data[["month_label", "revenue", "margin"]].assign(is_forecast=True),
                ],
                ignore_index=True,
            )
            with main_col:
                with st.container(border=True):
                    render_panel_header(
                        "Прогноз продаж",
                        "Оценка будущих месяцев на основе текущего тренда.",
                    )
                    forecast_fig = go.Figure()
                    hist = combined[~combined["is_forecast"]]
                    fut = combined[combined["is_forecast"]]

                    forecast_fig.add_trace(
                        go.Scatter(
                            x=hist["month_label"],
                            y=hist["revenue"],
                            name="Факт: выручка",
                            mode="lines+markers",
                            line=dict(color=PRIMARY_COLOR, width=3),
                        )
                    )
                    if margin_visible:
                        forecast_fig.add_trace(
                            go.Scatter(
                                x=hist["month_label"],
                                y=hist["margin"],
                                name="Факт: маржа",
                                mode="lines+markers",
                                line=dict(color=SECONDARY_COLOR, width=2),
                            )
                        )
                    forecast_fig.add_trace(
                        go.Scatter(
                            x=[hist["month_label"].iloc[-1]] + fut["month_label"].tolist(),
                            y=[hist["revenue"].iloc[-1]] + fut["revenue"].tolist(),
                            name="Прогноз: выручка",
                            mode="lines+markers",
                            line=dict(color=PRIMARY_COLOR, width=2, dash="dot"),
                            marker=dict(symbol="diamond", size=10),
                        )
                    )
                    if margin_visible and not fut["margin"].isna().all():
                        forecast_fig.add_trace(
                            go.Scatter(
                                x=[hist["month_label"].iloc[-1]] + fut["month_label"].tolist(),
                                y=[hist["margin"].iloc[-1]] + fut["margin"].tolist(),
                                name="Прогноз: маржа",
                                mode="lines+markers",
                                line=dict(color=SECONDARY_COLOR, width=2, dash="dot"),
                                marker=dict(symbol="diamond", size=10),
                            )
                        )
                    forecast_fig.update_layout(xaxis_title="Месяц", yaxis_title="Сумма", legend_title="")
                    polish_figure(forecast_fig, height=480)
                    st.plotly_chart(forecast_fig, use_container_width=True)

                    st.info(
                        f"Прогноз построен на основе линейного тренда по {len(monthly_summary)} месяцам."
                    )

                forecast_table = fut[["month_label", "revenue", "margin"]].copy()
                forecast_table.columns = ["Месяц (прогноз)", "Выручка", "Маржа"]
                forecast_table["Выручка"] = forecast_table["Выручка"].apply(format_money)
                forecast_table["Маржа"] = forecast_table["Маржа"].apply(lambda v: format_money(v) if not pd.isna(v) else "н/д")
                if not margin_visible and "Маржа" in forecast_table.columns:
                    forecast_table = forecast_table.drop(columns=["Маржа"])

                with st.container(border=True):
                    render_panel_header(
                        "Таблица прогноза",
                        "Помесячная расшифровка прогнозных значений.",
                    )
                    st.dataframe(forecast_table, use_container_width=True, hide_index=True)

    if active_advanced_screen == "Возвраты":
        return_product_summary = build_return_groups(data, product_analysis_column)
        return_monthly_summary = build_return_monthly_summary(data)
        return_salon_summary = build_return_groups(data, "salon") if "salon" in data.columns else pd.DataFrame()

        with main_col:
            with st.container(border=True):
                render_panel_header(
                    "Возвраты",
                    "Контроль возвратов и их влияния на выручку.",
                )
                if returns_overview["return_lines"] <= 0:
                    st.success("В текущем срезе возвраты не обнаружены.")
                else:
                    render_metric_cards(
                        [
                            {
                                "label": "Сумма возвратов",
                                "value": format_money(returns_overview["return_revenue"]),
                                "delta": "",
                            },
                            {
                                "label": "Доля возвратов",
                                "value": format_percent(returns_overview["return_share_pct"]),
                                "delta": "От валовой выручки",
                            },
                            {
                                "label": "Возвратных строк",
                                "value": format_number(returns_overview["return_lines"]),
                                "delta": "",
                            },
                            {
                                "label": "Товаров с возвратами",
                                "value": format_number(returns_overview["return_product_count"]),
                                "delta": "",
                            },
                        ]
                    )

                    returns_left, returns_right = st.columns(2, gap="medium")

                    with returns_left:
                        with st.container(border=True):
                            render_panel_header(
                                "Динамика возвратов",
                                "Изменение суммы и доли возвратов по месяцам.",
                            )
                            monthly_returns_chart = go.Figure()
                            monthly_returns_chart.add_trace(
                                go.Bar(
                                    x=return_monthly_summary["month_label"],
                                    y=return_monthly_summary["return_revenue"],
                                    name="Сумма возвратов",
                                    marker_color=CHART_DANGER_COLOR,
                                )
                            )
                            monthly_returns_chart.add_trace(
                                go.Scatter(
                                    x=return_monthly_summary["month_label"],
                                    y=return_monthly_summary["return_share_pct"],
                                    name="Доля возвратов, %",
                                    mode="lines+markers",
                                    line=dict(color=SECONDARY_COLOR, width=3),
                                    yaxis="y2",
                                )
                            )
                            monthly_returns_chart.update_layout(
                                xaxis_title="Месяц",
                                yaxis_title="Сумма возвратов",
                                yaxis2=dict(title="Доля возвратов, %", overlaying="y", side="right"),
                            )
                            polish_figure(monthly_returns_chart, height=420)
                            st.plotly_chart(monthly_returns_chart, use_container_width=True)

                    with returns_right:
                        with st.container(border=True):
                            render_panel_header(
                                "Товары с возвратами",
                                "Лидеры по сумме возвратов.",
                            )
                            top_return_items = return_product_summary.head(12).sort_values("return_revenue", ascending=True)
                            return_chart = px.bar(
                                top_return_items,
                                x="return_revenue",
                                y="group_name",
                                orientation="h",
                                color="return_share_pct",
                                color_continuous_scale=[CHART_SOFT_AMBER, CHART_ACCENT_COLOR, CHART_DANGER_COLOR],
                                labels={"group_name": "Товар", "return_revenue": "Сумма возвратов", "return_share_pct": "Доля возвратов, %"},
                                title="",
                            )
                            return_chart.update_layout(coloraxis_showscale=False, yaxis_title="")
                            polish_figure(return_chart, height=420)
                            st.plotly_chart(return_chart, use_container_width=True)

                    with st.container(border=True):
                        render_panel_header(
                            "Таблица возвратов по товарам",
                            "Подробная расшифровка по позициям.",
                        )
                        return_product_table = return_product_summary.copy()
                        if not return_product_table.empty:
                            return_product_table["last_return_date"] = pd.to_datetime(
                                return_product_table["last_return_date"], errors="coerce"
                            )
                        st.dataframe(
                            format_display_frame_for_role(
                                return_product_table,
                                current_user,
                                rename_map={
                                    "group_name": "Товар",
                                    "return_revenue": "Сумма возвратов",
                                    "return_margin": "Возврат по марже",
                                    "return_quantity": "Возврат по количеству",
                                    "return_lines": "Строк возврата",
                                    "return_share_pct": "Доля возвратов, %",
                                    "last_return_date": "Последний возврат",
                                },
                            ),
                            use_container_width=True,
                            hide_index=True,
                            height=460,
                        )
                        st.download_button(
                            "Скачать возвраты по товарам",
                            data=to_excel_report_bytes(
                                {
                                    "Возвраты": prepare_excel_report_frame(
                                        return_product_summary,
                                        current_user,
                                        {
                                            "group_name": "Товар",
                                            "return_revenue": "Сумма возвратов",
                                            "return_margin": "Возврат по марже",
                                            "return_quantity": "Возврат по количеству",
                                            "return_lines": "Строк возврата",
                                            "return_share_pct": "Доля возвратов, %",
                                            "last_return_date": "Последний возврат",
                                        },
                                    )
                                },
                                report_title="Возвраты по товарам",
                            ),
                            file_name="returns_by_product.xlsx",
                            mime=EXCEL_MIME,
                            key="dl_returns_csv"
                        )

                    if is_network_role(current_user["role"]) and not return_salon_summary.empty and len(return_salon_summary) > 1:
                        with st.container(border=True):
                            render_panel_header(
                                "Возвраты по салонам",
                                "Сравнение торговых точек по уровню возвратов.",
                            )
                            st.dataframe(
                                format_display_frame_for_role(
                                    return_salon_summary,
                                    current_user,
                                    rename_map={
                                        "group_name": "Салон",
                                        "return_revenue": "Сумма возвратов",
                                        "return_margin": "Возврат по марже",
                                        "return_quantity": "Возврат по количеству",
                                        "return_lines": "Строк возврата",
                                        "return_share_pct": "Доля возвратов, %",
                                        "last_return_date": "Последний возврат",
                                    },
                                ),
                                use_container_width=True,
                                hide_index=True,
                                height=360,
                            )

if active_screen == "Данные":
    visible_mapping = margin_safe_mapping(selected_mapping, current_user)
    main_col = st.container()
    side_control_col = main_col
    with main_col:
        render_section_intro(
            "Источники и выгрузки",
            "Здесь собраны исходные данные, служебная информация по сопоставлению колонок, журнал загрузок и сводка по месяцам. Эта вкладка нужна для проверки качества загрузки и состава данных.",
        )
        sample_path = Path("sample_sales_data.csv")
        if sample_path.exists():
            st.download_button(
                "Скачать безопасный шаблон CSV",
                data=sample_path.read_bytes(),
                file_name=sample_path.name,
                mime="text/csv",
                key="download_sales_data_template",
            )
        data_period_text = format_date_range_values(data["date"].min(), data["date"].max())
        archive_upload_count = len(manifest_view) if manifest_view is not None else 0
        data_salon_count = int(data["salon"].nunique()) if "salon" in data.columns else 1
        mapping_count = sum(1 for value in visible_mapping.values() if value) if visible_mapping else 0

        render_section_marker(
            "Паспорт данных",
            "Как читать эту вкладку",
            "Сначала проверьте паспорт набора данных, затем исходный файл и сопоставление колонок. После этого журнал загрузок покажет, что уже сохранено в архиве, а сводка по месяцам подтвердит, что периоды сложились корректно.",
        )
        render_workspace_band(
            [
                {"label": "Источник", "value": source_label, "meta": "Откуда сейчас построен анализ"},
                {"label": "Период данных", "value": data_period_text, "meta": "Фактический диапазон после фильтров"},
                {"label": "Строк в анализе", "value": format_number(overview["line_count"]), "meta": "Сколько операций прошло в текущую выборку"},
                {"label": "Салоны в наборе", "value": format_number(data_salon_count), "meta": "Сколько торговых точек участвует сейчас"},
                {"label": "Месяцев в выборке", "value": format_number(len(monthly_summary)), "meta": "Сколько периодов попало в итоговую аналитику"},
                {"label": "Архивных загрузок", "value": format_number(archive_upload_count), "meta": "Сколько файлов уже сохранено в архиве"},
            ]
        )
        render_journey_cards(
            [
                {
                    "title": "Проверьте, что загружено",
                    "body": "В блоке `Источник данных` теперь показывается только безопасный контекст по загрузке без содержимого файла. Используйте его, чтобы проверить источник, дату и наличие файла в архиве.",
                    "hint": "Исходная выгрузка больше не открывается прямо в интерфейсе.",
                },
                {
                    "title": "Сверьте поля аналитики",
                    "body": "В служебной информации видно, какие колонки файла стали датой, товаром, выручкой, себестоимостью и маржой. Это главный контроль качества распознавания 1С.",
                    "hint": "Особенно важны дата, товар и денежные поля.",
                },
                {
                    "title": "Подтвердите архив и периоды",
                    "body": "Журнал загрузок показывает, что уже сохранено в системе, а сводка по месяцам подтверждает, что временной ряд сложился корректно и без дыр в данных.",
                    "hint": "Эта проверка особенно полезна перед управленческими выводами.",
                },
            ]
        )
        render_snapshot_strip(
            [
                {
                    "label": "Распознано полей",
                    "value": format_number(mapping_count),
                    "hint": "Сколько колонок уже привязано к аналитической модели",
                },
                {
                    "label": "Архив выгрузок",
                    "value": format_number(archive_upload_count),
                    "hint": "Чем больше корректных файлов, тем сильнее анализ истории",
                },
                {
                    "label": "Глубина периода",
                    "value": format_number(len(monthly_summary)),
                    "hint": "Количество месяцев в текущем наборе данных",
                },
            ]
        )

        sku_cleanup_queue = build_sku_cleanup_queue(
            data,
            procurement_items,
            supplier_product_assignments,
            require_item_code=has_item_codes,
        )
        cleanup_issue_types = [
            "Без поставщика",
            "Без категории",
            "Без кода товара",
            "Нет в остатках",
            "Нулевой остаток",
        ]
        cleanup_counts = {
            issue: int(sku_cleanup_queue["issues"].fillna("").astype(str).str.contains(issue, regex=False).sum())
            if not sku_cleanup_queue.empty
            else 0
            for issue in cleanup_issue_types
        }

        with st.container(border=True):
            render_panel_header(
                "Очередь очистки SKU",
                "Автоматически собирает номенклатуру, где не хватает поставщика, категории, кода товара или связки с остатками. "
                "Это первый шаг к единому справочнику номенклатуры без ручного поиска проблем.",
            )
            render_snapshot_strip(
                [
                    {
                        "label": "SKU к разбору",
                        "value": format_number(len(sku_cleanup_queue)),
                        "hint": "позиции с хотя бы одной проблемой",
                        "tone": "warning" if not sku_cleanup_queue.empty else "success",
                    },
                    {
                        "label": "Без поставщика",
                        "value": format_number(cleanup_counts["Без поставщика"]),
                        "hint": "мешает заказам поставщикам",
                        "tone": "warning" if cleanup_counts["Без поставщика"] else "success",
                    },
                    {
                        "label": "Нет в остатках",
                        "value": format_number(cleanup_counts["Нет в остатках"]),
                        "hint": "не участвует в точном прогнозе закупки",
                        "tone": "warning" if cleanup_counts["Нет в остатках"] else "success",
                    },
                ]
            )
            if sku_cleanup_queue.empty:
                st.success("В текущем срезе не найдено проблемных SKU. Справочник выглядит аккуратно.")
            else:
                cleanup_filter = st.multiselect(
                    "Какие проблемы показать",
                    cleanup_issue_types,
                    default=[issue for issue in cleanup_issue_types if cleanup_counts[issue] > 0],
                    key="sku_cleanup_issue_filter",
                )
                cleanup_view = sku_cleanup_queue.copy()
                if cleanup_filter:
                    issue_mask = pd.Series(False, index=cleanup_view.index)
                    issue_text = cleanup_view["issues"].fillna("").astype(str)
                    for issue in cleanup_filter:
                        issue_mask = issue_mask | issue_text.str.contains(issue, regex=False)
                    cleanup_view = cleanup_view[issue_mask].copy()

                cleanup_rename_map = {
                    "issue_count": "Проблем",
                    "issues": "Что исправить",
                    "suggested_action": "Рекомендованное действие",
                    "product_key": "Ключ SKU",
                    "product": "Номенклатура",
                    "item_code": "Код товара",
                    "category": "Категория",
                    "effective_supplier": "Поставщик",
                    "sales_supplier": "Поставщик из продаж",
                    "manual_supplier": "Ручной поставщик",
                    "stock_supplier": "Поставщик из остатков",
                    "revenue": "Выручка",
                    "quantity": "Количество",
                    "line_count": "Строк продаж",
                    "stock_on_hand": "Остаток",
                    "stock_in_transit": "В пути",
                }
                cleanup_display_columns = [
                    "issue_count",
                    "issues",
                    "suggested_action",
                    "product",
                    "item_code",
                    "category",
                    "effective_supplier",
                    "revenue",
                    "quantity",
                    "line_count",
                    "stock_on_hand",
                    "stock_in_transit",
                ]
                st.caption(
                    f"Показано {format_number(len(cleanup_view.head(300)))} из "
                    f"{format_number(len(cleanup_view))} SKU. Полный список можно скачать в Excel."
                )
                st.dataframe(
                    format_display_frame_for_role(
                        cleanup_view.head(300),
                        current_user,
                        rename_map=cleanup_rename_map,
                        columns=cleanup_display_columns,
                    ),
                    use_container_width=True,
                    hide_index=True,
                    height=420,
                )
                st.download_button(
                    "Скачать очередь очистки SKU",
                    data=to_excel_report_bytes(
                        {
                            "Очередь SKU": prepare_excel_report_frame(
                                cleanup_view,
                                current_user,
                                cleanup_rename_map,
                                columns=[
                                    "issue_count",
                                    "issues",
                                    "suggested_action",
                                    "product_key",
                                    "product",
                                    "item_code",
                                    "category",
                                    "effective_supplier",
                                    "sales_supplier",
                                    "manual_supplier",
                                    "stock_supplier",
                                    "revenue",
                                    "quantity",
                                    "line_count",
                                    "stock_on_hand",
                                    "stock_in_transit",
                                ],
                            )
                        },
                        report_title="Очередь очистки SKU",
                    ),
                    file_name="sku_cleanup_queue.xlsx",
                    mime=EXCEL_MIME,
                    use_container_width=True,
                )

        with st.container(border=True):
            render_panel_header(
                "Источник данных",
                "Просмотр исходной выгрузки отключён. В этом блоке оставлен только безопасный контекст по источнику, чтобы сотрудники не скачивали и не открывали загруженный файл прямо из системы.",
            )
            st.info("Предпросмотр исходного файла и скачивание загруженной выгрузки отключены для всех ролей.")
            if not manifest_view.empty:
                st.dataframe(
                    format_display_frame(
                        manifest_view,
                        columns=["salon", "report_date", "source_filename", "uploaded_at"],
                    ),
                    use_container_width=True,
                    hide_index=True,
                    height=320,
                )
            else:
                st.caption("Когда в архиве появятся сохранённые загрузки, здесь будет видна их краткая история без содержимого файла.")

        with st.container(border=True):
            render_panel_header(
                "Как файл распознан",
                "Показывает, как именно колонки исходного файла были привязаны к полям аналитики. Для салона скрыты поля себестоимости и маржи, но сами расчёты для руководителя не теряются.",
            )
            if visible_mapping:
                st.caption(f"Распознано и привязано полей: {mapping_count}. Если что-то не совпало, ищите причину сначала здесь.")
                st.dataframe(
                    build_download_frame(visible_mapping),
                    use_container_width=True,
                    hide_index=True,
                    height=420,
                )
            else:
                st.info("Для архивного режима сопоставление колонок хранится у каждой загруженной выгрузки отдельно.")

        if not manifest_view.empty:
            with st.container(border=True):
                render_panel_header(
                    "Журнал загрузок",
                    "История всех сохранённых файлов по салонам и датам. Журнал нужен для контроля архива: какой файл уже лежит в системе, когда он был добавлен и какой салон он пополняет.",
                )
                manifest_table = manifest_view.copy()
                visible_columns = [column for column in ["salon", "report_date", "source_filename", "uploaded_at"] if column in manifest_table.columns]
                if visible_columns:
                    st.dataframe(
                        format_display_frame(manifest_table, columns=visible_columns),
                        use_container_width=True,
                        hide_index=True,
                        height=320,
                    )

        with st.container(border=True):
            render_panel_header(
                "Сводка по месяцам",
                "Главная проверочная таблица по периодам: выручка, себестоимость, маржа, количество и динамика. Если хотите понять, корректно ли сложились месяцы и откуда берётся тренд, сначала смотрите именно сюда.",
            )
            st.dataframe(
                format_display_frame_for_role(
                    monthly_summary,
                    current_user,
                    {
                        "month": "Месяц",
                        "month_label": "Код месяца",
                        "revenue": "Выручка",
                        "cost": "Себестоимость",
                        "margin": "Маржа",
                        "quantity": "Количество",
                        "product_count": "Товаров",
                        "revenue_change_pct": "Изменение выручки, %",
                        "margin_change_pct": "Изменение маржи, %",
                    },
                    columns=[
                        "month",
                        "revenue",
                        "cost",
                        "margin",
                        "quantity",
                        "product_count",
                        "revenue_change_pct",
                        "margin_change_pct",
                    ],
                ),
                use_container_width=True,
                hide_index=True,
                height=360,
            )
            st.download_button(
                "Скачать сводку по месяцам",
                data=to_excel_report_bytes(
                    {
                        "Сводка по месяцам": prepare_excel_report_frame(
                            monthly_summary,
                            current_user,
                            {
                                "month": "Месяц",
                                "month_label": "Код месяца",
                                "revenue": "Выручка",
                                "cost": "Себестоимость",
                                "margin": "Маржа",
                                "quantity": "Количество",
                                "product_count": "Товаров",
                                "revenue_change_pct": "Изменение выручки, %",
                                "margin_change_pct": "Изменение маржи, %",
                            },
                            columns=[
                                "month",
                                "revenue",
                                "cost",
                                "margin",
                                "quantity",
                                "product_count",
                                "revenue_change_pct",
                                "margin_change_pct",
                            ],
                        )
                    },
                    report_title="Сводка по месяцам",
                ),
                file_name="monthly_summary.xlsx",
                mime=EXCEL_MIME,
                use_container_width=True,
            )

    with side_control_col:
        st.markdown('<div class="nav-shell">', unsafe_allow_html=True)
        st.markdown('<div class="nav-title">Действия с данными</div>', unsafe_allow_html=True)
        st.info("Используйте эту панель для быстрой навигации или действий с архивом.")
        if st.button("Обновить данные", key="data_refresh_button", use_container_width=True):
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

if active_screen == "Управление" and can_manage_access(current_user):
    render_admin_tab(current_user, registered_salons)
